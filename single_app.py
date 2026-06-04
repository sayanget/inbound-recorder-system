# Copyright (c) 2026 Fan Yang. All rights reserved.
# 商业/营利性使用须事先书面许可；未经授权构成侵权，权利人保留依法主张全部救济之权利。
import sqlite3
from flask import Flask, request, jsonify, send_file, send_from_directory, session, redirect, Response, render_template
import os
import sys
import tempfile
import atexit

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))
except ImportError:
    pass
import csv
import io
import unicodedata
import re

# Fix Unicode output on Windows
if sys.platform == 'win32':
    try:
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
    except:
        pass
from datetime import datetime, timedelta, date, time as dtime
import pytz
import threading
import time
import schedule
from openpyxl import Workbook
import json
from queue import Queue
import calc_outsource_finance # 导入生产人工同步逻辑
import requests
import feishu_auth

try:
    import gofo_dms_auth as _gofo_dms_auth
except ImportError:
    _gofo_dms_auth = None

try:
    import psycopg2
    from psycopg2.extras import DictCursor
except ImportError:
    psycopg2 = None
    DictCursor = None
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json
import hashlib
import functools
import secrets

_cached_initial_admin_pw = None


def _initial_admin_password():
    """First-time admin password: INITIAL_ADMIN_PASSWORD env, or one random value (cached, logged once)."""
    global _cached_initial_admin_pw
    if _cached_initial_admin_pw is not None:
        return _cached_initial_admin_pw
    env_pw = os.environ.get("INITIAL_ADMIN_PASSWORD")
    if env_pw:
        _cached_initial_admin_pw = env_pw
        return env_pw
    generated = secrets.token_urlsafe(16)
    print(
        "[SECURITY] INITIAL_ADMIN_PASSWORD not set. "
        f"Generated default admin password (change immediately): {generated}"
    )
    _cached_initial_admin_pw = generated
    return generated


# 数据库抽象层 - 自动适配 SQLite/PostgreSQL
from database import get_db_connection, convert_sql, get_placeholder, USE_POSTGRES
from app_identity import identity_dict

app = Flask(__name__)
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_PERMANENT'] = True
app.config['SESSION_USE_SIGNER'] = True
app.config['SESSION_KEY_PREFIX'] = 'inbound_app:'


def _apply_trusted_proxy_headers():
    """在 Render 等反向代理后启用，使 request.remote_addr / X-Forwarded-For 可信。"""
    if (os.environ.get("TRUST_PROXY_HEADERS", "") or "").strip().lower() not in (
        "1",
        "true",
        "yes",
        "on",
    ):
        return
    try:
        from werkzeug.middleware.proxy_fix import ProxyFix

        app.wsgi_app = ProxyFix(
            app.wsgi_app,
            x_for=1,
            x_proto=1,
            x_host=1,
            x_port=1,
            x_prefix=1,
        )
    except Exception as e:
        print(f"[TRUST_PROXY_HEADERS] ProxyFix 未生效: {e}", flush=True)


_apply_trusted_proxy_headers()

try:
    from license_client import bootstrap_license_server_at_startup

    bootstrap_license_server_at_startup()
except Exception as _lic_boot_err:
    print(f"[LICENSE] startup bootstrap skipped: {_lic_boot_err}", flush=True)


def _license_request_exempt() -> bool:
    """不拦截的路径（健康检查、静态、登录、首次激活等）。"""
    p = request.path or ""
    if p in (
        "/ping",
        "/health",
        "/api/license_status",
        "/login",
        "/api/login",
        "/api/logout",
        "/api/check_login",
    ):
        return True
    if p.startswith("/api/admin/license/"):
        return True
    if p.startswith("/static"):
        return True
    if request.method == "OPTIONS" and p.startswith("/api/"):
        return True
    return False


@app.before_request
def _license_enforce_gate():
    """可选：LICENSE_ENFORCE=1 时向许可服务校验 LICENSE_DEVICE_TOKEN。"""
    if _license_request_exempt():
        return None
    ev = (os.environ.get("LICENSE_ENFORCE") or "").strip().lower()
    if ev not in ("1", "true", "yes", "on"):
        return None
    try:
        from license_client import resolve_device_token, verify_license

        base = (os.environ.get("LICENSE_SERVER_URL") or "").strip()
        tok = resolve_device_token()
    except Exception:
        base = (os.environ.get("LICENSE_SERVER_URL") or "").strip()
        tok = (os.environ.get("LICENSE_DEVICE_TOKEN") or "").strip()
    if not base or not tok:
        msg = "LICENSE_SERVER_URL 或 LICENSE_DEVICE_TOKEN 未配置"
        if request.path.startswith("/api/"):
            return jsonify({"ok": False, "error": "license", "reason": msg}), 503
        return (
            f"<h1>许可证未配置</h1><p>{msg}</p>",
            503,
            {"Content-Type": "text/html; charset=utf-8"},
        )
    try:
        ok, reason = verify_license()
    except Exception as e:
        ok, reason = False, str(e)
    if ok:
        return None
    hint = ""
    try:
        from license_client import license_server_url_effective, probe_license_server

        up, eff, err = probe_license_server()
        eff = eff or license_server_url_effective() or base
        if not up:
            hint = (
                f"<p>许可服务不可达。请运行 <code>run_license_server.bat</code>，"
                f"或用 <code>start_with_monitor.bat</code>（已设 LICENSE_AUTO_START_SERVER 时会自动起 8088）。"
                f"</p><p>配置: <code>{base}</code>"
                f"{(' · 探测: ' + err) if err else ''}</p>"
            )
        elif eff and eff.rstrip('/') != (base or '').rstrip('/'):
            hint = f"<p>实际连到: <code>{eff}</code>（内网 IP 已自动改用本机 127.0.0.1）</p>"
    except Exception:
        pass
    if request.path.startswith("/api/"):
        return jsonify({"ok": False, "error": "license", "reason": reason}), 403
    return (
        "<h1>许可证无效或已过期</h1><p>请检查 LICENSE_DEVICE_TOKEN 与许可服务状态。</p>"
        f"<p>原因: {reason}</p>{hint}",
        403,
        {"Content-Type": "text/html; charset=utf-8"},
    )


@app.before_request
def _telemetry_github_hook():
    """可选：首次请求时后台上报强关联信息到 GitHub Issue（须 TELEMETRY_GITHUB_ENABLE=1）。"""
    try:
        from telemetry_github import schedule_report_on_first_request
        schedule_report_on_first_request(request)
    except Exception as e:
        ev = (os.environ.get("TELEMETRY_GITHUB_ENABLE") or "").strip().lower()
        if ev in ("1", "true", "yes", "on"):
            print(f"[telemetry_github] hook 异常: {e}", flush=True)


@app.before_request
def _api_cors_preflight():
    """跨源打开静态页时，POST 上传等会先发 OPTIONS；必须响应否则浏览器报 Failed to fetch。"""
    if request.method == 'OPTIONS' and request.path.startswith('/api/'):
        r = Response('', status=204)
        r.headers['Access-Control-Allow-Origin'] = '*'
        r.headers['Access-Control-Allow-Methods'] = 'GET, POST, PUT, DELETE, PATCH, OPTIONS'
        r.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Requested-With'
        r.headers['Access-Control-Max-Age'] = '3600'
        return r


@app.after_request
def _api_cors_headers(response):
    if request.path.startswith('/api/'):
        response.headers.setdefault('Access-Control-Allow-Origin', '*')
    return response


@app.after_request
def _permissions_policy_header(response):
    # 显式授权 unload / beforeunload，避免 Chrome 对页面上任何 unload 监听器
    # (包括某些第三方脚本/扩展注入的 index.global.js 等)抛出
    # "[Violation] Permissions policy violation: unload is not allowed in this document."
    # unload=*       允许任何来源注册 unload 监听
    # 用 setdefault 防止覆盖下游代理可能已写的策略
    response.headers.setdefault('Permissions-Policy', 'unload=*')
    return response


@app.before_request
def _session_keep_alive():
    """已登录用户：持久会话 + 每次请求续期 Cookie，减少频繁重新登录。"""
    if session.get('user_id'):
        session.permanent = True
        session.modified = True


# 获取正确的数据库路径
def get_db_path():
    # 如果是打包后的exe环境，数据库在同级目录下
    if getattr(sys, 'frozen', False):
        # 打包后的exe环境
        return os.path.join(os.path.dirname(sys.executable), 'inbound.db')
    else:
        # 开发环境
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'inbound.db')

# 获取正确的静态文件目录
def get_static_dir():
    if getattr(sys, 'frozen', False):
        # 打包后的exe环境 - 静态文件被打包到exe中，需要使用特殊方法访问
        # PyInstaller会将数据文件放在_sys_meipass目录中
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, 'static')
        else:
            # 备用方案
            return os.path.join(os.path.dirname(sys.executable), 'static')
    else:
        # 开发环境 - 使用脚本所在目录的static子目录
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')

# 数据库连接辅助函数 - 兼容 SQLite 和 PostgreSQL
def get_db():
    """获取数据库连接 - 自动适配 SQLite/PostgreSQL"""
    if USE_POSTGRES:
        import psycopg2
        # 使用 DictCursor 替代 RealDictCursor
        # DictCursor 返回的行对象同时支持 索引访问(row[0]) 和 键访问(row['id'])
        # 这完美兼容了我们现有的代码 (混合使用了两种访问方式)
        from psycopg2.extras import DictCursor
        from database import DATABASE_URL
        
        # 创建原始连接
        original_conn = psycopg2.connect(DATABASE_URL, cursor_factory=DictCursor)
        
        # 创建连接包装类
        class ConnectionWrapper:
            def __init__(self, conn):
                self._conn = conn
            
            def cursor(self, *args, **kwargs):
                cur = self._conn.cursor(*args, **kwargs)
                original_execute = cur.execute
                
                def execute_wrapper(query, params=None):
                    # 自动转换 ? 为 %s
                    converted_query = query.replace('?', '%s')
                    return original_execute(converted_query, params)
                
                cur.execute = execute_wrapper
                return cur
            
            def commit(self):
                return self._conn.commit()
            
            def rollback(self):
                return self._conn.rollback()
            
            def close(self):
                return self._conn.close()
            
            def __enter__(self):
                return self
            
            def __exit__(self, exc_type, exc_val, exc_tb):
                if exc_type:
                    self._conn.rollback()
                else:
                    self._conn.commit()
                self._conn.close()
        
        return ConnectionWrapper(original_conn)
    else:
        import sqlite3
        # 增加 timeout 到 30 秒，减少 "database is locked" 错误
        conn = sqlite3.connect(DB_PATH, timeout=30)
        # 使用 Row 工厂,使 SQLite 结果可以像字典一样访问
        conn.row_factory = sqlite3.Row
        return conn

def convert_query_placeholders(query):
    """转换 SQL 查询占位符 - SQLite 的 ? 转为 PostgreSQL 的 %s"""
    if USE_POSTGRES:
        return query.replace('?', '%s')
    return query


def _parse_config_json_from_db(value, defaults):
    """sorting_schedule_config.config_json：SQLite 为 TEXT；PostgreSQL JSON/JSONB 可能已为 dict。"""
    if value is None:
        return defaults
    if isinstance(value, dict):
        return value
    if isinstance(value, str):
        try:
            return json.loads(value)
        except (json.JSONDecodeError, TypeError):
            return defaults
    if isinstance(value, (bytes, bytearray)):
        try:
            return json.loads(value.decode('utf-8'))
        except (json.JSONDecodeError, UnicodeDecodeError, TypeError):
            return defaults
    return defaults


def _db_row_get(row, key, default=None):
    """sqlite3.Row 无 .get()；PostgreSQL DictCursor 行可用下标。统一安全取值。"""
    try:
        v = row[key]
    except (KeyError, IndexError, TypeError):
        return default
    return default if v is None else v


# 入库录入件数 → 实到件数（统计、对比、与分拣对齐等统一口径；装载量 load_amount 不乘系数）
INBOUND_PIECES_ACTUAL_FACTOR = float(os.environ.get("INBOUND_PIECES_ACTUAL_FACTOR", "0.76"))
# CBS/CBT：仅按托盘数 × 本系数核算货量（默认 300 件/托，与 GOFO 托盘 344 系数分离）
INBOUND_CBS_CBT_PIECES_PER_PALLET = int(os.environ.get("INBOUND_CBS_CBT_PIECES_PER_PALLET", "300"))


def _sql_inbound_net_pieces_actual(prefix=""):
    """SQL 片段：GOFO 车型为 (录入件数 - 不计入件数) * 实到系数；CBS/CBT 为录入−排除（不乘系数）。"""
    p = prefix
    br = f"({p}pieces - COALESCE({p}excluded_pieces, 0))"
    scaled = f"({br} * {INBOUND_PIECES_ACTUAL_FACTOR})"
    return (
        f"(CASE WHEN {p}vehicle_type IN ('CBS', 'CBT') THEN CAST({br} AS REAL) "
        f"ELSE {scaled} END)"
    )


def _sql_inbound_sum_case_53g_zero():
    """SUM(CASE 53+G 货量不计 ELSE 实到件数 END)。"""
    inner = _sql_inbound_net_pieces_actual("")
    return (
        f"SUM(CASE WHEN vehicle_type = '53英尺' AND vehicle_no = 'G' "
        f"THEN 0 ELSE ({inner}) END)"
    )


def _py_inbound_actual_pieces(pieces, excluded_pieces=0):
    """单条记录实到件数（与 SQL 口径一致，不含 53+G 规则）。"""
    try:
        p = int(pieces or 0)
    except (TypeError, ValueError):
        p = 0
    try:
        e = int(excluded_pieces or 0)
    except (TypeError, ValueError):
        e = 0
    net = max(0, p - e)
    return float(net * INBOUND_PIECES_ACTUAL_FACTOR)


def _py_inbound_arrival_pieces(vehicle_type, vehicle_no, pieces, excluded_pieces=0):
    """单条展示/汇总用实到件数：53 英尺 + 车牌 G 为 0；CBS/CBT 为录入−排除（不乘系数）；其余 (录入−排除)×系数。"""
    if str(vehicle_type or "").strip() == "53英尺" and str(vehicle_no or "").strip() == "G":
        return 0.0
    if str(vehicle_type or "").strip() in ("CBS", "CBT"):
        try:
            p = int(pieces or 0)
        except (TypeError, ValueError):
            p = 0
        try:
            e = int(excluded_pieces or 0)
        except (TypeError, ValueError):
            e = 0
        return float(max(0, p - e))
    return _py_inbound_actual_pieces(pieces, excluded_pieces)


# 定义洛杉矶时区
LA_TZ = pytz.timezone('America/Los_Angeles')

def round_to_ten_thousand(value):
    """
    将数值向下取整到万位(千百十个位全部为0)
    例如: 1,232,342 -> 1,230,000
         987,654 -> 980,000
         45,678 -> 40,000
    """
    if value is None:
        return 0
    return (int(value) // 10000) * 10000

def _load_session_secret_key():
    """优先环境变量 SECRET_KEY；否则写入项目目录 .flask_session_secret，重启后会话仍有效。"""
    sk = os.environ.get("SECRET_KEY")
    if sk and sk.strip():
        return sk.strip()
    root = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(root, ".flask_session_secret")
    try:
        if os.path.isfile(path):
            with open(path, "r", encoding="utf-8") as f:
                s = f.read().strip()
                if len(s) >= 16:
                    return s
    except OSError:
        pass
    new = secrets.token_hex(32)
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(new)
        if hasattr(os, "chmod"):
            os.chmod(path, 0o600)
        print(f"[session] SECRET_KEY 已写入 {path}，重启后登录状态可保持", flush=True)
    except OSError as e:
        print(f"[session] 无法写入 {path}: {e}，本次使用内存密钥（重启后需重新登录）", flush=True)
    return new


app.secret_key = _load_session_secret_key()
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=30)
app.config['SESSION_COOKIE_SECURE'] = False
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

# 数据库路径配置
DB_PATH = os.environ.get('DATABASE_PATH') or get_db_path()

# 服务器配置
HOST = os.environ.get('HOST', '0.0.0.0')
PORT = int(os.environ.get('PORT', 8080))

# 应用初始化标志
_app_initialized = False
_background_jobs_lock_handle = None


def _bool_env(name: str, default: bool = False) -> bool:
    raw = os.environ.get(name)
    if raw is None:
        return default
    return str(raw).strip().lower() in ("1", "true", "yes", "on")


def _release_background_jobs_lock():
    global _background_jobs_lock_handle
    fh = _background_jobs_lock_handle
    if fh is None:
        return
    try:
        if os.name == "nt":
            import msvcrt  # type: ignore
            msvcrt.locking(fh.fileno(), msvcrt.LK_UNLCK, 1)
        else:
            import fcntl  # type: ignore
            fcntl.flock(fh.fileno(), fcntl.LOCK_UN)
    except Exception:
        pass
    try:
        fh.close()
    except Exception:
        pass
    _background_jobs_lock_handle = None


def _acquire_background_jobs_lock() -> bool:
    """Cross-process lock to ensure only one process starts background jobs."""
    global _background_jobs_lock_handle

    if _background_jobs_lock_handle is not None:
        return True

    # Optional kill switch for environments where schedulers are externalized.
    if _bool_env("ENABLE_BACKGROUND_JOBS", True) is False:
        print("[应用初始化] ENABLE_BACKGROUND_JOBS=0，跳过后台同步线程启动")
        return False

    lock_path = os.environ.get("BACKGROUND_JOBS_LOCKFILE")
    if not lock_path:
        lock_path = os.path.join(tempfile.gettempdir(), "inbound_python_source.background_jobs.lock")

    try:
        fh = open(lock_path, "a+", encoding="utf-8")
        if os.name == "nt":
            import msvcrt  # type: ignore
            fh.seek(0)
            msvcrt.locking(fh.fileno(), msvcrt.LK_NBLCK, 1)
        else:
            import fcntl  # type: ignore
            fcntl.flock(fh.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        fh.seek(0)
        fh.truncate()
        fh.write(f"pid={os.getpid()}\n")
        fh.flush()
        _background_jobs_lock_handle = fh
        atexit.register(_release_background_jobs_lock)
        print(f"[应用初始化] 获取后台任务锁成功: {lock_path} (pid={os.getpid()})")
        return True
    except Exception:
        try:
            fh.close()  # type: ignore[name-defined]
        except Exception:
            pass
        print(f"[应用初始化] 后台任务锁已被其他进程持有，当前进程不启动后台线程 (pid={os.getpid()})")
        return False


def initialize_app():
    """初始化应用 - 确保在 Gunicorn 环境下也能正确初始化"""
    global _app_initialized
    if not _app_initialized:
        try:
            print("[应用初始化] 开始初始化数据库...")
            init_db()
            print("[应用初始化] 数据库初始化完成")
            
            # 启动后台线程
            # 仅在非开发重新加载环境下启动, 或强制在生产环境下启动
            if not app.debug or os.environ.get('WERKZEUG_RUN_MAIN') == 'true':
                if not _acquire_background_jobs_lock():
                    _app_initialized = True
                    return
                print("[应用初始化] 启动后台同步线程...")
                
                # 1. 每日重置检查线程
                reset_thread = threading.Thread(target=daily_reset_check, daemon=True)
                reset_thread.start()

                # 2. 每日正午自动同步飞书运单数据
                sync_thread = threading.Thread(target=daily_feishu_sync_job, daemon=True)
                sync_thread.start()

                # 3. 每小时自动同步 Gofo 集包数据 (方案 1)
                gofo_sync_thread = threading.Thread(target=gofo_hourly_sync_job, daemon=True)
                gofo_sync_thread.start()
                
                # 4. 每日正午人工成本数据自动同步
                labor_sync_thread = threading.Thread(target=auto_sync_labor_data_job, daemon=True)
                labor_sync_thread.start()

                # 5. 每小时自动同步卡车约车数据
                truck_sync_thread = threading.Thread(target=truck_booking_hourly_sync_job, daemon=True)
                truck_sync_thread.start()

                # 6. 每小时自动同步每日集包 operlog 逐条（统计图「逐条日志」）
                packing_operlog_thread = threading.Thread(
                    target=daily_packing_operlog_hourly_sync_job, daemon=True
                )
                packing_operlog_thread.start()

                # 7. 每小时自动同步：CNO 小组分时明细 → 飞书电子表格「元数据」工作表
                feishu_wiki_meta_thread = threading.Thread(
                    target=feishu_wiki_sync_cno_labor_group_hourly_metadata_job,
                    daemon=True,
                )
                feishu_wiki_meta_thread.start()
                
                print("[应用初始化] 所有后台同步线程已启动")

            _app_initialized = True
        except Exception as e:
            print(f"[应用初始化] 错误: {e}")
            import traceback
            traceback.print_exc()
            # 不要抛出异常,让应用继续启动
            # 这样可以看到更详细的错误信息

# 获取正确的静态文件目录
def get_static_dir():
    if getattr(sys, 'frozen', False):
        # 打包后的exe环境 - 静态文件被打包到exe中，需要使用特殊方法访问
        # PyInstaller会将数据文件放在_sys_meipass目录中
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, 'static')
        else:
            # 备用方案
            return os.path.join(os.path.dirname(sys.executable), 'static')
    else:
        # 开发环境 - 使用脚本所在目录的static子目录
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')

# 定义洛杉矶时区
LA_TZ = pytz.timezone('America/Los_Angeles')

def round_to_ten_thousand(value):
    """
    将数值向下取整到万位(千百十个位全部为0)
    例如: 1,232,342 -> 1,230,000
         987,654 -> 980,000
         45,678 -> 40,000
    """
    if value is None:
        return 0
    return (int(value) // 10000) * 10000

# ============================================================================
# Server-Sent Events (SSE) Infrastructure for Real-Time Cross-Device Sync
# ============================================================================

# Global list of SSE client queues
sse_queues = []
sse_queues_lock = threading.Lock()

# Gofo Hourly Sync Status tracking
gofo_sync_status = {
    "last_sync_time": None,
    "status": "pending",
    "message": "Waiting for first sync...",
    "synced_count": 0,
    "last_pieces": 0,
    "last_hour": None,
    "manual_count": 0,
    "device_count": 0
}

def update_gofo_sync_status(status, message, synced_count=0, last_pieces=0, last_hour=None, manual_count=0, device_count=0):
    """更新全域 Gofo 同步狀態"""
    global gofo_sync_status
    now_la = datetime.now(LA_TZ)
    gofo_sync_status = {
        "last_sync_time": now_la.strftime('%Y-%m-%d %H:%M:%S'),
        "status": status,
        "message": message,
        "synced_count": synced_count,
        "last_pieces": last_pieces,
        "last_hour": last_hour,
        "manual_count": manual_count,
        "device_count": device_count
    }
    print(f"[GofoSyncStatus] {status}: {message}")

def log_gofo_sync_event(sync_type, status, message, synced_count=0, last_pieces=0, last_hour=None, manual_count=0, device_count=0):
    """將 Gofo 同步事件记录到数据库"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        now_la = datetime.now(LA_TZ).strftime('%Y-%m-%d %H:%M:%S')
        sql = """
            INSERT INTO gofo_sync_history (sync_time, sync_type, status, message, synced_count, last_pieces, last_hour, manual_count, device_count)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
        cursor.execute(sql, (now_la, sync_type, status, message, synced_count, last_pieces, last_hour, manual_count, device_count))
        conn.commit()
        conn.close()
    except Exception as e:
        print(f"[GofoSyncLog] Failed to log sync event: {e}")

def broadcast_update(event_type='refresh_stats', data=None):
    """
    向所有连接的 SSE 客户端广播更新事件
    
    Args:
        event_type: 事件类型 (默认: 'refresh_stats')
        data: 附加数据 (可选)
    """
    message = {
        'type': event_type,
        'data': data or {},
        'timestamp': datetime.now(LA_TZ).isoformat()
    }
    
    print(f"[SSE] Broadcasting {event_type} to {len(sse_queues)} clients")
    
    with sse_queues_lock:
        # 向所有连接的客户端发送消息
        dead_queues = []
        for q in sse_queues:
            try:
                q.put(message)
            except Exception as e:
                print(f"[SSE] Error sending to client: {e}")
                dead_queues.append(q)
        
        # 清理失败的队列
        for q in dead_queues:
            try:
                sse_queues.remove(q)
            except:
                pass

# ============================================================================

def init_db():
    """初始化数据库 - 自动适配 SQLite/PostgreSQL"""
    # 使用数据库抽象层
    with get_db_connection() as conn:
        cursor = conn.cursor()
        
        # 检查表是否存在 (PostgreSQL 和 SQLite 语法不同)
        if USE_POSTGRES:
            cursor.execute("""
                SELECT EXISTS (
                    SELECT FROM information_schema.tables 
                    WHERE table_name = 'inbound_records'
                )
            """)
            result = cursor.fetchone()
            # PostgreSQL RealDictCursor 返回字典,需要用列名访问
            need = not result['exists'] if USE_POSTGRES else not result[0]
        else:
            need = not os.path.exists(DB_PATH)
        
        if need:
            # 创建入库记录表
            sql = convert_sql("""CREATE TABLE inbound_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                dock_no INTEGER,
                vehicle_type TEXT,
                vehicle_no TEXT,
                unit TEXT,
                load_amount INTEGER,
                pieces INTEGER,
                time_slot TEXT,
                shift_type TEXT,
                remark TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                duration INTEGER,
                plate_excluded_load REAL DEFAULT 0,
                excluded_pieces INTEGER DEFAULT 0,
                business_type TEXT DEFAULT 'GOFO'
            );""")
            cursor.execute(sql)
        
            # 创建分拣记录表
            sql = convert_sql("""CREATE TABLE sorting_records (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                sorting_time DATETIME,
                pieces INTEGER,
                remark TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );""")
            cursor.execute(sql)
            
            # 创建操作日志表
            sql = convert_sql("""CREATE TABLE operation_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                operation_type TEXT,
                table_name TEXT,
                record_id INTEGER,
                old_data TEXT,
                new_data TEXT,
                operator TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );""")
            cursor.execute(sql)
            
            # 创建揽收预估数据表
            sql = convert_sql("""CREATE TABLE pickup_forecast (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                forecast_date DATE NOT NULL,
                forecast_amount INTEGER NOT NULL,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );""")
            cursor.execute(sql)
            
            # 创建用户表
            sql = convert_sql("""CREATE TABLE users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL,
                role TEXT DEFAULT 'user',
                is_active BOOLEAN DEFAULT 1,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );""")
            cursor.execute(sql)
            
            # 创建用户权限表
            sql = convert_sql("""CREATE TABLE user_permissions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                page_name TEXT NOT NULL,
                can_view BOOLEAN DEFAULT 0,
                can_edit BOOLEAN DEFAULT 0,
                can_delete BOOLEAN DEFAULT 0,
                FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE CASCADE
            );""")
            cursor.execute(sql)

            # 创建分拣排班配置表
            sql = convert_sql("""CREATE TABLE sorting_schedule_config (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                config_json TEXT NOT NULL,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );""")
            cursor.execute(sql)
        
        # 确保待发货表存在 (Migration for existing DBs)
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS pending_shipments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            record_date DATE NOT NULL,
            route_code TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
        cursor.execute(sql)

        # 分拣排班：人工班次 / 分拣机按日（与 sorting_schedule_config.config_json 同步写入）
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS sorting_manual_shift (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            config_id INTEGER NOT NULL,
            weekday INTEGER NOT NULL,
            shift_order INTEGER NOT NULL,
            start_h REAL NOT NULL,
            duration_h REAL NOT NULL,
            people INTEGER NOT NULL
        );""")
        cursor.execute(sql)
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS sorting_machine_day (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            config_id INTEGER NOT NULL,
            weekday INTEGER NOT NULL,
            lanes INTEGER NOT NULL,
            start_h REAL NOT NULL,
            hours_per_shift REAL NOT NULL,
            capacity_per_lane REAL NOT NULL
        );""")
        cursor.execute(sql)
        # V2: 按日期保存的分拣中心产能排班（移动端工具）
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS sorting_schedule_daily_plan (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            plan_date DATE UNIQUE NOT NULL,
            plan_json TEXT NOT NULL,
            report_text TEXT,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
        cursor.execute(sql)
        try:
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_sorting_manual_shift_cfg ON sorting_manual_shift(config_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_sorting_machine_day_cfg ON sorting_machine_day(config_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_sorting_daily_plan_date ON sorting_schedule_daily_plan(plan_date)")
        except Exception:
            pass

        # 确保 Gofo 同步历史记录表存在 (Migration for existing DBs)
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS gofo_sync_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sync_time DATETIME DEFAULT CURRENT_TIMESTAMP,
            sync_type TEXT, -- 'auto' or 'manual'
            status TEXT,    -- 'success' or 'error'
            message TEXT,
            synced_count INTEGER,
            last_pieces INTEGER,
            last_hour TEXT,
            manual_count INTEGER,
            device_count INTEGER
        );""")
        cursor.execute(sql)

        # 确保 Gofo 同步历史记录表有 manual_count 和 device_count 列 (Migration)
        try:
            cursor.execute("SELECT manual_count FROM gofo_sync_history LIMIT 1")
        except:
            if USE_POSTGRES:
                cursor.execute("ALTER TABLE gofo_sync_history ADD COLUMN manual_count INTEGER DEFAULT 0")
                cursor.execute("ALTER TABLE gofo_sync_history ADD COLUMN device_count INTEGER DEFAULT 0")
            else:
                cursor.execute("ALTER TABLE gofo_sync_history ADD COLUMN manual_count INTEGER DEFAULT 0")
                cursor.execute("ALTER TABLE gofo_sync_history ADD COLUMN device_count INTEGER DEFAULT 0")

        # CNO01（或指定目的站）按小时集包运单/袋数（popover 口径，供统计页曲线）
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS gofo_collect_destin_hourly (
            sorting_time TEXT NOT NULL,
            time_slot TEXT NOT NULL,
            destin_site TEXT NOT NULL,
            destin_id INTEGER,
            waybill_no_total INTEGER DEFAULT 0,
            package_no_total INTEGER DEFAULT 0,
            updated_at TEXT,
            PRIMARY KEY (sorting_time, time_slot, destin_site)
        );""")
        cursor.execute(sql)

        # CNO 直线窄带分拣机 AA–AD → 生产线 A–D，按 LA 整点小时 operatelog 产能 pieces=逐条 pieces_deduped=运单+类型+操作员去重
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS cno_narrowbelt_hourly (
            record_date TEXT NOT NULL,
            time_slot TEXT NOT NULL,
            line_code TEXT NOT NULL,
            pieces INTEGER NOT NULL DEFAULT 0,
            pieces_deduped INTEGER,
            synced_at TEXT,
            PRIMARY KEY (record_date, time_slot, line_code)
        );""")
        cursor.execute(sql)

        try:
            cursor.execute("SELECT pieces_deduped FROM cno_narrowbelt_hourly LIMIT 1")
        except Exception:
            if USE_POSTGRES:
                cursor.execute(
                    "ALTER TABLE cno_narrowbelt_hourly ADD COLUMN pieces_deduped INTEGER"
                )
            else:
                cursor.execute("ALTER TABLE cno_narrowbelt_hourly ADD COLUMN pieces_deduped INTEGER")

        # 有逐条件数但去重为 0：视为未写入双口径（旧 DEFAULT），清为 NULL 以便与「仅 pieces」旧数据区分；同步后会写入真实去重值
        try:
            cursor.execute(
                convert_query_placeholders(
                    "UPDATE cno_narrowbelt_hourly SET pieces_deduped = NULL "
                    "WHERE pieces > 0 AND pieces_deduped = 0"
                )
            )
        except Exception:
            pass

        # 劳务公司 Sorter（XX Sorter xx + DJ storing 01）分时产能
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS cno_labor_sorter_hourly (
            record_date TEXT NOT NULL,
            time_slot TEXT NOT NULL,
            company_code TEXT NOT NULL,
            pay_type TEXT NOT NULL,
            pieces INTEGER NOT NULL DEFAULT 0,
            pieces_deduped INTEGER,
            synced_at TEXT,
            PRIMARY KEY (record_date, time_slot, company_code, pay_type)
        );""")
        cursor.execute(sql)
        try:
            cursor.execute("SELECT pieces_deduped FROM cno_labor_sorter_hourly LIMIT 1")
        except Exception:
            if USE_POSTGRES:
                cursor.execute(
                    "ALTER TABLE cno_labor_sorter_hourly ADD COLUMN pieces_deduped INTEGER"
                )
            else:
                cursor.execute(
                    "ALTER TABLE cno_labor_sorter_hourly ADD COLUMN pieces_deduped INTEGER"
                )

        sql = convert_sql("""CREATE TABLE IF NOT EXISTS cno_labor_sorter_account_hourly (
            record_date TEXT NOT NULL,
            time_slot TEXT NOT NULL,
            company_code TEXT NOT NULL,
            account_label TEXT NOT NULL,
            pay_type TEXT NOT NULL,
            pieces INTEGER NOT NULL DEFAULT 0,
            pieces_deduped INTEGER,
            synced_at TEXT,
            PRIMARY KEY (record_date, time_slot, company_code, account_label)
        );""")
        cursor.execute(sql)

        # 劳务小组分时产能：运营日 × 统计窗口 × 整点 × 公司 × 组号（便于展示每组每小时票数）
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS cno_labor_group_hourly (
            anchor_date TEXT NOT NULL,
            stats_window TEXT NOT NULL,
            time_slot TEXT NOT NULL,
            company_code TEXT NOT NULL,
            group_no TEXT NOT NULL,
            pay_type TEXT NOT NULL,
            pieces INTEGER NOT NULL DEFAULT 0,
            pieces_deduped INTEGER,
            record_date_la TEXT,
            synced_at TEXT,
            PRIMARY KEY (anchor_date, stats_window, time_slot, company_code, group_no)
        );""")
        cursor.execute(sql)

        sql = convert_sql("""CREATE TABLE IF NOT EXISTS daily_packing_operlog_daily (
            anchor_date TEXT NOT NULL,
            stats_window TEXT NOT NULL DEFAULT 'calendar',
            manual_raw INTEGER NOT NULL DEFAULT 0,
            device_raw INTEGER NOT NULL DEFAULT 0,
            manual_dedup INTEGER NOT NULL DEFAULT 0,
            device_dedup INTEGER NOT NULL DEFAULT 0,
            synced_at TEXT,
            classifier_ver INTEGER NOT NULL DEFAULT 0,
            PRIMARY KEY (anchor_date, stats_window)
        );""")
        cursor.execute(sql)

        sql = convert_sql("""CREATE TABLE IF NOT EXISTS daily_packing_board_daily (
            anchor_date TEXT NOT NULL,
            stats_window TEXT NOT NULL DEFAULT 'calendar',
            manual_count INTEGER NOT NULL DEFAULT 0,
            device_count INTEGER NOT NULL DEFAULT 0,
            total_pieces INTEGER NOT NULL DEFAULT 0,
            synced_at TEXT,
            PRIMARY KEY (anchor_date, stats_window)
        );""")
        cursor.execute(sql)

        # DMS 运单管理查询（WaybillManageQuery）导入
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS gofo_waybill_manage_import (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            waybill_no TEXT NOT NULL,
            waybill_status TEXT,
            plan_origin_center TEXT,
            dest_station TEXT,
            courier_work_area_name TEXT,
            source_create_time TEXT,
            filter_create_date DATE NOT NULL,
            import_batch_id TEXT,
            raw_json TEXT,
            imported_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(waybill_no, filter_create_date)
        );""")
        cursor.execute(sql)

        # inbound_records：填写车牌时可选「不计入总托盘/货量」的装载量（与 load_amount 同单位）
        try:
            cursor.execute("SELECT plate_excluded_load FROM inbound_records LIMIT 1")
        except Exception:
            try:
                cursor.execute(
                    convert_query_placeholders(
                        "ALTER TABLE inbound_records ADD COLUMN plate_excluded_load REAL DEFAULT 0"
                    )
                )
                cursor.execute(
                    convert_query_placeholders(
                        "ALTER TABLE inbound_records ADD COLUMN excluded_pieces INTEGER DEFAULT 0"
                    )
                )
            except Exception as _e:
                print(f"[init_db] inbound_records plate_excluded_load migration: {_e}")

        try:
            cursor.execute("SELECT business_type FROM inbound_records LIMIT 1")
        except Exception:
            try:
                cursor.execute(
                    convert_query_placeholders(
                        "ALTER TABLE inbound_records ADD COLUMN business_type TEXT DEFAULT 'GOFO'"
                    )
                )
            except Exception as _e2:
                print(f"[init_db] inbound_records business_type migration: {_e2}")

        # 确保 outbound_records 表存在 (Migration for existing DBs that might be missing it)
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS outbound_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            record_date DATE NOT NULL,
            route_code TEXT NOT NULL,
            route_type TEXT,
            vehicle_count INTEGER DEFAULT 1,
            cost DECIMAL(10, 2),
            notes TEXT,
            created_by TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
        cursor.execute(sql)
        # 一车次一行：vehicle_count 一律为 1（历史/聚合残留的大于 1 也会改回）
        try:
            cursor.execute(convert_query_placeholders(
                "UPDATE outbound_records SET vehicle_count = 1 WHERE COALESCE(vehicle_count, 0) != 1"
            ))
        except Exception:
            pass
        if USE_POSTGRES:
            try:
                cursor.execute(
                    "ALTER TABLE outbound_records ALTER COLUMN vehicle_count SET DEFAULT 1"
                )
            except Exception:
                pass

        # 禁止 route_type='regular'：历史数据改为 branch，库层拒绝再写入 regular（应用层已 normalize）
        try:
            cursor.execute(
                convert_query_placeholders(
                    "UPDATE outbound_records SET route_type = 'branch' "
                    "WHERE LOWER(TRIM(COALESCE(route_type, ''))) = 'regular'"
                )
            )
        except Exception:
            pass
        if USE_POSTGRES:
            try:
                cursor.execute(
                    """
                    ALTER TABLE outbound_records ADD CONSTRAINT outbound_records_no_regular_route
                    CHECK (route_type IS NULL OR LOWER(TRIM(route_type)) <> 'regular')
                    """
                )
            except Exception:
                pass
        else:
            try:
                cursor.execute("DROP TRIGGER IF EXISTS outbound_records_reject_regular_bi")
                cursor.execute("DROP TRIGGER IF EXISTS outbound_records_reject_regular_bu")
                cursor.execute(
                    """
                    CREATE TRIGGER outbound_records_reject_regular_bi
                    BEFORE INSERT ON outbound_records
                    FOR EACH ROW
                    WHEN LOWER(TRIM(COALESCE(NEW.route_type, ''))) = 'regular'
                    BEGIN
                        SELECT RAISE(ABORT, 'route_type regular is not allowed');
                    END
                    """
                )
                cursor.execute(
                    """
                    CREATE TRIGGER outbound_records_reject_regular_bu
                    BEFORE UPDATE OF route_type ON outbound_records
                    FOR EACH ROW
                    WHEN LOWER(TRIM(COALESCE(NEW.route_type, ''))) = 'regular'
                    BEGIN
                        SELECT RAISE(ABORT, 'route_type regular is not allowed');
                    END
                    """
                )
            except Exception:
                pass

        # Migration: outbound_records.mt_number — used as the dedup key for the
        # incremental MT-based Google Sheet sync (K column rule). Filled only by
        # the new sync flow;手工录入 / truck_bookings 聚合写入仍为 NULL，互不影响。
        # PostgreSQL: ADD COLUMN IF NOT EXISTS supported (≥ 9.6).
        # SQLite: PRAGMA table_info → ADD COLUMN if missing.
        try:
            if USE_POSTGRES:
                cursor.execute(
                    "ALTER TABLE outbound_records ADD COLUMN IF NOT EXISTS mt_number TEXT"
                )
                cursor.execute(
                    "CREATE UNIQUE INDEX IF NOT EXISTS outbound_records_mt_number_uq "
                    "ON outbound_records(mt_number) WHERE mt_number IS NOT NULL"
                )
            else:
                cursor.execute("PRAGMA table_info(outbound_records)")
                cols = {row[1] for row in cursor.fetchall()}
                if "mt_number" not in cols:
                    cursor.execute(
                        "ALTER TABLE outbound_records ADD COLUMN mt_number TEXT"
                    )
                cursor.execute(
                    "CREATE UNIQUE INDEX IF NOT EXISTS outbound_records_mt_number_uq "
                    "ON outbound_records(mt_number) WHERE mt_number IS NOT NULL"
                )
        except Exception as _e:
            print(f"[init_db] outbound_records.mt_number migration: {_e}")

        sql = convert_sql("""CREATE TABLE IF NOT EXISTS consumables (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name VARCHAR(50) UNIQUE NOT NULL,
            unit VARCHAR(20) NOT NULL,
            safety_stock REAL NOT NULL,
            current_stock REAL DEFAULT 0.0,
            lead_time_days INTEGER NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
        cursor.execute(sql)

        # 确保 production_consumable_master 表存在
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS production_consumable_master (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name VARCHAR(50) UNIQUE NOT NULL,
            price REAL NOT NULL,
            batch VARCHAR(50) DEFAULT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
        cursor.execute(sql)

        # 确保 inventory_transactions 表存在 (Migration for existing DBs)
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS inventory_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            consumable_id INTEGER NOT NULL,
            type VARCHAR(10) NOT NULL,
            quantity REAL NOT NULL,
            related_warehouse_volume INTEGER,
            operator VARCHAR(50) NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            FOREIGN KEY(consumable_id) REFERENCES consumables(id)
        );""")
        cursor.execute(sql)

        # ====== 新增: 生产成本配置模块 ======
        # 1. 人工计时配置 (labor hourly)
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS config_labor_hourly (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            company_name VARCHAR(100) UNIQUE NOT NULL,
            hourly_rate REAL NOT NULL,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
        cursor.execute(sql)

        # 2. 计件费用配置 (labor piece rate)
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS config_labor_piece (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            operation_name VARCHAR(100) UNIQUE NOT NULL,
            piece_rate REAL NOT NULL,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
        cursor.execute(sql)

        # 3. 设备维护配置 (equipment hourly)
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS config_equipment_hourly (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            equipment_name VARCHAR(100) UNIQUE NOT NULL,
            hourly_rate REAL NOT NULL,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
        cursor.execute(sql)
        # ====================================

        # ====== 新增: 成本核算系统配置 ======
        # 1. 成本核算主表 (cost_main)
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS cost_main (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            period_start DATE NOT NULL,
            period_end DATE NOT NULL,
            direction VARCHAR(100) NOT NULL,
            total_transport_cost REAL DEFAULT 0,
            total_labor_cost REAL DEFAULT 0,
            total_consumable_cost REAL DEFAULT 0,
            total_other_cost REAL DEFAULT 0,
            total_cost REAL DEFAULT 0,
            total_pieces INTEGER DEFAULT 0,
            unit_cost REAL DEFAULT 0,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
        cursor.execute(sql)

        # 2. 流向工序计件配置表 (config_labor_price_flow)
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS config_labor_price_flow (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            direction VARCHAR(100) NOT NULL,
            operation_name VARCHAR(100) NOT NULL,
            piece_rate REAL NOT NULL,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(direction, operation_name)
        );""")
        cursor.execute(sql)

        # 3. 耗材分摊规则表 (config_consumable_split)
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS config_consumable_split (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            consumable_name VARCHAR(100) UNIQUE NOT NULL,
            split_method VARCHAR(50) NOT NULL DEFAULT 'weight',
            weight_json TEXT,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
        cursor.execute(sql)
        # 4. 卡车约车同步表 (truck_bookings)
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS truck_bookings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            pickup_date TEXT,
            destination VARCHAR(100),
            vendor VARCHAR(100),
            pickup_no VARCHAR(100) UNIQUE,
            cost REAL DEFAULT 0,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
        cursor.execute(sql)

        # 5. 系统配置表 (system_config)
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS system_config (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            config_key TEXT UNIQUE NOT NULL,
            config_value TEXT,
            description TEXT,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
        cursor.execute(sql)
        
        # 针对已存在表进行迁移 (Migration for existing columns)
        try:
            cursor.execute("SELECT description FROM system_config LIMIT 1")
        except:
            try:
                cursor.execute("ALTER TABLE system_config ADD COLUMN description TEXT")
                cursor.execute("ALTER TABLE system_config ADD COLUMN updated_at DATETIME DEFAULT CURRENT_TIMESTAMP")
            except:
                pass
        
        # Gofo token: never ship real credentials; set GOFO_ADMIN_TOKEN or configure in UI / DB.
        placeholder = get_placeholder()
        _gofo_default = os.environ.get("GOFO_ADMIN_TOKEN", "")
        cursor.execute(f"INSERT OR IGNORE INTO system_config (config_key, config_value, description) VALUES ({placeholder}, {placeholder}, {placeholder})", 
                      ("gofo_admin_token", _gofo_default, "Gofo API Admin Token"))
        cursor.execute(f"INSERT OR IGNORE INTO system_config (config_key, config_value, description) VALUES ({placeholder}, {placeholder}, {placeholder})",
                      ("license_device_token", "", "商业许可证激活后的 device_token（LICENSE_DEVICE_TOKEN 环境变量优先）"))
        cursor.execute(f"INSERT OR IGNORE INTO system_config (config_key, config_value, description) VALUES ({placeholder}, {placeholder}, {placeholder})",
                      ("last_stretch_film_deduct_date", "", "洛杉矶日历日期，上次自动扣减缠绕膜的日期（防多进程重复扣减）"))
        # ====================================

        # ====== 新增: 飞书数据同步备用表 ======
        sql = convert_sql("""CREATE TABLE IF NOT EXISTS feishu_raw_data (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            task_code VARCHAR(100) UNIQUE NOT NULL,
            record_date DATE NOT NULL,
            destination VARCHAR(100) NOT NULL,
            boxes_count INTEGER DEFAULT 0,
            tickets_count INTEGER DEFAULT 0,
            volume_load_rate VARCHAR(50),
            box_load_rate VARCHAR(50),
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
        cursor.execute(sql)
        # ====================================

        if need:
            # 插入默认管理员用户 (用户名: admin; 密码见 INITIAL_ADMIN_PASSWORD 或启动日志中的随机值)
            import hashlib
            admin_password = hashlib.sha256(_initial_admin_password().encode()).hexdigest()
            placeholder = get_placeholder()
            cursor.execute(f"INSERT INTO users (username, password_hash, role) VALUES ({placeholder}, {placeholder}, {placeholder})",
                        ("admin", admin_password, "admin"))
            
            # 为管理员用户设置默认权限
            cursor.execute(f"SELECT id FROM users WHERE username = {placeholder}", ("admin",))
            result = cursor.fetchone()
            admin_user_id = result[0] if USE_POSTGRES else result[0]
            
            # 管理员拥有所有页面的所有权限
            pages = ['index', 'sorting', 'history', 'statistics', 'logs', 'sorting-schedule', 'outbound-stats', 'consumables', 'operations_metrics']
            for page in pages:
                cursor.execute(f"""INSERT INTO user_permissions 
                    (user_id, page_name, can_view, can_edit, can_delete) 
                    VALUES ({placeholder}, {placeholder}, 1, 1, 1)""", 
                    (admin_user_id, page))

        # Ensure outbound-stats permission exists for all users (Migration)
        # This block runs regardless of 'need' to ensure existing databases are updated.
        placeholder = get_placeholder() # Re-get placeholder in case it's needed again
        cursor.execute(f"SELECT id FROM users WHERE username = {placeholder}", ("admin",))
        admin_res = cursor.fetchone()
        if admin_res:
             admin_uid = admin_res['id'] if USE_POSTGRES else admin_res[0]
             cursor.execute(f"SELECT id FROM user_permissions WHERE user_id = {placeholder} AND page_name = {placeholder}", (admin_uid, 'outbound-stats'))
             if not cursor.fetchone():
                 print("Migrating: Adding outbound-stats permission to users...")
                 # Get all users
                 cursor.execute("SELECT id, role FROM users")
                 all_users = cursor.fetchall()
                 for usr in all_users:
                     uid = usr['id'] if USE_POSTGRES else usr[0]
                     role = usr['role'] if USE_POSTGRES else usr[1]
                     # Admin gets full access, others get none by default (can be updated by admin)
                     can_access = 1 if role in ('admin', 'boss') else 0
                     cursor.execute(f"INSERT INTO user_permissions (user_id, page_name, can_view, can_edit, can_delete) VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})", 
                                   (uid, 'outbound-stats', can_access, can_access, can_access))
                                   
                     # Migrate consumables permission
                     cursor.execute(f"SELECT id FROM user_permissions WHERE user_id = {placeholder} AND page_name = {placeholder}", (uid, 'consumables'))
                     if not cursor.fetchone():
                         cursor.execute(f"INSERT INTO user_permissions (user_id, page_name, can_view, can_edit, can_delete) VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})", 
                                       (uid, 'consumables', can_access, can_access, can_access))
                                       
                     # Migrate cost_accounting permission
                     cursor.execute(f"SELECT id FROM user_permissions WHERE user_id = {placeholder} AND page_name = {placeholder}", (uid, 'cost_accounting'))
                     if not cursor.fetchone():
                         cursor.execute(f"INSERT INTO user_permissions (user_id, page_name, can_view, can_edit, can_delete) VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})", 
                                       (uid, 'cost_accounting', can_access, can_access, can_access))
        
        if not need: # This is the original 'else' block for `if need:`
            # 如果数据库已存在，检查并创建用户表和权限表（如果不存在）
            if USE_POSTGRES:
                cursor.execute("SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_name = 'users')")
                result = cursor.fetchone()
                table_exists = result['exists']
            else:
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='users'")
                table_exists = cursor.fetchone() is not None
            
            if not table_exists:
                # 创建用户表
                sql = convert_sql("""CREATE TABLE users (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    role TEXT DEFAULT 'user',
                    is_active BOOLEAN DEFAULT 1,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                );""")
                cursor.execute(sql)
                
                # 创建用户权限表
                sql = convert_sql("""CREATE TABLE user_permissions (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    user_id INTEGER NOT NULL,
                    page_name TEXT NOT NULL,
                    can_view BOOLEAN DEFAULT 0,
                    can_edit BOOLEAN DEFAULT 0,
                    can_delete BOOLEAN DEFAULT 0,
                    FOREIGN KEY (user_id) REFERENCES users (id) ON DELETE CASCADE
                );""")
                cursor.execute(sql)
                
                # 插入默认管理员用户 (密码: INITIAL_ADMIN_PASSWORD 或启动日志中的随机值)
                import hashlib
                admin_password = hashlib.sha256(_initial_admin_password().encode()).hexdigest()
                placeholder = get_placeholder()
                cursor.execute(f"INSERT INTO users (username, password_hash, role) VALUES ({placeholder}, {placeholder}, {placeholder})",
                            ("admin", admin_password, "admin"))
                
                # 为管理员用户设置默认权限
                cursor.execute(f"SELECT id FROM users WHERE username = {placeholder}", ("admin",))
                result = cursor.fetchone()
                admin_user_id = result['id'] if USE_POSTGRES else result[0]
                
                # 管理员拥有所有页面的所有权限
                pages = ['index', 'sorting', 'history', 'statistics', 'logs', 'sorting-schedule', 'outbound-stats', 'consumables', 'cost_accounting', 'operations_metrics']
                for page in pages:
                    cursor.execute(f"""INSERT INTO user_permissions 
                        (user_id, page_name, can_view, can_edit, can_delete) 
                        VALUES ({placeholder}, {placeholder}, 1, 1, 1)""", 
                        (admin_user_id, page))
            
            # ====== 新增: 生产成本配置模块 ======
            # 1. 人工计时配置 (labor hourly)
            sql = convert_sql("""CREATE TABLE IF NOT EXISTS config_labor_hourly (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                company_name VARCHAR(100) UNIQUE NOT NULL,
                hourly_rate REAL NOT NULL,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );""")
            cursor.execute(sql)

            # 2. 计件费用配置 (labor piece rate)
            sql = convert_sql("""CREATE TABLE IF NOT EXISTS config_labor_piece (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                operation_name VARCHAR(100) UNIQUE NOT NULL,
                piece_rate REAL NOT NULL,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );""")
            cursor.execute(sql)

            # 3. 设备维护配置 (equipment hourly)
            sql = convert_sql("""CREATE TABLE IF NOT EXISTS config_equipment_hourly (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                equipment_name VARCHAR(100) UNIQUE NOT NULL,
                hourly_rate REAL NOT NULL,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );""")
            cursor.execute(sql)
            # ====================================

            # ====== 新增: 成本核算系统配置 ======
            # 1. 成本核算主表 (cost_main)
            sql = convert_sql("""CREATE TABLE IF NOT EXISTS cost_main (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                period_start DATE NOT NULL,
                period_end DATE NOT NULL,
                direction VARCHAR(100) NOT NULL,
                total_transport_cost REAL DEFAULT 0,
                total_labor_cost REAL DEFAULT 0,
                total_consumable_cost REAL DEFAULT 0,
                total_other_cost REAL DEFAULT 0,
                total_cost REAL DEFAULT 0,
                total_pieces INTEGER DEFAULT 0,
                unit_cost REAL DEFAULT 0,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );""")
            cursor.execute(sql)

            # 2. 流向工序计件配置表 (config_labor_price_flow)
            sql = convert_sql("""CREATE TABLE IF NOT EXISTS config_labor_price_flow (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                direction VARCHAR(100) NOT NULL,
                operation_name VARCHAR(100) NOT NULL,
                piece_rate REAL NOT NULL,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(direction, operation_name)
            );""")
            cursor.execute(sql)

            # 3. 耗材分摊规则表 (config_consumable_split)
            sql = convert_sql("""CREATE TABLE IF NOT EXISTS config_consumable_split (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                consumable_name VARCHAR(100) UNIQUE NOT NULL,
                split_method VARCHAR(50) NOT NULL DEFAULT 'weight',
                weight_json TEXT,
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );""")
            cursor.execute(sql)
            # ====================================

            # ====== 新增: 飞书数据同步备用表 ======
            sql = convert_sql("""CREATE TABLE IF NOT EXISTS feishu_raw_data (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                task_code VARCHAR(100) UNIQUE NOT NULL,
                record_date DATE NOT NULL,
                destination VARCHAR(100) NOT NULL,
                boxes_count INTEGER DEFAULT 0,
                tickets_count INTEGER DEFAULT 0,
                volume_load_rate VARCHAR(50),
                box_load_rate VARCHAR(50),
                updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
            );""")
            cursor.execute(sql)
            # ====================================

            # 检查并创建揽收预估数据表（如果不存在）
            if USE_POSTGRES:
                cursor.execute("SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_name = 'pickup_forecast')")
                result = cursor.fetchone()
                table_exists = result['exists']
            else:
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='pickup_forecast'")
                table_exists = cursor.fetchone() is not None
            
            if not table_exists:
                sql = convert_sql("""CREATE TABLE pickup_forecast (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    forecast_date DATE NOT NULL,
                    forecast_amount INTEGER NOT NULL,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
                );""")
                cursor.execute(sql)
                
                # 创建分拣记录表
                sql = convert_sql("""CREATE TABLE IF NOT EXISTS sorting_records (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    sorting_time DATE,
                    pieces INTEGER,
                    remark TEXT,
                    time_slot TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                );""")
                cursor.execute(sql)

            # 检查并创建分拣排班配置表（如果不存在）
            if USE_POSTGRES:
                cursor.execute("SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_name = 'sorting_schedule_config')")
                result = cursor.fetchone()
                table_exists = result['exists']
            else:
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='sorting_schedule_config'")
                table_exists = cursor.fetchone() is not None
            
            if not table_exists:
                sql = convert_sql("""CREATE TABLE sorting_schedule_config (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    config_json TEXT NOT NULL,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
                );""")
                cursor.execute(sql)

            # 检查并创建出库记录表（如果不存在）
            if USE_POSTGRES:
                cursor.execute("SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_name = 'outbound_records')")
                result = cursor.fetchone()
                table_exists = result['exists']
            else:
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='outbound_records'")
                table_exists = cursor.fetchone() is not None
            
            if not table_exists:
                sql = convert_sql("""CREATE TABLE outbound_records (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    record_date TEXT NOT NULL,
                    route_code TEXT NOT NULL,
                    route_type TEXT NOT NULL,
                    vehicle_count INTEGER DEFAULT 1,
                    cost REAL DEFAULT 0,
                    notes TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    created_by TEXT,
                    UNIQUE(record_date, route_code)
                );""")
                cursor.execute(sql)

            # 检查并创建运费表（如果不存在）
            if USE_POSTGRES:
                cursor.execute("SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_name = 'freight_rates')")
                result = cursor.fetchone()
                table_exists = result['exists']
            else:
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='freight_rates'")
                table_exists = cursor.fetchone() is not None
            
            if not table_exists:
                sql = convert_sql("""CREATE TABLE freight_rates (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    route_code TEXT NOT NULL,
                    rate REAL NOT NULL,
                    effective_date DATETIME NOT NULL,
                    created_by TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    notes TEXT
                );""")
                cursor.execute(sql)
                
                # 创建索引以提高查询效率
                if USE_POSTGRES:
                    cursor.execute("CREATE INDEX idx_freight_route_date ON freight_rates(route_code, effective_date DESC)")
                else:
                    cursor.execute("CREATE INDEX idx_freight_route_date ON freight_rates(route_code, effective_date DESC)")
                
                # 插入初始运费数据
                initial_rates = [
                    ('ATL', 5100.00),
                    ('EWR', 6500.00),
                    ('MIA', 6800.00),
                    ('MCO', 6800.00),
                    ('DFW', 4000.00),
                    ('ORD', 4600.00),
                    ('IAH', 4000.00),
                    ('CLT', 6500.00)
                ]
                
                placeholder = get_placeholder()
                la_now = datetime.now(LA_TZ)
                for route_code, rate in initial_rates:
                    cursor.execute(
                        f"""INSERT INTO freight_rates (route_code, rate, effective_date, created_by, notes) 
                        VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})""",
                        (route_code, rate, la_now, 'system', 'Initial rate')
                    )

                # Generate pending_shipments table
                if USE_POSTGRES:
                    cursor.execute("SELECT EXISTS (SELECT FROM information_schema.tables WHERE table_name = 'pending_shipments')")
                    result = cursor.fetchone()
                    table_exists = result['exists']
                else:
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='pending_shipments'")
                    table_exists = cursor.fetchone() is not None

                if not table_exists:
                    sql = convert_sql("""CREATE TABLE pending_shipments (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        record_date DATE NOT NULL,
                        route_code TEXT NOT NULL,
                        quantity INTEGER NOT NULL,
                        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                    );""")
                    cursor.execute(sql)


def convert_utc_to_la(utc_time_str):
    """直接返回时间字符串，因为数据库中存储的已经是洛杉矶时间"""
    return utc_time_str

# 检查用户权限
def check_user_permission(page_name, permission_type='view'):
    if 'user_id' not in session:
        return False
    
    user_id = session['user_id']
    
    # 查询用户权限
    conn = get_db()
    query = f"""
        SELECT up.can_{permission_type}
        FROM user_permissions up
        WHERE up.user_id = ? AND up.page_name = ?
    """
    cursor = conn.cursor(); cursor.execute(query, (user_id, page_name))
    result = cursor.fetchone()
    conn.close()
    
    return result and (result[f'can_{permission_type}'] if USE_POSTGRES else result[0])

def daily_reset_check():
    """每日重置检查函数"""
    while True:
        try:
            # 获取洛杉矶当前时间
            la_now = datetime.now(LA_TZ)
            
            # 如果是午夜（0点）附近，执行重置
            if la_now.hour == 0 and la_now.minute == 0:
                print(f"执行每日重置: {la_now}")
                perform_daily_reset()
                
                # 等待一分钟，避免重复执行
                time.sleep(60)
            else:
                # 每分钟检查一次
                time.sleep(60)
        except Exception as e:
            print(f"每日重置检查出错: {e}")
            time.sleep(60)

def auto_deduct_stretch_film():
    """每日自动扣减缠绕膜 (80卷)。同一洛杉矶日历日仅执行一次（多 Gunicorn worker / 重复触发时只扣一次）。"""
    conn = None
    try:
        la_now = datetime.now(LA_TZ)
        today_la = la_now.strftime('%Y-%m-%d')
        la_now_str = la_now.strftime('%Y-%m-%d %H:%M:%S')
        placeholder = get_placeholder()

        conn = get_db()
        cursor = conn.cursor()
        if not USE_POSTGRES:
            conn.execute("BEGIN IMMEDIATE")

        # 旧库可能无此行；先占位，否则 UPDATE 影响 0 行会被误判为「本日已扣过」
        _desc = "洛杉矶日历日期，上次自动扣减缠绕膜的日期（防多进程重复扣减）"
        if USE_POSTGRES:
            cursor.execute(
                """
                INSERT INTO system_config (config_key, config_value, description)
                VALUES ('last_stretch_film_deduct_date', '', %s)
                ON CONFLICT (config_key) DO NOTHING
                """,
                (_desc,),
            )
        else:
            cursor.execute(
                f"INSERT OR IGNORE INTO system_config (config_key, config_value, description) VALUES ({placeholder}, {placeholder}, {placeholder})",
                ("last_stretch_film_deduct_date", "", _desc),
            )

        # 0) 原子抢占「本日扣减名额」：仅当 last_stretch_film_deduct_date != 今日 时更新成功，避免多进程各跑一遍
        cursor.execute(convert_query_placeholders(f"""
            UPDATE system_config SET config_value = {placeholder}, updated_at = {placeholder}
            WHERE config_key = 'last_stretch_film_deduct_date'
              AND (config_value IS NULL OR TRIM(config_value) = '' OR config_value != {placeholder})
        """), (today_la, la_now_str, today_la))
        if cursor.rowcount == 0:
            conn.rollback()
            conn.close()
            print(f"[AutoDeduct] 本日({today_la})已执行过缠绕膜自动扣减，跳过（防重复）")
            return

        print("[AutoDeduct] 正在执行每日缠绕膜自动扣减 (80卷)...")

        # 1) 查找缠绕膜 (ID通常为1，但为了安全使用名称查找)
        cursor.execute(f"SELECT id, current_stock, name FROM consumables WHERE name LIKE '%缠绕膜%' LIMIT 1")
        row = cursor.fetchone()

        if not row:
            print("[AutoDeduct] 警告: 未找到名称包含'缠绕膜'的耗材，跳过自动扣减（已回滚本日抢占，下次可重试）")
            conn.rollback()
            conn.close()
            return

        consumable_id = row['id'] if hasattr(row, 'keys') else row[0]
        current_stock = row['current_stock'] if hasattr(row, 'keys') else row[1]
        name = row['name'] if hasattr(row, 'keys') else row[2]

        # 2) 计算新库存 (假设每天固定消耗80卷)
        deduction_qty = 80.0
        next_stock = max(0, current_stock - deduction_qty)

        # 3) 更新库存
        cursor.execute(f"UPDATE consumables SET current_stock = {placeholder}, updated_at = {placeholder} WHERE id = {placeholder}",
                      (next_stock, la_now_str, consumable_id))

        # 4) 插入流水记录
        cursor.execute(f"""
            INSERT INTO inventory_transactions (consumable_id, type, quantity, related_warehouse_volume, operator, created_at)
            VALUES ({placeholder}, 'OUT', {placeholder}, NULL, 'SYSTEM', {placeholder})
        """, (consumable_id, deduction_qty, la_now_str))

        conn.commit()
        conn.close()
        conn = None
        print(f"[AutoDeduct] 成功扣减 {name}: {current_stock} -> {next_stock}")

    except Exception as e:
        print(f"[AutoDeduct] 错误: {e}")
        if conn is not None:
            try:
                conn.rollback()
            except Exception:
                pass
            try:
                conn.close()
            except Exception:
                pass

def perform_daily_reset():
    """执行每日重置 - 仅记录日志，并执行自动扣减逻辑"""
    try:
        # 执行缠绕膜自动扣减
        auto_deduct_stretch_film()
        
        print("每日重置检查完成 - 历史数据已永久保存")
    except Exception as e:
        print(f"每日重置执行出错: {e}")


@app.route('/')
def index():
    # 检查用户权限，所有用户都需要登录
    if 'user_id' not in session:
        return redirect('/login')
    
    # 已登录用户检查权限
    if not check_page_permission('index'):
        return redirect('/no_permission')
    
    # 有权限则返回主页
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'index.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404

@app.route('/sorting')
def sorting():
    # 检查用户权限
    if 'user_id' not in session:
        return redirect('/login')
    
    if not check_page_permission('sorting'):
        return redirect('/no_permission')
    
    # 返回分拣录入页面
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'sorting.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404

@app.route('/history')
def history():
    # 检查用户权限
    if 'user_id' not in session:
        return redirect('/login')
    
    if not check_page_permission('history'):
        return redirect('/no_permission')
    
    # 返回历史查询页面
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'history.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404

@app.route('/statistics')
def statistics():
    # 检查用户权限
    if 'user_id' not in session:
        return redirect('/login')
    
    if not check_page_permission('statistics'):
        return redirect('/no_permission')
    
    # 返回统计数据页面
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'statistics.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404


@app.route('/share/cno-labor-group-heatmap')
def share_cno_labor_group_heatmap():
    """小组 × 小时产能热力图独立分享页（需登录且有 statistics 权限）。"""
    if 'user_id' not in session:
        from urllib.parse import quote
        nxt = request.full_path if request.full_path else request.path
        return redirect('/login?next=' + quote(nxt, safe=''))
    if not check_page_permission('statistics'):
        return redirect('/no_permission')
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'cno_labor_group_heatmap_share.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    return f"File not found: {file_path}", 404


@app.route('/operations_metrics')
def operations_metrics():
    # 用户权限
    if 'user_id' not in session:
        return redirect('/login')
    
    if not check_page_permission('cost_accounting'):
        return redirect('/no_permission')
    
    # 返回成本核算页面 (原 cost_accounting，已迁移至 /operations_metrics)
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'cost_accounting.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404

@app.route('/sorting-schedule')
def sorting_schedule():
    # 检查用户权限
    if 'user_id' not in session:
        return redirect('/login')
    
    if not check_page_permission('sorting-schedule'):
        return redirect('/no_permission')
    
    # 返回分拣力排班上报页面
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'sorting-schedule.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404

@app.route('/outbound-stats')
def outbound_stats():
    # 检查用户权限
    if 'user_id' not in session:
        return redirect('/login')
    
    if not check_page_permission('outbound-stats'):
        return redirect('/no_permission')
    
    # 返回出库统计页面
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'outbound-stats.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404


@app.route('/route-distribution')
def route_distribution_page():
    """流向分布数据表（独立页面，复用 /api/outbound/records 数据源 + outbound-stats 权限）"""
    if 'user_id' not in session:
        return redirect('/login')

    if not check_page_permission('outbound-stats'):
        return redirect('/no_permission')

    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'route-distribution.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    return f"File not found: {file_path}", 404


@app.route('/route-map')
def route_map_page():
    """流向分布美国地图（Leaflet + OSM，复用 outbound-stats 权限与 records API）"""
    if 'user_id' not in session:
        return redirect('/login')

    if not check_page_permission('outbound-stats'):
        return redirect('/no_permission')

    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'route-map.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    return f"File not found: {file_path}", 404


@app.route('/schedule-packaging')
def schedule_packaging_page():
    if 'user_id' not in session:
        return redirect('/login')
    if not check_page_permission('outbound-stats'):
        return redirect('/no_permission')
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'schedule-packaging.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    return f"File not found: {file_path}", 404


@app.route('/api/analytics/schedule-vs-packaging')
def api_schedule_vs_packaging():
    """按运营日 06:00–次日 06:00（LA）聚合集包实际 vs 当前最新排班理论日产能。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not check_page_permission('outbound-stats'):
        return jsonify({'error': '无权限'}), 403
    start_s = request.args.get('start_date')
    end_s = request.args.get('end_date')
    if not start_s or not end_s:
        return jsonify({'error': '请提供 start_date 与 end_date（YYYY-MM-DD）'}), 400
    try:
        start_date = datetime.strptime(start_s, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_s, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': '日期格式无效'}), 400
    if start_date > end_date:
        return jsonify({'error': '开始日期不能晚于结束日期'}), 400

    try:
        defaults = {
            "manual": {"capacity": 3000, "hoursPerShift": 9, "schedule": [3, 3, 3, 3, 3, 3, 3]},
            "machine": {"capacity": 4500, "hoursPerShift": 6, "schedule": [4, 4, 4, 4, 4, 4, 4]},
            "night": {"capacity": 4500, "hoursPerShift": 6, "schedule": [4, 4, 4, 4, 4, 4, 4]},
        }
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(convert_query_placeholders(
            "SELECT config_json FROM sorting_schedule_config ORDER BY id DESC LIMIT 1"
        ))
        row = cursor.fetchone()
        if row:
            config = _parse_config_json_from_db(row[0], defaults)
        else:
            config = defaults

        mcfg = config.get('manual') if isinstance(config.get('manual'), dict) else {}
        msched = config.get('machine') if isinstance(config.get('machine'), dict) else {}

        schedule_manual = []
        schedule_machine = []
        for wd in range(7):
            schedule_manual.append(_manual_day_theoretical_pieces(mcfg, wd))
            schedule_machine.append(_machine_or_night_day_theoretical_pieces(msched, wd))

        dates = []
        actual_manual = []
        actual_device = []
        actual_pieces = []
        d = start_date
        while d <= end_date:
            dates.append(d.strftime('%Y-%m-%d'))
            next_d = d + timedelta(days=1)
            cursor.execute(convert_query_placeholders("""
                SELECT COALESCE(SUM(manual_count), 0), COALESCE(SUM(device_count), 0), COALESCE(SUM(pieces), 0)
                FROM sorting_records
                WHERE (sorting_time = ? AND time_slot >= '06:00')
                   OR (sorting_time = ? AND time_slot < '06:00')
            """), (d.strftime('%Y-%m-%d'), next_d.strftime('%Y-%m-%d')))
            r = cursor.fetchone()
            actual_manual.append(int(r[0]) if r and r[0] is not None else 0)
            actual_device.append(int(r[1]) if r and r[1] is not None else 0)
            actual_pieces.append(int(r[2]) if r and r[2] is not None else 0)
            d += timedelta(days=1)

        conn.close()

        sch_man = []
        sch_mac = []
        for ds in dates:
            wd = datetime.strptime(ds, '%Y-%m-%d').date().weekday()
            sch_man.append(schedule_manual[wd])
            sch_mac.append(schedule_machine[wd])

        return jsonify({
            'dates': dates,
            'series': {
                'schedule_manual': sch_man,
                'schedule_machine': sch_mac,
                'actual_manual': actual_manual,
                'actual_device': actual_device,
                'actual_pieces': actual_pieces,
            },
            'weekday_schedule': {
                'manual': schedule_manual,
                'machine': schedule_machine,
            },
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/statistics/daily_packing_split', methods=['GET'])
def api_statistics_daily_packing_split():
    """按 stats_window 聚合每日人工/设备集包件数；calendar=自然日；business=05–次日05；seventeen=17–次日17（本地）。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not check_page_permission('statistics'):
        return jsonify({'error': '无权限'}), 403
    try:
        window_mode = _parse_stats_window_param(request.args.get('stats_window'))

        def _parse_d(val):
            if not val:
                return None
            try:
                return datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
            except ValueError:
                return None

        start_in = _parse_d(request.args.get('start_date'))
        end_raw = request.args.get('end_date') or request.args.get('date')
        end_in = _parse_d(end_raw)

        if start_in and end_in and start_in <= end_in:
            span = (end_in - start_in).days + 1
            if span > 90:
                return jsonify({'error': '日期跨度不能超过 90 天'}), 400
            start_date, end_date = start_in, end_in
        else:
            days = request.args.get('days', type=int)
            if days is None or days < 1:
                days = 14
            days = min(days, 90)
            if end_in:
                end_date = end_in
            else:
                end_date = datetime.now(LA_TZ).date()
            start_date = end_date - timedelta(days=days - 1)
        count_mode = _parse_daily_packing_count_mode(request.args.get('count_mode'))
        sync_operlog = (request.args.get('sync_operlog') or '').strip().lower() in (
            '1', 'true', 'yes', 'on',
        )
        conn = get_db()
        cursor = conn.cursor()
        dates = []
        manual = []
        device = []
        total_pieces = []
        operlog_sync_failures = 0
        operlog_cache_misses = 0
        operlog_unsynced_dates = []
        board_fallback_days = 0
        board_cache_misses = 0
        hour_slots = []
        sync_board = (request.args.get('sync_board') or '').strip().lower() in (
            '1', 'true', 'yes', 'on',
        )
        import sync_daily_packing_board as _dp_board

        d = start_date
        while d <= end_date:
            dates.append(d.strftime('%Y-%m-%d'))
            if count_mode == 'board':
                board_res = None
                if sync_board:
                    board_res = _dp_board.sync_daily_packing_board_anchor(d, window_mode, force=True)
                else:
                    cached = _dp_board.read_daily_packing_board_anchor(d, window_mode)
                    if cached is not None:
                        board_res = {"success": True, **cached}
                    else:
                        board_cache_misses += 1
                        board_res = None
                if board_res and board_res.get("success"):
                    m = int(board_res.get("manual_count") or 0)
                    d0 = int(board_res.get("device_count") or 0)
                    tp = int(board_res.get("total_pieces") or 0)
                    if tp <= 0:
                        tp = m + d0
                    elif m + d0 > 0 and m + d0 != tp:
                        m = int(round(tp * m / (m + d0)))
                        d0 = tp - m
                    _, _, slots = _sorting_biz_day_manual_device_from_records(cursor, d, window_mode)
                else:
                    m, d0, slots = _sorting_biz_day_manual_device_from_records(cursor, d, window_mode)
                    tp = m + d0
                manual.append(m)
                device.append(d0)
                total_pieces.append(tp if board_res and board_res.get("success") else m + d0)
                hour_slots.append(slots)
            else:
                import sync_daily_packing_operlog as _dp_oper

                if sync_operlog:
                    sync_res = _dp_oper.sync_daily_packing_operlog_anchor(
                        d, window_mode, force=True
                    )
                else:
                    sync_res = _dp_oper.read_daily_packing_operlog_anchor(d, window_mode)
                    if (
                        not sync_res.get('success')
                        and count_mode in ('raw', 'deduped')
                        and (os.environ.get('STATS_OPERLOG_AUTO_FILL') or '1').strip().lower()
                        not in ('0', 'false', 'no', 'off')
                    ):
                        sync_res = _dp_oper.sync_daily_packing_operlog_anchor(
                            d, window_mode, force=True
                        )
                if not sync_res.get('success'):
                    if not sync_operlog:
                        operlog_cache_misses += 1
                        operlog_unsynced_dates.append(d.strftime('%Y-%m-%d'))
                        m, dv = 0, 0
                    else:
                        operlog_sync_failures += 1
                        m, dv = _sorting_biz_day_manual_device_split(cursor, d, window_mode)
                        board_fallback_days += 1
                elif count_mode == 'raw':
                    m = int(sync_res.get('manual_raw') or 0)
                    dv = int(sync_res.get('device_raw') or 0)
                else:
                    m = int(sync_res.get('manual_dedup') or 0)
                    dv = int(sync_res.get('device_dedup') or 0)
                manual.append(m)
                device.append(dv)
                total_pieces.append(m + dv)
                hour_slots.append(0)
            d += timedelta(days=1)
        conn.close()
        resp = jsonify({
            'dates': dates,
            'manual': manual,
            'device': device,
            'total_pieces': total_pieces,
            'count_mode': count_mode,
            'sync_operlog': sync_operlog,
            'operlog_sync_failures': operlog_sync_failures,
            'operlog_cache_misses': operlog_cache_misses,
            'operlog_unsynced_dates': operlog_unsynced_dates,
            'board_fallback_days': board_fallback_days,
            'board_cache_misses': board_cache_misses,
            'sync_board': sync_board,
            'hour_slots': hour_slots if count_mode == 'board' else [],
        })
        resp.headers['Cache-Control'] = 'no-store, max-age=0'
        return resp
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/statistics/daily_packing_operlog/sync', methods=['POST'])
def api_statistics_daily_packing_operlog_sync():
    """按单个运营锚点日拉取 operatelog（逐条不去重）并写入缓存。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not check_page_permission('statistics'):
        return jsonify({'error': '无权限'}), 403

    data = request.get_json(silent=True) or {}
    raw = (data.get('date') or request.args.get('date') or '').strip()[:10]
    if not raw:
        return jsonify({'success': False, 'error': '请选择运营日'}), 400
    try:
        d = datetime.strptime(raw, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'error': 'date 格式应为 YYYY-MM-DD'}), 400

    wm = _parse_stats_window_param(
        data.get('stats_window') or request.args.get('stats_window')
    )
    import sync_daily_packing_operlog as _dp_oper

    try:
        res = _dp_oper.sync_daily_packing_operlog_anchor(d, wm, force=True)
        if not res.get('success'):
            return jsonify({
                'success': False,
                'error': res.get('error') or 'operlog 同步失败',
                'date': raw,
                'stats_window': wm,
            }), 500
        manual = int(res.get('manual_raw') or 0)
        device = int(res.get('device_raw') or 0)
        raw_rows = int(res.get('raw_rows') or 0)
        period_begin = res.get('period_begin') or ''
        period_end = res.get('period_end') or ''
        warn = None
        if raw_rows <= 0:
            warn = (
                f'窗口 {period_begin} – {period_end} 内未拉到 scan 217 记录，'
                '请核对运营日、统计窗口与 Gofo 登录；勿与看板去重口径对比。'
            )
        msg = (
            f'{raw}（{wm}）{period_begin} – {period_end}：'
            f'日志 {raw_rows} 条，逐条 人工 {manual} / 设备 {device}，合计 {manual + device}'
        )
        return jsonify({
            'success': True,
            'date': raw,
            'stats_window': wm,
            'count_mode': 'raw',
            'period_begin': period_begin,
            'period_end': period_end,
            'raw_rows': raw_rows,
            'manual': manual,
            'device': device,
            'total': manual + device,
            'manual_dedup': int(res.get('manual_dedup') or 0),
            'device_dedup': int(res.get('device_dedup') or 0),
            'classifier_ver': int(res.get('classifier_ver') or 0),
            'warning': warn,
            'message': msg,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/statistics/center_collect/sync_day', methods=['POST'])
def api_statistics_center_collect_sync_day():
    """统计页：拉取单个 LA 日历日的集包看板 popover 并入库（供干线/支线占比）。"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': '未登录'}), 401
    if not check_page_permission('statistics'):
        return jsonify({'success': False, 'error': '无权限'}), 403
    data = request.get_json(silent=True) or {}
    date_str = (data.get('date') or request.args.get('date') or '').strip()[:10]
    if not date_str:
        return jsonify({'success': False, 'error': '请提供 date (YYYY-MM-DD)'}), 400
    try:
        import sync_center_collect as _cc
        result = _cc.fetch_center_collect_day(date_str)
        ok = bool(result.get('success'))
        return jsonify({
            'success': ok,
            'date': date_str,
            'stored_rows': int(result.get('stored_rows') or 0),
            'hours_fetched': int(result.get('hours_fetched') or 0),
            'hours_tried': int(result.get('hours_tried') or 0),
            'errors': result.get('errors') or [],
            'message': (
                f"{date_str} 看板 {result.get('hours_fetched', 0)}/"
                f"{result.get('hours_tried', 0)} 小时，入库 {result.get('stored_rows', 0)} 行"
            ),
            'detail': result,
        }), (200 if ok else 500)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e), 'date': date_str}), 500


@app.route('/api/statistics/center_collect/sync_hour', methods=['POST'])
def api_statistics_center_collect_sync_hour():
    """统计页：拉取单个 LA 日历日某一整点的集包看板并入库。"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': '未登录'}), 401
    if not check_page_permission('statistics'):
        return jsonify({'success': False, 'error': '无权限'}), 403
    data = request.get_json(silent=True) or {}
    date_str = (data.get('date') or request.args.get('date') or '').strip()[:10]
    if not date_str:
        return jsonify({'success': False, 'error': '请提供 date (YYYY-MM-DD)'}), 400
    hour_raw = data.get('hour')
    try:
        h = int(hour_raw)
    except (TypeError, ValueError):
        return jsonify({'success': False, 'error': 'hour 必须是 0~23 的整数'}), 400
    if not (0 <= h <= 23):
        return jsonify({'success': False, 'error': 'hour 必须是 0~23 的整数'}), 400
    try:
        import sync_center_collect as _cc
        result = _cc.fetch_center_collect_hour(date_str, h)
        ok = bool(result.get('success'))
        return jsonify({
            'success': ok,
            'date': date_str,
            'hour': h,
            'stored_rows': int(result.get('stored_rows') or 0),
            'message': f"{date_str} {h:02d}:00 入库 {result.get('stored_rows', 0)} 行",
            'detail': result,
        }), (200 if ok else 500)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e), 'date': date_str, 'hour': h}), 500


def _gofo_collect_biz_day_trunk_branch(
    cursor, d: datetime.date, window_mode: str = 'calendar', *, count_mode: str = 'board'
) -> tuple:
    """gofo_center_collect_stats 干线/支线；raw 用 package_cnt（扫包量），board 用 waybill_cnt。"""
    clause, binds = _record_date_hour_window_sql_binds(window_mode, d)
    use_package = _parse_daily_packing_count_mode(count_mode) == 'raw'
    cnt_col = 'package_cnt' if use_package else 'waybill_cnt'
    cursor.execute(
        convert_query_placeholders(
            f"""
            SELECT COALESCE(SUM(CASE WHEN destin_type = 1 THEN {cnt_col} ELSE 0 END), 0),
                   COALESCE(SUM(CASE WHEN destin_type = 2 THEN {cnt_col} ELSE 0 END), 0)
            FROM gofo_center_collect_stats
            WHERE {clause}
            """
        ),
        binds,
    )
    r = cursor.fetchone()
    return int(_db_row_get(r, 0, 0) or 0), int(_db_row_get(r, 1, 0) or 0)


def _calendar_dates_for_stats_anchor(d: datetime.date, window_mode: str) -> list:
    out = [d.strftime('%Y-%m-%d')]
    if window_mode in ('business', 'seventeen'):
        nxt = (d + timedelta(days=1)).strftime('%Y-%m-%d')
        if nxt not in out:
            out.append(nxt)
    return out


def _maybe_resync_center_collect_for_anchor(
    d: datetime.date, window_mode: str, *, do_sync: bool
) -> None:
    if not do_sync:
        return
    try:
        import sync_center_collect as _cc
    except Exception:
        return
    for ds in _calendar_dates_for_stats_anchor(d, window_mode):
        try:
            _cc.fetch_center_collect_day(ds)
        except Exception as ex:
            print(f"[center_collect] resync {ds} failed: {ex}")


def _biz_day_operlog_raw_total(
    d: datetime.date, window_mode: str, *, force_sync: bool = False
) -> tuple:
    """operatelog scan217 逐条合计（人工+设备）；返回 (total, success)。"""
    import sync_daily_packing_operlog as _dp_oper

    if force_sync:
        res = _dp_oper.sync_daily_packing_operlog_anchor(d, window_mode, force=True)
    else:
        res = _dp_oper.read_daily_packing_operlog_anchor(d, window_mode)
    if res.get('success'):
        m = int(res.get('manual_raw') or 0)
        dv = int(res.get('device_raw') or 0)
        return m + dv, True
    return 0, False


def _collect_biz_day_trunk_branch_aligned(
    cursor,
    d: datetime.date,
    window_mode: str = 'calendar',
    count_mode: str = 'board',
    *,
    force_sync_operlog: bool = False,
) -> tuple:
    """干线/支线：看板 destin 占比；柱高合计与 daily_packing_split 同源。

    count_mode=raw：合计=operatelog 逐条，占比用 package_cnt；board=看板 waybill + sorting 合计。"""
    tr, br = _gofo_collect_biz_day_trunk_branch(
        cursor, d, window_mode, count_mode=count_mode
    )
    cm = _parse_daily_packing_count_mode(count_mode)
    if cm == 'raw':
        total_sorting, _ok = _biz_day_operlog_raw_total(
            d, window_mode, force_sync=force_sync_operlog
        )
    else:
        total_sorting = _sorting_biz_day_manual_device_total(cursor, d, window_mode)
    if total_sorting <= 0:
        return 0, 0
    g = tr + br
    if g <= 0:
        return total_sorting, 0
    tr_adj = int(round(total_sorting * tr / g))
    br_adj = total_sorting - tr_adj
    return tr_adj, br_adj


def _sorting_biz_day_manual_device_from_records(cursor, d: datetime.date, window_mode: str = 'calendar') -> tuple:
    """sorting_records 按小时累加（可能因 popover 与 overview 口径不一致而偏离看板）。"""
    clause, binds = _sorting_slot_window_sql_binds(window_mode, d)
    cursor.execute(
        convert_query_placeholders(
            f"""
            SELECT COALESCE(SUM(manual_count), 0), COALESCE(SUM(device_count), 0),
                   COALESCE(SUM(pieces), 0), COUNT(*)
            FROM sorting_records
            WHERE {clause}
            """
        ),
        binds,
    )
    r = cursor.fetchone()
    m = int(_db_row_get(r, 0, 0) or 0)
    dv = int(_db_row_get(r, 1, 0) or 0)
    slots = int(_db_row_get(r, 3, 0) or 0)
    return m, dv, slots


def _sorting_biz_day_manual_device_split(cursor, d: datetime.date, window_mode: str = 'calendar') -> tuple:
    """锚点日 D 人工/设备：优先 Gofo overview 缓存，否则 sorting_records。"""
    import sync_daily_packing_board as _dp_board

    cached = _dp_board.read_daily_packing_board_anchor(d, window_mode)
    if cached is not None:
        return int(cached["manual_count"]), int(cached["device_count"])
    m, dv, _ = _sorting_biz_day_manual_device_from_records(cursor, d, window_mode)
    return m, dv


def _sorting_biz_day_manual_device_total(cursor, d: datetime.date, window_mode: str = 'calendar') -> int:
    """与 daily_packing_split 相同：锚点日 D 的 sorting_records 人工+设备件数合计。"""
    m, dv = _sorting_biz_day_manual_device_split(cursor, d, window_mode)
    return m + dv


def _parse_center_collect_sync_flags():
    sync_operlog = (request.args.get('sync_operlog') or '').strip().lower() in (
        '1', 'true', 'yes', 'on',
    )
    sync_collect = (request.args.get('sync_collect') or '').strip().lower() in (
        '1', 'true', 'yes', 'on',
    )
    return sync_operlog, sync_collect


@app.route('/api/statistics/daily_center_collect_split', methods=['GET'])
def api_statistics_daily_center_collect_split():
    """按 stats_window 聚合干线/支线；默认 raw=operlog 逐条合计 + 看板 package 占比；可 sync_operlog/sync_collect 重拉。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not check_page_permission('statistics'):
        return jsonify({'error': '无权限'}), 403
    try:
        window_mode = _parse_stats_window_param(request.args.get('stats_window'))

        def _parse_d(val):
            if not val:
                return None
            try:
                return datetime.strptime(str(val)[:10], '%Y-%m-%d').date()
            except ValueError:
                return None

        start_in = _parse_d(request.args.get('start_date'))
        end_raw = request.args.get('end_date') or request.args.get('date')
        end_in = _parse_d(end_raw)

        if start_in and end_in and start_in <= end_in:
            span = (end_in - start_in).days + 1
            if span > 90:
                return jsonify({'error': '日期跨度不能超过 90 天'}), 400
            start_date, end_date = start_in, end_in
        else:
            days = request.args.get('days', type=int)
            if days is None or days < 1:
                days = 14
            days = min(days, 90)
            if end_in:
                end_date = end_in
            else:
                end_date = datetime.now(LA_TZ).date()
            start_date = end_date - timedelta(days=days - 1)
        count_mode = _parse_daily_packing_count_mode(
            request.args.get('count_mode') or 'raw'
        )
        sync_operlog, sync_collect = _parse_center_collect_sync_flags()
        dates = []
        d = start_date
        while d <= end_date:
            dates.append(d.strftime('%Y-%m-%d'))
            d += timedelta(days=1)

        trunk_list: list = []
        branch_list: list = []
        conn = get_db()
        cursor = conn.cursor()
        try:
            d = start_date
            while d <= end_date:
                if sync_collect:
                    _maybe_resync_center_collect_for_anchor(
                        d, window_mode, do_sync=True
                    )
                tr, br = _collect_biz_day_trunk_branch_aligned(
                    cursor,
                    d,
                    window_mode,
                    count_mode,
                    force_sync_operlog=sync_operlog,
                )
                trunk_list.append(tr)
                branch_list.append(br)
                d += timedelta(days=1)
        except Exception:
            import traceback
            traceback.print_exc()
            trunk_list = [0] * len(dates)
            branch_list = [0] * len(dates)
        finally:
            conn.close()

        return jsonify({
            'dates': dates,
            'trunk': trunk_list,
            'branch': branch_list,
            'count_mode': count_mode,
            'sync_operlog': sync_operlog,
            'sync_collect': sync_collect,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def _parse_daily_packing_count_mode(val):
    """daily_packing_split count_mode：board=看板 sorting_records（默认）；raw=operatelog 逐条；deduped=operatelog 去重。"""
    s = (val or '').strip().lower()
    if s in ('raw', 'per_log', 'log', 'norepeat', 'no_dedup'):
        return 'raw'
    if s in ('deduped', 'dedup', 'operlog_dedup', 'log_dedup'):
        return 'deduped'
    if s in ('board', 'sorting', 'overview', 'gofo'):
        return 'board'
    return 'board'


def _parse_cno_narrowbelt_count_mode(val):
    """API count_mode：raw=operatelog 逐条计数；deduped=按 (运单, scanTypeStr, 操作员) 去重后计数。"""
    s = (val or '').strip().lower()
    if s in ('deduped', 'dedup', 'd', '1', 'true', 'yes'):
        return 'deduped'
    return 'raw'


def _build_cno_narrowbelt_hourly_series(anchor_date, window_mode: str = 'calendar', count_mode: str = 'raw'):
    """窄带 A–D 各时段件数；calendar=当日00–23；business=05–次日04（与 sorting_hourly 同源）。count_mode raw|deduped。"""
    if isinstance(anchor_date, datetime):
        anchor_date = anchor_date.date()
    next_cal = anchor_date + timedelta(days=1)
    anchor_str = anchor_date.strftime('%Y-%m-%d')
    next_str = next_cal.strftime('%Y-%m-%d')
    cm = _parse_cno_narrowbelt_count_mode(count_mode)

    if window_mode == 'business':
        labels = [f"{((5 + i) % 24):02d}:00" for i in range(24)]
    elif window_mode == 'seventeen':
        labels = [f"{((17 + i) % 24):02d}:00" for i in range(24)]
    else:
        labels = [f"{i:02d}:00" for i in range(24)]
    slot_to_idx = {s: i for i, s in enumerate(labels)}
    lines = {'A': [0] * 24, 'B': [0] * 24, 'C': [0] * 24, 'D': [0] * 24}
    dedup_column_cells = 0
    dedup_fallback_cells = 0

    def _norm_time_slot(val):
        s = str(val or '').strip()
        if not s:
            return ''
        if len(s) >= 8 and s[2] == ':' and s[5] == ':':
            s = f"{int(s[:2]):02d}:{s[3:5]}"
        elif ':' in s:
            parts = s.split(':')
            try:
                s = f"{int(parts[0]):02d}:{str(parts[1])[:2]}"
            except (ValueError, IndexError):
                return ''
        if len(s) == 5 and s[2] == ':':
            return s
        return ''

    clause, binds = _record_date_slot_window_sql_binds(window_mode, anchor_date)
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            convert_query_placeholders(
                f"""
                SELECT record_date, time_slot, line_code, pieces, pieces_deduped
                FROM cno_narrowbelt_hourly
                WHERE {clause}
                ORDER BY record_date, time_slot, line_code
                """
            ),
            binds,
        )
        for row in cursor.fetchall():
            rd = _db_row_get(row, 'record_date', '') or _db_row_get(row, 0, '')
            if hasattr(rd, 'strftime'):
                rd = rd.strftime('%Y-%m-%d')
            rd = str(rd)[:10]
            slot = _norm_time_slot(
                _db_row_get(row, 'time_slot', '') or _db_row_get(row, 1, '')
            )
            line = str(
                _db_row_get(row, 'line_code', '') or _db_row_get(row, 2, '')
            ).strip().upper()
            try:
                n_raw = int(_db_row_get(row, 'pieces', 0) or _db_row_get(row, 3, 0) or 0)
            except (TypeError, ValueError):
                n_raw = 0
            pd = _db_row_get(row, 'pieces_deduped', None)
            if pd is None:
                pd = _db_row_get(row, 4, None)
            try:
                n_dedup = int(pd) if pd is not None and pd != '' else None
            except (TypeError, ValueError):
                n_dedup = None
            used_dedup_fallback = False
            if cm == 'deduped':
                # NULL：尚未写入去重列，用逐条口径避免空白图
                # pieces_deduped=0 且 pieces>0：多为未双写同步；回退到逐条
                if n_dedup is None:
                    n = n_raw
                    used_dedup_fallback = True
                elif n_raw > 0 and n_dedup == 0:
                    n = n_raw
                    used_dedup_fallback = True
                else:
                    n = n_dedup
            else:
                n = n_raw
            if line not in lines or not slot:
                continue
            if window_mode == 'business':
                if rd == anchor_str and slot >= '05:00':
                    pass
                elif rd == next_str and slot < '05:00':
                    pass
                else:
                    continue
            elif window_mode == 'seventeen':
                if rd == anchor_str and slot >= '17:00':
                    pass
                elif rd == next_str and slot < '17:00':
                    pass
                else:
                    continue
            else:
                if rd == anchor_str:
                    pass
                else:
                    continue
            idx = slot_to_idx.get(slot)
            if idx is None:
                continue
            if cm == 'deduped':
                if used_dedup_fallback:
                    dedup_fallback_cells += 1
                else:
                    dedup_column_cells += 1
            lines[line][idx] = n
    finally:
        conn.close()

    return {
        'timezone': 'server_local',
        'stats_window': window_mode,
        'count_mode': cm,
        'dedup_column_cells': dedup_column_cells if cm == 'deduped' else 0,
        'dedup_fallback_cells': dedup_fallback_cells if cm == 'deduped' else 0,
        'date': anchor_str,
        'labels': labels,
        'lines': lines,
    }


def _norm_labor_time_slot(val):
    s = str(val or '').strip()
    if not s:
        return ''
    if len(s) >= 8 and s[2] == ':' and s[5] == ':':
        s = f"{int(s[:2]):02d}:{s[3:5]}"
    elif ':' in s:
        parts = s.split(':')
        try:
            s = f"{int(parts[0]):02d}:{str(parts[1])[:2]}"
        except (ValueError, IndexError):
            return ''
    if len(s) == 5 and s[2] == ':':
        return s
    return ''


def _labor_row_in_stats_window(rd, slot, anchor_str, next_str, window_mode):
    if window_mode == 'business':
        return (rd == anchor_str and slot >= '05:00') or (
            rd == next_str and slot < '05:00'
        )
    if window_mode == 'seventeen':
        return (rd == anchor_str and slot >= '17:00') or (
            rd == next_str and slot < '17:00'
        )
    return rd == anchor_str


def la_record_slot_to_operating_anchor(record_date: str, time_slot: str, window_mode: str):
    """洛杉矶日历日 + 整点时段 → 该 stats_window 下的运营锚点日 YYYY-MM-DD。"""
    try:
        d = datetime.strptime(str(record_date)[:10], '%Y-%m-%d').date()
    except ValueError:
        return None
    slot = _norm_labor_time_slot(time_slot)
    if not slot:
        return None
    try:
        h = int(slot[:2])
    except ValueError:
        return None
    if window_mode == 'business':
        if h < 5:
            d = d - timedelta(days=1)
    elif window_mode == 'seventeen':
        if h < 17:
            d = d - timedelta(days=1)
    return d.strftime('%Y-%m-%d')


def _labor_pick_piece_count(cm, n_raw, n_dedup):
    if cm == 'deduped':
        if n_dedup is None:
            return int(n_raw or 0), True
        if int(n_raw or 0) > 0 and int(n_dedup or 0) == 0:
            return int(n_raw or 0), True
        return int(n_dedup or 0), False
    return int(n_raw or 0), False


def _labor_pay_type_for_account(company: str, account: str) -> str:
    """与 sync_cno_labor_sorter_hourly 写入规则一致；读库时用当前规则覆盖历史 pay_type。"""
    try:
        from sync_cno_labor_sorter_hourly import classify_labor_pay_type

        return classify_labor_pay_type(company or '', account or '')
    except Exception:
        return 'piece'


def _build_cno_labor_sorter_group_summary(
    anchor_date, window_mode: str = 'calendar', count_mode: str = 'raw'
):
    """按公司汇总：计件/计时组数、合计件数、有产工作时长、产能指标。"""
    if isinstance(anchor_date, datetime):
        anchor_date = anchor_date.date()
    next_cal = anchor_date + timedelta(days=1)
    anchor_str = anchor_date.strftime('%Y-%m-%d')
    next_str = next_cal.strftime('%Y-%m-%d')
    cm = _parse_cno_narrowbelt_count_mode(count_mode)

    if window_mode == 'business':
        labels = [f"{((5 + i) % 24):02d}:00" for i in range(24)]
    elif window_mode == 'seventeen':
        labels = [f"{((17 + i) % 24):02d}:00" for i in range(24)]
    else:
        labels = [f"{i:02d}:00" for i in range(24)]
    slot_to_idx = {s: i for i, s in enumerate(labels)}

    clause, binds = _record_date_slot_window_sql_binds(window_mode, anchor_date)

    # (company, account) -> pieces；company -> {piece|hourly: 24 整点件数}（计薪按当前规则从账号重算）
    account_agg = {}
    company_hourly = {}
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            convert_query_placeholders(
                f"""
                SELECT record_date, time_slot, company_code, account_label, pay_type,
                       pieces, pieces_deduped
                FROM cno_labor_sorter_account_hourly
                WHERE {clause}
                """
            ),
            binds,
        )
        for row in cursor.fetchall():
            rd = _db_row_get(row, 'record_date', '') or _db_row_get(row, 0, '')
            if hasattr(rd, 'strftime'):
                rd = rd.strftime('%Y-%m-%d')
            rd = str(rd)[:10]
            slot = _norm_labor_time_slot(
                _db_row_get(row, 'time_slot', '') or _db_row_get(row, 1, '')
            )
            company = str(
                _db_row_get(row, 'company_code', '') or _db_row_get(row, 2, '')
            ).strip()
            account = str(
                _db_row_get(row, 'account_label', '') or _db_row_get(row, 3, '')
            ).strip()
            pay_type = _labor_pay_type_for_account(company, account)
            if not company or not account or pay_type not in ('piece', 'hourly'):
                continue
            if not slot or not _labor_row_in_stats_window(
                rd, slot, anchor_str, next_str, window_mode
            ):
                continue
            try:
                n_raw = int(_db_row_get(row, 'pieces', 0) or _db_row_get(row, 5, 0) or 0)
            except (TypeError, ValueError):
                n_raw = 0
            pd = _db_row_get(row, 'pieces_deduped', None)
            if pd is None:
                pd = _db_row_get(row, 6, None)
            try:
                n_dedup = int(pd) if pd is not None and pd != '' else None
            except (TypeError, ValueError):
                n_dedup = None
            n, _ = _labor_pick_piece_count(cm, n_raw, n_dedup)
            if n <= 0:
                continue
            idx = slot_to_idx.get(slot)
            if idx is not None:
                if company not in company_hourly:
                    company_hourly[company] = {
                        'piece': [0] * 24,
                        'hourly': [0] * 24,
                    }
                company_hourly[company][pay_type][idx] += int(n or 0)
            key = (company, account)
            if key not in account_agg:
                account_agg[key] = {'pay_type': pay_type, 'pieces': 0}
            account_agg[key]['pieces'] += n
            account_agg[key]['pay_type'] = pay_type
    finally:
        conn.close()

    all_companies = set(company_hourly.keys()) | {k[0] for k in account_agg}
    companies = {}
    for company in all_companies:
        ch = company_hourly.get(company) or {
            'piece': [0] * 24,
            'hourly': [0] * 24,
        }
        piece_arr = ch.get('piece') or [0] * 24
        hourly_arr = ch.get('hourly') or [0] * 24
        combined = [
            int(piece_arr[i] or 0) + int(hourly_arr[i] or 0) for i in range(24)
        ]
        piece_group_count = sum(
            1
            for (c, _a), info in account_agg.items()
            if c == company
            and info.get('pay_type') == 'piece'
            and int(info.get('pieces') or 0) > 0
        )
        hourly_group_count = sum(
            1
            for (c, _a), info in account_agg.items()
            if c == company
            and info.get('pay_type') == 'hourly'
            and int(info.get('pieces') or 0) > 0
        )
        group_count = piece_group_count + hourly_group_count
        piece_total_pieces = int(sum(piece_arr))
        hourly_total_pieces = int(sum(hourly_arr))
        total_pieces = piece_total_pieces + hourly_total_pieces
        piece_work_hours = sum(1 for v in piece_arr if int(v or 0) > 0)
        hourly_work_hours = sum(1 for v in hourly_arr if int(v or 0) > 0)
        active_hours = sum(1 for v in combined if int(v or 0) > 0)
        if group_count <= 0 and total_pieces <= 0:
            continue
        piece_pph = (
            round(piece_total_pieces / piece_work_hours, 1)
            if piece_work_hours > 0
            else 0.0
        )
        hourly_pph = (
            round(hourly_total_pieces / hourly_work_hours, 1)
            if hourly_work_hours > 0
            else 0.0
        )
        piece_pph_group = (
            round(piece_total_pieces / piece_group_count / piece_work_hours, 1)
            if piece_group_count > 0 and piece_work_hours > 0
            else 0.0
        )
        hourly_pph_group = (
            round(hourly_total_pieces / hourly_group_count / hourly_work_hours, 1)
            if hourly_group_count > 0 and hourly_work_hours > 0
            else 0.0
        )
        pieces_per_active_hour = (
            round(total_pieces / active_hours, 1) if active_hours > 0 else 0.0
        )
        pieces_per_hour_per_group = (
            round(total_pieces / group_count / active_hours, 1)
            if group_count > 0 and active_hours > 0
            else 0.0
        )
        companies[company] = {
            'company': company,
            'group_count': group_count,
            'piece_group_count': piece_group_count,
            'hourly_group_count': hourly_group_count,
            'piece_total_pieces': piece_total_pieces,
            'hourly_total_pieces': hourly_total_pieces,
            'piece_work_hours': piece_work_hours,
            'hourly_work_hours': hourly_work_hours,
            'total_pieces': total_pieces,
            'active_hours': active_hours,
            'work_hours': active_hours,
            'piece_pieces_per_active_hour': piece_pph,
            'hourly_pieces_per_active_hour': hourly_pph,
            'piece_pieces_per_hour_per_group': piece_pph_group,
            'hourly_pieces_per_hour_per_group': hourly_pph_group,
            'pieces_per_active_hour': pieces_per_active_hour,
            'pieces_per_hour_per_group': pieces_per_hour_per_group,
            'hourly_pieces': combined,
        }

    rows = sorted(
        companies.values(),
        key=lambda x: (-(x.get('total_pieces') or 0), x['company']),
    )
    tot_pieces = sum(r['total_pieces'] for r in rows)
    tot_active = sum(r['active_hours'] for r in rows)
    tot_groups = sum(r['group_count'] for r in rows)
    tot_piece_groups = sum(r.get('piece_group_count') or 0 for r in rows)
    tot_hourly_groups = sum(r.get('hourly_group_count') or 0 for r in rows)
    tot_piece_pieces = sum(r.get('piece_total_pieces') or 0 for r in rows)
    tot_hourly_pieces = sum(r.get('hourly_total_pieces') or 0 for r in rows)
    tot_piece_hours = sum(r.get('piece_work_hours') or 0 for r in rows)
    tot_hourly_hours = sum(r.get('hourly_work_hours') or 0 for r in rows)
    totals = {
        'group_count': tot_groups,
        'piece_group_count': tot_piece_groups,
        'hourly_group_count': tot_hourly_groups,
        'piece_total_pieces': tot_piece_pieces,
        'hourly_total_pieces': tot_hourly_pieces,
        'piece_work_hours': tot_piece_hours,
        'hourly_work_hours': tot_hourly_hours,
        'total_pieces': tot_pieces,
        'active_hours': tot_active,
        'work_hours': tot_active,
        'piece_pieces_per_active_hour': (
            round(tot_piece_pieces / tot_piece_hours, 1)
            if tot_piece_hours > 0
            else 0.0
        ),
        'hourly_pieces_per_active_hour': (
            round(tot_hourly_pieces / tot_hourly_hours, 1)
            if tot_hourly_hours > 0
            else 0.0
        ),
        'piece_pieces_per_hour_per_group': (
            round(
                sum(
                    (r.get('piece_pieces_per_hour_per_group') or 0)
                    * (r.get('piece_group_count') or 0)
                    for r in rows
                )
                / tot_piece_groups,
                1,
            )
            if tot_piece_groups > 0
            else 0.0
        ),
        'hourly_pieces_per_hour_per_group': (
            round(
                sum(
                    (r.get('hourly_pieces_per_hour_per_group') or 0)
                    * (r.get('hourly_group_count') or 0)
                    for r in rows
                )
                / tot_hourly_groups,
                1,
            )
            if tot_hourly_groups > 0
            else 0.0
        ),
        'pieces_per_active_hour': (
            round(tot_pieces / tot_active, 1) if tot_active > 0 else 0.0
        ),
        'pieces_per_hour_per_group': (
            round(
                sum(
                    (r['pieces_per_hour_per_group'] or 0) * (r['group_count'] or 0)
                    for r in rows
                )
                / tot_groups,
                1,
            )
            if tot_groups > 0
            else 0.0
        ),
    }
    return {'rows': rows, 'totals': totals, 'window_slots': labels}


def _build_cno_labor_sorter_hourly_series(
    anchor_date, window_mode: str = 'calendar', count_mode: str = 'raw'
):
    """劳务公司 Sorter 分时产能（GF 分计时/计件）；series[公司]={piece:[],hourly:[]}。"""
    if isinstance(anchor_date, datetime):
        anchor_date = anchor_date.date()
    next_cal = anchor_date + timedelta(days=1)
    anchor_str = anchor_date.strftime('%Y-%m-%d')
    next_str = next_cal.strftime('%Y-%m-%d')
    cm = _parse_cno_narrowbelt_count_mode(count_mode)

    if window_mode == 'business':
        labels = [f"{((5 + i) % 24):02d}:00" for i in range(24)]
    elif window_mode == 'seventeen':
        labels = [f"{((17 + i) % 24):02d}:00" for i in range(24)]
    else:
        labels = [f"{i:02d}:00" for i in range(24)]
    slot_to_idx = {s: i for i, s in enumerate(labels)}
    series = {}
    dedup_column_cells = 0
    dedup_fallback_cells = 0

    clause, binds = _record_date_slot_window_sql_binds(window_mode, anchor_date)
    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(
            convert_query_placeholders(
                f"""
                SELECT record_date, time_slot, company_code, account_label, pay_type,
                       pieces, pieces_deduped
                FROM cno_labor_sorter_account_hourly
                WHERE {clause}
                ORDER BY record_date, time_slot, company_code, account_label
                """
            ),
            binds,
        )
        for row in cursor.fetchall():
            rd = _db_row_get(row, 'record_date', '') or _db_row_get(row, 0, '')
            if hasattr(rd, 'strftime'):
                rd = rd.strftime('%Y-%m-%d')
            rd = str(rd)[:10]
            slot = _norm_labor_time_slot(
                _db_row_get(row, 'time_slot', '') or _db_row_get(row, 1, '')
            )
            company = str(
                _db_row_get(row, 'company_code', '') or _db_row_get(row, 2, '')
            ).strip()
            account = str(
                _db_row_get(row, 'account_label', '') or _db_row_get(row, 3, '')
            ).strip()
            pay_type = _labor_pay_type_for_account(company, account)
            if pay_type not in ('piece', 'hourly'):
                continue
            try:
                n_raw = int(_db_row_get(row, 'pieces', 0) or _db_row_get(row, 5, 0) or 0)
            except (TypeError, ValueError):
                n_raw = 0
            pd = _db_row_get(row, 'pieces_deduped', None)
            if pd is None:
                pd = _db_row_get(row, 6, None)
            try:
                n_dedup = int(pd) if pd is not None and pd != '' else None
            except (TypeError, ValueError):
                n_dedup = None
            n, used_dedup_fallback = _labor_pick_piece_count(cm, n_raw, n_dedup)
            if not company or not account or not slot:
                continue
            if not _labor_row_in_stats_window(rd, slot, anchor_str, next_str, window_mode):
                continue
            idx = slot_to_idx.get(slot)
            if idx is None:
                continue
            if company not in series:
                series[company] = {'piece': [0] * 24, 'hourly': [0] * 24}
            series[company][pay_type][idx] = (
                int(series[company][pay_type][idx] or 0) + int(n or 0)
            )
            if cm == 'deduped':
                if used_dedup_fallback:
                    dedup_fallback_cells += 1
                else:
                    dedup_column_cells += 1
    finally:
        conn.close()

    companies = sorted(
        c for c, buckets in series.items()
        if any(int(v or 0) > 0 for v in (buckets.get('piece') or []))
        or any(int(v or 0) > 0 for v in (buckets.get('hourly') or []))
    )

    return {
        'timezone': 'server_local',
        'stats_window': window_mode,
        'count_mode': cm,
        'dedup_column_cells': dedup_column_cells if cm == 'deduped' else 0,
        'dedup_fallback_cells': dedup_fallback_cells if cm == 'deduped' else 0,
        'date': anchor_str,
        'labels': labels,
        'companies': companies,
        'series': series,
    }


def _backfill_cno_labor_group_hourly_from_account(cursor, anchor_date, window_mode: str) -> int:
    """从账号分时表回填小组分时表（历史数据无 group_hourly 时）。"""
    if isinstance(anchor_date, datetime):
        anchor_date = anchor_date.date()
    anchor_str = anchor_date.strftime('%Y-%m-%d')
    synced_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    clause, binds = _record_date_slot_window_sql_binds(window_mode, anchor_date)
    cursor.execute(
        convert_query_placeholders(
            f"""
            SELECT record_date, time_slot, company_code, account_label, pay_type,
                   pieces, pieces_deduped
            FROM cno_labor_sorter_account_hourly
            WHERE {clause}
            """
        ),
        binds,
    )
    n = 0
    for row in cursor.fetchall():
        rd = _db_row_get(row, 'record_date', '') or _db_row_get(row, 0, '')
        if hasattr(rd, 'strftime'):
            rd = rd.strftime('%Y-%m-%d')
        rd = str(rd)[:10]
        slot = _norm_labor_time_slot(
            _db_row_get(row, 'time_slot', '') or _db_row_get(row, 1, '')
        )
        company = str(
            _db_row_get(row, 'company_code', '') or _db_row_get(row, 2, '')
        ).strip()
        group_no = str(
            _db_row_get(row, 'account_label', '') or _db_row_get(row, 3, '')
        ).strip()
        pay_type = _labor_pay_type_for_account(company, group_no)
        if not company or not group_no or pay_type not in ('piece', 'hourly') or not slot:
            continue
        anchor2 = la_record_slot_to_operating_anchor(rd, slot, window_mode)
        if anchor2 != anchor_str:
            continue
        try:
            n_raw = int(_db_row_get(row, 'pieces', 0) or _db_row_get(row, 5, 0) or 0)
        except (TypeError, ValueError):
            n_raw = 0
        pd = _db_row_get(row, 'pieces_deduped', None)
        if pd is None:
            pd = _db_row_get(row, 6, None)
        try:
            n_ded = int(pd) if pd is not None and pd != '' else None
        except (TypeError, ValueError):
            n_ded = None
        cursor.execute(
            convert_query_placeholders(
                """
                INSERT INTO cno_labor_group_hourly
                    (anchor_date, stats_window, time_slot, company_code, group_no,
                     pay_type, pieces, pieces_deduped, record_date_la, synced_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(anchor_date, stats_window, time_slot, company_code, group_no)
                DO UPDATE SET
                    pay_type = excluded.pay_type,
                    pieces = excluded.pieces,
                    pieces_deduped = excluded.pieces_deduped,
                    record_date_la = excluded.record_date_la,
                    synced_at = excluded.synced_at
                """
            ),
            (
                anchor_str,
                window_mode,
                slot,
                company,
                group_no,
                pay_type,
                n_raw,
                n_ded,
                rd,
                synced_at,
            ),
        )
        n += 1
    return n


def _aggregate_labor_group_hourly_from_account(
    cursor, anchor_date, window_mode: str, count_mode: str, labels, slot_to_idx
):
    """从 cno_labor_sorter_account_hourly 聚合小组×整点矩阵（与劳务汇总同源，不依赖 group_hourly 表）。"""
    if isinstance(anchor_date, datetime):
        anchor_date = anchor_date.date()
    anchor_str = anchor_date.strftime('%Y-%m-%d')
    next_str = (anchor_date + timedelta(days=1)).strftime('%Y-%m-%d')
    cm = _parse_cno_narrowbelt_count_mode(count_mode)
    matrix = {}
    clause, binds = _record_date_slot_window_sql_binds(window_mode, anchor_date)
    cursor.execute(
        convert_query_placeholders(
            f"""
            SELECT record_date, time_slot, company_code, account_label, pay_type,
                   pieces, pieces_deduped
            FROM cno_labor_sorter_account_hourly
            WHERE {clause}
            """
        ),
        binds,
    )
    for row in cursor.fetchall():
        rd = _db_row_get(row, 'record_date', '') or _db_row_get(row, 0, '')
        if hasattr(rd, 'strftime'):
            rd = rd.strftime('%Y-%m-%d')
        rd = str(rd)[:10]
        slot = _norm_labor_time_slot(
            _db_row_get(row, 'time_slot', '') or _db_row_get(row, 1, '')
        )
        company = str(
            _db_row_get(row, 'company_code', '') or _db_row_get(row, 2, '')
        ).strip()
        group_no = str(
            _db_row_get(row, 'account_label', '') or _db_row_get(row, 3, '')
        ).strip()
        pay_type = _labor_pay_type_for_account(company, group_no)
        if not company or not group_no or pay_type not in ('piece', 'hourly') or not slot:
            continue
        if not _labor_row_in_stats_window(rd, slot, anchor_str, next_str, window_mode):
            continue
        try:
            n_raw = int(_db_row_get(row, 'pieces', 0) or _db_row_get(row, 5, 0) or 0)
        except (TypeError, ValueError):
            n_raw = 0
        pd = _db_row_get(row, 'pieces_deduped', None)
        if pd is None:
            pd = _db_row_get(row, 6, None)
        try:
            n_ded = int(pd) if pd is not None and pd != '' else None
        except (TypeError, ValueError):
            n_ded = None
        n, _ = _labor_pick_piece_count(cm, n_raw, n_ded)
        idx = slot_to_idx.get(slot)
        if idx is None:
            continue
        key = (company, group_no, pay_type)
        if key not in matrix:
            matrix[key] = [0] * 24
        matrix[key][idx] += int(n or 0)
    return matrix


def _build_cno_labor_group_hourly_matrix(
    anchor_date, window_mode: str = 'calendar', count_mode: str = 'raw'
):
    """公司 × 组号 × 运营日各整点件数矩阵（用于小组每小时产能表）。"""
    if isinstance(anchor_date, datetime):
        anchor_date = anchor_date.date()
    anchor_str = anchor_date.strftime('%Y-%m-%d')
    cm = _parse_cno_narrowbelt_count_mode(count_mode)

    if window_mode == 'business':
        labels = [f"{((5 + i) % 24):02d}:00" for i in range(24)]
    elif window_mode == 'seventeen':
        labels = [f"{((17 + i) % 24):02d}:00" for i in range(24)]
    else:
        labels = [f"{i:02d}:00" for i in range(24)]
    slot_to_idx = {s: i for i, s in enumerate(labels)}

    matrix = {}
    backfilled = False
    conn = get_db()
    cursor = conn.cursor()
    try:
        matrix = _aggregate_labor_group_hourly_from_account(
            cursor, anchor_date, window_mode, cm, labels, slot_to_idx
        )
        if not any(sum(v) > 0 for v in matrix.values()):
            try:
                n_ins = _backfill_cno_labor_group_hourly_from_account(
                    cursor, anchor_date, window_mode
                )
                if n_ins:
                    conn.commit()
                    backfilled = True
                    matrix = _aggregate_labor_group_hourly_from_account(
                        cursor, anchor_date, window_mode, cm, labels, slot_to_idx
                    )
            except Exception:
                pass
    finally:
        conn.close()

    rows_out = []
    for (company, group_no, pay_type), hourly in sorted(matrix.items()):
        total = int(sum(hourly))
        if total <= 0:
            continue
        rows_out.append({
            'company': company,
            'group_no': group_no,
            'pay_type': pay_type,
            'hourly': hourly,
            'total': total,
        })

    return {
        'date': anchor_str,
        'stats_window': window_mode,
        'count_mode': cm,
        'labels': labels,
        'rows': rows_out,
        'backfilled': backfilled,
    }


@app.route('/api/statistics/cno_labor_sorter_hourly', methods=['GET'])
def api_statistics_cno_labor_sorter_hourly():
    """CNO 劳务公司 Sorter 分时产能（GF 计时/计件）；与窄带相同 stats_window、count_mode。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not check_page_permission('statistics'):
        return jsonify({'error': '无权限'}), 403
    raw = request.args.get('date')
    wm = _parse_stats_window_param(request.args.get('stats_window'))
    if raw:
        try:
            anchor = datetime.strptime(str(raw)[:10], '%Y-%m-%d').date()
        except ValueError:
            anchor = _default_stats_request_date(wm)
    else:
        anchor = _default_stats_request_date(wm)

    cm = _parse_cno_narrowbelt_count_mode(request.args.get('count_mode'))
    try:
        data = _build_cno_labor_sorter_hourly_series(anchor, wm, cm)
        data['group_summary'] = _build_cno_labor_sorter_group_summary(anchor, wm, cm)
        data['group_hourly_matrix'] = _build_cno_labor_group_hourly_matrix(
            anchor, wm, cm
        )
        data['sync_plan'] = _build_cno_operlog_sync_plan(anchor, wm)
        resp = jsonify(data)
        resp.headers['Cache-Control'] = 'no-store, max-age=0'
        return resp
    except Exception as e:
        labels = (
            [f"{((17 + i) % 24):02d}:00" for i in range(24)] if wm == 'seventeen'
            else ([f"{((5 + i) % 24):02d}:00" for i in range(24)] if wm == 'business'
                  else [f"{i:02d}:00" for i in range(24)]))
        return jsonify({
            'error': str(e),
            'date': anchor.strftime('%Y-%m-%d'),
            'count_mode': cm,
            'labels': labels,
            'companies': [],
            'series': {},
            'group_hourly_matrix': {
                'date': anchor.strftime('%Y-%m-%d'),
                'stats_window': wm,
                'count_mode': cm,
                'labels': labels,
                'rows': [],
            },
        }), 500


def _la_calendar_dates_for_stats_sync(anchor_date, window_mode: str):
    """统计窗口所需的 LA 日历日列表（business/seventeen 含锚点日+次日）。"""
    if isinstance(anchor_date, str):
        anchor_date = datetime.strptime(str(anchor_date)[:10], '%Y-%m-%d').date()
    dates = [anchor_date.strftime('%Y-%m-%d')]
    if window_mode in ('business', 'seventeen'):
        nxt = (anchor_date + timedelta(days=1)).strftime('%Y-%m-%d')
        if nxt not in dates:
            dates.append(nxt)
    return dates


def _la_calendar_day_hour_slots_to_sync(record_date_str: str) -> int:
    """某 LA 日历日 sync_la_calendar_day_hours 会尝试的整点小时数。"""
    record_date_str = (record_date_str or '').strip()[:10]
    d = datetime.strptime(record_date_str, '%Y-%m-%d').date()
    now = datetime.now(LA_TZ)
    if d > now.date():
        return 0
    if d < now.date():
        return 24
    day_start = LA_TZ.localize(datetime(d.year, d.month, d.day, 0, 0, 0))
    n = 0
    for h in range(24):
        slot = day_start + timedelta(hours=h)
        if slot.strftime('%Y-%m-%d') != record_date_str:
            break
        if slot > now:
            break
        n += 1
    return n


def _build_cno_operlog_sync_plan(anchor_date, window_mode: str = 'calendar'):
    """说明手动同步将拉取多少 LA 日历日 / 整点时段（与 sync_la_calendar_day_hours 一致）。"""
    if isinstance(anchor_date, str):
        anchor_date = datetime.strptime(str(anchor_date)[:10], '%Y-%m-%d').date()
    anchor_str = anchor_date.strftime('%Y-%m-%d')
    sync_dates = _la_calendar_dates_for_stats_sync(anchor_date, window_mode)
    hours_by_date = {
        d: _la_calendar_day_hour_slots_to_sync(d) for d in sync_dates
    }
    total_hours = sum(hours_by_date.values())
    today_la = datetime.now(LA_TZ).strftime('%Y-%m-%d')

    wm_labels = {
        'calendar': '自然日 0–24h',
        'business': '5:00–次日 5:00',
        'seventeen': '17:00–次日 17:00',
    }
    date_parts = []
    for d in sync_dates:
        ht = hours_by_date.get(d, 0)
        suffix = '今日至今' if d == today_la else ('满 24h' if ht >= 24 else f'{ht}h')
        date_parts.append(f'{d}（{suffix}）')

    day_count = len(sync_dates)
    return {
        'anchor_date': anchor_str,
        'stats_window': window_mode,
        'stats_window_label': wm_labels.get(window_mode, window_mode),
        'sync_dates': sync_dates,
        'day_count': day_count,
        'hours_estimate': total_hours,
        'hours_by_date': hours_by_date,
        'includes_today_partial': today_la in sync_dates,
        'summary': f'{day_count} 个 LA 日历日、约 {total_hours} 个整点时段',
        'detail': '、'.join(date_parts),
    }


def _start_cno_operlog_hourly_sync_background(
    date_str: str,
    log_prefix: str = 'CnoOperlogStatsSync',
    extra_dates=None,
):
    """后台拉取 LA 日历日 operatelog 分时（窄带 + 劳务 Sorter 同批写入）。"""
    date_str = (date_str or '').strip()[:10]
    if not date_str:
        date_str = datetime.now(LA_TZ).strftime('%Y-%m-%d')
    dates = [date_str]
    for d in extra_dates or []:
        d = (d or '').strip()[:10]
        if d and d not in dates:
            dates.append(d)

    def _worker():
        try:
            import importlib
            import sync_cno_narrowbelt_hourly as _nb

            importlib.reload(_nb)
            for d in dates:
                result = _nb.sync_la_calendar_day_hours(d)
                print(
                    f"[{log_prefix}] date={d} success={result.get('success')} "
                    f"hours={result.get('hours_attempted')} "
                    f"errs={len(result.get('errors') or [])}"
                )
                errs = result.get('errors') or []
                if errs:
                    print(f"[{log_prefix}] first errors: {errs[:5]}")
        except Exception as e:
            print(f"[{log_prefix}] FAILED dates={dates}: {e}")
            import traceback
            traceback.print_exc()

    threading.Thread(target=_worker, daemon=True).start()
    return dates


@app.route('/api/statistics/cno_labor_sorter_hourly/sync', methods=['POST'])
def api_statistics_cno_labor_sorter_hourly_sync():
    """统计页手动同步：拉取所选 LA 日历日 operatelog 分时（窄带 + 劳务 Sorter）。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录', 'success': False}), 401
    if not check_page_permission('statistics'):
        return jsonify({'error': '无权限', 'success': False}), 403

    data = request.get_json(silent=True) or {}
    date_str = (data.get('date') or request.args.get('date') or '').strip()[:10]
    wm = _parse_stats_window_param(
        data.get('stats_window') or request.args.get('stats_window')
    )
    if not date_str:
        date_str = _default_stats_request_date(wm).strftime('%Y-%m-%d')
    try:
        anchor = datetime.strptime(date_str, '%Y-%m-%d').date()
    except ValueError:
        anchor = _default_stats_request_date(wm)
        date_str = anchor.strftime('%Y-%m-%d')

    plan = _build_cno_operlog_sync_plan(anchor, wm)
    sync_dates = plan['sync_dates']
    _start_cno_operlog_hourly_sync_background(
        sync_dates[0],
        extra_dates=sync_dates[1:],
    )
    return jsonify({
        'success': True,
        'async': True,
        'date': date_str,
        'sync_dates': sync_dates,
        'sync_plan': plan,
        'message': (
            f'已提交后台同步：{plan["summary"]}（{plan["detail"]}）。'
            f'含窄带与劳务 Sorter；数据量大时约需数分钟，页面将自动刷新图表。'
        ),
    })


@app.route('/api/statistics/cno_labor_sorter_hourly/sync_plan', methods=['GET'])
def api_statistics_cno_labor_sorter_hourly_sync_plan():
    """返回当前所选日期/统计口径下，手动同步将覆盖的 LA 日历日与时段数。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not check_page_permission('statistics'):
        return jsonify({'error': '无权限'}), 403
    raw = request.args.get('date')
    wm = _parse_stats_window_param(request.args.get('stats_window'))
    if raw:
        try:
            anchor = datetime.strptime(str(raw)[:10], '%Y-%m-%d').date()
        except ValueError:
            anchor = _default_stats_request_date(wm)
    else:
        anchor = _default_stats_request_date(wm)
    return jsonify({'success': True, 'sync_plan': _build_cno_operlog_sync_plan(anchor, wm)})


def _append_cno_labor_sorter_hourly_csv_chart_section(w, data, wm, cm, write_header=True):
    """图表口径：按运营日 × 时段 × 公司 × 计薪类型。"""
    if write_header:
        w.writerow([
            'operating_day_anchor_la',
            'stats_window',
            'count_mode',
            'time_slot_la',
            'company_code',
            'pay_type',
            'pieces',
        ])
    for i, lab in enumerate(data['labels']):
        for company in data.get('companies') or []:
            buckets = (data.get('series') or {}).get(company) or {}
            for pay_type in ('piece', 'hourly'):
                if isinstance(buckets, list):
                    arr = buckets if pay_type == 'piece' else [0] * 24
                else:
                    arr = buckets.get(pay_type) or [0] * 24
                val = arr[i] if i < len(arr) else 0
                w.writerow([
                    data['date'],
                    data.get('stats_window', wm),
                    data.get('count_mode', cm),
                    lab,
                    company,
                    pay_type,
                    val,
                ])


def _append_cno_labor_sorter_group_summary_csv(w, gs):
    w.writerow([
        'company_code',
        'piece_groups',
        'piece_work_hours',
        'piece_total_pieces',
        'piece_pcs_per_active_hour',
        'piece_pcs_per_hour_per_group',
        'hourly_groups',
        'hourly_work_hours',
        'hourly_total_pieces',
        'hourly_pcs_per_active_hour',
        'hourly_pcs_per_hour_per_group',
    ])
    for row in gs.get('rows') or []:
        w.writerow([
            row.get('company'),
            row.get('piece_group_count', row.get('group_count')),
            row.get('piece_work_hours', 0),
            row.get('piece_total_pieces', 0),
            row.get('piece_pieces_per_active_hour', 0),
            row.get('piece_pieces_per_hour_per_group', 0),
            row.get('hourly_group_count', 0),
            row.get('hourly_work_hours', 0),
            row.get('hourly_total_pieces', 0),
            row.get('hourly_pieces_per_active_hour', 0),
            row.get('hourly_pieces_per_hour_per_group', 0),
        ])
    tot = gs.get('totals') or {}
    if tot:
        w.writerow([
            'TOTAL',
            tot.get('piece_group_count', tot.get('group_count')),
            tot.get('piece_work_hours', 0),
            tot.get('piece_total_pieces', 0),
            tot.get('piece_pieces_per_active_hour', 0),
            tot.get('piece_pieces_per_hour_per_group', 0),
            tot.get('hourly_group_count', 0),
            tot.get('hourly_work_hours', 0),
            tot.get('hourly_total_pieces', 0),
            tot.get('hourly_pieces_per_active_hour', 0),
            tot.get('hourly_pieces_per_hour_per_group', 0),
        ])


def _append_cno_labor_group_hourly_matrix_csv(w, matrix, write_header=True):
    """宽表：运营锚点日 × 公司 × 组号 × 各整点件数（与统计页矩阵一致）。"""
    labels = matrix.get('labels') or []
    if write_header:
        w.writerow(
            [
                'operating_day_anchor_la',
                'stats_window',
                'count_mode',
                'company_code',
                'group_no',
                'pay_type',
                'day_total',
            ]
            + list(labels)
        )
    anchor = matrix.get('date', '')
    sw = matrix.get('stats_window', '')
    cm = matrix.get('count_mode', 'raw')
    nlab = len(labels)
    for row in matrix.get('rows') or []:
        hourly = row.get('hourly') or []
        vals = [
            int(hourly[i]) if i < len(hourly) else 0 for i in range(nlab)
        ]
        w.writerow(
            [
                anchor,
                sw,
                cm,
                row.get('company'),
                row.get('group_no'),
                row.get('pay_type'),
                int(row.get('total') or 0),
            ]
            + vals
        )


def _append_cno_labor_sorter_account_slot_csv(
    w, cursor, start_date, end_date, wm, cm
):
    """库内原始：运营日 × 日历日 × 时段 × 账号。"""
    w.writerow([
        'operating_day_anchor_la',
        'stats_window',
        'count_mode',
        'record_date_la',
        'time_slot_la',
        'company_code',
        'account_label',
        'pay_type',
        'pieces',
    ])
    d0 = datetime.strptime(str(start_date)[:10], '%Y-%m-%d').date()
    d1 = datetime.strptime(str(end_date)[:10], '%Y-%m-%d').date()
    cur = d0
    while cur <= d1:
        anchor_str = cur.strftime('%Y-%m-%d')
        next_str = (cur + timedelta(days=1)).strftime('%Y-%m-%d')
        clause, binds = _record_date_slot_window_sql_binds(wm, cur)
        cursor.execute(
            convert_query_placeholders(
                f"""
                SELECT record_date, time_slot, company_code, account_label, pay_type,
                       pieces, pieces_deduped
                FROM cno_labor_sorter_account_hourly
                WHERE {clause}
                ORDER BY record_date, time_slot, company_code, account_label
                """
            ),
            binds,
        )
        for row in cursor.fetchall():
            rd = _db_row_get(row, 'record_date', '') or _db_row_get(row, 0, '')
            if hasattr(rd, 'strftime'):
                rd = rd.strftime('%Y-%m-%d')
            rd = str(rd)[:10]
            slot = _norm_labor_time_slot(
                _db_row_get(row, 'time_slot', '') or _db_row_get(row, 1, '')
            )
            if not slot or not _labor_row_in_stats_window(
                rd, slot, anchor_str, next_str, wm
            ):
                continue
            company = str(
                _db_row_get(row, 'company_code', '') or _db_row_get(row, 2, '')
            ).strip()
            account = str(
                _db_row_get(row, 'account_label', '') or _db_row_get(row, 3, '')
            ).strip()
            pay_type = str(
                _db_row_get(row, 'pay_type', '') or _db_row_get(row, 4, '')
            ).strip().lower()
            if pay_type not in ('piece', 'hourly'):
                continue
            try:
                n_raw = int(_db_row_get(row, 'pieces', 0) or _db_row_get(row, 5, 0) or 0)
            except (TypeError, ValueError):
                n_raw = 0
            pd = _db_row_get(row, 'pieces_deduped', None)
            if pd is None:
                pd = _db_row_get(row, 6, None)
            try:
                n_dedup = int(pd) if pd is not None and pd != '' else None
            except (TypeError, ValueError):
                n_dedup = None
            n, _ = _labor_pick_piece_count(cm, n_raw, n_dedup)
            w.writerow([
                anchor_str,
                wm,
                cm,
                rd,
                slot,
                company,
                account,
                pay_type,
                int(n or 0),
            ])
        cur += timedelta(days=1)


@app.route('/api/statistics/cno_labor_sorter_hourly/export', methods=['GET'])
def api_statistics_cno_labor_sorter_hourly_export():
    """导出劳务 Sorter 分时 CSV；口径与图表一致。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not check_page_permission('statistics'):
        return jsonify({'error': '无权限'}), 403
    raw = request.args.get('date')
    wm = _parse_stats_window_param(request.args.get('stats_window'))
    if raw:
        try:
            anchor = datetime.strptime(str(raw)[:10], '%Y-%m-%d').date()
        except ValueError:
            anchor = _default_stats_request_date(wm)
    else:
        anchor = _default_stats_request_date(wm)

    cm = _parse_cno_narrowbelt_count_mode(request.args.get('count_mode'))
    try:
        data = _build_cno_labor_sorter_hourly_series(anchor, wm, cm)
        data['group_summary'] = _build_cno_labor_sorter_group_summary(anchor, wm, cm)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

    buf = io.StringIO()
    w = csv.writer(buf)
    _append_cno_labor_sorter_hourly_csv_chart_section(w, data, wm, cm)
    w.writerow([])
    _append_cno_labor_sorter_group_summary_csv(w, data.get('group_summary') or {})
    w.writerow([])
    w.writerow(['# section', 'group_hourly_matrix'])
    try:
        matrix = _build_cno_labor_group_hourly_matrix(anchor, wm, cm)
        _append_cno_labor_group_hourly_matrix_csv(w, matrix, write_header=True)
    except Exception:
        w.writerow(
            [
                'operating_day_anchor_la',
                'stats_window',
                'count_mode',
                'company_code',
                'group_no',
                'pay_type',
                'day_total',
            ]
        )

    fn = f"cno_labor_sorter_hourly_{data['date']}_{data.get('count_mode', 'raw')}.csv"
    return Response(
        buf.getvalue().encode('utf-8-sig'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename={fn}'},
    )


@app.route('/api/statistics/cno_labor_sorter_hourly/export_range', methods=['GET'])
def api_statistics_cno_labor_sorter_hourly_export_range():
    """按 LA 运营日区间导出劳务 Sorter 分时（图表口径 + 日汇总 + 小组分时矩阵 + 账号时段明细）。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not check_page_permission('statistics'):
        return jsonify({'error': '无权限'}), 403

    start_raw = (request.args.get('start_date') or '').strip()[:10]
    end_raw = (request.args.get('end_date') or '').strip()[:10]
    if not start_raw or not end_raw:
        return jsonify({'success': False, 'error': '请提供 start_date 与 end_date'}), 400
    try:
        start_d = datetime.strptime(start_raw, '%Y-%m-%d').date()
        end_d = datetime.strptime(end_raw, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'success': False, 'error': '日期格式应为 YYYY-MM-DD'}), 400
    if end_d < start_d:
        return jsonify({'success': False, 'error': '结束日期不能早于开始日期'}), 400
    if (end_d - start_d).days > 62:
        return jsonify({'success': False, 'error': '区间最多 62 天'}), 400

    wm = _parse_stats_window_param(request.args.get('stats_window'))
    cm = _parse_cno_narrowbelt_count_mode(request.args.get('count_mode'))

    buf = io.StringIO()
    w = csv.writer(buf)
    conn = get_db()
    cursor = conn.cursor()
    try:
        w.writerow(['# section', 'hourly_by_company_slot'])
        cur = start_d
        while cur <= end_d:
            data = _build_cno_labor_sorter_hourly_series(cur, wm, cm)
            _append_cno_labor_sorter_hourly_csv_chart_section(
                w, data, wm, cm, write_header=(cur == start_d)
            )
            cur += timedelta(days=1)

        w.writerow([])
        w.writerow(['# section', 'daily_group_summary'])
        w.writerow(['operating_day_anchor_la'] + [
            'company_code',
            'piece_groups',
            'piece_work_hours',
            'piece_total_pieces',
            'hourly_groups',
            'hourly_work_hours',
            'hourly_total_pieces',
        ])
        cur = start_d
        while cur <= end_d:
            gs = _build_cno_labor_sorter_group_summary(cur, wm, cm)
            anchor_str = cur.strftime('%Y-%m-%d')
            for row in gs.get('rows') or []:
                w.writerow([
                    anchor_str,
                    row.get('company'),
                    row.get('piece_group_count', 0),
                    row.get('piece_work_hours', 0),
                    row.get('piece_total_pieces', 0),
                    row.get('hourly_group_count', 0),
                    row.get('hourly_work_hours', 0),
                    row.get('hourly_total_pieces', 0),
                ])
            cur += timedelta(days=1)

        w.writerow([])
        w.writerow(['# section', 'group_hourly_matrix'])
        cur = start_d
        first_matrix = True
        while cur <= end_d:
            matrix = _build_cno_labor_group_hourly_matrix(cur, wm, cm)
            _append_cno_labor_group_hourly_matrix_csv(
                w, matrix, write_header=first_matrix
            )
            first_matrix = False
            cur += timedelta(days=1)

        w.writerow([])
        w.writerow(['# section', 'account_slot_detail'])
        _append_cno_labor_sorter_account_slot_csv(
            w, cursor, start_raw, end_raw, wm, cm
        )
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()

    fn = (
        f"cno_labor_sorter_range_{start_raw}_{end_raw}_"
        f"{wm}_{cm}.csv"
    )
    return Response(
        buf.getvalue().encode('utf-8-sig'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename={fn}'},
    )


@app.route('/api/statistics/cno_labor_group_hourly', methods=['GET'])
def api_statistics_cno_labor_group_hourly():
    """小组 × 小时产能矩阵（热力图分享页专用）。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not check_page_permission('statistics'):
        return jsonify({'error': '无权限'}), 403
    raw = request.args.get('date')
    wm = _parse_stats_window_param(request.args.get('stats_window'))
    if raw:
        try:
            anchor = datetime.strptime(str(raw)[:10], '%Y-%m-%d').date()
        except ValueError:
            anchor = _default_stats_request_date(wm)
    else:
        anchor = _default_stats_request_date(wm)
    cm = _parse_cno_narrowbelt_count_mode(request.args.get('count_mode'))
    try:
        matrix = _build_cno_labor_group_hourly_matrix(anchor, wm, cm)
        resp = jsonify(matrix)
        resp.headers['Cache-Control'] = 'no-store, max-age=0'
        return resp
    except Exception as e:
        labels = (
            [f"{((17 + i) % 24):02d}:00" for i in range(24)] if wm == 'seventeen'
            else ([f"{((5 + i) % 24):02d}:00" for i in range(24)] if wm == 'business'
                  else [f"{i:02d}:00" for i in range(24)]))
        return jsonify({
            'error': str(e),
            'date': anchor.strftime('%Y-%m-%d'),
            'stats_window': wm,
            'count_mode': cm,
            'labels': labels,
            'rows': [],
        }), 500


@app.route('/api/statistics/cno_labor_group_hourly/export', methods=['GET'])
def api_statistics_cno_labor_group_hourly_export():
    """仅导出各小组每小时产能矩阵 CSV（与统计页表格列一致）。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not check_page_permission('statistics'):
        return jsonify({'error': '无权限'}), 403
    raw = request.args.get('date')
    wm = _parse_stats_window_param(request.args.get('stats_window'))
    if raw:
        try:
            anchor = datetime.strptime(str(raw)[:10], '%Y-%m-%d').date()
        except ValueError:
            anchor = _default_stats_request_date(wm)
    else:
        anchor = _default_stats_request_date(wm)
    cm = _parse_cno_narrowbelt_count_mode(request.args.get('count_mode'))
    try:
        matrix = _build_cno_labor_group_hourly_matrix(anchor, wm, cm)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(['# section', 'group_hourly_matrix'])
    _append_cno_labor_group_hourly_matrix_csv(w, matrix, write_header=True)
    fn = (
        f"cno_labor_group_hourly_{matrix.get('date', anchor.strftime('%Y-%m-%d'))}_"
        f"{wm}_{cm}.csv"
    )
    return Response(
        buf.getvalue().encode('utf-8-sig'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': f'attachment; filename={fn}'},
    )


@app.route('/api/statistics/cno_labor_group_hourly/feishu_sync', methods=['POST'])
def api_statistics_cno_labor_group_hourly_feishu_sync():
    """统计页手动同步：各小组每小时产能矩阵 → 飞书电子表格「元数据」工作表。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录', 'success': False}), 401
    if not check_page_permission('statistics'):
        return jsonify({'error': '无权限', 'success': False}), 403

    data = request.get_json(silent=True) or {}
    wm = _parse_stats_window_param(
        data.get('stats_window') or request.args.get('stats_window')
    )
    cm = _parse_cno_narrowbelt_count_mode(
        data.get('count_mode') or request.args.get('count_mode')
    )
    raw = (data.get('date') or request.args.get('date') or '').strip()[:10]
    if raw:
        try:
            anchor = datetime.strptime(raw, '%Y-%m-%d').date()
        except ValueError:
            anchor = _default_stats_request_date(wm)
    else:
        anchor = _default_stats_request_date(wm)

    try:
        info = feishu_sync_cno_labor_group_hourly_sheet_once(
            stats_window=wm, count_mode=cm, anchor_date=anchor
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

    return jsonify({
        'success': True,
        'operating_day': anchor.strftime('%Y-%m-%d'),
        'date': anchor.strftime('%Y-%m-%d'),
        'stats_window': wm,
        'count_mode': cm,
        'data_rows': info.get('data_rows'),
        'matrix_rows': info.get('matrix_rows'),
        'appended_today': info.get('appended_today'),
        'updated_range': info.get('updated_range'),
        'last_synced_at': datetime.now(LA_TZ).strftime('%Y-%m-%d %H:%M:%S'),
        'detail': info,
        'message': (
            f'已同步至飞书：运营日 {anchor.strftime("%Y-%m-%d")}，'
            f'本运营日 {info.get("matrix_rows") or 0} 个小组'
            f'（表内合计 {info.get("data_rows") or 0} 行数据）。'
        ),
    })


@app.route('/api/statistics/cno_labor_group_hourly/feishu_sync_status', methods=['GET'])
def api_statistics_cno_labor_group_hourly_feishu_sync_status():
    """返回飞书小组分时表最近一次成功同步时间。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not check_page_permission('statistics'):
        return jsonify({'error': '无权限'}), 403
    cfg_key = 'feishu_wiki_cno_labor_group_hourly_meta_last_synced_at'
    conn = get_db()
    cursor = conn.cursor()
    last = _get_system_config_value(cursor, cfg_key, '')
    conn.close()
    return jsonify({'success': True, 'last_synced_at': last or ''})


@app.route('/api/statistics/cno_narrowbelt_hourly', methods=['GET'])
def api_statistics_cno_narrowbelt_hourly():
    """CNO 直线窄带分拣机分时；与 sorting_hourly 相同 stats_window。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not check_page_permission('statistics'):
        return jsonify({'error': '无权限'}), 403
    raw = request.args.get('date')
    wm = _parse_stats_window_param(request.args.get('stats_window'))
    if raw:
        try:
            anchor = datetime.strptime(str(raw)[:10], '%Y-%m-%d').date()
        except ValueError:
            anchor = _default_stats_request_date(wm)
    else:
        anchor = _default_stats_request_date(wm)

    cm = _parse_cno_narrowbelt_count_mode(request.args.get('count_mode'))
    try:
        data = _build_cno_narrowbelt_hourly_series(anchor, wm, cm)
        resp = jsonify(data)
        resp.headers['Cache-Control'] = 'no-store, max-age=0'
        return resp
    except Exception as e:
        labels = (
            [f"{((17 + i) % 24):02d}:00" for i in range(24)] if wm == 'seventeen'
            else ([f"{((5 + i) % 24):02d}:00" for i in range(24)] if wm == 'business'
                  else [f"{i:02d}:00" for i in range(24)]))
        return jsonify({
            'error': str(e),
            'date': anchor.strftime('%Y-%m-%d'),
            'count_mode': _parse_cno_narrowbelt_count_mode(request.args.get('count_mode')),
            'labels': labels,
            'lines': {'A': [0] * 24, 'B': [0] * 24, 'C': [0] * 24, 'D': [0] * 24},
        }), 500


@app.route('/api/statistics/cno_narrowbelt_hourly/export', methods=['GET'])
def api_statistics_cno_narrowbelt_hourly_export():
    """导出窄带分时 CSV；口径与图表一致。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not check_page_permission('statistics'):
        return jsonify({'error': '无权限'}), 403
    raw = request.args.get('date')
    wm = _parse_stats_window_param(request.args.get('stats_window'))
    if raw:
        try:
            anchor = datetime.strptime(str(raw)[:10], '%Y-%m-%d').date()
        except ValueError:
            anchor = _default_stats_request_date(wm)
    else:
        anchor = _default_stats_request_date(wm)

    cm = _parse_cno_narrowbelt_count_mode(request.args.get('count_mode'))
    try:
        data = _build_cno_narrowbelt_hourly_series(anchor, wm, cm)
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([
        'operating_day_anchor_la',
        'time_slot_la',
        'line_a_pieces',
        'line_b_pieces',
        'line_c_pieces',
        'line_d_pieces',
    ])
    for i, lab in enumerate(data['labels']):
        w.writerow([
            data['date'],
            lab,
            data['lines']['A'][i],
            data['lines']['B'][i],
            data['lines']['C'][i],
            data['lines']['D'][i],
        ])

    fn = f"cno_narrowbelt_hourly_{data['date']}_{data.get('count_mode', 'raw')}.csv"
    return Response(
        buf.getvalue().encode('utf-8-sig'),
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': f'attachment; filename={fn}',
        },
    )


@app.route('/api/statistics/center_collect_week_comparison', methods=['GET'])
def api_statistics_center_collect_week_comparison():
    """自然周干线/支线堆叠；默认 raw 逐条 operlog + package 占比；与 daily_center_collect_split 同源。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not check_page_permission('statistics'):
        return jsonify({'error': '无权限'}), 403
    try:
        conn = get_db()
        cursor = conn.cursor()
        window_mode = _parse_stats_window_param(request.args.get('stats_window'))
        count_mode = _parse_daily_packing_count_mode(
            request.args.get('count_mode') or 'raw'
        )
        sync_operlog, sync_collect = _parse_center_collect_sync_flags()
        min_str = None
        try:
            cursor.execute(
                convert_query_placeholders(
                    "SELECT MIN(record_date) FROM gofo_center_collect_stats"
                )
            )
            mr = cursor.fetchone()
            min_str = _db_row_get(mr, 0, None) if mr else None
        except Exception:
            min_str = None

        if not min_str:
            conn.close()
            return jsonify([])

        try:
            min_date = datetime.strptime(str(min_str)[:10], '%Y-%m-%d').date()
        except ValueError:
            conn.close()
            return jsonify([])

        max_date = datetime.now(LA_TZ).date()
        end_raw = request.args.get('end_date') or request.args.get('date')
        if end_raw:
            try:
                ed = datetime.strptime(str(end_raw)[:10], '%Y-%m-%d').date()
                max_date = min(max_date, ed)
            except ValueError:
                pass

        def get_natural_week_range(date):
            weekday = date.weekday()
            week_start = date - timedelta(days=weekday)
            week_end = week_start + timedelta(days=6)
            return week_start, week_end

        data_first_monday, _ = get_natural_week_range(min_date)
        last_monday_to_include, _ = get_natural_week_range(max_date)
        loop_start = data_first_monday
        loop_end_monday = last_monday_to_include
        week_start_raw = request.args.get('week_start')
        week_end_raw = request.args.get('week_end')
        if week_start_raw:
            try:
                ws = datetime.strptime(str(week_start_raw)[:10], '%Y-%m-%d').date()
                wm, _ = get_natural_week_range(ws)
                loop_start = max(loop_start, wm)
            except ValueError:
                pass
        if week_end_raw:
            try:
                we = datetime.strptime(str(week_end_raw)[:10], '%Y-%m-%d').date()
                wm, _ = get_natural_week_range(we)
                loop_end_monday = min(loop_end_monday, wm)
            except ValueError:
                pass
        if loop_start > loop_end_monday:
            conn.close()
            return jsonify([])
        week_span = (loop_end_monday - loop_start).days // 7 + 1
        if week_span > 104:
            conn.close()
            return jsonify({'error': '周数区间过长（最多 104 周）'}), 400

        current_start = loop_start
        current_end = current_start + timedelta(days=6)
        weeks_data = []

        while current_start <= loop_end_monday:
            daily_data = []
            week_total_trunk = 0
            week_total_branch = 0

            for day_offset in range(7):
                current_day = current_start + timedelta(days=day_offset)

                if current_day > max_date:
                    daily_data.append({
                        'date': current_day.strftime('%Y-%m-%d'),
                        'weekday': current_day.weekday(),
                        'trunk': 0,
                        'branch': 0,
                        'total': 0,
                    })
                    continue

                if sync_collect:
                    _maybe_resync_center_collect_for_anchor(
                        current_day, window_mode, do_sync=True
                    )
                tr, br = _collect_biz_day_trunk_branch_aligned(
                    cursor,
                    current_day,
                    window_mode,
                    count_mode,
                    force_sync_operlog=sync_operlog,
                )
                tot = tr + br
                week_total_trunk += tr
                week_total_branch += br
                daily_data.append({
                    'date': current_day.strftime('%Y-%m-%d'),
                    'weekday': current_day.weekday(),
                    'trunk': tr,
                    'branch': br,
                    'total': tot,
                })

            week_total = week_total_trunk + week_total_branch
            total_change_percent = 0.0
            if weeks_data:
                last_tot = int(weeks_data[-1].get('week_total') or 0)
                if last_tot > 0:
                    total_change_percent = ((week_total - last_tot) / last_tot) * 100
                else:
                    total_change_percent = 100.0 if week_total > 0 else 0.0

            weeks_data.append({
                'week_label': f"{current_start.strftime('%m/%d')}-{current_end.strftime('%m/%d')}",
                'start_date': current_start.strftime('%Y-%m-%d'),
                'end_date': current_end.strftime('%Y-%m-%d'),
                'daily_data': daily_data,
                'week_total_trunk': week_total_trunk,
                'week_total_branch': week_total_branch,
                'week_total': week_total,
                'total_change_percent': round(total_change_percent, 2),
            })

            current_start = current_start + timedelta(days=7)
            current_end = current_end + timedelta(days=7)

        conn.close()

        if weeks_data and len(weeks_data) > 0:
            first_week_start = datetime.strptime(weeks_data[0]['start_date'], '%Y-%m-%d').date()
            if first_week_start < min_date:
                weeks_data = weeks_data[1:]

        if weeks_data and len(weeks_data) > 0:
            weeks_data[0]['total_change_percent'] = 0

        return jsonify({
            'weeks': weeks_data,
            'count_mode': count_mode,
            'sync_operlog': sync_operlog,
            'sync_collect': sync_collect,
        })

    except Exception as e:
        if 'conn' in locals():
            try:
                conn.close()
            except Exception:
                pass
        return jsonify({'error': f'获取干线支线周环比出错: {str(e)}'}), 500


@app.route('/api/statistics/packing_manual_device_week_comparison', methods=['GET'])
def api_statistics_packing_manual_device_week_comparison():
    """自然周（周一至周日）人工/设备堆叠；每日为 operatelog scan217 逐条（不去重），读 daily_packing_operlog_daily；周环比为相对上一完整自然周总件数。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not check_page_permission('statistics'):
        return jsonify({'error': '无权限'}), 403
    try:
        import sync_daily_packing_operlog as _dp_oper

        conn = get_db()
        cursor = conn.cursor()
        window_mode = _parse_stats_window_param(request.args.get('stats_window'))
        min_str = None
        try:
            cursor.execute(
                convert_query_placeholders(
                    "SELECT MIN(anchor_date) FROM daily_packing_operlog_daily WHERE stats_window = ?"
                ),
                (window_mode,),
            )
            mr = cursor.fetchone()
            min_str = _db_row_get(mr, 0, None) if mr else None
        except Exception:
            min_str = None

        if not min_str:
            conn.close()
            return jsonify([])

        try:
            min_date = datetime.strptime(str(min_str)[:10], '%Y-%m-%d').date()
        except ValueError:
            conn.close()
            return jsonify([])

        max_date = datetime.now(LA_TZ).date()
        end_raw = request.args.get('end_date') or request.args.get('date')
        if end_raw:
            try:
                ed = datetime.strptime(str(end_raw)[:10], '%Y-%m-%d').date()
                max_date = min(max_date, ed)
            except ValueError:
                pass

        def get_natural_week_range(date):
            weekday = date.weekday()
            week_start = date - timedelta(days=weekday)
            week_end = week_start + timedelta(days=6)
            return week_start, week_end

        data_first_monday, _ = get_natural_week_range(min_date)
        last_monday_to_include, _ = get_natural_week_range(max_date)
        loop_start = data_first_monday
        loop_end_monday = last_monday_to_include
        week_start_raw = request.args.get('week_start')
        week_end_raw = request.args.get('week_end')
        if week_start_raw:
            try:
                ws = datetime.strptime(str(week_start_raw)[:10], '%Y-%m-%d').date()
                wm, _ = get_natural_week_range(ws)
                loop_start = max(loop_start, wm)
            except ValueError:
                pass
        if week_end_raw:
            try:
                we = datetime.strptime(str(week_end_raw)[:10], '%Y-%m-%d').date()
                wm, _ = get_natural_week_range(we)
                loop_end_monday = min(loop_end_monday, wm)
            except ValueError:
                pass
        if loop_start > loop_end_monday:
            conn.close()
            return jsonify([])
        week_span = (loop_end_monday - loop_start).days // 7 + 1
        if week_span > 104:
            conn.close()
            return jsonify({'error': '周数区间过长（最多 104 周）'}), 400

        current_start = loop_start
        current_end = current_start + timedelta(days=6)
        weeks_data = []

        while current_start <= loop_end_monday:
            daily_data = []
            week_total_manual = 0
            week_total_device = 0

            for day_offset in range(7):
                current_day = current_start + timedelta(days=day_offset)

                if current_day > max_date:
                    daily_data.append({
                        'date': current_day.strftime('%Y-%m-%d'),
                        'weekday': current_day.weekday(),
                        'manual': 0,
                        'device': 0,
                        'total': 0,
                    })
                    continue

                sync_res = _dp_oper.read_daily_packing_operlog_anchor(
                    current_day, window_mode
                )
                if sync_res.get('success'):
                    mn = int(sync_res.get('manual_raw') or 0)
                    dv = int(sync_res.get('device_raw') or 0)
                else:
                    mn, dv = 0, 0
                tot = mn + dv
                week_total_manual += mn
                week_total_device += dv
                daily_data.append({
                    'date': current_day.strftime('%Y-%m-%d'),
                    'weekday': current_day.weekday(),
                    'manual': mn,
                    'device': dv,
                    'total': tot,
                })

            week_total = week_total_manual + week_total_device
            total_change_percent = 0.0
            if weeks_data:
                last_tot = int(weeks_data[-1].get('week_total') or 0)
                if last_tot > 0:
                    total_change_percent = ((week_total - last_tot) / last_tot) * 100
                else:
                    total_change_percent = 100.0 if week_total > 0 else 0.0

            weeks_data.append({
                'week_label': f"{current_start.strftime('%m/%d')}-{current_end.strftime('%m/%d')}",
                'start_date': current_start.strftime('%Y-%m-%d'),
                'end_date': current_end.strftime('%Y-%m-%d'),
                'daily_data': daily_data,
                'week_total_manual': week_total_manual,
                'week_total_device': week_total_device,
                'week_total': week_total,
                'total_change_percent': round(total_change_percent, 2),
            })

            current_start = current_start + timedelta(days=7)
            current_end = current_end + timedelta(days=7)

        conn.close()

        if weeks_data and len(weeks_data) > 0:
            first_week_start = datetime.strptime(weeks_data[0]['start_date'], '%Y-%m-%d').date()
            if first_week_start < min_date:
                weeks_data = weeks_data[1:]

        if weeks_data and len(weeks_data) > 0:
            weeks_data[0]['total_change_percent'] = 0

        return jsonify(weeks_data)

    except Exception as e:
        if 'conn' in locals():
            try:
                conn.close()
            except Exception:
                pass
        return jsonify({'error': f'获取人工设备集包周环比出错: {str(e)}'}), 500


@app.route('/consumables')
def consumables():
    # 检查用户权限
    if 'user_id' not in session:
        return redirect('/login')
    
    if not check_page_permission('consumables'):
        return redirect('/no_permission')
    
    # 返回耗材进销存页面
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'consumables.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404

# [已迁移] /cost_accounting 页面已移至 /operations_metrics
@app.route('/cost_accounting')
def cost_accounting():
    return redirect('/operations_metrics')

def _save_sorting_shift_rows(cursor, config_id, data):
    """将人工多班次与分拣机按日排班写入结构化表（与 config_json 一致）。"""
    if not config_id or not isinstance(data, dict):
        return
    ph = get_placeholder()
    manual = data.get('manual') or {}
    machine = data.get('machine') or {}

    shifts_by_day = manual.get('shiftsByDay')
    if isinstance(shifts_by_day, list):
        for wd, shifts in enumerate(shifts_by_day):
            if wd > 6:
                break
            if not isinstance(shifts, list):
                continue
            for order, s in enumerate(shifts):
                if not isinstance(s, dict):
                    continue
                try:
                    start_h = float(s.get('start', 17.5))
                    duration_h = float(s.get('hours', 0))
                    people = int(s.get('people', 0))
                except (TypeError, ValueError):
                    continue
                if duration_h <= 0:
                    continue
                cursor.execute(
                    f"""INSERT INTO sorting_manual_shift (config_id, weekday, shift_order, start_h, duration_h, people)
                        VALUES ({ph}, {ph}, {ph}, {ph}, {ph}, {ph})""",
                    (config_id, wd, order, start_h, duration_h, people),
                )

    sched = machine.get('schedule') or [0] * 7
    starts = machine.get('startTimes') or [17.5] * 7
    try:
        hps = float(machine.get('hoursPerShift', 6))
    except (TypeError, ValueError):
        hps = 6.0
    try:
        cap = float(machine.get('capacity', 4500))
    except (TypeError, ValueError):
        cap = 4500.0
    for wd in range(7):
        try:
            lanes = int(sched[wd]) if wd < len(sched) else 0
        except (TypeError, ValueError):
            lanes = 0
        try:
            st = float(starts[wd]) if wd < len(starts) else 17.5
        except (TypeError, ValueError):
            st = 17.5
        cursor.execute(
            f"""INSERT INTO sorting_machine_day (config_id, weekday, lanes, start_h, hours_per_shift, capacity_per_lane)
                VALUES ({ph}, {ph}, {ph}, {ph}, {ph}, {ph})""",
            (config_id, wd, lanes, st, hps, cap),
        )


@app.route('/api/sorting-schedule', methods=['GET'])
def api_get_sorting_schedule():
    """获取分拣排班配置"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # 获取最新的配置
        if USE_POSTGRES:
             cursor.execute("SELECT config_json FROM sorting_schedule_config ORDER BY id DESC LIMIT 1")
        else:
             cursor.execute("SELECT config_json FROM sorting_schedule_config ORDER BY id DESC LIMIT 1")
             
        result = cursor.fetchone()
        
        if result:
            raw = result['config_json'] if USE_POSTGRES else result[0]
            cfg = _parse_config_json_from_db(raw, {})
            return jsonify(cfg)
        else:
            return jsonify({})
    except Exception as e:
        print(f"Error getting sorting schedule: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()

@app.route('/api/sorting-schedule', methods=['POST'])
def api_save_sorting_schedule():
    """保存分拣排班配置"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    # Check edit permission
    if not check_user_permission('sorting-schedule', 'edit'):
        return jsonify({'error': '无编辑权限'}), 403
        
    data = request.json
    if not data:
        return jsonify({'error': '无数据'}), 400
        
    conn = get_db()
    cursor = conn.cursor()
    
    try:
        # 保存配置
        config_json = json.dumps(data)
        current_time = datetime.now(pytz.timezone('America/Los_Angeles')).strftime('%Y-%m-%d %H:%M:%S')
        placeholder = get_placeholder()

        if USE_POSTGRES:
            cursor.execute(
                f"INSERT INTO sorting_schedule_config (config_json, updated_at) VALUES ({placeholder}, {placeholder}) RETURNING id",
                (config_json, current_time),
            )
            row = cursor.fetchone()
            config_id = row['id'] if row else None
        else:
            cursor.execute(
                f"INSERT INTO sorting_schedule_config (config_json, updated_at) VALUES ({placeholder}, {placeholder})",
                (config_json, current_time),
            )
            config_id = cursor.lastrowid

        _save_sorting_shift_rows(cursor, config_id, data)
        conn.commit()

        return jsonify({'success': True, 'config_id': config_id})
    except Exception as e:
        conn.rollback()
        print(f"Error saving sorting schedule: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


def _v2_format_start_hhmm(val):
    """小数小时 -> HH:MM"""
    st = _v2_parse_start_hours(val, 17.5)
    h = int(st)
    m = int(round((st - h) * 60))
    if m >= 60:
        h += 1
        m -= 60
    if h >= 24:
        h = 23
        m = 59
    return f"{h:02d}:{m:02d}"


def _default_v2_plan_date_la():
    """产能排班针对次日班次；未传日期时与前端日期控件默认一致（洛杉矶日历次日）。"""
    return (datetime.now(LA_TZ) + timedelta(days=1)).strftime('%Y-%m-%d')


def _build_v2_schedule_report(data, plan_date=None):
    """生成 V2 文本报告：当日逐条排班 + 汇总（与前端一览表一致）。"""
    plan = data or {}
    resources = plan.get('resources') if isinstance(plan, dict) else []
    if not isinstance(resources, list):
        resources = []
    # 无全局班次时长：按每单元 8 小时估算班次总产能（与前端一致）
    default_shift_h = 8.0
    try:
        default_shift_h = float(plan.get('defaultShiftHours', plan.get('shiftHours', 8)))
    except (TypeError, ValueError, AttributeError):
        default_shift_h = 8.0
    if default_shift_h <= 0:
        default_shift_h = 8.0

    eq_cap = 3100
    man_cap_per_person = 310
    company_set = set()
    total_headcount = 0
    hourly_total = 0
    active_lines = 0
    active_stations = 0

    lines_out = []
    stations_out = []
    company_start_times = {}

    for r in resources:
        if not isinstance(r, dict):
            continue
        r_type = str(r.get('type', '') or '').strip()
        name = str(r.get('name', '') or '').strip() or ('资源' + str(r.get('id', '')))
        company = str(r.get('company', '') or '').strip()
        tstr = _v2_format_start_hhmm(r.get('startTime'))
        try:
            hc = int(float(r.get('headcount', 0)))
        except (TypeError, ValueError):
            hc = 0
        hc = max(0, hc)

        if not company:
            # 文本报告不列出未激活点位
            continue

        company_start_times.setdefault(company, set()).add(tstr)
        company_set.add(company)
        if r_type == 'line':
            active_lines += 1
            hourly_total += eq_cap
            total_headcount += hc
            lines_out.append(
                f"  {name}  开工 {tstr}  {company}  人数 {hc}"
            )
        else:
            active_stations += 1
            total_headcount += hc
            hr_cap = hc * man_cap_per_person
            hourly_total += hr_cap
            stations_out.append(
                f"  {name}  开工 {tstr}  {company}  人数 {hc}"
            )

    total_capacity = int(round(hourly_total * default_shift_h))
    date_text = str(plan_date or _default_v2_plan_date_la())

    lines_section = '\n'.join(lines_out) if lines_out else '  （本段无激活点位）'
    stations_section = '\n'.join(stations_out) if stations_out else '  （本段无激活点位）'

    if company_start_times:
        company_time_lines = []
        for comp in sorted(company_start_times.keys()):
            times_sorted = sorted(company_start_times[comp])
            company_time_lines.append(f"  {comp}：{'、'.join(times_sorted)}")
        company_times_section = '\n'.join(company_time_lines)
    else:
        company_times_section = '  （无激活劳务公司）'

    summary_extra = ''
    if not company_set:
        summary_extra = '\n提示：当前无任何点位选择劳务公司，汇总产能为 0；请在卡片中选择劳务公司以激活排班。'

    return (
        "【分拣中心产能排班】\n"
        f"排班日期：{date_text}\n"
        f"估算：预计班次总产能按每单元 {default_shift_h:g} 小时折算。\n"
        "\n—— 设备分拣线 ——\n"
        f"{lines_section}\n"
        "\n—— 人工分拣台 ——\n"
        f"{stations_section}\n"
        "\n—— 各公司开工时间 ——\n"
        f"{company_times_section}\n"
        "\n—— 汇总 ——\n"
        f"劳务公司个数：{len(company_set)}\n"
        f"启用生产线 / 分拣台：{active_lines} / {active_stations}\n"
        f"总人数（产线+分拣台）：{total_headcount}\n"
        f"预计班次总产能：{total_capacity:,} 件\n"
        f"{summary_extra}"
    ).rstrip()


@app.route('/api/sorting-schedule/v2/daily', methods=['GET'])
def api_get_sorting_schedule_v2_daily():
    """按天查询 V2 排班配置。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401

    plan_date = (request.args.get('date') or '').strip()
    if not plan_date:
        plan_date = _default_v2_plan_date_la()

    conn = get_db()
    cursor = conn.cursor()
    ph = get_placeholder()
    try:
        cursor.execute(
            f"SELECT plan_json, report_text, updated_at FROM sorting_schedule_daily_plan WHERE plan_date = {ph} LIMIT 1",
            (plan_date,),
        )
        row = cursor.fetchone()
        if not row:
            return jsonify({'success': True, 'date': plan_date, 'data': None, 'report_text': ''})

        if hasattr(row, 'keys'):
            raw = row['plan_json']
            report_text = row['report_text'] or ''
            updated_at = row['updated_at']
        else:
            raw = row[0]
            report_text = row[1] or ''
            updated_at = row[2]

        data = _parse_config_json_from_db(raw, {})
        # 历史数据或旧版本可能未写入 report_text，按 plan 即时生成
        rt = str(report_text or '').strip()
        if not rt and isinstance(data, dict) and data.get('resources'):
            report_text = _build_v2_schedule_report(data, plan_date=plan_date)
        else:
            report_text = str(report_text or '')

        return jsonify({
            'success': True,
            'date': plan_date,
            'data': data,
            'report_text': report_text,
            'updated_at': str(updated_at) if updated_at is not None else None,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/sorting-schedule/v2/daily', methods=['POST'])
def api_save_sorting_schedule_v2_daily():
    """按天保存 V2 排班配置，并生成文本报告。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not check_user_permission('sorting-schedule', 'edit'):
        return jsonify({'error': '无编辑权限'}), 403

    payload = request.json or {}
    plan_date = str(payload.get('date') or '').strip()
    if not plan_date:
        plan_date = _default_v2_plan_date_la()
    plan = payload.get('plan') if isinstance(payload, dict) else {}
    if not isinstance(plan, dict):
        return jsonify({'error': '无效的排班数据'}), 400

    generated_report_text = _build_v2_schedule_report(plan, plan_date=plan_date)
    manual_report_text = str(payload.get('report_text') or '').strip()
    report_text = manual_report_text if manual_report_text else generated_report_text
    plan_json = json.dumps(plan, ensure_ascii=False)
    now_str = datetime.now(LA_TZ).strftime('%Y-%m-%d %H:%M:%S')

    conn = get_db()
    cursor = conn.cursor()
    ph = get_placeholder()
    try:
        if USE_POSTGRES:
            cursor.execute(
                f"""
                INSERT INTO sorting_schedule_daily_plan (plan_date, plan_json, report_text, updated_at)
                VALUES ({ph}, {ph}, {ph}, {ph})
                ON CONFLICT (plan_date) DO UPDATE
                SET plan_json = EXCLUDED.plan_json,
                    report_text = EXCLUDED.report_text,
                    updated_at = EXCLUDED.updated_at
                """,
                (plan_date, plan_json, report_text, now_str),
            )
        else:
            cursor.execute(
                f"""
                INSERT INTO sorting_schedule_daily_plan (plan_date, plan_json, report_text, updated_at)
                VALUES ({ph}, {ph}, {ph}, {ph})
                ON CONFLICT(plan_date) DO UPDATE SET
                    plan_json = excluded.plan_json,
                    report_text = excluded.report_text,
                    updated_at = excluded.updated_at
                """,
                (plan_date, plan_json, report_text, now_str),
            )
        conn.commit()
        return jsonify({
            'success': True,
            'date': plan_date,
            'report_text': report_text,
            'generated_report_text': generated_report_text,
            'manual_report_used': bool(manual_report_text),
        })
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/users', methods=['GET'])
def get_users():
    """获取所有用户列表 - 仅限管理员"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    # 检查是否为管理员
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
    result = cursor.fetchone()
    
    current_role = result['role'] if USE_POSTGRES else result[0]
    if current_role not in ('admin', 'boss'):
        conn.close()
        return jsonify({'error': '无权访问'}), 403
    
    # 获取所有用户
    cursor.execute("SELECT id, username, role, is_active, created_at FROM users ORDER BY id")
    users = cursor.fetchall()
    conn.close()
    
    user_list = []
    for user in users:
        user_list.append({
            'id': user['id'] if USE_POSTGRES else user[0],
            'username': user['username'] if USE_POSTGRES else user[1],
            'role': user['role'] if USE_POSTGRES else user[2],
            'is_active': user['is_active'] if USE_POSTGRES else user[3],
            'created_at': user['created_at'] if USE_POSTGRES else user[4]
        })
    
    return jsonify(user_list)

@app.route('/api/users/role', methods=['POST'])
def update_user_role():
    """更新用户角色 - 仅限管理员"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    # 检查是否为管理员
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
    result = cursor.fetchone()
    
    current_role = result['role'] if USE_POSTGRES else result[0]
    if current_role not in ('admin', 'boss'):
        conn.close()
        return jsonify({'error': '无权访问'}), 403
        
    data = request.json
    target_user_id = data.get('user_id')
    new_role = data.get('role')
    
    if not target_user_id or not new_role:
        return jsonify({'error': '参数不完整'}), 400
        
    if new_role not in ['admin', 'user', 'viewer']:
        return jsonify({'error': '无效的角色类型'}), 400
        
    try:
        # 更新用户角色
        cursor.execute("UPDATE users SET role = ? WHERE id = ?", (new_role, target_user_id))
        
        # 如果是管理员，赋予所有权限
        if new_role == 'admin':
            pages = ['index', 'sorting', 'history', 'statistics', 'logs', 'sorting-schedule', 'outbound-stats']
            for page in pages:
                # 检查权限是否存在
                cursor.execute("SELECT id FROM user_permissions WHERE user_id = ? AND page_name = ?", (target_user_id, page))
                if cursor.fetchone():
                    cursor.execute("UPDATE user_permissions SET can_view=1, can_edit=1, can_delete=1 WHERE user_id = ? AND page_name = ?", (target_user_id, page))
                else:
                    cursor.execute("INSERT INTO user_permissions (user_id, page_name, can_view, can_edit, can_delete) VALUES (?, ?, 1, 1, 1)", (target_user_id, page))
        elif new_role == 'viewer':
             # 视图者只能查看
            pages = ['index', 'sorting', 'history', 'statistics', 'logs', 'sorting-schedule', 'outbound-stats']
            for page in pages:
                cursor.execute("SELECT id FROM user_permissions WHERE user_id = ? AND page_name = ?", (target_user_id, page))
                if cursor.fetchone():
                    cursor.execute("UPDATE user_permissions SET can_view=1, can_edit=0, can_delete=0 WHERE user_id = ? AND page_name = ?", (target_user_id, page))
                else:
                    cursor.execute("INSERT INTO user_permissions (user_id, page_name, can_view, can_edit, can_delete) VALUES (?, ?, 1, 0, 0)", (target_user_id, page))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500

@app.route('/react-dashboard')
def react_dashboard():
    """React统计仪表板 - 使用React最佳实践构建"""
    try:
        # 暂时跳过权限检查以排除登录问题干扰调试
        # if 'user_id' not in session:
        #     return redirect('/login')
        
        # if not check_page_permission('statistics'):
        #     return redirect('/no_permission')
        
        print(f"DEBUG: Processing /react-dashboard request from {request.remote_addr}")
        
        # 返回React仪表板页面
        static_dir = get_static_dir()
        file_path = os.path.join(static_dir, 'react-dashboard', 'dist', 'index.html')
        print(f"DEBUG: Serving dashboard file from: {file_path}")
        
        if os.path.exists(file_path):
            # 使用send_file更加健壮，并禁用缓存以方便调试
            response = send_file(file_path, mimetype='text/html')
            response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
            response.headers['Pragma'] = 'no-cache'
            response.headers['Expires'] = '0'
            return response
        else:
            print(f"ERROR: File not found at {file_path}")
            return f"Dashboard file not found at: {file_path}", 404
            
    except Exception as e:
        print(f"ERROR: Exception in react_dashboard: {e}")
        import traceback
        traceback.print_exc()
        return f"Server Error: {str(e)}", 500

@app.route('/api/outbound/stats', methods=['GET'])
def get_outbound_stats():
    """获取出库统计数据（按路线汇总）"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        return jsonify({'error': '缺少日期参数'}), 400
    
    conn = get_db()
    cursor = conn.cursor()
    placeholder = get_placeholder()
    
    try:
        cursor.execute(f"""
            SELECT route_code, route_type, 
                   SUM(vehicle_count) as total_vehicles, 
                   SUM(cost) as total_cost,
                   COUNT(*) as record_count
            FROM outbound_records
            WHERE record_date BETWEEN {placeholder} AND {placeholder}
            GROUP BY route_code, route_type
            ORDER BY route_type, route_code
        """, (start_date, end_date))
        
        records = cursor.fetchall()
        result = []
        
        for record in records:
            result.append({
                'route_code': record['route_code'] if USE_POSTGRES else record[0],
                'route_type': record['route_type'] if USE_POSTGRES else record[1],
                'total_vehicles': record['total_vehicles'] if USE_POSTGRES else record[2],
                'total_cost': record['total_cost'] if USE_POSTGRES else record[3],
                'record_count': record['record_count'] if USE_POSTGRES else record[4]
            })
        
        return jsonify(result)
    except Exception as e:
        print(f"Error getting outbound stats: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        conn.close()


def _tms_shuttle_la_calendar_date() -> str:
    """与 DMS 一致的洛杉矶日历日（短驳业务日）。"""
    la = pytz.timezone('America/Los_Angeles')
    return datetime.now(la).strftime('%Y-%m-%d')


def _tms_shuttle_sync_date(day_str: str) -> dict:
    """按指定 YYYY-MM-DD 执行 sync_day（LA 业务日或用户在页面上选择的日期）。"""
    import importlib.util
    base_dir = os.path.dirname(os.path.abspath(__file__))
    script_path = os.path.join(base_dir, 'scripts', 'sync_tms_shuttle_completed.py')
    if not os.path.exists(script_path):
        return {'success': False, 'error': f'sync 脚本不存在: {script_path}'}
    spec = importlib.util.spec_from_file_location('sync_tms_shuttle_completed', script_path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.sync_day(day_str)


def _tms_shuttle_pivot_use_business_axis() -> bool:
    """默认 False：统计 pivot 按自然日 00:00～23:59、列 00～23 点。
    设 GOFO_TMS_SHUTTLE_PIVOT_BUSINESS_DAY=1 则与短驳同步业务窗（默认 05:00～次日 04:59）及业务列顺序一致。"""
    return (os.environ.get('GOFO_TMS_SHUTTLE_PIVOT_BUSINESS_DAY') or '').strip().lower() in (
        '1', 'true', 'yes', 'on',
    )


def _tms_shuttle_pivot_natural_calendar_bounds(day_str: str):
    """所选日历日 00:00:00～23:59:59（与时钟 0～23 点列一致）。"""
    d = datetime.strptime(day_str.strip(), '%Y-%m-%d').date()
    lo = datetime.combine(d, dtime.min)
    hi = datetime.combine(d, dtime(23, 59, 59))
    return lo, hi


def _tms_shuttle_pivot_candidate_record_dates(date_str: str):
    """自然日 pivot：业务日 sync 下「日历 D」的发车可能在 record_date=D 或 D-1，两行都拉。"""
    d = datetime.strptime(date_str.strip(), '%Y-%m-%d').date()
    prev = (d - timedelta(days=1)).strftime('%Y-%m-%d')
    return date_str, prev


def _tms_shuttle_pivot_day_start_hour() -> int:
    """与 sync_tms_shuttle_completed 一致：默认 5→05:00～次日 04:59:59。"""
    raw = (os.environ.get('GOFO_TMS_SHUTTLE_DAY_START_HOUR') or '5').strip().lower()
    if raw in ('0', 'calendar', 'midnight'):
        return 0
    try:
        h = int(raw)
    except ValueError:
        h = 5
    return max(0, min(23, h))


def _tms_shuttle_pivot_window_bounds(day_str: str):
    d = datetime.strptime(day_str.strip(), '%Y-%m-%d').date()
    h0 = _tms_shuttle_pivot_day_start_hour()
    if h0 <= 0:
        lo = datetime.combine(d, dtime.min)
        hi = datetime.combine(d, dtime(23, 59, 59))
        return lo, hi
    lo = datetime.combine(d, dtime(h0, 0, 0))
    hi = lo + timedelta(days=1) - timedelta(seconds=1)
    return lo, hi


def _tms_shuttle_pivot_normalize_date_iso(d_val):
    if d_val is None:
        return None
    ds = str(d_val).strip().replace('/', '-').split()[0]
    if not ds:
        return None
    for fmt in ('%Y-%m-%d', '%m-%d-%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(ds, fmt).date().strftime('%Y-%m-%d')
        except ValueError:
            continue
    return None


def _tms_shuttle_normalize_date_time_cols(d_val, t_val):
    """兼容拆分表里整段 datetime 误存在日期列、时间列为空的旧数据（ISO T / 空格 / 毫秒）。"""
    if d_val is None:
        return None, t_val
    d_s = str(d_val).strip()
    t_s = (str(t_val).strip() if t_val is not None else '')
    if t_s:
        return d_val, t_val
    u = d_s.replace('/', '-')
    if len(u) >= 11 and u[10] in 'Tt':
        u = u[:10] + ' ' + u[11:]
    if '.' in u:
        u = u.split('.')[0].strip()
    parts = u.split(None, 1)
    if len(parts) == 2 and ':' in parts[1]:
        return parts[0], parts[1]
    return d_val, t_val


def _tms_shuttle_pivot_parse_depart_dt(d_val, t_val):
    """拆分表 actual_departure_date + actual_departure_time → 本地 naive datetime。"""
    d_val, t_val = _tms_shuttle_normalize_date_time_cols(d_val, t_val)
    if d_val is None or t_val is None:
        return None
    ds = str(d_val).strip().replace('/', '-')
    ts = str(t_val).strip()
    if not ds or not ts:
        return None
    if '.' in ts:
        ts = ts.split('.')[0].strip()
    pairs = []
    for df in ('%Y-%m-%d', '%m-%d-%Y', '%d-%m-%Y'):
        for tf in ('%H:%M:%S', '%H:%M'):
            pairs.append((df, tf))
    cand_dts = [ds]
    if ' ' in ds:
        cand_dts.append(ds.split()[0])
    cand_ts = [ts]
    if len(ts) >= 8 and ts.count(':') >= 1:
        cand_ts.append(ts[:8])
    for dpart in cand_dts:
        for tpart in cand_ts:
            for df, tf in pairs:
                try:
                    return datetime.strptime(f'{dpart} {tpart}', f'{df} {tf}')
                except ValueError:
                    continue
    return None


def _tms_shuttle_pivot_business_hour_labels(h0: int):
    """24 列：自 h0 起顺排，如 5→ 05:00…23:00, 00:00…04:00。"""
    if h0 <= 0:
        return [f'{h:02d}:00' for h in range(24)]
    return [f'{(h0 + i) % 24:02d}:00' for i in range(24)]


def _tms_shuttle_pivot_business_hour_index(full_hour: int, h0: int):
    if h0 <= 0:
        return full_hour if 0 <= full_hour <= 23 else None
    if full_hour >= h0:
        return full_hour - h0
    return full_hour + (24 - h0)


def _check_tms_shuttle_api_permission() -> bool:
    """统计页短驳 pivot / outbound-stats 立即同步共用。"""
    return bool(
        check_page_permission('outbound-stats')
        or check_page_permission('statistics')
    )


@app.route('/api/tms/shuttle-completed/pivot', methods=['GET'])
def get_tms_shuttle_pivot():
    """短驳 pivot：默认自然日 00:00～23:59，列 00～23 点；可选业务窗见 GOFO_TMS_SHUTTLE_PIVOT_BUSINESS_DAY。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not _check_tms_shuttle_api_permission():
        return jsonify({'error': '无权限'}), 403

    date_str = (request.args.get('date') or '').strip()
    if not date_str:
        date_str = _tms_shuttle_la_calendar_date()
    try:
        datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': '日期格式无效，应为 YYYY-MM-DD'}), 400

    sw_arg = request.args.get('stats_window')
    if sw_arg is not None and str(sw_arg).strip() != '':
        wm = _parse_stats_window_param(sw_arg)
    else:
        wm = 'business' if _tms_shuttle_pivot_use_business_axis() else 'calendar'

    if wm == 'business':
        h0 = 5
    elif wm == 'seventeen':
        h0 = 17
    else:
        h0 = 0

    hours = _tms_shuttle_pivot_business_hour_labels(h0)
    d_parse = datetime.strptime(date_str, '%Y-%m-%d').date()
    win_lo, win_hi_excl = _stats_period_bounds(d_parse, wm)
    win_hi = win_hi_excl - timedelta(seconds=1)
    use_shift_axis = wm in ('business', 'seventeen')
    if wm == 'business':
        where_sql = "record_date = ?"
        date_params = (date_str,)
    elif wm == 'seventeen':
        d_next = (d_parse + timedelta(days=1)).strftime('%Y-%m-%d')
        where_sql = "record_date IN (?, ?)"
        date_params = (date_str, d_next)
    else:
        d0, d1 = _tms_shuttle_pivot_candidate_record_dates(date_str)
        where_sql = "record_date IN (?, ?)"
        date_params = (d0, d1)

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(convert_query_placeholders(
            f"""
            SELECT destination, actual_departure_date, actual_departure_time,
                   actual_arrival_date, actual_arrival_time, record_date
            FROM gofo_tms_shuttle_split
            WHERE {where_sql}
            """
        ), date_params)
        rows = cursor.fetchall()
    except Exception as e:
        msg = str(e)
        if 'no such table' in msg.lower() or 'undefinedtable' in msg.lower() or 'does not exist' in msg.lower():
            return jsonify({
                'date': date_str, 'hours': hours,
                'destinations': [], 'grand_total': 0,
                'note': '表 gofo_tms_shuttle_split 不存在，请先运行同步',
                'pivot_mode': 'natural' if wm == 'calendar' else wm,
                'day_start_hour': h0,
                'departure_window': {
                    'start': win_lo.strftime('%Y-%m-%d %H:%M:%S'),
                    'end': win_hi.strftime('%Y-%m-%d %H:%M:%S'),
                },
            })
        print(f"Error get_tms_shuttle_pivot: {e}")
        return jsonify({'error': msg}), 500
    finally:
        try:
            conn.close()
        except Exception:
            pass

    def _hour_from_clock(s):
        if not s:
            return None
        parts = str(s).strip().split(':')
        if not parts or not str(parts[0]).isdigit():
            return None
        hi = int(parts[0])
        return hi if 0 <= hi <= 23 else None

    matrix = {}
    dest_orphans = {}
    skipped_outside = 0

    for r in rows:
        try:
            dest = r['destination']
            dep_d = r['actual_departure_date']
            dep_t = r['actual_departure_time']
            arr_d = r['actual_arrival_date']
            arr_t = r['actual_arrival_time']
            rec_date = r['record_date']
        except Exception:
            dest = r[0]
            dep_d = r[1]
            dep_t = r[2]
            arr_d = r[3] if len(r) > 3 else None
            arr_t = r[4] if len(r) > 4 else None
            rec_date = r[5] if len(r) > 5 else None

        dest = (dest or '').strip() or '未知'
        dt_dep = _tms_shuttle_pivot_parse_depart_dt(dep_d, dep_t)
        dt_use = dt_dep
        if dt_use is None and arr_d and arr_t:
            dt_use = _tms_shuttle_pivot_parse_depart_dt(arr_d, arr_t)

        hh_idx = None
        if dt_use is not None:
            if dt_use < win_lo or dt_use > win_hi:
                skipped_outside += 1
                continue
            if use_shift_axis:
                hh_idx = _tms_shuttle_pivot_business_hour_index(dt_use.hour, h0)
            else:
                hh_idx = dt_use.hour
        else:
            if not use_shift_axis:
                dnorm = _tms_shuttle_pivot_normalize_date_iso(dep_d) or _tms_shuttle_pivot_normalize_date_iso(arr_d)
                if dnorm is not None:
                    if dnorm != date_str:
                        continue
                else:
                    if rec_date != date_str:
                        continue

            t = dep_t if (dep_t and str(dep_t).strip()) else arr_t
            clock_h = _hour_from_clock(t) if t else None
            if clock_h is None:
                dest_orphans[dest] = dest_orphans.get(dest, 0) + 1
                continue
            hh_idx = _tms_shuttle_pivot_business_hour_index(clock_h, h0) if use_shift_axis else clock_h

        bucket = matrix.setdefault(dest, [0] * 24)
        bucket[hh_idx] += 1

    all_names = set(matrix.keys()) | set(dest_orphans.keys())
    total_orphans = sum(dest_orphans.values())
    if total_orphans > 0:
        hours.append('无时间')
        for name in all_names:
            hourly = matrix.setdefault(name, [0] * 24)
            ox = dest_orphans.get(name, 0)
            hourly.append(ox)

    out_rows = []
    for name in all_names:
        hourly = matrix.get(name, [0] * 24)
        out_rows.append({'name': name, 'hourly': hourly, 'total': sum(hourly)})
    out_rows.sort(key=lambda x: (-x['total'], x['name']))
    grand = sum(r['total'] for r in out_rows)
    return jsonify({
        'date': date_str,
        'hours': hours,
        'destinations': out_rows,
        'grand_total': grand,
        'pivot_mode': 'natural' if wm == 'calendar' else wm,
        'day_start_hour': h0,
        'departure_window': {
            'start': win_lo.strftime('%Y-%m-%d %H:%M:%S'),
            'end': win_hi.strftime('%Y-%m-%d %H:%M:%S'),
        },
        'skipped_outside_window': skipped_outside,
    })


@app.route('/api/tms/shuttle-completed/sync', methods=['POST'])
def post_tms_shuttle_sync():
    """触发 TMS 短驳同步。Body/Query 可选 date=YYYY-MM-DD；省略则用洛杉矶当日（与 DMS 一致）。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not _check_tms_shuttle_api_permission():
        return jsonify({'error': '无权限'}), 403
    body = request.get_json(silent=True) or {}
    day = (body.get('date') or request.args.get('date') or '').strip()
    if not day:
        day = _tms_shuttle_la_calendar_date()
    else:
        try:
            datetime.strptime(day, '%Y-%m-%d')
        except ValueError:
            return jsonify({'error': '日期格式无效，应为 YYYY-MM-DD'}), 400
    try:
        res = _tms_shuttle_sync_date(day)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    if not res.get('success'):
        return jsonify(res), 500
    return jsonify({
        'success': True,
        'date': res.get('date'),
        'fetched': res.get('fetched'),
        'fetched_before_row_filter': res.get('fetched_before_row_filter'),
        'fetched_before_calendar_filter': res.get('fetched_before_calendar_filter'),
        'departure_window_filter': res.get('departure_window_filter'),
        'stored': res.get('stored'),
        'total_expected': res.get('total_expected'),
        'actual_departure_window': res.get('actual_departure_window'),
        'day_start_hour': res.get('day_start_hour'),
    })


@app.route('/api/tms/shuttle-completed', methods=['GET'])
def get_tms_shuttle_completed():
    """读取 gofo_tms_shuttle_split 表（始发地/目的地 + 拆分后的实际发车/到车日期与时间）。

    数据由 scripts/sync_tms_shuttle_completed.py 同步：
      - gofo_tms_shuttle_completed       字段全集（含 raw_json）
      - gofo_tms_shuttle_split           本接口使用的展示子集（日期/时间已拆分入库）
    """
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if not _check_tms_shuttle_api_permission():
        return jsonify({'error': '无权限'}), 403

    date_str = (request.args.get('date') or '').strip()
    if not date_str:
        date_str = _tms_shuttle_la_calendar_date()
    try:
        datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': '日期格式无效，应为 YYYY-MM-DD'}), 400

    conn = get_db()
    cursor = conn.cursor()
    try:
        cursor.execute(convert_query_placeholders(
            """
            SELECT task_no, place_of_origin, destination,
                   actual_departure_date, actual_departure_time,
                   actual_arrival_date,   actual_arrival_time
            FROM gofo_tms_shuttle_split
            WHERE record_date = ?
            ORDER BY (actual_departure_date IS NULL),
                     actual_departure_date,
                     (actual_departure_time IS NULL),
                     actual_departure_time
            """
        ), (date_str,))
        rows = cursor.fetchall()
    except Exception as e:
        msg = str(e)
        if 'no such table' in msg.lower() or 'undefinedtable' in msg.lower() or 'does not exist' in msg.lower():
            return jsonify({
                'success': True,
                'date': date_str,
                'rows': [],
                'note': '表 gofo_tms_shuttle_split 不存在，请先运行 scripts/sync_tms_shuttle_completed.py'
            })
        print(f"Error get_tms_shuttle_completed: {e}")
        return jsonify({'error': msg}), 500
    finally:
        try:
            conn.close()
        except Exception:
            pass

    def _v(row, key, idx):
        try:
            return row[key]
        except Exception:
            return row[idx]

    out = []
    for r in rows:
        dd = _v(r, 'actual_departure_date', 3)
        dt = _v(r, 'actual_departure_time', 4)
        ad = _v(r, 'actual_arrival_date', 5)
        at = _v(r, 'actual_arrival_time', 6)
        dd, dt = _tms_shuttle_normalize_date_time_cols(dd, dt)
        ad, at = _tms_shuttle_normalize_date_time_cols(ad, at)
        out.append({
            'task_no': _v(r, 'task_no', 0),
            'place_of_origin': _v(r, 'place_of_origin', 1),
            'destination': _v(r, 'destination', 2),
            'actual_departure_date': dd,
            'actual_departure_time_only': dt,
            'actual_arrival_date': ad,
            'actual_arrival_time_only': at,
        })

    return jsonify({'success': True, 'date': date_str, 'rows': out})


@app.route('/tabler-dashboard')
def tabler_dashboard():
    return render_template('tabler_dashboard.html')

@app.route('/shadcn')
def shadcn_dashboard():
    # Reuse the React app index.html, handled by client-side routing
    path = os.path.join('static', 'react-dashboard', 'dist', 'index.html')
    if os.path.exists(path):
        with open(path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {
            'Content-Type': 'text/html; charset=utf-8',
            'Cache-Control': 'no-store, no-cache, must-revalidate, max-age=0',
            'Pragma': 'no-cache',
            'Expires': '0'
        }
    return "React build not found. Please run 'npm run build' in static/react-dashboard.", 404

@app.route('/ping')
def ping():
    """健康检查接口"""
    return "pong", 200


@app.route('/api/app_identity')
def api_app_identity():
    """返回应用指纹、版权与商业使用声明（含制作人 Fan Yang）。"""
    return jsonify(identity_dict()), 200


@app.route("/api/license_status")
def api_license_status():
    """许可证配置与在线校验结果（不强制要求 LICENSE_ENFORCE）。"""
    try:
        from license_client import license_status

        return jsonify(license_status()), 200
    except Exception as e:
        return jsonify({"ok": False, "error": str(e)}), 500


def _admin_license_guard():
    if 'user_id' not in session:
        return jsonify({
            'success': False,
            'error': '未登录',
            'hint': '请先用管理员账号登录主系统（/login），再打开 /admin 激活',
        }), 401
    if session.get('role') not in ('admin', 'boss'):
        return jsonify({'success': False, 'error': '权限不足'}), 403
    return None


@app.route('/api/admin/license/status', methods=['GET'])
def admin_license_status():
    """管理员查看商业授权详情（含指纹，不含 token）。"""
    err = _admin_license_guard()
    if err is not None:
        return err
    try:
        from license_client import license_status

        st = license_status()
        st['server_url_configured'] = bool(
            (os.environ.get('LICENSE_SERVER_URL') or '').strip()
        )
        return jsonify({'success': True, 'data': st}), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/license/activate', methods=['POST'])
def admin_license_activate():
    """用许可证密钥激活本实例（device_token 写入 system_config，除非环境变量已固定 token）。"""
    err = _admin_license_guard()
    if err is not None:
        return err
    data = request.get_json(silent=True) or {}
    license_key = (data.get('license_key') or '').strip()
    if not license_key:
        return jsonify({'success': False, 'error': '请提供 license_key'}), 400
    if not (os.environ.get('LICENSE_SERVER_URL') or '').strip():
        return jsonify({'success': False, 'error': '未配置 LICENSE_SERVER_URL'}), 400
    try:
        from license_client import activate_license, device_fingerprint, invalidate_verify_cache

        ok, reason, payload = activate_license(license_key, device_fingerprint())
        if not ok:
            return jsonify({'success': False, 'error': reason or 'activate_failed'}), 400
        device_token = (payload.get('device_token') or '').strip()
        if not device_token:
            return jsonify({'success': False, 'error': '许可服务未返回 device_token'}), 502
        stored_in = 'env'
        if not (os.environ.get('LICENSE_DEVICE_TOKEN') or '').strip():
            conn = get_db()
            cursor = conn.cursor()
            if USE_POSTGRES:
                cursor.execute(
                    """
                    INSERT INTO system_config (config_key, config_value, description)
                    VALUES (%s, %s, %s)
                    ON CONFLICT (config_key) DO UPDATE SET
                        config_value = EXCLUDED.config_value,
                        updated_at = CURRENT_TIMESTAMP
                    """,
                    (
                        'license_device_token',
                        device_token,
                        '商业许可证 device_token',
                    ),
                )
            else:
                cursor.execute(
                    "INSERT OR REPLACE INTO system_config (config_key, config_value, description) VALUES (?, ?, ?)",
                    ('license_device_token', device_token, '商业许可证 device_token'),
                )
            conn.commit()
            conn.close()
            stored_in = 'database'
        invalidate_verify_cache()
        return jsonify({
            'success': True,
            'message': '激活成功',
            'stored_in': stored_in,
            'label': payload.get('label'),
            'expires_at': payload.get('expires_at'),
            'token_from_env': stored_in == 'env',
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/dashboard-assets/<path:filename>')
def serve_dashboard_assets(filename):
    """Serve React Dashboard assets from a custom path to avoid conflicts"""
    static_dir = get_static_dir()
    # Path to the dist folder (assets are inside dist/assets)
    # The URL request will contain 'assets/filename.js'
    # So we join dist + filename
    dist_dir = os.path.join(static_dir, 'react-dashboard', 'dist')
    file_path = os.path.join(dist_dir, filename)
    
    if os.path.exists(file_path):
        mimetype = None
        if filename.endswith('.js'):
            mimetype = 'application/javascript'
        elif filename.endswith('.css'):
            mimetype = 'text/css'
        
        return send_file(file_path, mimetype=mimetype)
    else:
        return f"Asset not found: {file_path}", 404



@app.route('/debug_pallet_chart.html')
def debug_pallet_chart():
    # 返回调试页面
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'debug_pallet_chart.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404

@app.route('/pallet_chart_test.html')
def pallet_chart_test():
    # 返回托盘图表测试页面
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'pallet_chart_test.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404

@app.route('/logs')
def logs_page():
    # 未登录：浏览器应跳转登录页（避免返回 JSON 被当成「页面内容」）
    if 'user_id' not in session:
        return redirect('/login')

    # 与导航一致：admin/boss 可看；其余用户查库权限
    if session.get('role') not in ('admin', 'boss'):
        if not check_user_permission('logs', 'view'):
            return redirect('/no_permission')

    # 返回操作日志查询页面
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'logs.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404

# 登录页面
@app.route('/login')
def login_page():
    # 返回静登录页面
    static_dir = get_static_dir()
    return send_from_directory(static_dir, 'login.html')

# 管理员后台页面
@app.route('/admin')
def admin_page():
    # 返回管理员后台页面
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'admin.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404

# 无权限提示页面
@app.route('/no_permission')
def no_permission():
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'no_permission.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404

# 检查重复记录的API端点
@app.route('/api/check_duplicate', methods=['POST'])
def check_duplicate():
    data = request.json
    dock_no = data.get("dock_no")
    vehicle_type = data.get("vehicle_type")
    
    # Car / Van / CBS / CBT 不按道口做短时重复提示
    if vehicle_type in ('Car', 'Van', 'CBS', 'CBT'):
        return jsonify({"is_duplicate": False})
    
    if not dock_no:
        return jsonify({"is_duplicate": False})
    
    conn = get_db()
    cursor = conn.cursor(); cursor.execute("""
        SELECT id, vehicle_type, vehicle_no, created_at 
        FROM inbound_records 
        WHERE dock_no = ? AND vehicle_type NOT IN ('Car', 'Van', 'CBS', 'CBT')
        ORDER BY created_at DESC 
        LIMIT 1
    """, (dock_no,))
    
    last_record = cursor.fetchone()
    conn.close()
    
    if not last_record:
        return jsonify({"is_duplicate": False})
    
    # 计算时间差
    try:
        if isinstance(last_record[3], str):
            last_time = datetime.strptime(last_record[3], '%Y-%m-%d %H:%M:%S')
        else:
            last_time = last_record[3]
        current_time = datetime.now(LA_TZ).replace(tzinfo=None)
        time_diff_seconds = (current_time - last_time).total_seconds()
        time_diff_minutes = int(time_diff_seconds / 60)
        
        if time_diff_minutes < 25:
            return jsonify({
                "is_duplicate": True,
                "time_diff_minutes": time_diff_minutes,
                "last_record": {
                    "id": last_record[0],
                    "vehicle_type": last_record[1],
                    "vehicle_no": last_record[2] or "无",
                    "created_at": last_record[3]
                }
            })
    except Exception as e:
        print(f"检查重复记录时出错: {e}")
        return jsonify({"is_duplicate": False})
    
    return jsonify({"is_duplicate": False})


def _parse_plate_excluded_load(raw):
    try:
        v = float(raw)
        return max(0.0, v)
    except (TypeError, ValueError):
        return 0.0


def _excluded_pieces_for_plate_load(vehicle_type, plate_excluded_load):
    v = _parse_plate_excluded_load(plate_excluded_load)
    if v <= 0:
        return 0
    if vehicle_type in ('16英尺', '26英尺', '53英尺'):
        return int(round(v * 344))
    if vehicle_type in ('Car', 'Van'):
        return int(round(v * 172))
    if vehicle_type in ('CBS', 'CBT'):
        return int(round(v * INBOUND_CBS_CBT_PIECES_PER_PALLET))
    return 0


def _apply_plate_exclusion_to_record(data):
    """填写车牌时可录入不计入统计的装载量；写入 plate_excluded_load / excluded_pieces。"""
    vn = str(data.get('vehicle_no') or '').strip()
    pel = _parse_plate_excluded_load(data.get('plate_excluded_load'))
    if not vn:
        data['plate_excluded_load'] = 0.0
        data['excluded_pieces'] = 0
        return
    vt = data.get('vehicle_type') or ''
    ep = _excluded_pieces_for_plate_load(vt, pel)
    try:
        pieces = int(data.get('pieces') or 0)
    except (TypeError, ValueError):
        pieces = 0
    if ep > pieces:
        ep = max(0, pieces)
    data['plate_excluded_load'] = float(pel) if pel else 0.0
    data['excluded_pieces'] = int(ep)


def _normalize_inbound_business_type(data):
    """业务类型：GOFO（默认）/ CBS / CBT。"""
    if not isinstance(data, dict):
        return 'GOFO'
    raw = data.get('business_type') or data.get('businessType') or 'GOFO'
    bt = str(raw).strip().upper()
    if bt not in ('GOFO', 'CBS', 'CBT'):
        bt = 'GOFO'
    data['business_type'] = bt
    return bt


def _apply_vehicle_type_defaults(data):
    bt = _normalize_inbound_business_type(data)
    if bt in ('CBS', 'CBT'):
        data['vehicle_type'] = bt
        data['unit'] = '托盘'
        load_raw = data.get('load_amount', 0)
        try:
            la = int(load_raw)
        except (TypeError, ValueError):
            la = 0
        data['load_amount'] = la
        data['pieces'] = la * INBOUND_CBS_CBT_PIECES_PER_PALLET if la > 0 else 0
        data.setdefault('dock_no', 0)
        try:
            data['dock_no'] = int(data.get('dock_no') or 0)
        except (TypeError, ValueError):
            data['dock_no'] = 0
        data['vehicle_no'] = str(data.get('vehicle_no') or '').strip()
        return

    vt = data.get("vehicle_type", "")
    if vt == "16英尺":
        data["unit"] = "托盘"
        load_amount = data.get("load_amount", 0)
        if not load_amount or load_amount == 0:
            data["load_amount"] = 6
            data["pieces"] = 6 * 344
        else:
            data["pieces"] = int(load_amount) * 344
    elif vt == "26英尺":
        data["unit"] = "托盘"
        data["load_amount"] = 12
        data["pieces"] = 12 * 344
    elif vt == "Car":
        data["unit"] = "篮筐"
        data["load_amount"] = 1
        data["pieces"] = 1 * 172
    elif vt == "Van":
        data["unit"] = "篮筐"
        data["load_amount"] = 9
        data["pieces"] = 9 * 172
    elif vt == "53英尺":
        data.setdefault("unit", "托盘")
        load_amount = data.get("load_amount", 0)
        if not load_amount or load_amount == 0:
            data["load_amount"] = 24
            data["pieces"] = 24 * 344
        elif load_amount > 0:
            data["pieces"] = load_amount * 344
        elif data.get("pieces") and data["pieces"] > 0:
            data["load_amount"] = data["pieces"] // 344


def _insert_inbound_record_core(data, current_time, *, broadcast=True):
    """在已执行车型规则与 plate_exclusion 后写入一条 inbound 记录。current_time 决定 created_at、班次、默认时间段。"""
    conn = get_db()
    current_time_str = current_time.strftime('%Y-%m-%d %H:%M:%S')
    if current_time.hour < 17:
        shift_type = "早班"
    else:
        shift_type = "晚班"
    time_slot = data.get("time_slot")
    if not time_slot or time_slot == "" or time_slot is None:
        time_slot = str(current_time.hour)
    dock_no = data.get("dock_no")
    vehicle_type = data.get("vehicle_type")
    if dock_no and vehicle_type not in ('Car', 'Van', 'CBS', 'CBT'):
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, created_at, vehicle_type FROM inbound_records 
            WHERE dock_no = ? AND vehicle_type NOT IN ('Car', 'Van', 'CBS', 'CBT')
            ORDER BY created_at DESC 
            LIMIT 1
        """, (dock_no,))
        last_record = cursor.fetchone()
        if last_record:
            try:
                last_id = last_record[0]
                last_vehicle_type = last_record[2]
                if isinstance(last_record[1], str):
                    last_time = datetime.strptime(last_record[1], '%Y-%m-%d %H:%M:%S')
                else:
                    last_time = last_record[1]
                time_diff_seconds = (current_time - last_time).total_seconds()
                last_duration = int(time_diff_seconds / 60)
                if last_duration < 0:
                    last_duration = 0
                conn.cursor().execute("""
                    UPDATE inbound_records 
                    SET duration = ? 
                    WHERE id = ?
                """, (last_duration, last_id))
                print(f"[INFO] 更新记录ID {last_id} ({last_vehicle_type}) 的时长为 {last_duration} 分钟")
            except Exception as e:
                print(f"计算并更新上一条记录时长时出错: {e}")
    pel_ins = data.get("plate_excluded_load", 0)
    epc_ins = data.get("excluded_pieces", 0)
    try:
        pel_ins = float(pel_ins) if pel_ins is not None else 0.0
    except (TypeError, ValueError):
        pel_ins = 0.0
    try:
        epc_ins = int(epc_ins) if epc_ins is not None else 0
    except (TypeError, ValueError):
        epc_ins = 0
    insert_sql = """INSERT INTO inbound_records
        (dock_no, vehicle_type, vehicle_no, unit, load_amount, pieces, time_slot, shift_type, remark, created_at, duration, plate_excluded_load, excluded_pieces, business_type)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, NULL, ?, ?, ?)"""
    insert_params = (
        data.get("dock_no"),
        data.get("vehicle_type"),
        data.get("vehicle_no"),
        data.get("unit"),
        data.get("load_amount"),
        data.get("pieces"),
        time_slot,
        shift_type,
        data.get("remark"),
        current_time_str,
        pel_ins,
        epc_ins,
        data.get("business_type") or "GOFO",
    )
    cursor = conn.cursor()
    if USE_POSTGRES:
        cursor.execute(
            convert_query_placeholders(insert_sql + " RETURNING id"),
            insert_params,
        )
        _row = cursor.fetchone()
        new_id = int(_row[0]) if _row is not None else None
    else:
        cursor.execute(convert_query_placeholders(insert_sql), insert_params)
        new_id = cursor.lastrowid
    conn.commit()
    conn.close()
    if broadcast:
        broadcast_update('refresh_stats', {'action': 'add', 'id': new_id})
    return new_id


INBOUND_IMPORT_VEHICLE_TYPES = frozenset({'16英尺', '26英尺', 'Car', 'Van', '53英尺', '其他', 'CBS', 'CBT'})


def _normalize_import_vehicle_type(raw):
    """
    将导入中的车型简写规范为系统内车型名。
    常见：Excel 数字列 26、53 或文本「26」「53」表示 26英尺、53英尺卡车。
    """
    if raw is None:
        return ''
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        try:
            n = int(round(float(raw)))
            if n == 16:
                return '16英尺'
            if n == 26:
                return '26英尺'
            if n == 53:
                return '53英尺'
        except (TypeError, ValueError):
            pass
    s = str(raw).strip()
    if not s:
        return ''
    if s in INBOUND_IMPORT_VEHICLE_TYPES:
        return s
    # 英文大小写、中文俗称（面包车即 Van）
    slx = s.casefold().replace(' ', '').replace('_', '')
    if slx == 'van':
        return 'Van'
    if slx == 'car':
        return 'Car'
    if s in ('面包车', '厢式车'):
        return 'Van'
    try:
        f = float(s.replace(',', ''))
        if abs(f - round(f)) < 1e-9:
            n = int(round(f))
            if n == 16:
                return '16英尺'
            if n == 26:
                return '26英尺'
            if n == 53:
                return '53英尺'
    except (TypeError, ValueError):
        pass
    sl = s.casefold().replace(' ', '')
    for prefix, canon in (('16', '16英尺'), ('26', '26英尺'), ('53', '53英尺')):
        if sl == prefix + 'ft' or sl == prefix + '-ft' or sl == prefix + 'foot':
            return canon
    if s.isdigit():
        if s == '16':
            return '16英尺'
        if s == '26':
            return '26英尺'
        if s == '53':
            return '53英尺'
    return s


# 表头至少包含以下三列；其余列可省略，导入后按车型规则与录入时间自动补全。
# 内部标准列名（解析后统一为这三项；文件中第三列可写「时间」「時間」等，见别名）
INBOUND_IMPORT_REQUIRED_HEADERS = ('码头号', '车辆类型', '录入时间')
# 下载模板用表头（与 Excel 常见写法一致：第三列为「时间」）
INBOUND_IMPORT_TEMPLATE_HEADERS = ('码头号', '车辆类型', '时间')
INBOUND_IMPORT_OPTIONAL_HEADERS = (
    '车牌号', '装载量', '时间段', '备注', '不计入统计装载量', '件数',
)
INBOUND_IMPORT_MAX_ROWS = 2000

# 表头别名（Excel 另存、英文列名、少字）：归一化后映射到内部标准列名
INBOUND_HEADER_ALIASES = {
    '码头号': (
        '码头号', 'dock', 'dock_no', 'dockno', '码头编号', 'dock no', 'dock_id', 'dockid',
        'docknumber', 'dock_number', 'mt', 'mt_no',
    ),
    '车辆类型': (
        '车辆类型', 'vehicle_type', 'vehicletype', 'vehicle type', '车型', '类型', 'type', 'vt',
        'vtype', 'vehicle', 'cartype', 'car_type',
    ),
    '录入时间': (
        '录入时间', '录入時間', 'entry_time', 'entrytime', 'entry time', '时间', '時間', 'time', '录入时刻', '时刻',
        '录时间', 'logtime', 'record_time', 'recordtime', 'datetime', 'date_time',
    ),
    '车牌号': ('车牌号', '车牌', 'plate', 'license_plate', 'vehicle_no', 'vehicleno'),
    '装载量': ('装载量', 'load', 'load_amount', '装载', 'amount'),
    '时间段': ('时间段', 'slot', 'time_slot', '时段'),
    '备注': ('备注', 'remark', 'note', 'notes', 'memo'),
    '不计入统计装载量': ('不计入统计装载量', 'plate_excluded', 'excluded_load', '不计入统计', '排除装载'),
    '件数': ('件数', 'pieces', 'piece', 'pcs', 'qty'),
}

_INBOUND_HEADER_LOOKUP = None


def _strip_invisible_chars(s: str) -> str:
    """去掉零宽字符等，避免 Excel 复制表头后无法匹配。"""
    if not s:
        return ''
    return re.sub(r'[\u200b-\u200d\ufeff\u2060\u00ad]', '', s)


def _normalize_header_token(s):
    if s is None:
        return ''
    t = _strip_invisible_chars(str(s)).strip().lstrip('\ufeff').replace('\u00a0', ' ')
    t = unicodedata.normalize('NFKC', t)
    t = t.strip(' "\'').strip()
    return t


def _header_match_key(s):
    """用于别名匹配：英文忽略大小写与下划线空格。"""
    t = _normalize_header_token(s)
    if not t:
        return ''
    t = t.replace('\u3000', ' ').replace('_', ' ').replace(' ', '')
    if all(ord(c) < 128 for c in t):
        return t.casefold()
    return t


def _inbound_header_lookup_dict():
    global _INBOUND_HEADER_LOOKUP
    if _INBOUND_HEADER_LOOKUP is None:
        d = {}
        for canon, variants in INBOUND_HEADER_ALIASES.items():
            for v in variants:
                k = _header_match_key(v)
                if k and k not in d:
                    d[k] = canon
        for canon in INBOUND_IMPORT_REQUIRED_HEADERS + INBOUND_IMPORT_OPTIONAL_HEADERS:
            k = _header_match_key(canon)
            if k and k not in d:
                d[k] = canon
        _INBOUND_HEADER_LOOKUP = d
    return _INBOUND_HEADER_LOOKUP


def _canonical_inbound_header(raw):
    """将文件中的表头列名转为内部标准名；无法识别则返回 None。"""
    lk = _inbound_header_lookup_dict()
    k = _header_match_key(raw)
    if k:
        hit = lk.get(k)
        if hit:
            return hit
    t = _normalize_header_token(raw)
    if t in ('时间', '時間', '录入时间', '录入時間'):
        return '录入时间'
    # 模糊：Excel 简写、少字、英文混排（不用「含时间」泛匹配，避免「时间戳」等误判）
    tc = t.replace(' ', '').replace('\u3000', '')
    tl = tc.casefold()
    if ('码头' in tc and '号' in tc) or tc == '码头' or ('dock' in tl and 'no' in tl) or tl == 'dock':
        return '码头号'
    if ('车辆' in tc and '类型' in tc) or ('车' in tc and '类型' in tc) or ('vehicle' in tl and 'type' in tl):
        return '车辆类型'
    if '时间段' in tc or (tc.startswith('时间') and '段' in tc):
        return '时间段'
    if ('录入' in tc and '时间' in tc) or tc in ('时间', '時間', '时刻'):
        return '录入时间'
    if 'entry' in tl and 'time' in tl:
        return '录入时间'
    if tl in ('time', 'entry_time', 'entrytime', 'datetime'):
        return '录入时间'
    return None


def _inbound_row_to_canonical(row):
    """CSV 行 dict -> 标准列名 -> 值（后者覆盖前者）。"""
    out = {}
    for k, v in row.items():
        if k is None:
            continue
        ck = _canonical_inbound_header(_normalize_csv_cell_key(k))
        if ck:
            out[ck] = v
    return out


def _strip_leading_blank_lines(text: str) -> str:
    lines = text.splitlines()
    while lines and not lines[0].strip():
        lines.pop(0)
    return '\n'.join(lines)


def _prepare_inbound_csv_text(text: str) -> str:
    """去掉前导空行；Excel 欧洲区常见首行 Sep=; 需跳过，否则整行被当成表头。"""
    text = _strip_leading_blank_lines(text or '')
    lines = text.splitlines()
    if lines:
        first = lines[0].strip()
        if re.match(r'^sep\s*=\s*\S', first, re.I):
            lines = lines[1:]
    return '\n'.join(lines)


def _fix_inbound_csv_if_header_in_one_cell(text: str) -> str:
    """
    表头被写进一个单元格时，首行解析为单列但单元格内含「码头号,车辆类型,...」。
    将首行展开为多列标准 CSV 行。
    """
    req = set(INBOUND_IMPORT_REQUIRED_HEADERS)
    lines = text.splitlines()
    if not lines:
        return text
    first = lines[0]
    try:
        row0 = next(csv.reader(io.StringIO(first)))
    except (StopIteration, csv.Error):
        return text
    if len(row0) != 1:
        return text
    cell = str(row0[0] or '')
    if ',' not in cell:
        return text
    try:
        inner = next(csv.reader(io.StringIO(cell)))
    except (StopIteration, csv.Error):
        return text
    if len(inner) < 3 or not _header_row_matches_required(inner, req):
        return text
    buf = io.StringIO()
    csv.writer(buf).writerow(inner)
    lines[0] = buf.getvalue().rstrip('\r\n')
    return '\n'.join(lines)


def _normalize_csv_cell_key(key):
    s = _strip_invisible_chars(str(key or '')).strip().lstrip('\ufeff')
    s = s.replace('\u00a0', ' ')
    s = unicodedata.normalize('NFKC', s).strip()
    return s


def _decode_inbound_csv_bytes(raw: bytes) -> str:
    """Excel 常见为 UTF-8、系统 ANSI(GBK)；依次尝试。"""
    if not raw:
        return ''
    if len(raw) >= 2 and raw[:2] in (b'\xff\xfe', b'\xfe\xff'):
        try:
            return raw.decode('utf-16')
        except UnicodeDecodeError:
            pass
    last_err = None
    for enc in ('utf-8-sig', 'utf-8', 'gb18030', 'gbk', 'cp936'):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError as e:
            last_err = e
    if last_err:
        raise last_err
    return ''


def _header_row_matches_required(fieldnames, req):
    if not fieldnames:
        return False
    fn = [_normalize_csv_cell_key(h) for h in fieldnames if h is not None]
    fn = [x for x in fn if x != '']
    canon = {_canonical_inbound_header(h) for h in fn}
    canon.discard(None)
    return req.issubset(canon)


def _try_dict_reader_inbound(sub_text: str, req: set, delims: tuple) -> object:
    """在一段文本上尝试分隔符，表头须满足必填列。成功返回 csv.DictReader，否则 None。"""
    if not (sub_text or '').strip():
        return None
    for delim in delims:
        r = csv.DictReader(io.StringIO(sub_text), delimiter=delim)
        if not r.fieldnames:
            continue
        if _header_row_matches_required(r.fieldnames, req):
            return r
    sub_lines = sub_text.splitlines()
    if not sub_lines:
        return None
    first = sub_lines[0]
    for delim in delims:
        try:
            row0 = next(csv.reader(io.StringIO(first), delimiter=delim))
        except (StopIteration, csv.Error):
            continue
        if len(row0) < 3:
            continue
        if _header_row_matches_required(row0, req):
            return csv.DictReader(io.StringIO(sub_text), delimiter=delim)
    return None


def _csv_dict_reader_for_inbound(text: str):
    """
    自动选择分隔符（逗号 / 分号 / Tab / 全角逗号 / 竖线），使必填表头能正确识别。
    支持跳过前若干行（Excel 第 1 行常为标题「入库明细」等，真正表头在第二行）。
    """
    text = _prepare_inbound_csv_text(text)
    text = _fix_inbound_csv_if_header_in_one_cell(text)
    if not text.strip():
        return None
    req = set(INBOUND_IMPORT_REQUIRED_HEADERS)
    delims = (',', ';', '\t', '\uff0c', '|')
    lines = text.splitlines()
    max_skip = min(12, len(lines))
    for skip in range(max_skip):
        sub = '\n'.join(lines[skip:])
        hit = _try_dict_reader_inbound(sub, req, delims)
        if hit is not None:
            return hit
    try:
        sample = text[:8192]
        dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|\uff0c')
        r = csv.DictReader(io.StringIO(text), dialect=dialect)
        if r.fieldnames and _header_row_matches_required(r.fieldnames, req):
            return r
    except (csv.Error, AttributeError):
        pass
    return csv.DictReader(io.StringIO(text))


def _parse_import_datetime_on_date(import_date_str, time_cell):
    """import_date_str: 页面「导入日期」YYYY-MM-DD；time_cell: 行内「录入时间」必填，支持 HH:MM 或 HH:MM:SS。"""
    d = (import_date_str or '').strip()
    if not d:
        raise ValueError('导入日期为空')
    t_raw = (time_cell or '').strip() if time_cell is not None else ''
    if not t_raw:
        raise ValueError('录入时间必填')
    if t_raw.count(':') == 1:
        t_raw = t_raw + ':00'
    elif t_raw.count(':') == 0:
        raise ValueError('录入时间格式须为 HH:MM 或 HH:MM:SS')
    return datetime.strptime(f"{d} {t_raw}", '%Y-%m-%d %H:%M:%S')


def _parse_dock_no_cell(raw):
    if raw is None or str(raw).strip() == '':
        raise ValueError('码头号必填')
    s = str(raw).strip()
    try:
        return int(float(s))
    except (TypeError, ValueError):
        raise ValueError(f'码头号无效: {raw!r}')


def _parse_float_cell(raw, default=0.0):
    if raw is None or str(raw).strip() == '':
        return default
    return float(str(raw).strip())


def _parse_optional_int_cell(raw):
    if raw is None or str(raw).strip() == '':
        return None
    return int(float(str(raw).strip()))


def _csv_row_to_inbound_dict(norm_row, import_date_str):
    """norm_row: 表头 -> 值（已去 BOM）。返回 (data_dict, created_at datetime)。"""
    vt = _normalize_import_vehicle_type(norm_row.get('车辆类型'))
    if not vt or vt not in INBOUND_IMPORT_VEHICLE_TYPES:
        raise ValueError(f'车辆类型无效或为空: {vt!r}')
    dock_no = _parse_dock_no_cell(norm_row.get('码头号'))
    plate = (norm_row.get('车牌号') or '').strip()
    pel = _parse_float_cell(norm_row.get('不计入统计装载量'), 0.0)
    remark = norm_row.get('备注')
    if remark is None:
        remark = ''
    else:
        remark = str(remark).strip()
    ts_slot = norm_row.get('时间段')
    if ts_slot is not None and str(ts_slot).strip() != '':
        time_slot = str(ts_slot).strip()
    else:
        time_slot = None
    load_raw = norm_row.get('装载量')
    if load_raw is None or str(load_raw).strip() == '':
        load_amount = 0
    else:
        load_amount = _parse_float_cell(load_raw, 0.0)
    pieces_cell = _parse_optional_int_cell(norm_row.get('件数'))
    data = {
        'dock_no': dock_no,
        'vehicle_type': vt,
        'vehicle_no': plate,
        'load_amount': load_amount,
        'remark': remark,
        'plate_excluded_load': pel,
    }
    if time_slot is not None:
        data['time_slot'] = time_slot
    if vt == '其他':
        data['pieces'] = pieces_cell if pieces_cell is not None else 0
    elif pieces_cell is not None:
        data['pieces'] = pieces_cell
    created_at = _parse_import_datetime_on_date(import_date_str, norm_row.get('录入时间'))
    return data, created_at


def _inbound_row_looks_like_data(row):
    """首行无表头时：第一列像码头号且第二列为合法车型。"""
    if len(row) < 3:
        return False
    try:
        _parse_dock_no_cell(row[0])
    except ValueError:
        return False
    return _normalize_import_vehicle_type(str(row[1])) in INBOUND_IMPORT_VEHICLE_TYPES


def _inbound_positional_parse(text, import_date_str):
    """
    表头无法识别时的兜底：固定列序为 码头号、车辆类型、时间（第4列起依次对应可选列）。
    首行若已是数据（无表头），则从第一行开始导入。
    """
    text = _prepare_inbound_csv_text(text)
    text = _fix_inbound_csv_if_header_in_one_cell(text)
    lines = [ln for ln in text.splitlines() if ln.strip()]
    delims = (',', ';', '\t', '\uff0c', '|')
    opt_keys = INBOUND_IMPORT_OPTIONAL_HEADERS
    for skip in range(min(16, len(lines))):
        sub = '\n'.join(lines[skip:])
        for delim in delims:
            try:
                rows = list(csv.reader(io.StringIO(sub), delimiter=delim))
            except csv.Error:
                continue
            if not rows or len(rows[0]) < 3:
                continue
            start_i = 0 if _inbound_row_looks_like_data(rows[0]) else 1
            if start_i >= len(rows):
                continue
            parsed = []
            errors = []
            for i in range(start_i, len(rows)):
                row = rows[i]
                row_index = skip + i + 1
                if len(row) < 3:
                    continue
                if all((c is None or str(c).strip() == '') for c in row[:3]):
                    continue
                norm = {'码头号': row[0], '车辆类型': row[1], '录入时间': row[2]}
                for j, ok in enumerate(opt_keys):
                    idx = 3 + j
                    if len(row) > idx and row[idx] is not None and str(row[idx]).strip() != '':
                        norm[ok] = row[idx]
                try:
                    data, created_at = _csv_row_to_inbound_dict(norm, import_date_str)
                    parsed.append((created_at, row_index, data))
                except ValueError as e:
                    errors.append({"row": row_index, "message": str(e)})
            if len(parsed) + len(errors) > INBOUND_IMPORT_MAX_ROWS:
                return None
            if parsed or errors:
                return parsed, errors
    return None


@app.route('/api/record', methods=['POST'])
def record():
    data = request.json if isinstance(request.json, dict) else {}
    _normalize_inbound_business_type(data)
    if data.get('business_type') == 'GOFO' and data.get('vehicle_type') is not None:
        data['vehicle_type'] = _normalize_import_vehicle_type(data['vehicle_type'])
    _apply_vehicle_type_defaults(data)
    if data.get('business_type') in ('CBS', 'CBT'):
        try:
            la = int(data.get('load_amount') or 0)
        except (TypeError, ValueError):
            la = 0
        if la <= 0:
            return jsonify({"success": False, "error": "CBS/CBT 须填写大于 0 的整数托盘数"}), 400
    _apply_plate_exclusion_to_record(data)
    new_id = _insert_inbound_record_core(data, datetime.now(LA_TZ).replace(tzinfo=None), broadcast=True)
    pel_ins = data.get("plate_excluded_load", 0)
    epc_ins = data.get("excluded_pieces", 0)
    try:
        pel_ins = float(pel_ins) if pel_ins is not None else 0.0
    except (TypeError, ValueError):
        pel_ins = 0.0
    try:
        epc_ins = int(epc_ins) if epc_ins is not None else 0
    except (TypeError, ValueError):
        epc_ins = 0
    return jsonify({
        "success": True,
        "record_id": new_id,
        "plate_excluded_load": pel_ins,
        "excluded_pieces": epc_ins,
    })


@app.route('/api/inbound_import_template', methods=['GET'])
def inbound_import_template():
    buf = io.StringIO()
    buf.write('\ufeff')
    w = csv.writer(buf, lineterminator='\r\n')
    w.writerow(list(INBOUND_IMPORT_TEMPLATE_HEADERS))
    w.writerow(['1', '16英尺', '09:15'])
    body = buf.getvalue()
    return Response(
        body.encode('utf-8'),
        mimetype='text/csv; charset=utf-8',
        headers={'Content-Disposition': 'attachment; filename="inbound_import_template.csv"'},
    )


@app.route('/api/inbound_import', methods=['POST'])
def inbound_import():
    import_date = (request.form.get('import_date') or '').strip()
    if not import_date:
        return jsonify({"success": False, "error": "请填写导入日期 import_date（YYYY-MM-DD）"}), 400
    try:
        datetime.strptime(import_date, '%Y-%m-%d')
    except ValueError:
        return jsonify({"success": False, "error": "导入日期格式应为 YYYY-MM-DD"}), 400
    f = request.files.get('file')
    if not f or not f.filename:
        return jsonify({"success": False, "error": "请上传 CSV 文件"}), 400
    raw = f.read()
    if not raw:
        return jsonify({"success": False, "error": "文件为空"}), 400
    try:
        text = _decode_inbound_csv_bytes(raw)
    except UnicodeDecodeError:
        return jsonify({"success": False, "error": "无法解析文件编码，请用 Excel「另存为 CSV UTF-8」或记事本保存为 UTF-8"}), 400
    reader = _csv_dict_reader_for_inbound(text)
    req = set(INBOUND_IMPORT_REQUIRED_HEADERS)
    fnames = []
    if reader and reader.fieldnames:
        fnames = [_normalize_csv_cell_key(h) for h in reader.fieldnames if h is not None]
    canon_set = {_canonical_inbound_header(h) for h in fnames}
    canon_set.discard(None)
    use_positional = (not reader or not reader.fieldnames or not req.issubset(canon_set))

    parsed = []
    errors = []
    if use_positional:
        res = _inbound_positional_parse(text, import_date)
        if not res:
            opt_hint = '可选追加列: ' + ','.join(INBOUND_IMPORT_OPTIONAL_HEADERS)
            return jsonify({
                "success": False,
                "error": (
                    "无法按表头识别列；已尝试按列序（第1列码头、第2列车型、第3列时间）解析仍失败。"
                    "请确认 CSV 为三列及以上，且另存为 UTF-8 或 Excel 简体中文 CSV。"
                    + opt_hint
                ),
            }), 400
        parsed, errors = res
    else:
        row_index = 1
        for row in reader:
            row_index += 1
            if row_index - 1 > INBOUND_IMPORT_MAX_ROWS:
                return jsonify({"success": False, "error": f"单次最多处理 {INBOUND_IMPORT_MAX_ROWS} 行"}), 400
            if not row or all((v is None or str(v).strip() == '') for v in row.values()):
                continue
            norm = _inbound_row_to_canonical(row)
            try:
                data, created_at = _csv_row_to_inbound_dict(norm, import_date)
                parsed.append((created_at, row_index, data))
            except ValueError as e:
                errors.append({"row": row_index, "message": str(e)})
    if not parsed and not errors:
        return jsonify({"success": False, "error": "没有有效数据行"}), 400
    if errors:
        return jsonify({"success": False, "errors": errors, "imported": 0}), 400
    parsed.sort(key=lambda x: (x[0], x[1]))
    imported = 0
    try:
        for created_at, _ridx, data in parsed:
            data = dict(data)
            _normalize_inbound_business_type(data)
            _apply_vehicle_type_defaults(data)
            _apply_plate_exclusion_to_record(data)
            _insert_inbound_record_core(data, created_at, broadcast=False)
            imported += 1
    except Exception as e:
        print(f"[ERROR] inbound_import 写入失败: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e), "imported": imported}), 500
    broadcast_update('refresh_stats', {'action': 'batch_import', 'count': imported})
    return jsonify({"success": True, "imported": imported, "errors": []})

@app.route('/api/record/<int:record_id>', methods=['PUT'])
def update_record(record_id):
    print(f"[DEBUG] PUT /api/record/{record_id} 被调用")
    data = request.json
    print(f"[DEBUG] 接收到的数据: {data}")
    if isinstance(data, dict) and data.get('vehicle_type') is not None:
        data['vehicle_type'] = _normalize_import_vehicle_type(data['vehicle_type'])
    
    # 获取当前系统时间并自动判断班次类型
    current_time = datetime.now(LA_TZ).replace(tzinfo=None)
    
    # 自动判断班次类型：17点之前是早班，17点之后是晚班
    if current_time.hour < 17:
        shift_type = "早班"
    else:
        shift_type = "晚班"
    
    # 对于53英尺车辆，如果没有输入装载量，则默认为24托盘
    vt = data.get("vehicle_type", "")
    if vt == "16英尺":
        data.setdefault("unit", "托盘")
        load_amount = data.get("load_amount", 0)
        if not load_amount or load_amount == 0:
            data["load_amount"] = 6
            data["pieces"] = 6 * 344
        else:
            data["pieces"] = int(load_amount) * 344
    elif vt in ("CBS", "CBT"):
        data.setdefault("unit", "托盘")
        load_amount = data.get("load_amount", 0)
        try:
            la = int(load_amount)
        except (TypeError, ValueError):
            la = 0
        if la > 0:
            data["pieces"] = la * INBOUND_CBS_CBT_PIECES_PER_PALLET
    elif vt == "53英尺":
        data.setdefault("unit", "托盘")
        load_amount = data.get("load_amount", 0)
        if not load_amount or load_amount == 0:
            # 默认24托盘
            data["load_amount"] = 24
            data["pieces"] = 24 * 344  # 8256件
        elif load_amount > 0:
            # 用户输入了装载量，自动计算件数
            data["pieces"] = load_amount * 344
        # 如果已有件数但没有装载量，也可以反向计算装载量
        elif data.get("pieces") and data["pieces"] > 0:
            data["load_amount"] = data["pieces"] // 344

    if vt in ("CBS", "CBT"):
        data["business_type"] = vt
    else:
        data["business_type"] = str(data.get("business_type") or "GOFO").strip().upper()
        if data["business_type"] not in ("GOFO", "CBS", "CBT"):
            data["business_type"] = "GOFO"

    _apply_plate_exclusion_to_record(data)

    conn = None
    try:
        conn = get_db()
        print(f"[DEBUG] 数据库连接成功: {DB_PATH}")
        
        # 获取修改前的数据
        old_record_cur = conn.cursor(); old_record_cur.execute("SELECT * FROM inbound_records WHERE id=?", (record_id,))
        old_record = old_record_cur.fetchone()
        print(f"[DEBUG] 原始记录: {old_record}")
        
        cursor = conn.cursor(); cursor.execute(convert_query_placeholders("""UPDATE inbound_records SET
            dock_no=?, vehicle_type=?, vehicle_no=?, unit=?, load_amount=?, pieces=?, time_slot=?, shift_type=?, remark=?, duration=?, plate_excluded_load=?, excluded_pieces=?, business_type=?
            WHERE id=?"""),
            (data.get("dock_no"), data.get("vehicle_type"), data.get("vehicle_no"),
             data.get("unit"), data.get("load_amount"), data.get("pieces"),
             data.get("time_slot"), shift_type, data.get("remark"), data.get("duration"),
             data.get("plate_excluded_load", 0), data.get("excluded_pieces", 0),
             data.get("business_type") or "GOFO", record_id))
        
        print(f"[DEBUG] 更新操作影响的行数: {cursor.rowcount}")
        
        # 如果记录被成功更新，记录日志
        if cursor.rowcount > 0:
            # 获取修改后的数据
            new_record_cur = conn.cursor(); new_record_cur.execute("SELECT * FROM inbound_records WHERE id=?", (record_id,))
            new_record = new_record_cur.fetchone()
            
            # 记录操作日志
            column_names = [description[0] for description in old_record_cur.description]
            old_data = dict(zip(column_names, old_record)) if old_record else {}
            new_data = dict(zip(column_names, new_record)) if new_record else {}
            
            # 删除游标对象，避免序列化错误
            if 'cursor' in old_data:
                del old_data['cursor']
            if 'cursor' in new_data:
                del new_data['cursor']
            
            conn.cursor().execute("""INSERT INTO operation_logs 
                (operation_type, table_name, record_id, old_data, new_data)
                VALUES (?, ?, ?, ?, ?)""",
                ('edit', 'inbound_records', record_id, 
                 json.dumps(old_data, default=str), 
                 json.dumps(new_data, default=str)))
            
            
            conn.commit()
            
            # Broadcast update to all connected SSE clients
            broadcast_update('refresh_stats', {'action': 'update', 'id': record_id})

            if conn:
                conn.close()
            print("[DEBUG] 记录更新成功")
            return jsonify({"success": True})
        else:
            conn.commit()
            if conn:
                conn.close()
            print("[DEBUG] 记录未找到")
            return jsonify({"success": False, "error": "记录未找到"}), 404
    except Exception as e:
        print(f"[DEBUG] 更新记录时发生错误: {str(e)}")
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass

@app.route('/api/record/<int:record_id>', methods=['GET'])
def get_record(record_id):
    """获取单个入库记录的详细信息"""
    try:
        conn = get_db()
        cursor = conn.cursor(); cursor.execute(convert_query_placeholders("SELECT * FROM inbound_records WHERE id=?"), (record_id,))
        record = cursor.fetchone()
        
        if record:
            # 获取列名
            column_names = [description[0] for description in cursor.description]
            # 创建记录字典
            record_dict = dict(zip(column_names, record))
            conn.close()
            return jsonify(record_dict)
        else:
            conn.close()
            return jsonify({"error": "记录未找到"}), 404
    except Exception as e:
        if 'conn' in locals():
            conn.close()
        return jsonify({"error": str(e)}), 500

@app.route('/api/record/<int:record_id>', methods=['DELETE'])
def delete_record(record_id):
    conn = None
    try:
        conn = get_db()
        
        # 获取删除前的数据
        old_record_cur = conn.cursor(); old_record_cur.execute("SELECT * FROM inbound_records WHERE id=?", (record_id,))
        old_record = old_record_cur.fetchone()
        
        cursor = conn.cursor(); cursor.execute(convert_query_placeholders("DELETE FROM inbound_records WHERE id=?"), (record_id,))
        
        # 如果记录被成功删除，记录日志
        if cursor.rowcount > 0:
            # 记录操作日志
            column_names = [description[0] for description in old_record_cur.description]
            old_data = dict(zip(column_names, old_record)) if old_record else {}
            
            # 删除游标对象，避免序列化错误
            if 'cursor' in old_data:
                del old_data['cursor']
            
            conn.cursor().execute("""INSERT INTO operation_logs 
                (operation_type, table_name, record_id, old_data, new_data)
                VALUES (?, ?, ?, ?, ?)""",
                ('delete', 'inbound_records', record_id, 
                 json.dumps(old_data, default=str), 
                 json.dumps({})))  # 删除操作没有新数据
            
            conn.commit()
            
            # 删除相关的操作日志
            conn.cursor().execute("DELETE FROM operation_logs WHERE table_name='inbound_records' AND record_id=?", (record_id,))
            conn.commit()
            conn.close()
            
            # Broadcast update to all connected SSE clients
            broadcast_update('refresh_stats', {'action': 'delete', 'id': record_id})
            
            return jsonify({"success": True})
        else:
            conn.close()
            return jsonify({"success": False, "error": "记录未找到"}), 404
            
    except Exception as e:
        if conn:
            conn.close()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass

@app.route('/api/list')
def list_data():
    print("[DEBUG] /api/list 路由被调用")
    print("[DEBUG] 函数开始执行")
    conn=get_db()
    print(f"[DEBUG] 数据库连接成功: {DB_PATH}")
    
    # 获取当前日期和昨天日期
    current_date = datetime.now(LA_TZ).date()
    yesterday_date = current_date - timedelta(days=1)
    print(f"[DEBUG] 当前系统日期: {current_date}")
    print(f"[DEBUG] 昨天日期: {yesterday_date}")

    # 计算当天和昨天的查询时间范围
    # 当天00:00:00
    today_start = datetime.combine(current_date, datetime.min.time())
    # 今天23:59:59
    today_end = datetime.combine(current_date, datetime.max.time())
    # 昨天00:00:00
    yesterday_start = datetime.combine(yesterday_date, datetime.min.time())
    # 昨天23:59:59
    yesterday_end = datetime.combine(yesterday_date, datetime.max.time())
    
    # 次日00:00:00的时间（系统时间，用于上限）
    next_day_start = datetime.combine(current_date + timedelta(days=1), datetime.min.time())

    print(f"[DEBUG] 查询时间范围: {yesterday_start} 到 {next_day_start}")

    # 查询当天和昨天的所有记录
    # 修改：排除车牌号包含'G'的53英尺车辆
    cur = conn.cursor(); cur.execute("""
        SELECT ir.id, ir.dock_no, ir.vehicle_type, ir.vehicle_no, ir.unit, ir.load_amount,
               ir.pieces, ir.time_slot, ir.shift_type, ir.remark, ir.created_at, ir.created_by,
               u.username as created_by_username, ir.duration,
               ir.plate_excluded_load, ir.excluded_pieces,
               COALESCE(ir.business_type, 'GOFO') as business_type
        FROM inbound_records ir
        LEFT JOIN users u ON ir.created_by = u.id
        WHERE 
            (ir.created_at >= ? AND ir.created_at <= ?) OR (ir.created_at >= ? AND ir.created_at <= ?)
        ORDER BY ir.created_at DESC""", (
            yesterday_start.strftime('%Y-%m-%d %H:%M:%S'),
            yesterday_end.strftime('%Y-%m-%d %H:%M:%S'),
            today_start.strftime('%Y-%m-%d %H:%M:%S'),
            today_end.strftime('%Y-%m-%d %H:%M:%S')
        ))
    
    raw_rows = cur.fetchall()
    print(f"[DEBUG] 数据库查询返回记录数: {len(raw_rows)}")
    
    rows=[{
        "id":r[0], "dock_no":r[1], "vehicle_type":r[2], "vehicle_no":r[3],
        "unit":r[4], "load_amount":r[5], "pieces":r[6],
        "time_slot":r[7], "shift_type":r[8], "remark":r[9],
        "created_at":r[10],  # 数据库中存储的是系统时间，直接返回
        "created_by":r[11],  # 创建者用户ID
        "created_by_username":r[12] or "未知用户",  # 创建者用户名
        "duration":r[13],  # 时长(分钟)
        "plate_excluded_load": r[14] if len(r) > 14 else 0,
        "excluded_pieces": r[15] if len(r) > 15 else 0,
        "business_type": (r[16] if len(r) > 16 else None) or "GOFO",
        "pieces_actual": round(
            _py_inbound_arrival_pieces(
                r[2], r[3], r[6], r[15] if len(r) > 15 else 0
            ),
            2,
        ),
    } for r in raw_rows]
    
    print(f"[DEBUG] 处理后返回记录数: {len(rows)}")
    conn.close()
    print(f"[DEBUG] 返回JSON数据: {jsonify(rows)}")
    return jsonify(rows)

# 新增API：获取按时间段分组的入库数据


@app.route('/api/list/check_updates')
def check_list_updates():
    try:
        current_date = datetime.now(LA_TZ).date()
        yesterday_date = current_date - timedelta(days=1)
        today_start = datetime.combine(current_date, datetime.min.time())
        today_end = datetime.combine(current_date, datetime.max.time())
        yesterday_start = datetime.combine(yesterday_date, datetime.min.time())
        yesterday_end = datetime.combine(yesterday_date, datetime.max.time())

        conn = get_db()
        cur = conn.cursor()
        cur.execute(convert_query_placeholders("""
            SELECT MAX(id) as max_id, COUNT(*) as row_count 
            FROM inbound_records ir 
            WHERE 
            (ir.created_at >= ? AND ir.created_at <= ?) OR 
            (ir.created_at >= ? AND ir.created_at <= ?)
        """), (
            yesterday_start.strftime('%Y-%m-%d %H:%M:%S'),
            yesterday_end.strftime('%Y-%m-%d %H:%M:%S'),
            today_start.strftime('%Y-%m-%d %H:%M:%S'),
            today_end.strftime('%Y-%m-%d %H:%M:%S')
        ))
        res = cur.fetchone()
        conn.close()

        max_id = res['max_id'] if res and hasattr(res, 'keys') else (res[0] if res else 0)
        row_count = res['row_count'] if res and hasattr(res, 'keys') else (res[1] if res else 0)

        return jsonify({
            'latest_id': max_id or 0,
            'count': row_count or 0
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/inbound_hourly')
def inbound_hourly_data():
    # 获取日期参数，默认为今天
    date_str = request.args.get('date')
    window_mode = _parse_stats_window_param(request.args.get('stats_window'))

    conn=get_db()

    if date_str:
        try:
            request_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            conn.close()
            return jsonify({"error": "日期格式无效，请使用YYYY-MM-DD格式"}), 400
    else:
        request_date = _default_stats_request_date(window_mode)

    period_start, period_end = _stats_period_bounds(request_date, window_mode)

    cur = conn.cursor(); cur.execute(f"""
        SELECT 
            ir.time_slot, 
            SUM({_sql_inbound_net_pieces_actual('ir.')}) as total_pieces,
            SUM(CASE 
                WHEN ir.vehicle_type IN ('26英尺', '53英尺') THEN (ir.load_amount - COALESCE(ir.plate_excluded_load, 0))
                ELSE 0 
            END) as total_load_amount
        FROM inbound_records ir
        WHERE 
            ir.created_at >= ? AND ir.created_at < ? AND ir.time_slot IS NOT NULL
        GROUP BY ir.time_slot
        ORDER BY ir.time_slot""", (
            period_start.strftime('%Y-%m-%d %H:%M:%S'),
            period_end.strftime('%Y-%m-%d %H:%M:%S')
        ))
    rows=[{
        "time_slot": r[0],
        "total_pieces": int(r[1]) if r[1] else 0,
        "total_load_amount": r[2] if r[2] else 0
    } for r in cur.fetchall()]
    return jsonify(rows)


@app.route('/api/inbound_hourly_avg')
def inbound_hourly_avg_data():
    """历史全量数据中，每个时间段的日均入库件数（按天聚合后求均值）。"""
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute(f"""
            SELECT time_slot, AVG(day_pieces) as avg_pieces
            FROM (
                SELECT DATE(created_at) as d, time_slot, SUM({_sql_inbound_net_pieces_actual('')}) as day_pieces
                FROM inbound_records
                WHERE time_slot IS NOT NULL AND time_slot != ''
                GROUP BY DATE(created_at), time_slot
            ) t
            GROUP BY time_slot
            ORDER BY time_slot
        """)
        rows = [{
            "time_slot": r[0],
            "avg_pieces": float(r[1]) if r[1] is not None else 0.0
        } for r in cur.fetchall()]
        return jsonify(rows)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()

@app.route('/api/pallet_hourly')
def pallet_hourly_data():
    date_str = request.args.get('date')
    window_mode = _parse_stats_window_param(request.args.get('stats_window'))

    conn=get_db()

    if date_str:
        try:
            request_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            conn.close()
            return jsonify({"error": "日期格式无效，请使用YYYY-MM-DD格式"}), 400
    else:
        request_date = _default_stats_request_date(window_mode)

    period_start, period_end = _stats_period_bounds(request_date, window_mode)

    cur = conn.cursor(); cur.execute("""
        SELECT time_slot, SUM(load_amount - COALESCE(plate_excluded_load, 0)) as total_load_amount, COUNT(*) as count
        FROM inbound_records 
        WHERE 
            created_at >= ? AND created_at < ? AND (vehicle_type = '16英尺' OR vehicle_type = '26英尺' OR vehicle_type = '53英尺')
            AND NOT (vehicle_type = '53英尺' AND vehicle_no = 'G')
        GROUP BY time_slot
        ORDER BY time_slot""", (
            period_start.strftime('%Y-%m-%d %H:%M:%S'),
            period_end.strftime('%Y-%m-%d %H:%M:%S')
        ))
    current_day_rows=[{
        "time_slot": r[0] if r[0] else '未指定',
        "total_load_amount": r[1] if r[1] else 0,
        "count": r[2] if r[2] else 0
    } for r in cur.fetchall()]
    
    # 查询历史数据：所有历史数据按时段分组
    historical_cur = conn.cursor(); historical_cur.execute("""
        SELECT 
            time_slot,
            SUM(load_amount - COALESCE(plate_excluded_load, 0)) as total_load_amount,
            COUNT(*) as total_count,
            COUNT(DISTINCT DATE(created_at)) as days_count
        FROM inbound_records
        WHERE vehicle_type IN ('16英尺', '26英尺', '53英尺')
            AND NOT (vehicle_type = '53英尺' AND vehicle_no = 'G')
            AND time_slot IS NOT NULL
            AND time_slot != ''
        GROUP BY time_slot
        ORDER BY CAST(time_slot AS INTEGER)
    """)
    
    # 计算每天平均值：总和 ÷ 天数
    historical_avg_rows = []
    for r in historical_cur.fetchall():
        time_slot = r[0]
        total_load = r[1] if r[1] else 0
        days_count = r[3] if r[3] else 1  # 避免除以0
        avg_per_day = round(total_load / days_count, 2) if days_count > 0 else 0
        
        historical_avg_rows.append({
            "time_slot": time_slot,
            "avg_load_amount": avg_per_day,
            "total_load_amount": int(total_load),
            "total_count": r[2] if r[2] else 0,
            "days_count": days_count
        })
    
    conn.close()
    
    # 返回包含当天数据和历史平均值的结构
    return jsonify({
        'current_day': current_day_rows,
        'historical_avg': historical_avg_rows
    })



# 新增API：获取按时间段分组的分拣数据
@app.route('/api/sorting_hourly')
def sorting_hourly_data():
    date_str = request.args.get('date')
    window_mode = _parse_stats_window_param(request.args.get('stats_window'))

    conn=get_db()

    if date_str:
        try:
            request_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            conn.close()
            return jsonify({"error": "日期格式无效，请使用YYYY-MM-DD格式"}), 400
    else:
        request_date = _default_stats_request_date(window_mode)

    clause, binds = _sorting_slot_window_sql_binds(window_mode, request_date)

    cur = conn.cursor(); cur.execute(f"""SELECT 
                            time_slot, 
                            SUM(pieces) as total_pieces,
                            SUM(manual_count) as manual_total,
                            SUM(device_count) as device_total
                        FROM sorting_records 
                        WHERE {clause}
                        GROUP BY time_slot
                        ORDER BY time_slot""", binds)
    rows=[{
        "time_slot": r[0],
        "total_pieces": int(r[1]) if r[1] else 0,
        "manual_count": int(r[2]) if r[2] else 0,
        "device_count": int(r[3]) if r[3] else 0
    } for r in cur.fetchall()]
    _cno_site = os.environ.get("GOFO_CNO01_SITE", "CNO01").strip() or "CNO01"
    try:
        cur.execute(
            f"""SELECT time_slot, COALESCE(waybill_no_total, 0)
               FROM gofo_collect_destin_hourly
               WHERE destin_site = ?
                 AND ({clause})""",
            (_cno_site,) + binds,
        )
        cno_map = {r[0]: int(r[1]) for r in cur.fetchall()}
    except Exception:
        cno_map = {}
    for row in rows:
        row["cno01_waybill"] = int(cno_map.get(row["time_slot"], 0))
    _have_slots = {r["time_slot"] for r in rows}
    for _slot, _w in cno_map.items():
        if _slot not in _have_slots:
            rows.append(
                {
                    "time_slot": _slot,
                    "total_pieces": 0,
                    "manual_count": 0,
                    "device_count": 0,
                    "cno01_waybill": int(_w),
                }
            )
            _have_slots.add(_slot)
    conn.close()
    return jsonify(rows)


@app.route('/api/sorting_hourly_avg')
def sorting_hourly_avg_data():
    """历史全量数据中，每个时间段的日均分拣件数（按天聚合后求均值）。"""
    conn = get_db()
    try:
        cur = conn.cursor()
        cur.execute("""
            SELECT time_slot, AVG(day_pieces) as avg_pieces
            FROM (
                SELECT sorting_time as d, time_slot, SUM(pieces) as day_pieces
                FROM sorting_records
                WHERE time_slot IS NOT NULL AND time_slot != ''
                GROUP BY sorting_time, time_slot
            ) t
            GROUP BY time_slot
            ORDER BY time_slot
        """)
        rows = [{
            "time_slot": r[0],
            "avg_pieces": float(r[1]) if r[1] is not None else 0.0
        } for r in cur.fetchall()]
        return jsonify(rows)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        conn.close()

@app.route('/api/history')
def get_history():
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    # Legacy support: if only 'date' is provided
    if not start_date_str and not end_date_str:
        date_str = request.args.get('date')
        if not date_str:
            return jsonify({"error": "请提供日期参数"}), 400
        start_date_str = date_str
        end_date_str = date_str
    
    # If one is provided but not the other, use the one provided for both (or handle as single day)
    if not start_date_str:
        start_date_str = end_date_str
    if not end_date_str:
        end_date_str = start_date_str
    
    conn = get_db()
    
    try:
        # 解析请求的日期
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # 获取洛杉矶当前日期
        la_tz = pytz.timezone('America/Los_Angeles')
        la_today = datetime.now(la_tz).date()
        
        # 检查是否选择了未来日期（明天或以后）
        # 检查是否选择了未来日期（明天或以后）
        # 只要开始日期在今天之前（或今天）即可。如果结束日期是未来，可以限制为今天。
        if start_date > la_today:
            conn.close()
            return jsonify({
                "error": "不能查询未来日期的数据",
                "records": [],
                "sorting_records": [],
                "stats": {
                    "total_vehicles": 0,
                    "total_pieces": 0,
                    "record_count": 0,
                    "vehicle_stats": []
                }
            }), 400
        
        # 计算次日日期
        # 计算次日日期（用于结束日期的界限）
        next_day_after_end = end_date + timedelta(days=1)
        
        # 构建日期范围查询条件（使用自然日）
        # 开始日期00:00:00的时间（系统时间）
        query_start = datetime.combine(start_date, datetime.min.time())
        
        # 结束日期次日00:00:00的时间（系统时间，用于上限）
        query_end = datetime.combine(next_day_after_end, datetime.min.time())
        
        # 查询指定日期的入库记录（查询当天00:00之后到次日00:00之前的所有记录）
        inbound_query = """
            SELECT id, dock_no, vehicle_type, vehicle_no, unit, load_amount,
                   pieces, time_slot, shift_type, remark, created_at, duration, is_synced,
                   COALESCE(excluded_pieces, 0)
            FROM inbound_records 
            WHERE 
                created_at >= ? AND created_at < ?
            ORDER BY created_at DESC
        """
        inbound_cur = conn.cursor(); inbound_cur.execute(inbound_query, (
            query_start.strftime('%Y-%m-%d %H:%M:%S'), 
            query_end.strftime('%Y-%m-%d %H:%M:%S')
        ))
        inbound_rows = [{
            "id": r[0], "dock_no": r[1], "vehicle_type": r[2], "vehicle_no": r[3],
            "unit": r[4], "load_amount": r[5], "pieces": r[6],
            "time_slot": r[7], "shift_type": r[8], "remark": r[9],
            "created_at": r[10],  # 数据库中存储的是系统时间，直接返回
            "duration": r[11],  # 时长(分钟)
            "is_synced": r[12],
            "excluded_pieces": r[13],
            "pieces_actual": round(
                _py_inbound_arrival_pieces(r[2], r[3], r[6], r[13]),
                2,
            ),
        } for r in inbound_cur.fetchall()]
        
        # 查询指定日期的分拣记录（按照分拣日期逻辑查询，查询当天00:00之后到次日00:00之前的所有记录）
        sorting_query = """
            SELECT id, sorting_time, pieces, remark, created_at, time_slot
            FROM sorting_records 
            WHERE 
                sorting_time >= ? AND sorting_time < ?
            ORDER BY created_at DESC
        """
        sorting_cur = conn.cursor(); sorting_cur.execute(sorting_query, (
            query_start.strftime('%Y-%m-%d'), 
            query_end.strftime('%Y-%m-%d')
        ))
        sorting_rows = [{
            "id": r[0], "sorting_time": r[1], "pieces": r[2], "remark": r[3],
            "created_at": r[4], "time_slot": r[5]
        } for r in sorting_cur.fetchall()]
        
        # 计算统计信息（总件数=实到件数口径）
        total_vehicles = len(inbound_rows)
        total_pieces = int(
            round(
                sum(
                    _py_inbound_arrival_pieces(
                        record.get("vehicle_type"),
                        record.get("vehicle_no"),
                        record.get("pieces"),
                        record.get("excluded_pieces"),
                    )
                    for record in inbound_rows
                )
            )
        )
        record_count = len(inbound_rows) + len(sorting_rows)
        
        # [新增] 计算预计时段到达车次 (下一个小时的时间段均值) - 用于摘要卡片替换
        predicted_next_hour_vehicles = 0
        try:
            now_la = datetime.now(la_tz)
            next_hour_la = now_la + timedelta(hours=1)
            next_time_slot = str(next_hour_la.hour)
            
            predict_query = """
                SELECT AVG(daily_count) FROM (
                    SELECT COUNT(*) as daily_count 
                    FROM inbound_records 
                    WHERE time_slot = ? 
                    GROUP BY DATE(created_at)
                )
            """
            predict_cur = conn.cursor()
            predict_cur.execute(predict_query, (next_time_slot,))
            predict_result = predict_cur.fetchone()
            predicted_val = predict_result[0] if predict_result else 0
            predicted_next_hour_vehicles = int(round(float(predicted_val))) if predicted_val else 0
        except Exception as e:
            print(f"Error calculating predicted vehicles in history: {e}")
            pass
        
        # 各车型统计（件数为实到口径）
        vehicle_stats = {}
        for record in inbound_rows:
            vehicle_type = record.get("vehicle_type", "未知")
            if vehicle_type not in vehicle_stats:
                vehicle_stats[vehicle_type] = {"count": 0, "pieces": 0.0}
            vehicle_stats[vehicle_type]["count"] += 1
            vehicle_stats[vehicle_type]["pieces"] += _py_inbound_arrival_pieces(
                record.get("vehicle_type"),
                record.get("vehicle_no"),
                record.get("pieces"),
                record.get("excluded_pieces"),
            )
        
        # 转换为列表格式
        vehicle_stats_list = [
            {
                "vehicle_type": vt,
                "count": stats["count"],
                "total_pieces": int(round(stats["pieces"])),
            }
            for vt, stats in vehicle_stats.items()
        ]
        
        conn.close()
        
        return jsonify({
            "records": inbound_rows,
            "sorting_records": sorting_rows,
            "stats": {
                "total_vehicles": total_vehicles,
                "total_pieces": total_pieces,
                "record_count": record_count,
                "predicted_next_hour_vehicles": predicted_next_hour_vehicles,
                "vehicle_stats": vehicle_stats_list
            }
        })
    except Exception as e:
        conn.close()
        return jsonify({"error": f"处理历史记录查询时出错: {str(e)}"}), 500

@app.route('/api/history/export')
def export_history_excel():
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    # Allow exporting all if no date provided? Better restrict to a range or single day default
    if not start_date_str:
        return jsonify({"error": "请提供开始日期"}), 400
        
    if not end_date_str:
        end_date_str = start_date_str
        
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({"error": "日期格式无效"}), 400
        
    conn = get_db()
    
    try:
        # Calculate existing logic to fetch records
        next_day_after_end = end_date + timedelta(days=1)
        query_start = datetime.combine(start_date, datetime.min.time())
        query_end = datetime.combine(next_day_after_end, datetime.min.time())
        
        # Fetch Inbound Records
        inbound_query = """
            SELECT id, dock_no, vehicle_type, vehicle_no, unit, load_amount,
                   pieces, time_slot, shift_type, remark, created_at, duration,
                   COALESCE(excluded_pieces, 0)
            FROM inbound_records 
            WHERE created_at >= ? AND created_at < ?
            ORDER BY created_at DESC
        """
        cursor = conn.cursor()
        cursor.execute(inbound_query, (
            query_start.strftime('%Y-%m-%d %H:%M:%S'), 
            query_end.strftime('%Y-%m-%d %H:%M:%S')
        ))
        inbound_rows = cursor.fetchall()
        
        # Keep connection open for summary query
        
        # Create Excel Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "入库记录"
        
        # Headers（件数：录入 + 实到=录入×系数）
        headers = ['ID', '码头', '车型', '车牌', 'Unit', '装载', '录入件数', '实到件数', '时间段', '班次', '备注', '创建时间', '时长(分钟)']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
        # Data
        for row_num, row in enumerate(inbound_rows, 2):
            # 0..11 同上, 12: excluded_pieces
            pa = round(
                _py_inbound_arrival_pieces(row[2], row[3], row[6], row[12] if len(row) > 12 else 0),
                2,
            )
            ws.cell(row=row_num, column=1, value=row[0])
            ws.cell(row=row_num, column=2, value=row[1])
            ws.cell(row=row_num, column=3, value=row[2])
            ws.cell(row=row_num, column=4, value=row[3])
            ws.cell(row=row_num, column=5, value=row[4])
            ws.cell(row=row_num, column=6, value=row[5])
            ws.cell(row=row_num, column=7, value=row[6])
            ws.cell(row=row_num, column=8, value=pa)
            ws.cell(row=row_num, column=9, value=row[7])
            ws.cell(row=row_num, column=10, value=row[8])
            ws.cell(row=row_num, column=11, value=row[9])
            ws.cell(row=row_num, column=12, value=str(row[10]))
            ws.cell(row=row_num, column=13, value=row[11])
            
        # Adjust column widths
        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 15
            
        # Generate Daily Summary Sheet
        ws_summary = wb.create_sheet("每日汇总")
        
        # Summary Headers
        summary_headers = ['日期', '总车次', '总件数']
        for col_num, header in enumerate(summary_headers, 1):
            cell = ws_summary.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
        # Group data by date (using Python to avoid another DB query if possible, or just query again for simplicity/reliability)
        # Let's query again to let DB handle aggregation
        _ns = _sql_inbound_net_pieces_actual("")
        summary_query = f"""
            SELECT date(created_at) as record_date, COUNT(*) as daily_count,
                   SUM(CASE WHEN vehicle_type = '53英尺' AND vehicle_no = 'G' THEN 0
                       ELSE ({_ns}) END) as daily_pieces
            FROM inbound_records
            WHERE created_at >= ? AND created_at < ?
            GROUP BY date(created_at)
            ORDER BY date(created_at)
        """
        # Note: SQLite uses strftime('%Y-%m-%d', created_at) or date(created_at) depending on version/storage
        # If Postgres: date(created_at)
        # Both support date() function usually. Let's try to be generic or use our convert_sql if needed.
        # But get_db() handles connections. `date()` is standard enough.
        
        cursor_summary = conn.cursor()
        cursor_summary.execute(summary_query, (
            query_start.strftime('%Y-%m-%d %H:%M:%S'), 
            query_end.strftime('%Y-%m-%d %H:%M:%S')
        ))
        summary_rows = cursor_summary.fetchall()
        
        for row_num, row in enumerate(summary_rows, 2):
            ws_summary.cell(row=row_num, column=1, value=str(row[0])) # date
            ws_summary.cell(row=row_num, column=2, value=row[1])      # count
            ws_summary.cell(row=row_num, column=3, value=row[2])      # pieces

        # Adjust summary column widths
        for col_num in range(1, len(summary_headers) + 1):
            col_letter = get_column_letter(col_num)
            ws_summary.column_dimensions[col_letter].width = 15
            
        conn.close()
            
        # Generate response
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"history_records_{start_date_str}_to_{end_date_str}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        if 'conn' in locals():
            conn.close()
        print(f"Export error: {e}")
        return jsonify({"error": str(e)}), 500

def _operation_type_to_ui_level(operation_type):
    """与前端筛选一致：INSERT 等 → INFO，UPDATE → WARNING，DELETE → ERROR"""
    if not operation_type:
        return 'INFO'
    u = str(operation_type).strip().upper()
    if u == 'DELETE':
        return 'ERROR'
    if u == 'UPDATE':
        return 'WARNING'
    return 'INFO'


def _require_logs_api_access():
    """与 /logs 页面一致：须登录；admin/boss 放行，否则需 logs 页 view 权限。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if session.get('role') not in ('admin', 'boss'):
        if not check_user_permission('logs', 'view'):
            return jsonify({'error': '权限不足'}), 403
    return None


@app.route('/api/logs')
def get_operation_logs():
    """获取操作日志列表（分页；字段与 logs.html 约定：logs[].timestamp / level / message）"""
    auth = _require_logs_api_access()
    if auth is not None:
        return auth
    try:
        page = max(1, int(request.args.get('page', 1)))
        per_page = min(100, max(1, int(request.args.get('per_page', 20))))
        date_from = (request.args.get('date_from') or '').strip()
        date_to = (request.args.get('date_to') or '').strip()
        level_filter = (request.args.get('level') or '').strip().upper()

        ph = get_placeholder()
        conditions = []
        params = []

        if date_from:
            conditions.append(f"created_at >= {ph}")
            params.append(f"{date_from} 00:00:00")
        if date_to:
            conditions.append(f"created_at <= {ph}")
            params.append(f"{date_to} 23:59:59")

        if level_filter == 'INFO':
            conditions.append(f"(UPPER(COALESCE(operation_type, '')) NOT IN ('UPDATE', 'DELETE'))")
        elif level_filter == 'WARNING':
            conditions.append(f"(UPPER(COALESCE(operation_type, '')) = 'UPDATE')")
        elif level_filter == 'ERROR':
            conditions.append(f"(UPPER(COALESCE(operation_type, '')) = 'DELETE')")

        where_sql = " WHERE " + " AND ".join(conditions) if conditions else ""

        conn = get_db()
        cursor = conn.cursor()

        count_sql = f"SELECT COUNT(*) FROM operation_logs{where_sql}"
        cursor.execute(convert_query_placeholders(count_sql), tuple(params))
        total = cursor.fetchone()[0]
        total = int(total) if total is not None else 0

        offset = (page - 1) * per_page
        data_sql = f"""
            SELECT id, operation_type, table_name, record_id, old_data, new_data, operator, created_at
            FROM operation_logs
            {where_sql}
            ORDER BY created_at DESC
            LIMIT {ph} OFFSET {ph}
        """
        cursor.execute(convert_query_placeholders(data_sql), tuple(params + [per_page, offset]))

        rows = cursor.fetchall()
        result_logs = []

        for log in rows:
            if hasattr(log, 'keys'):
                log_id = log['id']
                operation_type = log['operation_type']
                table_name = log['table_name']
                record_id = log['record_id']
                old_data = log['old_data']
                new_data = log['new_data']
                operator = _db_row_get(log, 'operator', None)
                created_at = log['created_at']
            else:
                log_id, operation_type, table_name, record_id, old_data, new_data, operator, created_at = log

            try:
                old_data_parsed = json.loads(old_data) if old_data else {}
            except Exception:
                old_data_parsed = {"raw_data": old_data}

            try:
                new_data_parsed = json.loads(new_data) if new_data else {}
            except Exception:
                new_data_parsed = {"raw_data": new_data}

            ui_level = _operation_type_to_ui_level(operation_type)
            ts = created_at
            if hasattr(ts, 'strftime'):
                ts = ts.strftime('%Y-%m-%d %H:%M:%S')
            elif ts is not None:
                ts = str(ts)

            rid = record_id if record_id is not None else '-'
            msg = f"{operation_type or '-'} {table_name or '-'} #{rid}"
            if operator:
                msg += f" · {operator}"

            result_logs.append({
                "id": log_id,
                "operation_type": operation_type,
                "table_name": table_name,
                "record_id": record_id,
                "old_data": old_data_parsed,
                "new_data": new_data_parsed,
                "created_at": created_at,
                "timestamp": ts or 'N/A',
                "level": ui_level,
                "message": msg,
            })

        conn.close()

        pages = max(1, (total + per_page - 1) // per_page) if total else 1
        pagination = {
            "page": page,
            "pages": pages,
            "has_prev": page > 1 and total > 0,
            "has_next": page < pages and total > 0,
            "prev_num": page - 1,
            "next_num": page + 1,
        }

        return jsonify({"logs": result_logs, "pagination": pagination})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/logs/<int:log_id>')
def get_operation_log_detail(log_id):
    """获取单个操作日志详情"""
    auth = _require_logs_api_access()
    if auth is not None:
        return auth
    try:
        conn = get_db()
        cursor = conn.cursor()
        ph = get_placeholder()

        # 查询指定ID的操作日志
        cursor.execute(convert_query_placeholders(f"""
            SELECT id, operation_type, table_name, record_id, old_data, new_data, created_at 
            FROM operation_logs 
            WHERE id = {ph}
        """), (log_id,))
        
        log = cursor.fetchone()
        if not log:
            conn.close()
            return jsonify({"error": "日志未找到"}), 404
        
        log_id, operation_type, table_name, record_id, old_data, new_data, created_at = log
        
        # 解析JSON数据
        try:
            old_data_parsed = json.loads(old_data) if old_data else {}
        except:
            old_data_parsed = {"raw_data": old_data}
            
        try:
            new_data_parsed = json.loads(new_data) if new_data else {}
        except:
            new_data_parsed = {"raw_data": new_data}
        
        result = {
            "id": log_id,
            "operation_type": operation_type,
            "table_name": table_name,
            "record_id": record_id,
            "old_data": old_data_parsed,
            "new_data": new_data_parsed,
            "created_at": created_at
        }
        
        conn.close()
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/sorting', methods=['POST'])
def add_sorting_record():
    data = request.json
    
    conn=get_db()
    
    # 获取当前洛杉矶时间
    la_tz = pytz.timezone('America/Los_Angeles')
    current_la_time = datetime.now(la_tz)
    current_la_time_str = current_la_time.strftime('%Y-%m-%d %H:%M:%S')
    
    conn.cursor().execute("""INSERT INTO sorting_records
        (sorting_time, pieces, remark, time_slot, created_at, manual_count, device_count)
        VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (data.get("sorting_time"), data.get("pieces"), data.get("remark"), data.get("time_slot"), current_la_time_str, data.get("manual_count"), data.get("device_count")))
    conn.commit()
    conn.close()
    return jsonify({"success":True})


def _enrich_hourly_results_gofo_popover(hourly_results):
    """用 popover 覆盖 pieces 并写入 CNO01 字段（与 perform_gofo_hourly_sync 一致）。"""
    _use_popover = os.environ.get("GOFO_USE_POPOVER_HOURLY", "1").strip().lower() not in ("0", "false", "no")
    if not _use_popover:
        return
    try:
        from gofo_popover_collect import (
            la_hour_window_strings,
            popover_destin_hour_totals,
            sum_popover_collect_for_window,
        )

        _cno_destin_id = int(os.environ.get("GOFO_CNO01_DESTIN_ID", "17"))
        for res in hourly_results:
            try:
                hs, he = la_hour_window_strings(res["date"], res["slot"])
                pop = sum_popover_collect_for_window(hs, he)
                if pop is not None:
                    wsum, psum = pop
                    wsum = int(wsum)
                    psum = int(psum)
                    res["pieces"] = wsum
                    res["popover_pkg"] = psum
                    # 人工/设备保持 Gofo overview 原值，不按运单数比例缩放（与看板 collectArtificial/Device 一致）
                else:
                    res["popover_pkg"] = None
                cno = popover_destin_hour_totals(hs, he, _cno_destin_id)
                if cno is not None:
                    res["cno01_waybill"] = int(cno[0])
                    res["cno01_pkg"] = int(cno[1])
                else:
                    res["cno01_waybill"] = None
                    res["cno01_pkg"] = None
            except Exception as ex:
                print(f"[GofoPopover] hour {res.get('slot')} skip: {ex}")
                res["popover_pkg"] = None
                res["cno01_waybill"] = None
    except Exception as ex:
        print(f"[GofoPopover] module skip: {ex}")


def _write_hourly_results_sorting_and_cno(
    hourly_results,
    remark,
    current_la_time_str,
    *,
    gofo_hourly_stats=None,
):
    """
    写入 sorting_records 与 gofo_collect_destin_hourly。
    gofo_hourly_stats: None，或 dict(report_hour, collect_total, collect_artificial, collect_device) 以插入 gofo_hourly_stats。
    返回 {synced_count, synced_hour, pieces}
    """
    conn = get_db()
    cursor = conn.cursor()
    synced_count = 0
    last_hour_saved = ""
    last_pieces_saved = 0
    _cno_site = os.environ.get("GOFO_CNO01_SITE", "CNO01").strip() or "CNO01"
    _cno_destin_id = int(os.environ.get("GOFO_CNO01_DESTIN_ID", "17"))
    try:
        if gofo_hourly_stats is not None:
            cursor.execute(
                """
                INSERT INTO gofo_hourly_stats (report_hour, collect_total, collect_artificial, collect_device)
                VALUES (?, ?, ?, ?)
                """,
                (
                    gofo_hourly_stats["report_hour"],
                    gofo_hourly_stats["collect_total"],
                    gofo_hourly_stats["collect_artificial"],
                    gofo_hourly_stats["collect_device"],
                ),
            )
        for res in hourly_results:
            rmk = remark
            if res.get("popover_pkg") is not None:
                rmk += f" popoverPkg={res['popover_pkg']}"
            cursor.execute(
                """
                SELECT id FROM sorting_records
                WHERE sorting_time = ? AND time_slot = ?
                """,
                (res["date"], res["slot"]),
            )
            row = cursor.fetchone()
            if row:
                cursor.execute(
                    """
                    UPDATE sorting_records SET pieces = ?, manual_count = ?, device_count = ?, remark = ?
                    WHERE id = ?
                    """,
                    (res["pieces"], res["manual"], res["device"], rmk, row[0]),
                )
            else:
                cursor.execute(
                    """
                    INSERT INTO sorting_records (sorting_time, pieces, remark, time_slot, created_at, manual_count, device_count)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        res["date"],
                        res["pieces"],
                        rmk,
                        res["slot"],
                        current_la_time_str,
                        res["manual"],
                        res["device"],
                    ),
                )

            if res.get("cno01_waybill") is not None:
                try:
                    if USE_POSTGRES:
                        cursor.execute(
                            """
                            INSERT INTO gofo_collect_destin_hourly
                            (sorting_time, time_slot, destin_site, destin_id, waybill_no_total, package_no_total, updated_at)
                            VALUES (%s,%s,%s,%s,%s,%s,%s)
                            ON CONFLICT (sorting_time, time_slot, destin_site) DO UPDATE SET
                            waybill_no_total = EXCLUDED.waybill_no_total,
                            package_no_total = EXCLUDED.package_no_total,
                            updated_at = EXCLUDED.updated_at
                            """,
                            (
                                res["date"],
                                res["slot"],
                                _cno_site,
                                _cno_destin_id,
                                res["cno01_waybill"],
                                res.get("cno01_pkg") or 0,
                                current_la_time_str,
                            ),
                        )
                    else:
                        cursor.execute(
                            """
                            INSERT OR REPLACE INTO gofo_collect_destin_hourly
                            (sorting_time, time_slot, destin_site, destin_id, waybill_no_total, package_no_total, updated_at)
                            VALUES (?,?,?,?,?,?,?)
                            """,
                            (
                                res["date"],
                                res["slot"],
                                _cno_site,
                                _cno_destin_id,
                                res["cno01_waybill"],
                                res.get("cno01_pkg") or 0,
                                current_la_time_str,
                            ),
                        )
                except Exception as _e:
                    print(f"[gofo_collect_destin_hourly] upsert skip: {_e}")

            synced_count += 1
            last_hour_saved = res["slot"]
            last_pieces_saved = res["pieces"]

        conn.commit()
    except Exception as e:
        conn.rollback()
        raise e
    finally:
        conn.close()

    return {
        "synced_count": synced_count,
        "synced_hour": last_hour_saved,
        "pieces": last_pieces_saved,
    }


def perform_gofo_hourly_sync():
    """核心同步逻辑：从 Gofo 抓取每小时集包数据并存入数据库"""
    from gofo_config import get_gofo_token
    token = get_gofo_token()
    BASE_HEADERS = {
        "Admin-Token": token,
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "User-Agent": "Mozilla/5.0",
        "channel-id": "us",
        "lang": "zh"
    }
    
    # Get Time Range
    la_tz = pytz.timezone('America/Los_Angeles')
    now_la = datetime.now(la_tz)
    today_str = now_la.strftime('%Y-%m-%d')
    start_time = f"{today_str} 00:00:00"
    end_time = now_la.strftime('%Y-%m-%d %H:%M:%S')

    # Trend Lookback (last 24 hours) to ensure historical corrections
    trend_start_dt = now_la - timedelta(hours=24)
    trend_start_str = trend_start_dt.strftime('%Y-%m-%d %H:00:00')

    # 1. Fetch Overview (for gofo_hourly_stats - Today's Total)
    overview_url = "https://dms.gofoexpress.com/prod-api/dbu_report/common/magic/center/board/overview"
    overview_payload = {
        "centerIds": [596],
        "startTime": start_time,
        "endTime": end_time,
        "groupType": 2
    }
    
    # Trend Payload (Wider range for sorting_records)
    trend_payload = {
        "centerIds": [596],
        "startTime": trend_start_str,
        "endTime": end_time,
        "groupType": 2
    }
    
    max_retries = 3
    retry_delay = 2
    overview_json = {}
    
    for attempt in range(max_retries):
        try:
            overview_res = requests.post(overview_url, headers=BASE_HEADERS, json=overview_payload, timeout=20)
            overview_json = overview_res.json()
            break
        except requests.exceptions.RequestException as e:
            if attempt == max_retries - 1:
                raise Exception(f"Gofo API 连接失败 (DNS/网络错误): {str(e)}")
            time.sleep(retry_delay)
    
    if overview_json.get('code') == 401:
        raise Exception("Gofo API 登录失效 (Token Expired). 请更新 Token。")
        
    overview_data = overview_json.get('data') or {}
    
    # Parse "Pallet Count / Piece Count" format - Take pieces (second part)
    def parse_gofo_cnt(val):
        if val is None: return 0
        if isinstance(val, (int, float)): return int(val)
        if not isinstance(val, str) or not val.strip(): return 0
        if '/' in val: 
            try:
                return int(val.split('/')[-1].replace(',', '').strip())
            except:
                return 0
        try:
            return int(val.replace(',', '').strip())
        except:
            return 0

    collect_total = parse_gofo_cnt(overview_data.get('collectTotalCnt'))
    collect_art = parse_gofo_cnt(overview_data.get('collectTotalCntArtificial'))
    collect_device = parse_gofo_cnt(overview_data.get('collectTotalCntDevice'))
    
    # 2. Fetch Trend (for sorting_records)
    chart_url = "https://dms.gofoexpress.com/prod-api/dbu_report/common/magic/center/board/operation/chart_v2"
    chart_json = {}
    
    for attempt in range(max_retries):
        try:
            chart_res = requests.post(chart_url, headers=BASE_HEADERS, json=trend_payload, timeout=20) # Use wider range
            chart_json = chart_res.json()
            break
        except requests.exceptions.RequestException as e:
            if attempt == max_retries - 1:
                raise Exception(f"Gofo API 连接图表数据失败: {str(e)}")
            time.sleep(retry_delay)
    
    if chart_json.get('code') == 401:
        raise Exception("Gofo API 登录失效 (Token Expired). 请更新 Token。")
        
    chart_data = chart_json.get('data') or []
    
    if not chart_data:
        raise Exception("未找到趋势数据 (No trend data found)")

    # 3. Pre-fetch all hourly data before opening database connection
    # This avoids holding a database lock while making slow network requests
    hourly_results = []
    
    for i in range(len(chart_data)):
        item = chart_data[i]
        report_hour_str = item.get('hour')
        if not report_hour_str:
            continue
            
        try:
            hour_start = f"{report_hour_str}:00:00"
            hour_end = f"{report_hour_str}:59:59"
            
            hour_overview_payload = {
                "centerIds": [596],
                "startTime": hour_start,
                "endTime": hour_end,
                "groupType": 2
            }
            hour_json = {}
            for attempt in range(max_retries):
                try:
                    hour_res = requests.post(overview_url, headers=BASE_HEADERS, json=hour_overview_payload, timeout=15)
                    if not hour_res.ok:
                        break
                    hour_json = hour_res.json()
                    break
                except requests.exceptions.RequestException:
                    if attempt < max_retries - 1:
                        time.sleep(1)
            
            if not hour_json or hour_json.get('code') == 401:
                continue
                
            hour_overview = hour_json.get('data') or {}
            if not hour_overview:
                continue
            
            def parse_hourly_cnt(val):
                if val is None: return 0
                if isinstance(val, (int, float)): return int(val)
                if not isinstance(val, str) or not val.strip(): return 0
                if '/' in val: 
                    try:
                        return int(val.split('/')[-1].replace(',', '').strip())
                    except:
                        return 0
                try:
                    return int(val.replace(',', '').strip())
                except:
                    return 0

            hourly_pieces = parse_hourly_cnt(hour_overview.get('collectTotalCnt'))
            hourly_manual = parse_hourly_cnt(hour_overview.get('collectTotalCntArtificial'))
            hourly_device = parse_hourly_cnt(hour_overview.get('collectTotalCntDevice'))

            # Sanity Check Fallback: If total pieces is 0 but children have data, use sum
            if hourly_pieces == 0 and (hourly_manual > 0 or hourly_device > 0):
                hourly_pieces = hourly_manual + hourly_device

            # Parse timestamp
            report_dt = datetime.strptime(report_hour_str, '%Y-%m-%d %H')
            target_date = report_dt.strftime('%Y-%m-%d')
            target_slot = report_dt.strftime('%H:00')
            
            hourly_results.append({
                "date": target_date,
                "slot": target_slot,
                "pieces": hourly_pieces,
                "manual": hourly_manual,
                "device": hourly_device,
                "report_hour": report_hour_str
            })
        except Exception as e:
            print(f"Error fetching data for hour {report_hour_str}: {e}")

    _enrich_hourly_results_gofo_popover(hourly_results)

    latest_report = chart_data[-1]
    report_hour = latest_report.get("hour")
    base_remark = f"Auto-synced from Gofo ({now_la.strftime('%H:%M')})"
    current_la_time_str = now_la.strftime("%Y-%m-%d %H:%M:%S")
    wr = _write_hourly_results_sorting_and_cno(
        hourly_results,
        base_remark,
        current_la_time_str,
        gofo_hourly_stats={
            "report_hour": report_hour,
            "collect_total": collect_total,
            "collect_artificial": collect_art,
            "collect_device": collect_device,
        },
    )
    result = {
        **wr,
        "total_today": collect_total,
        "manual_today": collect_art,
        "device_today": collect_device,
    }
    try:
        import sync_daily_packing_board as _dp_board

        anchor = now_la.date()
        for wm in ("calendar", "business", "seventeen"):
            threading.Thread(
                target=_dp_board.sync_daily_packing_board_anchor,
                args=(anchor, wm),
                kwargs={"force": True},
                daemon=True,
            ).start()
    except Exception as _ex:
        print(f"[daily_packing cache] post-sync skip: {_ex}")
    return result


def perform_gofo_backfill_range(start_date_str, end_date_str):
    """按日历日从 Gofo 拉取 chart_v2 + 逐小时 overview，可选 popover 覆盖件数并写 CNO01（与 perform_gofo_hourly_sync 一致），
    回填 sorting_records 与 gofo_collect_destin_hourly。用于更正统计图某区间或补今天数据。"""
    try:
        d0 = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        d1 = datetime.strptime(end_date_str, '%Y-%m-%d').date()
    except ValueError:
        raise ValueError('日期须为 YYYY-MM-DD')
    if d0 > d1:
        raise ValueError('开始日期不能晚于结束日期')

    from gofo_config import get_gofo_token
    token = get_gofo_token()
    BASE_HEADERS = {
        "Admin-Token": token,
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "User-Agent": "Mozilla/5.0",
        "channel-id": "us",
        "lang": "zh",
    }
    overview_url = "https://dms.gofoexpress.com/prod-api/dbu_report/common/magic/center/board/overview"
    chart_url = "https://dms.gofoexpress.com/prod-api/dbu_report/common/magic/center/board/operation/chart_v2"
    la_tz = pytz.timezone('America/Los_Angeles')
    now_la = datetime.now(la_tz)
    current_la_time_str = now_la.strftime('%Y-%m-%d %H:%M:%S')
    remark = f"Gofo backfill {start_date_str}~{end_date_str} ({now_la.strftime('%m-%d %H:%M')})"

    def parse_hourly_cnt(val):
        if val is None:
            return 0
        if isinstance(val, (int, float)):
            return int(val)
        if not isinstance(val, str) or not val.strip():
            return 0
        if '/' in val:
            try:
                return int(val.split('/')[-1].replace(',', '').strip())
            except Exception:
                return 0
        try:
            return int(val.replace(',', '').strip())
        except Exception:
            return 0

    max_retries = 3
    hourly_results = []
    d = d0
    while d <= d1:
        date_str = d.strftime('%Y-%m-%d')
        trend_payload = {
            "centerIds": [596],
            "startTime": f"{date_str} 00:00:00",
            "endTime": f"{date_str} 23:59:59",
            "groupType": 2,
        }
        chart_json = {}
        for attempt in range(max_retries):
            try:
                chart_res = requests.post(chart_url, headers=BASE_HEADERS, json=trend_payload, timeout=30)
                chart_json = chart_res.json()
                break
            except requests.exceptions.RequestException as e:
                if attempt == max_retries - 1:
                    raise Exception(f"Gofo chart_v2 请求失败 ({date_str}): {e}") from e
                time.sleep(1)
        if chart_json.get('code') == 401:
            raise Exception("Gofo API 登录失效 (Token Expired). 请更新 Token。")
        chart_data = chart_json.get('data') or []
        if not chart_data:
            print(f"[Gofo backfill] 无趋势数据: {date_str}")

        for i in range(len(chart_data)):
            item = chart_data[i]
            report_hour_str = item.get('hour')
            if not report_hour_str:
                continue
            try:
                hour_start = f"{report_hour_str}:00:00"
                hour_end = f"{report_hour_str}:59:59"
                hour_overview_payload = {
                    "centerIds": [596],
                    "startTime": hour_start,
                    "endTime": hour_end,
                    "groupType": 2,
                }
                hour_json = {}
                for attempt in range(max_retries):
                    try:
                        hour_res = requests.post(overview_url, headers=BASE_HEADERS, json=hour_overview_payload, timeout=15)
                        if not hour_res.ok:
                            break
                        hour_json = hour_res.json()
                        break
                    except requests.exceptions.RequestException:
                        if attempt < max_retries - 1:
                            time.sleep(1)
                if not hour_json or hour_json.get('code') == 401:
                    continue
                hour_overview = hour_json.get('data') or {}
                if not hour_overview:
                    continue
                hourly_pieces = parse_hourly_cnt(hour_overview.get('collectTotalCnt'))
                hourly_manual = parse_hourly_cnt(hour_overview.get('collectTotalCntArtificial'))
                hourly_device = parse_hourly_cnt(hour_overview.get('collectTotalCntDevice'))
                if hourly_pieces == 0 and (hourly_manual > 0 or hourly_device > 0):
                    hourly_pieces = hourly_manual + hourly_device
                report_dt = datetime.strptime(report_hour_str, '%Y-%m-%d %H')
                target_date = report_dt.strftime('%Y-%m-%d')
                target_slot = report_dt.strftime('%H:00')
                hourly_results.append({
                    "date": target_date,
                    "slot": target_slot,
                    "pieces": hourly_pieces,
                    "manual": hourly_manual,
                    "device": hourly_device,
                    "report_hour": report_hour_str,
                })
            except Exception as e:
                print(f"[Gofo backfill] hour {report_hour_str}: {e}")

        d += timedelta(days=1)
        time.sleep(0.25)

    _enrich_hourly_results_gofo_popover(hourly_results)
    wr = _write_hourly_results_sorting_and_cno(
        hourly_results,
        remark,
        current_la_time_str,
        gofo_hourly_stats=None,
    )

    try:
        import sync_daily_packing_board as _dp_board

        d_cache = d0
        while d_cache <= d1:
            for wm in ("calendar", "business", "seventeen"):
                try:
                    _dp_board.sync_daily_packing_board_anchor(d_cache, wm, force=True)
                except Exception as _bc:
                    print(f"[daily_packing board] backfill cache {d_cache} {wm}: {_bc}")
            d_cache += timedelta(days=1)
    except Exception as _bx:
        print(f"[daily_packing board] backfill cache skip: {_bx}")

    return {
        "success": True,
        "synced_count": wr["synced_count"],
        "start_date": start_date_str,
        "end_date": end_date_str,
        "hour_slots": len(hourly_results),
        "synced_hour": wr.get("synced_hour"),
        "pieces": wr.get("pieces"),
    }


def perform_gofo_backfill_today():
    """按洛杉矶「今天」的日历日重拉 chart_v2 + 每小时 overview + popover/CNO01，写 sorting_records 与 gofo_collect_destin_hourly。"""
    la_tz = pytz.timezone("America/Los_Angeles")
    today_str = datetime.now(la_tz).strftime("%Y-%m-%d")
    return perform_gofo_backfill_range(today_str, today_str)


@app.route('/api/gofo/sync_hourly', methods=['POST'])
def sync_gofo_hourly():
    try:
        result = perform_gofo_hourly_sync()
        try:
            import sync_cno_narrowbelt_hourly as _cno_nb

            _cno_nb.sync_today_la_hours()
        except Exception as _nb_e:
            print(f"[GofoManualSync] cno narrowbelt: {_nb_e}")
        synced_count = result.get('synced_count', 0)
        pieces = result.get('pieces', 0)
        hour = result.get('synced_hour')
        manual_today = result.get('manual_today', 0)
        device_today = result.get('device_today', 0)
        
        update_gofo_sync_status(
            "success", 
            f"Manual sync: {synced_count} hours", 
            synced_count,
            pieces,
            hour,
            manual_today,
            device_today
        )
        
        # 记录到历史记录
        log_gofo_sync_event("manual", "success", f"Manual sync completed: {synced_count} hours", synced_count, pieces, hour, manual_today, device_today)
        
        return jsonify({
            "success": True, 
            **result
        })
    except Exception as e:
        error_msg = str(e)
        update_gofo_sync_status("error", f"Manual sync error: {error_msg}")
        log_gofo_sync_event("manual", "error", f"Manual sync failed: {error_msg}")
        return jsonify({"success": False, "error": error_msg}), 500

@app.route('/api/gofo/sync_status', methods=['GET'])
def get_gofo_sync_status():
    global gofo_sync_status
    return jsonify(gofo_sync_status)

@app.route('/api/gofo/sync_history', methods=['GET'])
def get_gofo_sync_history():
    """获取 Gofo 同步历史记录（表 gofo_sync_history）。支持按日期、类型筛选。

    Query: start_date, end_date (YYYY-MM-DD), sync_type (auto|manual), limit (默认 50，最大 500)
    """
    try:
        limit = request.args.get('limit', 50, type=int)
        if not limit or limit < 1:
            limit = 50
        limit = min(limit, 500)

        start_date = request.args.get('start_date', type=str)
        end_date = request.args.get('end_date', type=str)
        sync_type = request.args.get('sync_type', type=str)

        if start_date:
            try:
                datetime.strptime(start_date.strip(), "%Y-%m-%d")
            except ValueError:
                return jsonify({"error": "Invalid start_date, use YYYY-MM-DD"}), 400
        if end_date:
            try:
                datetime.strptime(end_date.strip(), "%Y-%m-%d")
            except ValueError:
                return jsonify({"error": "Invalid end_date, use YYYY-MM-DD"}), 400
        if start_date and end_date:
            if datetime.strptime(start_date, "%Y-%m-%d") > datetime.strptime(end_date, "%Y-%m-%d"):
                return jsonify({"error": "start_date must be <= end_date"}), 400

        conditions = []
        params = []

        if start_date:
            conditions.append("sync_time >= ?")
            params.append(f"{start_date.strip()} 00:00:00")

        if end_date:
            end_dt = datetime.strptime(end_date.strip(), "%Y-%m-%d") + timedelta(days=1)
            conditions.append("sync_time < ?")
            params.append(end_dt.strftime("%Y-%m-%d %H:%M:%S"))

        if sync_type in ('auto', 'manual'):
            conditions.append("sync_type = ?")
            params.append(sync_type)

        where_sql = (" WHERE " + " AND ".join(conditions)) if conditions else ""
        sql = f"""
            SELECT id, sync_time, sync_type, status, message, synced_count, last_pieces, last_hour, manual_count, device_count
            FROM gofo_sync_history
            {where_sql}
            ORDER BY sync_time DESC
            LIMIT ?
        """
        params.append(limit)

        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(sql, tuple(params))
        rows = cursor.fetchall()

        history = []
        for r in rows:
            history.append({
                "id": _db_row_get(r, 'id'),
                "sync_time": _db_row_get(r, 'sync_time'),
                "sync_type": _db_row_get(r, 'sync_type'),
                "status": _db_row_get(r, 'status'),
                "message": _db_row_get(r, 'message'),
                "synced_count": _db_row_get(r, 'synced_count'),
                "last_pieces": _db_row_get(r, 'last_pieces'),
                "last_hour": _db_row_get(r, 'last_hour'),
                "manual_count": _db_row_get(r, 'manual_count', 0),
                "device_count": _db_row_get(r, 'device_count', 0),
            })

        conn.close()
        return jsonify(history)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/sorting', methods=['GET'])
def get_sorting_records():
    conn=get_db()
    cur = conn.cursor(); cur.execute("""SELECT id, sorting_time, pieces, remark, created_at, time_slot, manual_count, device_count
                        FROM sorting_records ORDER BY sorting_time DESC, time_slot DESC""")
    rows=[{
        "id":r[0], "sorting_time":r[1], "pieces":r[2], "remark":r[3],
        "created_at":convert_utc_to_la(r[4]), "time_slot":r[5],
        "manual_count": r[6], "device_count": r[7]
    } for r in cur.fetchall()]
    return jsonify(rows)
    
@app.route('/api/sorting/export', methods=['GET'])
def export_sorting_records():
    """将分拣记录导出为 Excel"""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute("""
            SELECT id, sorting_time, pieces, manual_count, device_count, remark, created_at, time_slot
            FROM sorting_records 
            ORDER BY sorting_time DESC, time_slot DESC
        """)
        records = cur.fetchall()
        
        # 创建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "分拣记录"
        
        # 表头
        headers = ['ID', '分拣日期', '时间段', '总件数', '人工集包', '设备集包', '备注', '创建时间']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
        # 填充数据
        for row_num, row in enumerate(records, 2):
            # 数据库字段索引: 0:id, 1:sorting_time, 2:pieces, 3:manual_count, 4:device_count, 5:remark, 6:created_at, 7:time_slot
            ws.cell(row=row_num, column=1, value=row[0]) # ID
            ws.cell(row=row_num, column=2, value=row[1]) # 分拣日期
            ws.cell(row=row_num, column=3, value=row[7]) # 时间段
            ws.cell(row=row_num, column=4, value=row[2]) # 总件数
            ws.cell(row=row_num, column=5, value=row[3]) # 人工集包
            ws.cell(row=row_num, column=6, value=row[4]) # 设备集包
            ws.cell(row=row_num, column=7, value=row[5]) # 备注
            ws.cell(row=row_num, column=8, value=str(row[6])) # 创建时间
            
        # 自动调整列宽
        for col_num in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = 15
            
        conn.close()
        
        # 将 Excel 写入内存流
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 获取当前日期用于文件名
        today_str = datetime.now().strftime('%Y%m%d')
        filename = f"sorting_records_{today_str}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        if 'conn' in locals():
            conn.close()
        print(f"Export sorting records error: {e}")
        return jsonify({"error": str(e)}), 500

@app.route('/api/sorting/<int:record_id>', methods=['PUT'])
def update_sorting_record(record_id):
    data = request.json
    conn = None
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # 1. Get old data for logging
        cursor.execute("SELECT sorting_time, pieces, remark, time_slot, manual_count, device_count FROM sorting_records WHERE id=?", (record_id,))
        old_row = cursor.fetchone()
        if not old_row:
            conn.close()
            return jsonify({"success": False, "error": "Record not found"}), 404
            
        old_data = {
            "sorting_time": old_row[0],
            "pieces": old_row[1],
            "remark": old_row[2],
            "time_slot": old_row[3],
            "manual_count": old_row[4],
            "device_count": old_row[5]
        }
        
        # 2. Update record
        new_data = {
            "sorting_time": data.get("sorting_time", old_data["sorting_time"]),
            "pieces": data.get("pieces", old_data["pieces"]),
            "remark": data.get("remark", old_data["remark"]),
            "time_slot": data.get("time_slot", old_data["time_slot"]),
            "manual_count": data.get("manual_count", old_data["manual_count"]),
            "device_count": data.get("device_count", old_data["device_count"])
        }
        
        cursor.execute("""
            UPDATE sorting_records 
            SET sorting_time = ?, pieces = ?, remark = ?, time_slot = ?, manual_count = ?, device_count = ?
            WHERE id = ?
        """, (new_data["sorting_time"], new_data["pieces"], new_data["remark"], new_data["time_slot"], new_data["manual_count"], new_data["device_count"], record_id))
        
        # 3. Log operation
        cursor.execute("""
            INSERT INTO operation_logs (operation_type, table_name, record_id, old_data, new_data)
            VALUES (?, ?, ?, ?, ?)
        """, ('update', 'sorting_records', record_id, json.dumps(old_data, default=str), json.dumps(new_data, default=str)))
        
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    except Exception as e:
        if conn: conn.close()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/sorting/<int:record_id>', methods=['DELETE'])
def delete_sorting_record(record_id):
    conn = None
    try:
        conn = get_db()
        
        # 获取删除前的数据
        old_record_cur = conn.cursor(); old_record_cur.execute("SELECT * FROM sorting_records WHERE id=?", (record_id,))
        old_record = old_record_cur.fetchone()
        
        cursor = conn.cursor()
        cursor.execute("DELETE FROM sorting_records WHERE id=?", (record_id,))
        
        # 如果记录被成功删除，记录日志
        if cursor.rowcount > 0:
            # 记录操作日志
            column_names = [description[0] for description in old_record_cur.description]
            old_data = dict(zip(column_names, old_record)) if old_record else {}
            
            # 删除游标对象，避免序列化错误
            old_data.pop('cursor', None)
            
            conn.cursor().execute("""INSERT INTO operation_logs 
                (operation_type, table_name, record_id, old_data, new_data)
                VALUES (?, ?, ?, ?, ?)""",
                ('delete', 'sorting_records', record_id, 
                 json.dumps(old_data, default=str), 
                 json.dumps({})))  # 删除操作没有新数据
        
            conn.commit()
            conn.close()
            return jsonify({"success": True})
        else:
            conn.commit()
            conn.close()
            return jsonify({"success": False, "error": "记录未找到"}), 404
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass

# 获取揽收预估与实际入库对比数据
@app.route('/api/forecast_vs_actual')
def forecast_vs_actual():
    conn = get_db()
    
    # 查询所有有预估数据的日期，按日期升序排列
    forecast_dates_cur = conn.cursor(); forecast_dates_cur.execute("""
        SELECT forecast_date, forecast_amount 
        FROM pickup_forecast 
        ORDER BY forecast_date ASC
    """)
    
    forecast_records = forecast_dates_cur.fetchall()
    
    dates = []
    forecast_data = []
    actual_data = []
    difference_percent = []
    
    # 遍历所有有预估数据的日期
    for record in forecast_records:
        date_str = record[0]
        forecast_val = record[1]
        
        dates.append(date_str)
        forecast_data.append(forecast_val)
        
        # 查询该日期的实际入库数据
        try:
            # 兼容处理: PostgreSQL 可能直接返回 date 对象, SQLite 返回字符串
            if isinstance(date_str, str):
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
            else:
                date = date_str # 假设是 date 或 datetime 对象
            # 运营日: 05:00 到 次日 05:00 (洛杉矶时间)
            next_day = date + timedelta(days=1)
            req_5am_la = LA_TZ.localize(datetime.combine(date, datetime.min.time().replace(hour=5)))
            next_5am_la = LA_TZ.localize(datetime.combine(next_day, datetime.min.time().replace(hour=5)))
            
            # 转换到数据库使用的本地时区
            day_start = req_5am_la.astimezone(LA_TZ)
            day_end = next_5am_la.astimezone(LA_TZ)
            
            _net_pf = _sql_inbound_net_pieces_actual("")
            actual_cur = conn.cursor(); actual_cur.execute(f"""
                SELECT SUM(CASE 
                    WHEN vehicle_type = '53英尺' AND vehicle_no = 'G' THEN 0 
                    ELSE ({_net_pf})
                END) as total_pieces
                FROM inbound_records
                WHERE created_at >= ? AND created_at < ?
            """, (day_start.strftime('%Y-%m-%d %H:%M:%S'), 
                  day_end.strftime('%Y-%m-%d %H:%M:%S')))
            
            actual_row = actual_cur.fetchone()
            actual_val = int(actual_row[0]) if actual_row and actual_row[0] else 0
            
            # Use actual volume directly
            actual_val = int(actual_val)
            
            actual_data.append(actual_val)
            
            # 计算差异百分比
            if forecast_val > 0:
                diff_pct = round((actual_val - forecast_val) / forecast_val * 100, 1)
            else:
                diff_pct = 0
            difference_percent.append(diff_pct)
        except Exception as e:
            print(f"处理日期 {date_str} 时出错: {e}")
            actual_data.append(0)
            difference_percent.append(0)
    
    conn.close()
    
    return jsonify({
        "dates": dates,
        "forecast": forecast_data,
        "actual": actual_data,
        "difference_percent": difference_percent
    })

# ============================================================================
# SSE Endpoint for Real-Time Updates
# ============================================================================

@app.route('/api/sse/updates')
def sse_updates():
    """
    Server-Sent Events 端点，用于推送实时更新到所有连接的客户端
    支持跨设备、跨浏览器的实时同步
    """
    def event_stream():
        # 为这个客户端创建一个队列
        q = Queue()
        with sse_queues_lock:
            sse_queues.append(q)
            client_count = len(sse_queues)
        
        print(f"[SSE] New client connected. Total clients: {client_count}")
        
        try:
            # 发送初始连接消息
            yield f"data: {json.dumps({'type': 'connected', 'message': 'SSE connection established'})}\n\n"
            
            # 持续监听队列中的事件
            while True:
                try:
                    # 等待事件，超时30秒发送心跳
                    message = q.get(timeout=30)
                    yield f"data: {json.dumps(message)}\n\n"
                except:
                    # 超时，发送心跳保持连接
                    yield f"data: {json.dumps({'type': 'heartbeat', 'timestamp': datetime.now(LA_TZ).isoformat()})}\n\n"
        finally:
            # 客户端断开时，移除队列
            with sse_queues_lock:
                if q in sse_queues:
                    sse_queues.remove(q)
                    client_count = len(sse_queues)
            print(f"[SSE] Client disconnected. Remaining clients: {client_count}")
    
    return Response(event_stream(), mimetype='text/event-stream', headers={
        'Cache-Control': 'no-cache',
        'X-Accel-Buffering': 'no',  # Disable nginx buffering
        'Connection': 'keep-alive'
    })

# ============================================================================
# 人工多班次产能（重叠时段叠加）与统一分拣小时产能
# ============================================================================

def _manual_ppp_from_config(mcfg):
    if not isinstance(mcfg, dict):
        mcfg = {}
    _ppp = mcfg.get('piecesPerPersonPerHour')
    if _ppp is None:
        _pps = mcfg.get('peoplePerStation')
        _cap_legacy = mcfg.get('capacity', 600)
        try:
            _cap_legacy = float(_cap_legacy)
        except (TypeError, ValueError):
            _cap_legacy = 600.0
        if _pps and float(_pps) > 0:
            _ppp = _cap_legacy / float(_pps)
        else:
            _ppp = _cap_legacy
    try:
        return float(_ppp)
    except (TypeError, ValueError):
        return 600.0


def _manual_shifts_for_weekday(mcfg, weekday_index):
    """weekday_index 0=Mon … 6=Sun，返回 [{start, hours, people}, …]。"""
    if not isinstance(mcfg, dict):
        mcfg = {}
    sbd = mcfg.get('shiftsByDay')
    if isinstance(sbd, list) and len(sbd) > weekday_index:
        raw = sbd[weekday_index]
        if isinstance(raw, list) and len(raw) > 0:
            out = []
            for s in raw:
                if not isinstance(s, dict):
                    continue
                try:
                    start = float(s.get('start', 17.5))
                    hours = float(s.get('hours', 0))
                    people = int(s.get('people', 0))
                except (TypeError, ValueError):
                    continue
                if hours <= 0 or people < 0:
                    continue
                out.append({'start': start, 'hours': hours, 'people': people})
            if out:
                return out
    schedule = mcfg.get('schedule') or [0] * 7
    start_times = mcfg.get('startTimes') or [17.5] * 7
    try:
        hps = float(mcfg.get('hoursPerShift', 9))
    except (TypeError, ValueError):
        hps = 9.0
    if weekday_index < 0 or weekday_index >= len(schedule):
        return []
    try:
        pe = int(schedule[weekday_index]) if schedule[weekday_index] is not None else 0
    except (TypeError, ValueError):
        pe = 0
    try:
        st = float(start_times[weekday_index]) if weekday_index < len(start_times) else 17.5
    except (TypeError, ValueError):
        st = 17.5
    if pe <= 0:
        return []
    return [{'start': st, 'hours': hps, 'people': pe}]


def _manual_day_theoretical_pieces(mcfg, weekday_index):
    """按 5 分钟步长积分当日人工理论件数（与产能计划页多班次逻辑一致）。"""
    ppp = _manual_ppp_from_config(mcfg)
    shifts = _manual_shifts_for_weekday(mcfg, weekday_index)
    if not shifts:
        return 0
    step_h = 5.0 / 60.0
    total = 0.0
    t = 0.0
    while t < 24.0 - 1e-9:
        cap = 0.0
        for sh in shifts:
            st = float(sh['start'])
            hrs = float(sh['hours'])
            pe = int(sh['people'])
            if hrs > 0 and t >= st - 1e-9 and t < st + hrs - 1e-9:
                cap += pe * ppp
        total += cap * step_h
        t += step_h
    return int(round(total))


def _machine_or_night_day_theoretical_pieces(subcfg, weekday_index):
    """分拣机/夜班：道数 × 每班小时 × 每小时产能（件/日）。"""
    if not isinstance(subcfg, dict):
        subcfg = {}
    lines = subcfg.get('machineLines')
    if isinstance(lines, list) and len(lines) > 0:
        try:
            cap = float(subcfg.get('capacity', 4500))
        except (TypeError, ValueError):
            cap = 4500.0
        total = 0.0
        for ln in lines:
            if not isinstance(ln, dict):
                continue
            try:
                hrs = float(ln.get('hours', 0))
            except (TypeError, ValueError):
                hrs = 0.0
            if hrs <= 0:
                continue
            total += cap * hrs
        return int(round(total))
    schedule = subcfg.get('schedule') or [0] * 7
    try:
        lanes = int(schedule[weekday_index]) if 0 <= weekday_index < len(schedule) and schedule[weekday_index] is not None else 0
    except (TypeError, ValueError):
        lanes = 0
    try:
        hours = float(subcfg.get('hoursPerShift', 6))
    except (TypeError, ValueError):
        hours = 6.0
    try:
        cap = float(subcfg.get('capacity', 4500))
    except (TypeError, ValueError):
        cap = 4500.0
    return int(lanes * hours * cap)


def manual_hourly_capacity_at(dt_check, req_date, mcfg, la_tz, ppp):
    """dt_check 为 LA 时区 aware。同一时刻多班次重叠则产能相加（件/小时）。"""
    if not isinstance(mcfg, dict):
        mcfg = {}
    cap = 0.0
    ppp = float(ppp)
    prev_date = req_date - timedelta(days=1)
    for base_date in (prev_date, req_date):
        shifts = _manual_shifts_for_weekday(mcfg, base_date.weekday())
        for sh in shifts:
            sh_start = float(sh['start'])
            hrs = float(sh['hours'])
            pe = int(sh['people'])
            hr = int(sh_start)
            mn = int(round((sh_start - hr) * 60))
            if mn >= 60:
                mn = 59
            start_abs = la_tz.localize(datetime.combine(base_date, datetime.min.time().replace(hour=hr, minute=mn)))
            end_abs = start_abs + timedelta(hours=hrs)
            if start_abs <= dt_check < end_abs:
                cap += pe * ppp
    return cap


def sorting_total_hourly_at(dt_check, req_date, config, la_tz):
    """人工 + 分拣机 + 夜班，件/小时（与 get_statistics 中原逻辑一致，人工改为多班次叠加）。"""
    mcfg = config.get('manual') if isinstance(config.get('manual'), dict) else {}
    ppp = _manual_ppp_from_config(mcfg)
    manual_part = manual_hourly_capacity_at(dt_check, req_date, mcfg, la_tz, ppp)

    msched = config.get('machine', {}) or {}
    nsched = config.get('night', {}) or {}
    machine_schedule = msched.get('schedule', [4, 4, 4, 4, 4, 2, 2])
    night_schedule_val = nsched.get('schedule', [0] * 7)
    if not isinstance(night_schedule_val, list):
        night_schedule_val = [0] * 7
    try:
        machine_hours = float(msched.get('hoursPerShift', 6))
    except (TypeError, ValueError):
        machine_hours = 6.0
    try:
        night_hours = float(nsched.get('hoursPerShift', 6))
    except (TypeError, ValueError):
        night_hours = 6.0
    try:
        machine_cap = float(msched.get('capacity', 4500))
    except (TypeError, ValueError):
        machine_cap = 4500.0
    try:
        night_cap = float(nsched.get('capacity', 4500))
    except (TypeError, ValueError):
        night_cap = 4500.0

    dow = req_date.weekday()
    machine_lanes = int(machine_schedule[dow]) if dow < len(machine_schedule) else 0
    night_lanes = int(night_schedule_val[dow]) if dow < len(night_schedule_val) else 0

    ngt_start = la_tz.localize(datetime.combine(req_date, datetime.min.time().replace(hour=23)))
    machine_part = 0.0
    machine_lines = msched.get('machineLines')
    if isinstance(machine_lines, list) and len(machine_lines) > 0:
        for ln in machine_lines:
            if not isinstance(ln, dict):
                continue
            try:
                st = float(ln.get('start', 17.5))
            except (TypeError, ValueError):
                st = 17.5
            try:
                hrs = float(ln.get('hours', machine_hours))
            except (TypeError, ValueError):
                hrs = machine_hours
            if hrs <= 0:
                continue
            hr = int(st)
            mn = int(round((st - hr) * 60))
            if mn >= 60:
                mn = 59
            mac_start = la_tz.localize(datetime.combine(req_date, datetime.min.time().replace(hour=hr, minute=mn)))
            if mac_start <= dt_check < mac_start + timedelta(hours=hrs):
                machine_part += machine_cap
    else:
        mac_start = la_tz.localize(datetime.combine(req_date, datetime.min.time().replace(hour=17, minute=30)))
        if mac_start <= dt_check < mac_start + timedelta(hours=machine_hours):
            machine_part += machine_lanes * machine_cap
    night_part = 0.0
    if ngt_start <= dt_check < ngt_start + timedelta(hours=night_hours):
        night_part += night_lanes * night_cap

    return manual_part + machine_part + night_part


def simulate_sorting_completion_time(remaining_pieces, start_time, config, la_tz):
    """按 10 分钟步长用 sorting_total_hourly_at 推演完成时刻。"""
    if remaining_pieces <= 0:
        return start_time
    pieces_left = float(remaining_pieces)
    current = start_time
    if current.tzinfo is None:
        current = la_tz.localize(current)
    else:
        current = current.astimezone(la_tz)
    step = timedelta(minutes=10)
    max_loops = 7 * 24 * 6
    loops = 0
    while pieces_left > 1e-6 and loops < max_loops:
        loops += 1
        req_date = current.date()
        cap = sorting_total_hourly_at(current, req_date, config, la_tz)
        if cap <= 0:
            current += step
            continue
        chunk_hours = step.total_seconds() / 3600.0
        chunk_pieces = cap * chunk_hours
        if chunk_pieces >= pieces_left:
            hours_needed = pieces_left / cap
            return current + timedelta(hours=hours_needed)
        pieces_left -= chunk_pieces
        current += step
    return current


def calculate_estimated_completion_time(remaining_pieces, start_time, manual_people, machine_lanes, night_lanes, manual_cap, machine_cap, night_cap):
    """
    根据实际排班逻辑计算预计完成时间（按自然时间累加）
    
    计算方式：
    1. 计算各时段需要的工作小时数
    2. 将所有工作小时数累加
    3. 从当前时间开始，加上总工作小时数（自然时间流逝）
    
    实际班次时间:
    - 人工: 17:30-23:30 (6小时)
    - 机器: 17:30-00:00 (6.5小时)
    - 夜班: 23:30-05:00 (5.5小时)
    
    时段产能:
    - 17:30-23:00: 人工+机器 ((manual_people * manual_cap) + (machine_lanes * machine_cap))；manual_cap 为单人每小时件数
    - 23:00-23:30: 人工+机器 ((manual_people * manual_cap) + (machine_lanes * machine_cap))
    - 23:30-00:00: 机器 ((machine_lanes * machine_cap))
    - 00:00-06:00: 夜班 (night_lanes * night_cap)
    - 06:00-17:30: 休息 (0)
    """
    if remaining_pieces <= 0:
        return start_time
    
    current_time = start_time
    pieces_left = remaining_pieces
    total_work_hours = 0  # 累计需要的工作小时数
    
    # 防止死循环：最多推演7天
    max_loops = 7 * 10
    loops = 0
    
    while pieces_left > 0 and loops < max_loops:
        loops += 1
        
        # 获取当前时间的时、分
        h = current_time.hour
        m = current_time.minute
        
        # 确定当前所处的时段、产能及下一个时间点
        capacity = 0
        next_time = None
        
        if h == 17 and m >= 30:
            # 17:30-18:00: 人工+机器
            capacity = (manual_people * manual_cap) + (machine_lanes * machine_cap)
            next_time = current_time.replace(hour=18, minute=0, second=0, microsecond=0)
            
        elif h >= 18 and h < 23:
            # 18:00-23:00: 人工+机器
            capacity = (manual_people * manual_cap) + (machine_lanes * machine_cap)
            next_time = current_time.replace(hour=23, minute=0, second=0, microsecond=0)
            
        elif h == 23 and m < 30:
            # 23:00-23:30: 人工+机器
            capacity = (manual_people * manual_cap) + (machine_lanes * machine_cap)
            next_time = current_time.replace(hour=23, minute=30, second=0, microsecond=0)
            
        elif (h == 23 and m >= 30) or h == 0:
            # 23:30-00:00: 仅机器 (夜班0点开始)
            if h == 23:
                capacity = machine_lanes * machine_cap
                next_time = (current_time + timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
            else:
                # 00:00-06:00: 仅夜班
                capacity = night_lanes * night_cap
                next_time = current_time.replace(hour=6, minute=0, second=0, microsecond=0)
                
        elif h > 0 and h < 6:
            # 00:00-06:00: 仅夜班
            capacity = night_lanes * night_cap
            next_time = current_time.replace(hour=6, minute=0, second=0, microsecond=0)
            
        elif (h >= 6 and h < 17) or (h == 17 and m < 30):
            # 06:00-17:30: 休息
            capacity = 0
            next_time = current_time.replace(hour=17, minute=30, second=0, microsecond=0)
            
        else:
            # 其他情况（不应该发生）
            capacity = 0
            next_time = current_time + timedelta(hours=1)
        
        # 确保 next_time 比 current_time 晚
        if next_time <= current_time:
            next_time += timedelta(days=1)
        
        # 计算该时段的时长（小时）
        available_hours = (next_time - current_time).total_seconds() / 3600.0
        
        if capacity > 0:
            potential_production = available_hours * capacity
            
            if pieces_left <= potential_production:
                # 能在这个时段内完成
                hours_needed = pieces_left / capacity
                total_work_hours += hours_needed
                pieces_left = 0
                break
            else:
                # 依然做不完，扣减所有可能的产出，跳到下一时段
                pieces_left -= potential_production
                total_work_hours += available_hours
                current_time = next_time
        else:
            # 休息时间，无产出，直接跳过（不累加工作小时）
            current_time = next_time
    
    # 从开始时间加上总工作小时数（自然时间流逝）
    return start_time + timedelta(hours=total_work_hours)


def _v2_parse_start_hours(val, default=17.5):
    """V2 开工时间：支持小数小时或 'HH:MM' 字符串。"""
    if val is None:
        return float(default)
    if isinstance(val, (int, float)):
        try:
            x = float(val)
            return x if x >= 0 else float(default)
        except (TypeError, ValueError):
            return float(default)
    s = str(val).strip()
    if not s:
        return float(default)
    if ':' in s:
        parts = s.split(':', 1)
        try:
            hh = int(parts[0])
            mm = int(parts[1]) if len(parts) > 1 else 0
            return max(0.0, hh + mm / 60.0)
        except (TypeError, ValueError):
            return float(default)
    try:
        x = float(s)
        return x if x >= 0 else float(default)
    except (TypeError, ValueError):
        return float(default)


def _v2_plan_total_headcount(plan):
    """与产能计划看板「总人数」一致：已激活产线人数 + 已激活分拣台人数。"""
    if not isinstance(plan, dict):
        return 0
    resources = plan.get('resources')
    if not isinstance(resources, list):
        return 0
    total = 0
    for r in resources:
        if not isinstance(r, dict):
            continue
        company = str(r.get('company') or '').strip()
        if not company:
            continue
        try:
            hc = int(float(r.get('headcount', 0)))
        except (TypeError, ValueError):
            hc = 0
        total += max(0, hc)
    return total


def _v2_earliest_active_start_datetime(request_date, daily_plan, la_tz):
    """当天 V2 已激活资源的最晚不早于的开工时刻（LA，自然日 request_date）。"""
    if not isinstance(daily_plan, dict):
        return None
    resources = daily_plan.get('resources')
    if not isinstance(resources, list):
        return None
    earliest = None
    for r in resources:
        if not isinstance(r, dict):
            continue
        if not str(r.get('company') or '').strip():
            continue
        st = _v2_parse_start_hours(r.get('startTime'), 17.5)
        h = int(st)
        m = int(round((st - h) * 60))
        if m >= 60:
            h += 1
            m -= 60
        if h >= 24:
            h = 23
            m = 59
        dt = la_tz.localize(datetime.combine(request_date, datetime.min.time().replace(hour=h, minute=m)))
        if earliest is None or dt < earliest:
            earliest = dt
    return earliest


def _v2_plan_to_schedule_config(plan, request_date):
    """将 V2 当天排班（resources + 每单元开工时间）映射为统计计算使用的统一 config。"""
    if not isinstance(plan, dict):
        return None
    resources = plan.get('resources')
    if not isinstance(resources, list):
        return None

    shift_hours = 8.0
    try:
        sh = float(plan.get('defaultShiftHours', plan.get('shiftHours', 8)))
        if sh > 0:
            shift_hours = sh
    except (TypeError, ValueError):
        pass

    active_lines = 0
    machine_lines_meta = []
    station_shifts = []

    for r in resources:
        if not isinstance(r, dict):
            continue
        company = str(r.get('company') or '').strip()
        if not company:
            continue
        r_type = str(r.get('type') or '').strip()
        st = _v2_parse_start_hours(r.get('startTime'), 17.5)
        if r_type == 'line':
            active_lines += 1
            machine_lines_meta.append({'start': st, 'hours': shift_hours})
        elif r_type == 'station':
            try:
                hc = int(float(r.get('headcount', 0)))
            except (TypeError, ValueError):
                hc = 0
            hc = max(0, hc)
            if hc > 0:
                station_shifts.append({'start': st, 'hours': shift_hours, 'people': hc})

    dow = request_date.weekday()  # 0=Mon ... 6=Sun
    machine_schedule = [0] * 7
    machine_schedule[dow] = active_lines
    manual_shifts = [[] for _ in range(7)]
    if station_shifts:
        manual_shifts[dow] = station_shifts

    return {
        "manual": {
            "piecesPerPersonPerHour": 310,
            "defaultShiftHours": shift_hours,
            "shiftsByDay": manual_shifts,
        },
        "machine": {
            "capacity": 3100,
            "hoursPerShift": shift_hours,
            "schedule": machine_schedule,
            "startTimes": [17.5] * 7,
            "machineLines": machine_lines_meta,
        },
        "night": {"capacity": 0, "hoursPerShift": 0, "schedule": [0] * 7},
    }


def _parse_stats_window_param(raw):
    """calendar = 本地自然日 00:00–次日 00:00；business = 当日 05:00–次日 05:00；
    seventeen = 当日 17:00–次日 17:00（均与 inbound_records.created_at 同一 naive 本地时钟）。"""
    if not raw:
        return 'calendar'
    r = str(raw).strip().lower().replace('-', '_')
    if r in ('business', 'biz', 'b', '5', '05', 'shift', 'operational'):
        return 'business'
    if r in ('seventeen', '17', '17_17', '17h', 'ops17', 'day17'):
        return 'seventeen'
    return 'calendar'


def _default_stats_request_date(window_mode):
    now = datetime.now(LA_TZ)
    d = now.date()
    if window_mode == 'business' and now.hour < 5:
        return d - timedelta(days=1)
    if window_mode == 'seventeen' and now.hour < 17:
        return d - timedelta(days=1)
    return d


def _stats_period_bounds(request_date, window_mode):
    next_d = request_date + timedelta(days=1)
    if window_mode == 'business':
        start = datetime.combine(request_date, datetime.min.time().replace(hour=5))
        end = datetime.combine(next_d, datetime.min.time().replace(hour=5))
    elif window_mode == 'seventeen':
        start = datetime.combine(request_date, datetime.min.time().replace(hour=17))
        end = datetime.combine(next_d, datetime.min.time().replace(hour=17))
    else:
        start = datetime.combine(request_date, datetime.min.time())
        end = datetime.combine(next_d, datetime.min.time())
    return start, end


def _sorting_slot_window_sql_binds(window_mode: str, d: date):
    """sorting_time + time_slot，与 stats_window 一致。"""
    ds = d.strftime('%Y-%m-%d')
    nxt = (d + timedelta(days=1)).strftime('%Y-%m-%d')
    if window_mode == 'business':
        sql = (
            "(sorting_time = ? AND time_slot >= '05:00') OR (sorting_time = ? AND time_slot < '05:00')"
        )
        return sql, (ds, nxt)
    if window_mode == 'seventeen':
        sql = (
            "(sorting_time = ? AND time_slot >= '17:00') OR (sorting_time = ? AND time_slot < '17:00')"
        )
        return sql, (ds, nxt)
    return "sorting_time = ?", (ds,)


def _record_date_hour_window_sql_binds(window_mode: str, d: date):
    """record_date + record_hour（签入/集包看板等）。"""
    ds = d.strftime('%Y-%m-%d')
    nxt = (d + timedelta(days=1)).strftime('%Y-%m-%d')
    if window_mode == 'business':
        sql = (
            "(record_date = ? AND record_hour >= '05:00') OR (record_date = ? AND record_hour < '05:00')"
        )
        return sql, (ds, nxt)
    if window_mode == 'seventeen':
        sql = (
            "(record_date = ? AND record_hour >= '17:00') OR (record_date = ? AND record_hour < '17:00')"
        )
        return sql, (ds, nxt)
    return "record_date = ?", (ds,)


def _record_date_slot_window_sql_binds(window_mode: str, d: date):
    """record_date + time_slot（cno_narrowbelt_hourly）。"""
    ds = d.strftime('%Y-%m-%d')
    nxt = (d + timedelta(days=1)).strftime('%Y-%m-%d')
    if window_mode == 'business':
        sql = (
            "(record_date = ? AND time_slot >= '05:00') OR (record_date = ? AND time_slot < '05:00')"
        )
        return sql, (ds, nxt)
    if window_mode == 'seventeen':
        sql = (
            "(record_date = ? AND time_slot >= '17:00') OR (record_date = ? AND time_slot < '17:00')"
        )
        return sql, (ds, nxt)
    return "record_date = ?", (ds,)


@app.route('/api/stats')
def get_statistics():
    # 获取日期参数，默认为今天
    date_str = request.args.get('date')
    window_mode = _parse_stats_window_param(request.args.get('stats_window'))
    
    # 必须与 /api/record 等写入路径一致：USE_POSTGRES 时读 PostgreSQL，禁止只读本地 SQLite 导致统计恒为旧数据/空
    conn = get_db()

    # 获取洛杉矶当前日期
    la_tz = pytz.timezone('America/Los_Angeles')
    
    if date_str:
        # 如果提供了日期参数，使用指定日期
        try:
            request_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            if 'conn' in locals(): conn.close()
            return jsonify({"error": "日期格式无效，请使用YYYY-MM-DD格式"}), 400
    else:
        # 与 inbound_records.created_at 一致：入库使用 datetime.now() 的「服务器本地」naive 时间，
        # 统计窗口必须与同一套本地时间对齐（calendar / business 5–5 均由 _stats_period_bounds 计算）。
        request_date = _default_stats_request_date(window_mode)

    period_start, period_end = _stats_period_bounds(request_date, window_mode)
    
    # 查询属于当前统计窗口的记录（与 created_at 一致）
    records_query = """
        SELECT id, created_at, vehicle_type, time_slot FROM inbound_records 
        WHERE 
            created_at >= ? AND created_at < ?
    """
    records_cur = conn.cursor(); records_cur.execute(records_query, (
        period_start.strftime('%Y-%m-%d %H:%M:%S'), 
        period_end.strftime('%Y-%m-%d %H:%M:%S')
    ))
    records = records_cur.fetchall()
    
    # 总车次和总货物量（查询当天00:00之后到次日00:00之前的所有记录）
    # 特殊规则: 53英尺车牌号为G的车辆,货量不计入总计,但车次计入；件数为实到=(录入-排除)*系数
    _net = _sql_inbound_net_pieces_actual("")
    total_query = f"""
        SELECT COUNT(*) as total_vehicles, 
               SUM(CASE 
                   WHEN vehicle_type = '53英尺' AND vehicle_no = 'G' THEN 0 
                   ELSE ({_net})
               END) as total_pieces 
        FROM inbound_records 
        WHERE 
            created_at >= ? AND created_at < ?
    """
    total_cur = conn.cursor(); total_cur.execute(total_query, (
        period_start.strftime('%Y-%m-%d %H:%M:%S'), 
        period_end.strftime('%Y-%m-%d %H:%M:%S')
    ))
    total_result = total_cur.fetchone()
    total_vehicles = (total_result['total_vehicles'] if USE_POSTGRES else total_result[0]) if total_result else 0
    # 注意: SQLite 中 SUM 可能返回 None, PostgreSQL 也可能
    total_pieces_val = total_result['total_pieces'] if USE_POSTGRES else total_result[1]
    total_pieces = int(total_pieces_val) if total_pieces_val else 0
    
    # 托盘总数（查询当天00:00之后到次日00:00之前，车辆类型为26英尺或53英尺的装载量总和）
    # 特殊规则: 53英尺车牌号为G的车辆,装载量不计入
    pallet_query = """
        SELECT SUM(load_amount - COALESCE(plate_excluded_load, 0)) as total_pallets
        FROM inbound_records 
        WHERE 
            created_at >= ? AND created_at < ? 
            AND (vehicle_type = '16英尺' OR vehicle_type = '26英尺' OR vehicle_type = '53英尺')
            AND NOT (vehicle_type = '53英尺' AND vehicle_no = 'G')
    """
    pallet_cur = conn.cursor(); pallet_cur.execute(pallet_query, (
        period_start.strftime('%Y-%m-%d %H:%M:%S'), 
        period_end.strftime('%Y-%m-%d %H:%M:%S')
    ))
    pallet_result = pallet_cur.fetchone()
    pallet_val = pallet_result['total_pallets'] if USE_POSTGRES else pallet_result[0]
    total_pallets = int(pallet_val) if pallet_val else 0

    # 当日「车牌不计入统计」装载量与件数合计（不计入件数也折算为实到件数口径）
    excluded_totals_query = f"""
        SELECT COALESCE(SUM(plate_excluded_load), 0) AS sum_plate,
               COALESCE(SUM(excluded_pieces * {INBOUND_PIECES_ACTUAL_FACTOR}), 0) AS sum_ex_pieces
        FROM inbound_records
        WHERE created_at >= ? AND created_at < ?
    """
    ex_cur = conn.cursor()
    ex_cur.execute(excluded_totals_query, (
        period_start.strftime('%Y-%m-%d %H:%M:%S'),
        period_end.strftime('%Y-%m-%d %H:%M:%S'),
    ))
    ex_row = ex_cur.fetchone()
    total_plate_excluded_load = float(ex_row[0] or 0) if ex_row else 0.0
    total_excluded_pieces_stat = int(ex_row[1] or 0) if ex_row else 0
    
    # 各车型统计（查询当天00:00之后到次日00:00之前的所有记录）
    # 特殊规则: 53英尺车牌号为G的车辆,车次计入但货量不计入
    vehicle_stats_query = f"""
        SELECT vehicle_type, 
               COUNT(*) as count, 
               SUM(CASE 
                   WHEN vehicle_type = '53英尺' AND vehicle_no = 'G' THEN 0 
                   ELSE ({_net})
               END) as total_pieces 
        FROM inbound_records 
        WHERE 
            created_at >= ? AND created_at < ?
        GROUP BY vehicle_type
    """
    vehicle_stats_cur = conn.cursor(); vehicle_stats_cur.execute(vehicle_stats_query, (
        period_start.strftime('%Y-%m-%d %H:%M:%S'), 
        period_end.strftime('%Y-%m-%d %H:%M:%S')
    ))
    vehicle_stats = [{
        "vehicle_type": r[0],
        "count": r[1],
        "total_pieces": int(r[2]) if r[2] else 0
    } for r in vehicle_stats_cur.fetchall()]
    
    # 初始化统计变量
    vehicles_19_to_20 = 0
    # 统计19:00-20:00时间段各车型到车数量
    vehicles_19_to_20_by_type = {}
    
    # 统计20:00-21:00时间段记录
    vehicles_20_to_21 = 0
    # 统计20:00-21:00时间段各车型到车数量
    vehicles_20_to_21_by_type = {}
    
    # 统计超过24:00的记录（即次日00:00之后的记录）
    vehicles_after_24 = 0

    # 统计当天分拣完成的货物总量（查询当天00:00之后到次日00:00之前的所有记录）
    sorting_query = """
        SELECT SUM(pieces) as total_sorted
        FROM sorting_records
        WHERE sorting_time >= ? AND sorting_time < ?
    """
    sorting_cur = conn.cursor(); sorting_cur.execute(sorting_query, (
        period_start.strftime('%Y-%m-%d %H:%M:%S'),
        period_end.strftime('%Y-%m-%d %H:%M:%S')
    ))
    sorting_result = sorting_cur.fetchone()
    sorting_val = sorting_result['total_sorted'] if USE_POSTGRES else sorting_result[0]
    total_sorted_pieces = int(sorting_val) if sorting_val else 0
    
    # 处理每条记录，用于统计特定时间段的数据（按录入的时间段time_slot统计）
    for record in records:
        record_id, created_at_str, vehicle_type, time_slot = record
        # 基于录入的时间段进行统计
        if time_slot:
            try:
                time_slot_int = int(time_slot)
                # 检查是否在19:00-20:00之间（录入的时间段为19）
                if time_slot_int == 19:
                    vehicles_19_to_20 += 1
                    
                    # 统计19:00-20:00时间段各车型到车数量
                    if vehicle_type not in vehicles_19_to_20_by_type:
                        vehicles_19_to_20_by_type[vehicle_type] = 0
                    vehicles_19_to_20_by_type[vehicle_type] += 1
                
                # 检查是否在20:00-21:00之间（录入的时间段为20）
                if time_slot_int == 20:
                    vehicles_20_to_21 += 1
                    
                    # 统计20:00-21:00时间段各车型到车数量
                    if vehicle_type not in vehicles_20_to_21_by_type:
                        vehicles_20_to_21_by_type[vehicle_type] = 0
                    vehicles_20_to_21_by_type[vehicle_type] += 1
                
                # 检查是否是超过24:00的记录（录入的时间段为24或更大）
                if time_slot_int >= 24:
                    vehicles_after_24 += 1
            except ValueError:
                # 如果time_slot不是有效的整数，跳过这条记录
                pass
        else:
            # 如果没有录入时间段，仍然使用原来基于创建时间的逻辑作为后备
            if created_at_str:
                try:
                    # 将UTC时间字符串转换为datetime对象
                    if isinstance(created_at_str, str):
                        utc_time = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M:%S')
                    else:
                        utc_time = created_at_str
                    utc_time = pytz.utc.localize(utc_time)
                    # 转换为系统本地时间（修改这里）
                    local_time = utc_time.astimezone(LA_TZ)
                    
                    # 检查是否在19:00-20:00之间（系统本地时间）
                    if local_time.hour == 19:
                        vehicles_19_to_20 += 1
                        
                        # 统计19:00-20:00时间段各车型到车数量
                        if vehicle_type not in vehicles_19_to_20_by_type:
                            vehicles_19_to_20_by_type[vehicle_type] = 0
                        vehicles_19_to_20_by_type[vehicle_type] += 1
                    
                    # 检查是否在20:00-21:00之间（系统本地时间）
                    if local_time.hour == 20:
                        vehicles_20_to_21 += 1
                        
                        # 统计20:00-21:00时间段各车型到车数量
                        if vehicle_type not in vehicles_20_to_21_by_type:
                            vehicles_20_to_21_by_type[vehicle_type] = 0
                        vehicles_20_to_21_by_type[vehicle_type] += 1
                    
                    # 检查是否是明天的记录（即次日00:00之后的记录）
                    if local_time.date() > request_date:
                        vehicles_after_24 += 1
                except Exception as e:
                    print(f"处理记录 {record_id} 的时间时出错: {e}")
    
    
    # === 计算趋势数据 (环比上周同一统计窗口) ===
    prev_date = request_date - timedelta(days=7)
    prev_start, prev_end = _stats_period_bounds(prev_date, window_mode)
    
    # 查询上周同一天的总车次和货量
    trend_query = f"""
        SELECT COUNT(*) as total_vehicles, 
               SUM(CASE 
                   WHEN vehicle_type = '53英尺' AND vehicle_no = 'G' THEN 0 
                   ELSE ({_net})
               END) as total_pieces 
        FROM inbound_records 
        WHERE 
            created_at >= ? AND created_at < ?
    """
    trend_cur = conn.cursor(); trend_cur.execute(trend_query, (
        prev_start.strftime('%Y-%m-%d %H:%M:%S'), 
        prev_end.strftime('%Y-%m-%d %H:%M:%S')
    ))
    trend_result = trend_cur.fetchone()
    
    prev_vehicles = (trend_result['total_vehicles'] if USE_POSTGRES else trend_result[0]) if trend_result else 0
    prev_pieces_val = trend_result['total_pieces'] if USE_POSTGRES else trend_result[1]
    prev_pieces = int(prev_pieces_val) if prev_pieces_val else 0

    # [新增] 查询上周同一天的托盘总数
    trend_pallet_query = """
        SELECT SUM(load_amount - COALESCE(plate_excluded_load, 0)) as total_pallets
        FROM inbound_records 
        WHERE 
            created_at >= ? AND created_at < ? 
            AND (vehicle_type = '16英尺' OR vehicle_type = '26英尺' OR vehicle_type = '53英尺')
            AND NOT (vehicle_type = '53英尺' AND vehicle_no = 'G')
    """
    trend_pallet_cur = conn.cursor(); trend_pallet_cur.execute(trend_pallet_query, (
        prev_start.strftime('%Y-%m-%d %H:%M:%S'), 
        prev_end.strftime('%Y-%m-%d %H:%M:%S')
    ))
    trend_pallet_result = trend_pallet_cur.fetchone()
    prev_pallets_val = trend_pallet_result['total_pallets'] if USE_POSTGRES else trend_pallet_result[0]
    prev_pallets = int(prev_pallets_val) if prev_pallets_val else 0

    # [新增] 查询上周同一天的晚班数据 (用于计算晚班趋势)
    trend_records_query = """
        SELECT id, created_at, vehicle_type, time_slot FROM inbound_records 
        WHERE 
            created_at >= ? AND created_at < ?
    """
    trend_records_cur = conn.cursor(); trend_records_cur.execute(trend_records_query, (
        prev_start.strftime('%Y-%m-%d %H:%M:%S'), 
        prev_end.strftime('%Y-%m-%d %H:%M:%S')
    ))
    prev_records = trend_records_cur.fetchall()

    prev_night_shift_total = 0
    prev_vehicles_19_to_20 = 0
    prev_vehicles_20_to_21 = 0
    prev_vehicles_after_24 = 0

    for record in prev_records:
        record_id, created_at_str, vehicle_type, time_slot = record
        if time_slot:
            try:
                time_slot_int = int(time_slot)
                if time_slot_int == 19:
                    prev_vehicles_19_to_20 += 1
                    prev_night_shift_total += 1
                elif time_slot_int == 20:
                    prev_vehicles_20_to_21 += 1
                    prev_night_shift_total += 1
                elif time_slot_int >= 24:
                    prev_vehicles_after_24 += 1
                    prev_night_shift_total += 1
            except ValueError:
                pass
        else:
            # 后备逻辑
            if created_at_str:
                try:
                    if isinstance(created_at_str, str):
                        utc_time = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M:%S')
                    else:
                        utc_time = created_at_str
                    utc_time = pytz.utc.localize(utc_time)
                    local_time = utc_time.astimezone(LA_TZ)
                    
                    if local_time.date() > prev_date: # 次日00:00以后
                         prev_vehicles_after_24 += 1
                         prev_night_shift_total += 1
                    elif local_time.hour == 19:
                         prev_vehicles_19_to_20 += 1
                         prev_night_shift_total += 1
                    elif local_time.hour == 20:
                         prev_vehicles_20_to_21 += 1
                         prev_night_shift_total += 1
                except:
                    pass
    
    # 计算增长率
    def calculate_trend(current, previous):
        if previous == 0:
            return 100 if current > 0 else 0
        return ((current - previous) / previous) * 100

    pieces_trend = calculate_trend(total_pieces, prev_pieces)
    vehicles_trend = calculate_trend(total_vehicles, prev_vehicles)
    pallets_trend = calculate_trend(total_pallets, prev_pallets)

    # 计算今日晚班总数
    today_night_shift_total = vehicles_19_to_20 + vehicles_20_to_21 + vehicles_after_24
    night_shift_trend = calculate_trend(today_night_shift_total, prev_night_shift_total)
    
    # 计算各时段趋势
    vehicles_19_to_20_trend = calculate_trend(vehicles_19_to_20, prev_vehicles_19_to_20)
    vehicles_20_to_21_trend = calculate_trend(vehicles_20_to_21, prev_vehicles_20_to_21)
    vehicles_after_24_trend = calculate_trend(vehicles_after_24, prev_vehicles_after_24)
    
    # 计算当前每小时综合分拣产能（人工支持多班次重叠叠加）
    sorting_capacity_hourly = 0
    schedule_config_for_estimate = {
        "manual": {"capacity": 3000, "schedule": [5, 5, 5, 4, 4, 4, 1], "hoursPerShift": 9, "startTimes": [17.5] * 7},
        "machine": {"capacity": 4500, "schedule": [4, 4, 4, 4, 4, 2, 2], "hoursPerShift": 6},
        "night": {"capacity": 4500, "hoursPerShift": 6, "schedule": [0] * 7},
    }
    manual_people = 0
    machine_lanes = 0
    night_lanes = 0
    manual_cap = 600.0
    machine_cap = 4500.0
    night_cap = 4500.0
    v2_daily_plan = {}
    try:
        schedule_cursor = conn.cursor()
        defaults = {
            "manual": {"capacity": 3000, "schedule": [5, 5, 5, 4, 4, 4, 1], "hoursPerShift": 9, "startTimes": [17.5] * 7},
            "machine": {"capacity": 4500, "schedule": [4, 4, 4, 4, 4, 2, 2], "hoursPerShift": 6},
            "night": {"capacity": 4500, "hoursPerShift": 6, "schedule": [0] * 7},
        }
        zero_config = {
            "manual": {"piecesPerPersonPerHour": 310, "defaultShiftHours": 8, "shiftsByDay": [[] for _ in range(7)]},
            "machine": {"capacity": 3100, "hoursPerShift": 8, "schedule": [0] * 7, "startTimes": [17.5] * 7},
            "night": {"capacity": 0, "hoursPerShift": 0, "schedule": [0] * 7},
        }

        # 优先来源：V2 按天保存的排班（/sorting-schedule 保存当天配置）
        config = None
        try:
            schedule_cursor.execute(
                "SELECT plan_json FROM sorting_schedule_daily_plan WHERE plan_date = ? LIMIT 1",
                (request_date.strftime('%Y-%m-%d'),),
            )
            daily_row = schedule_cursor.fetchone()
            if daily_row and daily_row[0]:
                v2_daily_plan = _parse_config_json_from_db(daily_row[0], {})
                config = _v2_plan_to_schedule_config(v2_daily_plan, request_date)
        except Exception as e:
            print(f"Warning: read v2 daily plan failed: {e}")
            config = None

        # 若当天未保存 V2 排班：按 0 处理，不回退旧版全局配置
        if config is None:
            config = zero_config

        schedule_config_for_estimate = config
        day_of_week = request_date.weekday()

        def get_val(cfg, key1, key2, default_val):
            return cfg.get(key1, defaults.get(key1, {})).get(key2, default_val)

        machine_schedule = get_val(config, 'machine', 'schedule', defaults['machine']['schedule'])
        night_schedule_val = config.get('night', {}).get('schedule', [0] * 7)
        if not isinstance(night_schedule_val, list):
            night_schedule_val = [0] * 7
        machine_lanes = machine_schedule[day_of_week] if day_of_week < len(machine_schedule) else 0
        night_lanes = night_schedule_val[day_of_week] if day_of_week < len(night_schedule_val) else 0

        _mcfg = config.get('manual', {}) if isinstance(config.get('manual'), dict) else {}
        manual_cap = _manual_ppp_from_config(_mcfg)
        try:
            machine_cap = float(get_val(config, 'machine', 'capacity', 4500))
            night_cap = float(get_val(config, 'night', 'capacity', 4500))
        except (TypeError, ValueError):
            machine_cap = 4500.0
            night_cap = 4500.0

        now_la = datetime.now(la_tz)
        if manual_cap > 0:
            mh = manual_hourly_capacity_at(now_la, request_date, _mcfg, la_tz, manual_cap)
            manual_people = int(round(mh / manual_cap))
        else:
            manual_people = 0

        sorting_capacity_hourly = sorting_total_hourly_at(now_la, request_date, schedule_config_for_estimate, la_tz)
    except Exception as e:
        print(f"Error calculating sorting capacity: {e}")
        sorting_capacity_hourly = 15000

    earliest_sort_start_la = _v2_earliest_active_start_datetime(request_date, v2_daily_plan, la_tz)

    def get_capacity_at_time(dt_check, req_date):
        return sorting_total_hourly_at(dt_check, req_date, schedule_config_for_estimate, la_tz)

    # 计算托盘余量 (1托盘 = 344件)
    PIECES_PER_PALLET = 344
    pallet_capacity_hourly = sorting_capacity_hourly / PIECES_PER_PALLET if sorting_capacity_hourly > 0 else 0
    
    # 计算已分拣托盘数：使用实时计算（基于时间和产能）
    
    # 获取当前时间（系统本地时间，注意这里可能需要时区转换，但上下文里 mixed usage）
    # Use LA time for consistency with capacity check
    now = datetime.now(la_tz)
    
    # 从当日分拣排班最早开工时刻起积分（无 V2 排班则按 0:00）；产能为 0 的时段不计入
    integration_start = (
        earliest_sort_start_la
        if earliest_sort_start_la is not None
        else la_tz.localize(datetime.combine(request_date, datetime.min.time()))
    )
    
    theoretical_sorted_pieces = 0
    
    if now > integration_start:
        # Integrate capacity from start to now
        # Step in 10-minute intervals for better precision than hourly but cheaper than minute
        curr = integration_start
        step_minutes = 10
        while curr < now:
            # Calculate capacity for this interval
            # We assume capacity is constant for the interval based on start of interval
            cap_hourly = get_capacity_at_time(curr, request_date)
            
            # Determine duration of this step (might be partial at the end)
            next_step = curr + timedelta(minutes=step_minutes)
            if next_step > now:
                duration_hours = (now - curr).total_seconds() / 3600
            else:
                duration_hours = step_minutes / 60
            
            theoretical_sorted_pieces += cap_hourly * duration_hours
            curr = next_step
            
    theoretical_sorted_pallets = theoretical_sorted_pieces / PIECES_PER_PALLET if theoretical_sorted_pieces > 0 else 0

    # 实际已分拣托盘数（从分拣记录表）
    recorded_sorted_pallets = total_sorted_pieces / PIECES_PER_PALLET if total_sorted_pieces > 0 else 0
    
    # 使用两者中的较大值（以实际记录为准，但不低于理论值）
    sorted_pallets = max(theoretical_sorted_pallets, recorded_sorted_pallets)
    
    # 计算托盘余量（不允许为负数，四舍五入为整数）
    remaining_pallets = round(max(0, total_pallets - sorted_pallets))
    
    # 计算预计完成时间：剩余件数 = 到货件数 − 已分拣件数；按当天排班产能与开工时刻推演
    estimated_completion_timestamp = None
    remaining_pieces_for_est = max(0, total_pieces - total_sorted_pieces)

    if remaining_pieces_for_est > 0:
        now_local_naive = datetime.now()
        is_current_business_context = period_start <= now_local_naive < period_end

        if is_current_business_context:
            try:
                limit_now = datetime.now(la_tz)
                sim_start = limit_now
                if earliest_sort_start_la is not None and sim_start < earliest_sort_start_la:
                    sim_start = earliest_sort_start_la
                est_time = simulate_sorting_completion_time(
                    remaining_pieces_for_est,
                    sim_start,
                    schedule_config_for_estimate,
                    la_tz,
                )
                estimated_completion_timestamp = est_time.isoformat()
            except Exception as e:
                print(f"Error calculating estimated time: {e}")
    
    # ============================================================================
    # Enhanced Sorting Progress Calculations
    # ============================================================================
    
    # Calculate sorting progress percentage
    sorting_progress_percentage = 0
    if total_pallets > 0:
        sorting_progress_percentage = min(100, (sorted_pallets / total_pallets) * 100)
    
    # Calculate remaining duration and format completion time
    remaining_duration_minutes = None
    estimated_completion_time_formatted = None
    
    if estimated_completion_timestamp:
        try:
            est_time = datetime.fromisoformat(estimated_completion_timestamp.replace('Z', '+00:00'))
            if est_time.tzinfo is None:
                est_time = la_tz.localize(est_time)
            else:
                est_time = est_time.astimezone(la_tz)
            _now_la = datetime.now(la_tz)
            remaining_duration = (est_time - _now_la).total_seconds() / 60
            remaining_duration_minutes = max(0, int(remaining_duration))
            
            # Format completion time in local timezone
            estimated_completion_time_formatted = est_time.strftime('%Y-%m-%d %H:%M')
        except Exception as e:
            print(f"Error formatting completion time: {e}")
            pass
    
    # Check if sorting is currently active (17:30-05:00)
    is_sorting_active = False
    _now_check = datetime.now()
    _current_hour = _now_check.hour
    _current_minute = _now_check.minute
    if (_current_hour > 17 or (_current_hour == 17 and _current_minute >= 30)) and _current_hour < 24:
        is_sorting_active = True
    elif 0 <= _current_hour < 5:
        is_sorting_active = True
    
    # Calculate current sorting rate (pieces/hour in last hour)
    current_sorting_rate = 0
    try:
        one_hour_ago = datetime.now() - timedelta(hours=1)
        rate_query = """
            SELECT SUM(pieces) as recent_sorted
            FROM sorting_records
            WHERE sorting_time >= ?
        """
        rate_cur = conn.cursor()
        rate_cur.execute(rate_query, (one_hour_ago.strftime('%Y-%m-%d %H:%M:%S'),))
        rate_result = rate_cur.fetchone()
        rate_val = rate_result['recent_sorted'] if USE_POSTGRES else rate_result[0]
        current_sorting_rate = int(rate_val) if rate_val else 0
    except Exception as e:
        print(f"Error calculating sorting rate: {e}")
        pass
    # [新增] 计算预计时段到达车次 (下一个小时的时间段均值)
    predicted_next_hour_vehicles = 0
    now_la = datetime.now(la_tz)
    next_hour_la = now_la + timedelta(hours=1)
    next_time_slot = str(next_hour_la.hour)
    
    try:
        # 查询历史上该时段的平均到车数
        # 思路：按日期分组，统计每天该时段的到车数，最后取平均值
        predict_query = """
            SELECT AVG(daily_count) FROM (
                SELECT COUNT(*) as daily_count 
                FROM inbound_records 
                WHERE time_slot = ? 
                GROUP BY DATE(created_at)
            )
        """
        predict_cur = conn.cursor()
        predict_cur.execute(predict_query, (next_time_slot,))
        predict_result = predict_cur.fetchone()
        predicted_val = predict_result[0] if predict_result else 0
        predicted_next_hour_vehicles = int(round(float(predicted_val))) if predicted_val else 0
    except Exception as e:
        print(f"Error calculating predicted vehicles: {e}")
        pass

    cbs_today_pallets = cbs_today_pieces = cbt_today_pallets = cbt_today_pieces = 0
    try:
        cbc = conn.cursor()
        cbc.execute(convert_query_placeholders("""
            SELECT
                COALESCE(SUM(CASE WHEN vehicle_type = 'CBS' THEN load_amount ELSE 0 END), 0),
                COALESCE(SUM(CASE WHEN vehicle_type = 'CBS' THEN (pieces - COALESCE(excluded_pieces, 0)) ELSE 0 END), 0),
                COALESCE(SUM(CASE WHEN vehicle_type = 'CBT' THEN load_amount ELSE 0 END), 0),
                COALESCE(SUM(CASE WHEN vehicle_type = 'CBT' THEN (pieces - COALESCE(excluded_pieces, 0)) ELSE 0 END), 0)
            FROM inbound_records
            WHERE created_at >= ? AND created_at < ?
        """), (
            period_start.strftime('%Y-%m-%d %H:%M:%S'),
            period_end.strftime('%Y-%m-%d %H:%M:%S'),
        ))
        rw = cbc.fetchone()
        if rw is not None:
            cbs_today_pallets = int(float(rw[0] or 0))
            cbs_today_pieces = int(float(rw[1] or 0))
            cbt_today_pallets = int(float(rw[2] or 0))
            cbt_today_pieces = int(float(rw[3] or 0))
    except Exception as e_cbc:
        print(f"[stats] CBS/CBT aggregate: {e_cbc}")

    conn.close()
    
    return jsonify({
        "estimated_completion_timestamp": estimated_completion_timestamp,
        "calculated_sorted_pieces": int(theoretical_sorted_pieces),
        "total_vehicles": total_vehicles,
        "total_pieces": total_pieces,
        # 矫正总量 = 当日入库实到件数合计（与 total_pieces 一致，与历史核对口径一致；不用成本表覆盖）
        "correctedTotal": int(total_pieces),
        "total_pallets": total_pallets,
        "vehicle_stats": vehicle_stats,
        "vehicles_19_to_20": vehicles_19_to_20,
        "vehicles_19_to_20_by_type": vehicles_19_to_20_by_type,
        "vehicles_20_to_21": vehicles_20_to_21,
        "vehicles_20_to_21_by_type": vehicles_20_to_21_by_type,
        "vehicles_after_24": vehicles_after_24,
        # 趋势数据
        "pieces_trend": round(pieces_trend, 1),
        "vehicles_trend": round(vehicles_trend, 1),
        "pallets_trend": round(pallets_trend, 1),
        "night_shift_trend": round(night_shift_trend, 1),
        "vehicles_19_to_20_trend": round(vehicles_19_to_20_trend, 1),
        "vehicles_20_to_21_trend": round(vehicles_20_to_21_trend, 1),
        "vehicles_after_24_trend": round(vehicles_after_24_trend, 1),
        "prev_pieces": prev_pieces,
        "prev_vehicles": prev_vehicles,
        "sorting_capacity_hourly": sorting_capacity_hourly,
        "total_sorted_pieces": total_sorted_pieces,
        "sorted_pieces": int(max(theoretical_sorted_pieces, total_sorted_pieces)),
        # 托盘相关数据
        "pallet_capacity_hourly": round(pallet_capacity_hourly, 2),
        "sorted_pallets": round(sorted_pallets, 1),
        "remaining_pallets": round(remaining_pallets, 1),
        # 增强的分拣进度数据
        "sorting_progress_percentage": round(sorting_progress_percentage, 1),
        "remaining_duration_minutes": remaining_duration_minutes,
        "estimated_completion_time_formatted": estimated_completion_time_formatted,
        "current_sorting_rate": current_sorting_rate,
        "is_sorting_active": is_sorting_active,
        "predicted_next_hour_vehicles": predicted_next_hour_vehicles,
        "next_time_slot": next_time_slot,
        "total_plate_excluded_load": round(total_plate_excluded_load, 4),
        "total_excluded_pieces": total_excluded_pieces_stat,
        "cbs_today_pallets": cbs_today_pallets,
        "cbs_today_pieces": cbs_today_pieces,
        "cbt_today_pallets": cbt_today_pallets,
        "cbt_today_pieces": cbt_today_pieces,
        "stats_window": window_mode,
        "stats_window_start": period_start.strftime('%Y-%m-%d %H:%M:%S'),
        "stats_window_end_exclusive": period_end.strftime('%Y-%m-%d %H:%M:%S'),
    })



@app.route('/api/daily_trend')
def get_daily_trend():
    """获取每日货物趋势数据（显示所有有记录的日期）"""
    try:
        window_mode = _parse_stats_window_param(request.args.get('stats_window'))
        conn = get_db()
        
        # 查询数据库中所有有记录的日期（按日期分组）
        # 使用DATE函数提取日期部分
        dates_query = """
            SELECT DISTINCT DATE(created_at) as record_date
            FROM inbound_records
            ORDER BY record_date ASC
        """
        cursor = conn.cursor(); cursor.execute(dates_query)
        # 兼容处理：确保转换为字符串
        raw_dates = [row[0] for row in cursor.fetchall()]
        record_dates = []
        for d in raw_dates:
            if isinstance(d, (datetime, date)):
                record_dates.append(d.strftime('%Y-%m-%d'))
            else:
                record_dates.append(str(d))
        
        # 如果没有记录，返回空数组
        if not record_dates:
            conn.close()
            return jsonify([])
        
        # 美国联邦假期（2025-2026年）
        US_FEDERAL_HOLIDAYS = {
            # 2025
            '2025-01-01': "New Year's Day",
            '2025-01-20': "Martin Luther King Jr. Day",
            '2025-02-17': "Presidents' Day",
            '2025-05-26': "Memorial Day",
            '2025-07-04': "Independence Day",
            '2025-09-01': "Labor Day",
            '2025-10-13': "Columbus Day",
            '2025-11-11': "Veterans Day",
            '2025-11-27': "Thanksgiving",
            '2025-12-25': "Christmas Day",
            # 2026
            '2026-01-01': "New Year's Day",
            '2026-01-19': "Martin Luther King Jr. Day",
            '2026-02-16': "Presidents' Day",
            '2026-05-25': "Memorial Day",
            '2026-07-03': "Independence Day (Observed)",
            '2026-07-04': "Independence Day",
            '2026-09-07': "Labor Day",
            '2026-10-12': "Columbus Day",
            '2026-11-11': "Veterans Day",
            '2026-11-26': "Thanksgiving",
            '2026-12-25': "Christmas Day"
        }
        
        # 准备结果数组
        result = []
        
        _net_dt = _sql_inbound_net_pieces_actual("")
        # 为每个有记录的日期查询数据
        for date_str in record_dates:
            target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
            day_start, day_end = _stats_period_bounds(target_date, window_mode)

            # 查询当天的货物总量、车次总数和托盘总数
            query = f"""
                SELECT 
                    SUM(CASE 
                        WHEN vehicle_type = '53英尺' AND vehicle_no = 'G' THEN 0 
                        ELSE ({_net_dt})
                    END) as total_pieces,
                    COUNT(*) as total_vehicles,
                    SUM(CASE 
                        WHEN (vehicle_type IN ('26英尺', '53英尺')) AND NOT (vehicle_type = '53英尺' AND vehicle_no = 'G') 
                        THEN (load_amount - COALESCE(plate_excluded_load, 0))
                        ELSE 0 
                    END) as total_pallets
                FROM inbound_records
                WHERE created_at >= ? AND created_at < ?
            """
            cursor = conn.cursor(); cursor.execute(query, (
                day_start.strftime('%Y-%m-%d %H:%M:%S'),
                day_end.strftime('%Y-%m-%d %H:%M:%S')
            ))
            row = cursor.fetchone()
            # 直接使用实际数值，不取整
            total_pieces = int(row[0]) if row[0] else 0
            total_vehicles = int(row[1]) if row[1] else 0
            total_pallets = int(row[2]) if row[2] else 0
            
            # 获取星期几 (0=Monday, 6=Sunday)
            weekday = target_date.weekday()
            weekday_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
            
            # 检查是否是假期
            is_holiday = date_str in US_FEDERAL_HOLIDAYS
            holiday_name = US_FEDERAL_HOLIDAYS.get(date_str, '')
            
            result.append({
                'date': target_date.strftime('%Y-%m-%d'),
                'weekday': weekday_names[weekday],
                'weekday_num': weekday,  # 0-6
                'is_holiday': is_holiday,
                'holiday_name': holiday_name,
                'total_pieces': total_pieces,
                'total_vehicles': total_vehicles,
                'total_pallets': total_pallets
            })
        
        conn.close()
        return jsonify(result)
    
    except Exception as e:
        if 'conn' in locals():
            conn.close()
        return jsonify({'error': f'获取每日趋势数据出错: {str(e)}'}), 500


@app.route('/api/weekly_pallet_trend')
def get_weekly_pallet_trend():
    """获取每周托盘趋势数据（按自然周汇总）"""
    try:
        window_mode = _parse_stats_window_param(request.args.get('stats_window'))
        conn = get_db()
        
        # 查询数据库中所有有记录的日期
        dates_query = """
            SELECT DISTINCT DATE(created_at) as record_date
            FROM inbound_records
            ORDER BY record_date ASC
        """
        cursor = conn.cursor()
        cursor.execute(dates_query)
        raw_dates = [row[0] for row in cursor.fetchall()]
        
        if not raw_dates:
            conn.close()
            return jsonify([])
        
        # 转换日期格式
        record_dates = []
        for d in raw_dates:
            if isinstance(d, (datetime, date)):
                record_dates.append(d if isinstance(d, date) else d.date())
            else:
                record_dates.append(datetime.strptime(str(d), '%Y-%m-%d').date())
        
        # 获取最小和最大日期
        min_date = min(record_dates)
        max_date = max(record_dates)
        end_raw = request.args.get('end_date') or request.args.get('date')
        if end_raw:
            try:
                ed = datetime.strptime(str(end_raw)[:10], '%Y-%m-%d').date()
                max_date = min(max_date, ed)
            except ValueError:
                pass
        
        # 辅助函数：获取自然周的起止日期
        def get_natural_week_range(date_obj):
            """获取自然周的起止时间（周一到周日）"""
            weekday = date_obj.weekday()
            week_start = date_obj - timedelta(days=weekday)
            week_end = week_start + timedelta(days=6)
            return week_start, week_end
        
        # 对齐最小日期到该周的周一
        current_start, current_end = get_natural_week_range(min_date)
        
        weeks_data = []
        
        # 循环直到包含当前日期
        while current_start <= max_date:
            # 查询该周每一天的托盘数据
            daily_data = []
            week_total_pallets = 0
            
            # 循环7天（周一到周日）
            for day_offset in range(7):
                current_day = current_start + timedelta(days=day_offset)

                if current_day > max_date:
                    daily_data.append({
                        'date': current_day.strftime('%Y-%m-%d'),
                        'weekday': current_day.weekday(),
                        'pallets': 0
                    })
                    continue

                day_start_dt, day_end_dt = _stats_period_bounds(current_day, window_mode)

                # 查询当天托盘数据
                day_query = """
                    SELECT SUM(CASE 
                        WHEN (vehicle_type IN ('26英尺', '53英尺')) AND NOT (vehicle_type = '53英尺' AND vehicle_no = 'G') 
                        THEN (load_amount - COALESCE(plate_excluded_load, 0))
                        ELSE 0 
                    END) as total_pallets
                    FROM inbound_records
                    WHERE created_at >= ? AND created_at < ?
                """

                cursor = conn.cursor()
                cursor.execute(day_query, (
                    day_start_dt.strftime('%Y-%m-%d %H:%M:%S'),
                    day_end_dt.strftime('%Y-%m-%d %H:%M:%S')
                ))
                
                row = cursor.fetchone()
                day_pallets = int(row[0]) if row[0] else 0
                
                # 累加到周总计
                week_total_pallets += day_pallets
                
                # 添加到每日数据数组
                daily_data.append({
                    'date': current_day.strftime('%Y-%m-%d'),
                    'weekday': current_day.weekday(),  # 0=周一, 6=周日
                    'pallets': day_pallets
                })
            
            # 只添加有数据的周
            if week_total_pallets > 0:
                weeks_data.append({
                    'week_label': f"{current_start.strftime('%m/%d')}-{current_end.strftime('%m/%d')}",
                    'start_date': current_start.strftime('%Y-%m-%d'),
                    'end_date': current_end.strftime('%Y-%m-%d'),
                    'daily_data': daily_data,  # 每天的详细数据
                    'week_total_pallets': week_total_pallets
                })
            
            # 移动到下一周
            current_start = current_start + timedelta(days=7)
            current_end = current_end + timedelta(days=7)
        
        conn.close()
        return jsonify(weeks_data)
    
    except Exception as e:
        if 'conn' in locals():
            conn.close()
        return jsonify({'error': f'获取每周托盘趋势数据出错: {str(e)}'}), 500



@app.route('/api/week_comparison')
def get_week_comparison():
    """获取所有周的对比数据，包含每周内每天的详细数据（使用自然周：周一00:00到周日23:59）"""
    try:
        conn = get_db()
        window_mode = _parse_stats_window_param(request.args.get('stats_window'))

        
        # 1. 获取最小日期（第一条记录的时间）
        query = "SELECT MIN(created_at) FROM inbound_records"
        cursor = conn.cursor(); cursor.execute(query)
        min_res = cursor.fetchone()
        min_str = (min_res['min'] if USE_POSTGRES else min_res[0]) if min_res else None
        
        if not min_str:
            conn.close()
            return jsonify([])  # 没有数据
            
        # 解析最小日期
        try:
            if hasattr(min_str, 'date'): # 检查是否为 datetime/date 对象
                min_date = min_str.date() if hasattr(min_str, 'date') and callable(min_str.date) else min_str
                # 如果是 datetime 对象，它有 date() 方法；如果是 date 对象，它就是 date
                if isinstance(min_date, datetime):
                     min_date = min_date.date()
            else:
                # 字符串处理
                min_date = datetime.strptime(str(min_str).split(' ')[0], '%Y-%m-%d').date()
        except Exception as e:
            print(f"解析最小日期出错: {e}, 使用当前日期")
            min_date = datetime.now(LA_TZ).date()
            
        # 设置最大日期为今天（确保显示到本周）；可选 end_date 截断
        max_date = datetime.now(LA_TZ).date()
        end_raw = request.args.get('end_date') or request.args.get('date')
        if end_raw:
            try:
                ed = datetime.strptime(str(end_raw)[:10], '%Y-%m-%d').date()
                max_date = min(max_date, ed)
            except ValueError:
                pass
        
        # 辅助函数：获取自然周的起止日期
        def get_natural_week_range(date):
            """获取自然周的起止时间（周一到周日）"""
            # weekday() 返回0-6，0是周一
            weekday = date.weekday()
            week_start = date - timedelta(days=weekday)
            week_end = week_start + timedelta(days=6)
            return week_start, week_end

        data_first_monday, _ = get_natural_week_range(min_date)
        last_monday_to_include, _ = get_natural_week_range(max_date)
        loop_start = data_first_monday
        loop_end_monday = last_monday_to_include
        week_start_raw = request.args.get('week_start')
        week_end_raw = request.args.get('week_end')
        if week_start_raw:
            try:
                ws = datetime.strptime(str(week_start_raw)[:10], '%Y-%m-%d').date()
                wm, _ = get_natural_week_range(ws)
                loop_start = max(loop_start, wm)
            except ValueError:
                pass
        if week_end_raw:
            try:
                we = datetime.strptime(str(week_end_raw)[:10], '%Y-%m-%d').date()
                wm, _ = get_natural_week_range(we)
                loop_end_monday = min(loop_end_monday, wm)
            except ValueError:
                pass
        if loop_start > loop_end_monday:
            conn.close()
            return jsonify([])
        week_span = (loop_end_monday - loop_start).days // 7 + 1
        if week_span > 104:
            conn.close()
            return jsonify({'error': '周数区间过长（最多 104 周）'}), 400

        current_start = loop_start
        current_end = current_start + timedelta(days=6)

        weeks_data = []

        _net_wk = _sql_inbound_net_pieces_actual("")
        # 循环直到包含当前日期
        while current_start <= loop_end_monday:
            # 查询该周每一天的数据
            daily_data = []
            week_total_pieces = 0
            week_total_vehicles = 0
            
            # 循环7天（周一到周日）
            for day_offset in range(7):
                current_day = current_start + timedelta(days=day_offset)

                if current_day > max_date:
                    daily_data.append({
                        'date': current_day.strftime('%Y-%m-%d'),
                        'weekday': current_day.weekday(),
                        'pieces': 0,
                        'vehicles': 0
                    })
                    continue

                day_start_dt, day_end_dt = _stats_period_bounds(current_day, window_mode)

                # 查询当天数据
                day_query = f"""
                    SELECT 
                        COUNT(*) as vehicle_count,
                        SUM(CASE 
                            WHEN vehicle_type = '53英尺' AND vehicle_no = 'G' THEN 0 
                            ELSE ({_net_wk})
                        END) as total_pieces
                    FROM inbound_records
                    WHERE created_at >= ? AND created_at < ?
                """

                cursor = conn.cursor(); cursor.execute(day_query, (
                    day_start_dt.strftime('%Y-%m-%d %H:%M:%S'),
                    day_end_dt.strftime('%Y-%m-%d %H:%M:%S')
                ))
                
                row = cursor.fetchone()
                day_vehicles = row[0] if row[0] else 0
                # 直接使用实际数值，不取整
                day_pieces = int(row[1]) if row[1] else 0
                
                # Use raw pieces
                day_pieces = int(day_pieces)
                
                # 累加到周总计
                week_total_vehicles += day_vehicles
                week_total_pieces += day_pieces
                
                # 添加到每日数据数组
                daily_data.append({
                    'date': current_day.strftime('%Y-%m-%d'),
                    'weekday': current_day.weekday(),  # 0=周一, 6=周日
                    'pieces': day_pieces,
                    'vehicles': day_vehicles
                })
            
            # 计算环比（如果不是第一周）
            pieces_change_percent = 0
            vehicles_change_percent = 0
            
            if weeks_data:
                last_week = weeks_data[-1]
                last_pieces = last_week['week_total_pieces']
                last_vehicles = last_week['week_total_vehicles']
                
                if last_pieces > 0:
                    pieces_change_percent = ((week_total_pieces - last_pieces) / last_pieces) * 100
                else:
                    pieces_change_percent = 100 if week_total_pieces > 0 else 0
                    
                if last_vehicles > 0:
                    vehicles_change_percent = ((week_total_vehicles - last_vehicles) / last_vehicles) * 100
                else:
                    vehicles_change_percent = 100 if week_total_vehicles > 0 else 0
            
            weeks_data.append({
                'week_label': f"{current_start.strftime('%m/%d')}-{current_end.strftime('%m/%d')}",
                'start_date': current_start.strftime('%Y-%m-%d'),
                'end_date': current_end.strftime('%Y-%m-%d'),
                'daily_data': daily_data,  # 每天的详细数据
                'week_total_pieces': week_total_pieces,
                'week_total_vehicles': week_total_vehicles,
                'pieces_change_percent': round(pieces_change_percent, 2),
                'vehicles_change_percent': round(vehicles_change_percent, 2)
            })
            
            # 移动到下一周
            current_start = current_start + timedelta(days=7)
            current_end = current_end + timedelta(days=7)
            
        conn.close()
        
        # 过滤掉第一周（如果不完整）
        # 判断标准：如果第一周的起始日期早于数据库中的最小日期，说明这周是不完整的
        if weeks_data and len(weeks_data) > 0:
            first_week = weeks_data[0]
            first_week_start = datetime.strptime(first_week['start_date'], '%Y-%m-%d').date()
            
            # 如果第一周的周一早于数据库最小日期，说明第一周不完整
            if first_week_start < min_date:
                print(f"过滤掉不完整的第一周: {first_week['week_label']}")
                weeks_data = weeks_data[1:]  # 移除第一周
        
        # [Option A] 将显示的第一个周的环比设为 0%
        if weeks_data and len(weeks_data) > 0:
            weeks_data[0]['pieces_change_percent'] = 0
            weeks_data[0]['vehicles_change_percent'] = 0
        
        return jsonify(weeks_data)
    
    except Exception as e:
        if 'conn' in locals():
            conn.close()
        return jsonify({'error': f'获取周环比数据出错: {str(e)}'}), 500




@app.route('/api/export_csv')
def export_csv():
    try:
        date_str = request.args.get('date')
        window_mode = _parse_stats_window_param(request.args.get('stats_window'))

        if not date_str:
            date_str = _default_stats_request_date(window_mode).strftime('%Y-%m-%d')

        conn = get_db()

        request_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        today_start, next_day_start = _stats_period_bounds(request_date, window_mode)
        
        _net_ex = _sql_inbound_net_pieces_actual("")
        # 总车次和总货物量（查询当天00:00之后到次日00:00之前的所有记录）
        total_query = f"""
            SELECT COUNT(*) as total_vehicles, 
                   SUM(CASE WHEN vehicle_type = '53英尺' AND vehicle_no = 'G' THEN 0 
                       ELSE ({_net_ex}) END) as total_pieces 
            FROM inbound_records 
            WHERE 
                created_at >= ? AND created_at < ?
        """
        total_cur = conn.cursor(); total_cur.execute(total_query, (
            today_start.strftime('%Y-%m-%d %H:%M:%S'), 
            next_day_start.strftime('%Y-%m-%d %H:%M:%S')
        ))
        total_result = total_cur.fetchone()
        total_vehicles = total_result[0] if total_result[0] else 0
        total_pieces = int(total_result[1]) if total_result[1] else 0
        
        # 托盘总数（查询当天00:00之后到次日00:00之前，车辆类型为26英尺或53英尺的装载量总和）
        pallet_query = """
            SELECT SUM(load_amount - COALESCE(plate_excluded_load, 0)) as total_pallets
            FROM inbound_records 
            WHERE 
                created_at >= ? AND created_at < ? AND (vehicle_type = '16英尺' OR vehicle_type = '26英尺' OR vehicle_type = '53英尺')
                AND NOT (vehicle_type = '53英尺' AND vehicle_no = 'G')
        """
        pallet_cur = conn.cursor(); pallet_cur.execute(pallet_query, (
            today_start.strftime('%Y-%m-%d %H:%M:%S'), 
            next_day_start.strftime('%Y-%m-%d %H:%M:%S')
        ))
        pallet_result = pallet_cur.fetchone()
        total_pallets = int(pallet_result[0]) if pallet_result[0] else 0
        
        # 各车型统计（查询当天00:00之后到次日00:00之前的所有记录）
        vehicle_stats_query = f"""
            SELECT vehicle_type, COUNT(*) as count, 
                   SUM(CASE WHEN vehicle_type = '53英尺' AND vehicle_no = 'G' THEN 0 
                       ELSE ({_net_ex}) END) as total_pieces 
            FROM inbound_records 
            WHERE 
                created_at >= ? AND created_at < ?
            GROUP BY vehicle_type
        """
        vehicle_stats_cur = conn.cursor(); vehicle_stats_cur.execute(vehicle_stats_query, (
            today_start.strftime('%Y-%m-%d %H:%M:%S'), 
            next_day_start.strftime('%Y-%m-%d %H:%M:%S')
        ))
        vehicle_stats = [{
            "vehicle_type": r[0],
            "count": r[1],
            "total_pieces": int(r[2]) if r[2] else 0
        } for r in vehicle_stats_cur.fetchall()]
        
        # 查询属于指定自然日的记录（查询当天00:00之后到次日00:00之前的所有记录）
        records_query = """
            SELECT id, created_at, vehicle_type, time_slot FROM inbound_records 
            WHERE 
                created_at >= ? AND created_at < ?
        """
        records_cur = conn.cursor(); records_cur.execute(records_query, (
            today_start.strftime('%Y-%m-%d %H:%M:%S'), 
            next_day_start.strftime('%Y-%m-%d %H:%M:%S')
        ))
        records = records_cur.fetchall()
        
        # 初始化统计变量
        vehicles_19_to_20 = 0
        vehicles_20_to_21 = 0
        vehicles_after_24 = 0
        
        # 处理每条记录，用于统计特定时间段的数据（按录入的时间段time_slot统计）
        for record in records:
            record_id, created_at_str, vehicle_type, time_slot = record
            # 基于录入的时间段进行统计
            if time_slot:
                try:
                    time_slot_int = int(time_slot)
                    # 检查是否在19:00-20:00之间（录入的时间段为19）
                    if time_slot_int == 19:
                        vehicles_19_to_20 += 1
                    # 检查是否在20:00-21:00之间（录入的时间段为20）
                    if time_slot_int == 20:
                        vehicles_20_to_21 += 1
                    # 检查是否是超过24:00的记录（录入的时间段为24或更大）
                    if time_slot_int >= 24:
                        vehicles_after_24 += 1
                except ValueError:
                    # 如果time_slot不是有效的整数，跳过这条记录
                    pass
            else:
                # 如果没有录入时间段，仍然使用原来基于创建时间的逻辑作为后备
                if created_at_str:
                    try:
                        # 将UTC时间字符串转换为datetime对象
                        if isinstance(created_at_str, str):
                            utc_time = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M:%S')
                        else:
                            utc_time = created_at_str # 已经是 datetime 对象
                        utc_time = pytz.utc.localize(utc_time)
                        # 转换为系统本地时间（修改这里）
                        local_time = utc_time.astimezone(LA_TZ)
                        
                        # 检查是否在19:00-20:00之间（系统本地时间）
                        if local_time.hour == 19:
                            vehicles_19_to_20 += 1
                        # 检查是否在20:00-21:00之间（系统本地时间）
                        if local_time.hour == 20:
                            vehicles_20_to_21 += 1
                        # 检查是否是明天的记录（即次日00:00之后的记录）
                        if local_time.date() > request_date:
                            vehicles_after_24 += 1
                    except Exception as e:
                        print(f"处理记录 {record_id} 的时间时出错: {e}")
        
        conn.close()
        
        # 创建CSV内容
        import csv
        import io
        
        # 创建内存中的CSV文件
        output = io.StringIO()
        writer = csv.writer(output, quoting=csv.QUOTE_MINIMAL)
        
        # 写入BOM以支持UTF-8编码的Excel打开
        output.write('\ufeff')
        
        # 写入统计摘要标题
        writer.writerow(['统计摘要'])
        writer.writerow(['总车次', '总货物量', '托盘总数', '19:00-20:00到车数', '20:00-21:00到车数', '超过24:00到车数'])
        
        # 写入统计摘要数据
        writer.writerow([total_vehicles, total_pieces, total_pallets, vehicles_19_to_20, vehicles_20_to_21, vehicles_after_24])
        
        # 添加空行分隔
        writer.writerow([])
        
        # 写入各车型统计标题
        writer.writerow(['各车型统计'])
        writer.writerow(['车型', '车次', '货物总量'])
        
        # 写入各车型统计数据
        for stat in vehicle_stats:
            writer.writerow([stat["vehicle_type"], stat["count"], stat["total_pieces"]])
        
        # 获取CSV内容
        csv_content = output.getvalue()
        output.close()
        
        # 创建响应对象
        from flask import Response
        filename = f"inbound_stats_summary_{date_str}.csv"  # 使用英文文件名避免编码问题
        
        response = Response(
            csv_content,
            mimetype='text/csv',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        )
        return response
        
    except Exception as e:
        print(f"导出CSV文件出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"导出失败: {str(e)}"}), 500

@app.route('/api/export_excel')
def export_excel():
    try:
        # 获取查询日期参数，如果没有则使用当天
        date_str = request.args.get('date')
        
        if not date_str:
            # 获取系统当前日期（使用洛杉矶时间）
            date_str = datetime.now(LA_TZ).strftime('%Y-%m-%d')
        
        conn = get_db()
        
        # 解析请求的日期
        request_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # 计算次日日期
        next_date = request_date + timedelta(days=1)
        
        # 构建日期范围查询条件（使用自然日而不是洛杉矶时区时间）
        # 当天00:00:00的时间（系统时间）
        today_start = datetime.combine(request_date, datetime.min.time())
        
        # 次日00:00:00的时间（系统时间，用于上限）
        next_day_start = datetime.combine(next_date, datetime.min.time())
        
        # 查询指定日期的入库记录（查询当天00:00之后到次日00:00之前的所有记录）
        inbound_query = """
            SELECT id, dock_no, vehicle_type, vehicle_no, unit, load_amount,
                   pieces, time_slot, shift_type, remark, created_at, duration,
                   COALESCE(excluded_pieces, 0)
            FROM inbound_records 
            WHERE 
                created_at >= ? AND created_at < ?
            ORDER BY created_at DESC
        """
        inbound_cur = conn.cursor(); inbound_cur.execute(inbound_query, (
            today_start.strftime('%Y-%m-%d %H:%M:%S'), 
            next_day_start.strftime('%Y-%m-%d %H:%M:%S')
        ))
        inbound_rows = [{
            "id": r[0], "dock_no": r[1], "vehicle_type": r[2], "vehicle_no": r[3],
            "unit": r[4], "load_amount": r[5], "pieces": r[6],
            "time_slot": r[7], "shift_type": r[8], "remark": r[9],
            "created_at": r[10],  # 数据库中存储的是系统时间，直接返回
            "duration": r[11],  # 时长(分钟)
            "excluded_pieces": r[12],
            "pieces_actual": round(
                _py_inbound_arrival_pieces(r[2], r[3], r[6], r[12]),
                2,
            ),
        } for r in inbound_cur.fetchall()]
        
        # 查询指定日期的分拣记录（按照自然日逻辑查询，查询当天00:00之后到次日00:00之前的所有记录）
        sorting_query = """
            SELECT id, sorting_time, pieces, remark, created_at, time_slot
            FROM sorting_records 
            WHERE 
                created_at >= ? AND created_at < ?
            ORDER BY created_at DESC
        """
        sorting_cur = conn.cursor(); sorting_cur.execute(sorting_query, (
            today_start.strftime('%Y-%m-%d %H:%M:%S'), 
            next_day_start.strftime('%Y-%m-%d %H:%M:%S')
        ))
        sorting_rows = [{
            "id": r[0], "sorting_time": r[1], "pieces": r[2], "remark": r[3],
            "created_at": r[4], "time_slot": r[5]
        } for r in sorting_cur.fetchall()]
        
        conn.close()
        
        # 创建Excel工作簿
        wb = Workbook()
        
        # 创建入库记录工作表
        ws1 = wb.active
        ws1.title = "入库记录"
        
        # 添加表头
        inbound_headers = ['ID', '码头号', '车辆类型', '车牌号', '单位', '装载量', '录入件数', '实到件数', '时间段', '班次类型', '备注', '创建时间', '时长(分钟)']
        ws1.append(inbound_headers)
        
        # 设置表头样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for col in range(1, len(inbound_headers) + 1):
            cell = ws1.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # 添加入库数据
        for row_dict in inbound_rows:
            row_data = [
                row_dict["id"], row_dict["dock_no"], row_dict["vehicle_type"], row_dict["vehicle_no"],
                row_dict["unit"], row_dict["load_amount"], row_dict["pieces"],
                row_dict.get("pieces_actual", 0),
                row_dict["time_slot"], row_dict["shift_type"], row_dict["remark"], row_dict["created_at"],
                row_dict.get("duration", ""),
            ]
            ws1.append(row_data)
        
        # 调整列宽
        for column in ws1.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws1.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # 创建分拣记录工作表（只有在有数据时才创建）
        if sorting_rows:
            ws2 = wb.create_sheet("分拣记录")
            
            # 添加表头
            sorting_headers = ['ID', '分拣日期', '件数', '时间段', '备注', '创建时间']
            ws2.append(sorting_headers)
            
            # 设置表头样式
            for col in range(1, len(sorting_headers) + 1):
                cell = ws2.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # 添加分拣数据
            for row_dict in sorting_rows:
                # 按照表头顺序构造行数据
                row_data = [
                    row_dict["id"], row_dict["sorting_time"], row_dict["pieces"],
                    row_dict["time_slot"], row_dict["remark"], row_dict["created_at"]
                ]
                ws2.append(row_data)
            
            # 调整列宽
            for column in ws2.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws2.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # 保存Excel文件
        filename = f"inbound_stats_{date_str}.xlsx"  # 使用英文文件名避免编码问题
        filepath = os.path.join(os.path.dirname(__file__), filename)
        wb.save(filepath)
        
        # 返回Excel文件，使用更直接的方法避免文件名编码问题
        from flask import Response
        import os
        
        # 读取文件内容
        with open(filepath, 'rb') as f:
            file_content = f.read()
        
        # 创建响应对象，明确设置Content-Type和Content-Disposition
        response = Response(
            file_content,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        )
        return response
        
    except Exception as e:
        print(f"导出Excel文件出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"导出失败: {str(e)}"}), 500

# 新增API：导出最近记录
@app.route('/api/export_recent_records')
def export_recent_records():
    try:
        conn = get_db()
        
        # 获取系统当前日期 (LA时间)
        today = datetime.now(LA_TZ).date()
        
        # 计算次日日期
        next_date = today + timedelta(days=1)
        
        # 构建日期范围查询条件（使用自然日）
        # 当天00:00:00的时间（系统时间）
        today_start = datetime.combine(today, datetime.min.time())
        
        # 次日00:00:00的时间（系统时间，用于上限）
        next_day_start = datetime.combine(next_date, datetime.min.time())
        
        # 查询属于当前自然日的记录（查询当天00:00之后到次日00:00之前的所有记录）
        cur = conn.cursor(); cur.execute("""
            SELECT id, dock_no, vehicle_type, vehicle_no, unit, load_amount,
                   pieces, time_slot, shift_type, remark, created_at,
                   COALESCE(excluded_pieces, 0)
            FROM inbound_records 
            WHERE 
                created_at >= ? AND created_at < ?
            ORDER BY created_at DESC""", (
                today_start.strftime('%Y-%m-%d %H:%M:%S'), 
                next_day_start.strftime('%Y-%m-%d %H:%M:%S')
            ))
        
        rows = [{
            "id": r[0], "dock_no": r[1], "vehicle_type": r[2], "vehicle_no": r[3],
            "unit": r[4], "load_amount": r[5], "pieces": r[6],
            "time_slot": r[7], "shift_type": r[8], "remark": r[9],
            "created_at": r[10],
            "excluded_pieces": r[11],
            "pieces_actual": round(
                _py_inbound_arrival_pieces(r[2], r[3], r[6], r[11]),
                2,
            ),
        } for r in cur.fetchall()]
        
        conn.close()
        
        # 创建Excel工作簿
        wb = Workbook()
        
        # 创建工作表
        ws = wb.active
        ws.title = "最近记录"
        
        # 添加表头
        headers = ['ID', '码头号', '车辆类型', '车牌号', '单位', '装载量', '录入件数', '实到件数', '时间段', '班次类型', '备注', '创建时间']
        ws.append(headers)
        
        # 设置表头样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # 添加数据
        for row_dict in rows:
            row_data = [
                row_dict["id"], row_dict["dock_no"], row_dict["vehicle_type"], row_dict["vehicle_no"],
                row_dict["unit"], row_dict["load_amount"], row_dict["pieces"],
                row_dict.get("pieces_actual", 0),
                row_dict["time_slot"], row_dict["shift_type"], row_dict["remark"], row_dict["created_at"]
            ]
            ws.append(row_data)
        
        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # 保存Excel文件
        filename = f"recent_records_{today.strftime('%Y-%m-%d')}.xlsx"
        filepath = os.path.join(os.path.dirname(__file__), filename)
        wb.save(filepath)
        
        # 返回Excel文件
        from flask import Response
        
        # 读取文件内容
        with open(filepath, 'rb') as f:
            file_content = f.read()
        
        # 创建响应对象，明确设置Content-Type和Content-Disposition
        response = Response(
            file_content,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        )
        return response
        
    except Exception as e:
        print(f"导出最近记录Excel文件出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"导出失败: {str(e)}"}), 500

# 用户登录API
# 权限检查辅助函数
def check_page_permission(page_name):
    """检查用户是否有页面访问权限"""
    if 'user_id' not in session:
        return False
    
    # 管理员有所有权限
    if session.get('role') in ('admin', 'boss'):
        return True
    
    # 查询用户权限
    try:
        conn = get_db()
        cursor = conn.cursor(); cursor.execute("""
            SELECT can_view FROM user_permissions
            WHERE user_id = ? AND page_name = ?
        """, (session['user_id'], page_name))
        
        result = cursor.fetchone()
        conn.close()
        
        return result and bool(result['can_view'] if USE_POSTGRES else result[0])
    except:
        return False

def get_first_accessible_page(user_id, role):
    """获取用户第一个有权限访问的页面"""
    print(f"[DEBUG] get_first_accessible_page called: user_id={user_id}, role={role}")
    
    # 管理员默认跳转到首页
    if role in ('admin', 'boss'):
        print(f"[DEBUG] User is admin, redirecting to /")
        return '/'
    
    # 页面优先级顺序
    page_priority = [
        ('outbound-stats', '/outbound-stats'), # Prioritize outbound-stats as requested
        ('index', '/'),
        ('sorting', '/sorting'),
        ('history', '/history'),
        ('statistics', '/statistics'),
        ('logs', '/logs')
    ]
    
    try:
        conn = get_db()
        for page_name, page_url in page_priority:
            cursor = conn.cursor(); cursor.execute("""
                SELECT can_view FROM user_permissions
                WHERE user_id = ? AND page_name = ? AND can_view = 1
            """, (user_id, page_name))
            
            result = cursor.fetchone()
            print(f"[DEBUG] Checking {page_name}: result={result}")
            
            if result:
                print(f"[DEBUG] Found accessible page: {page_name} -> {page_url}")
                conn.close()
                return page_url
        
        conn.close()
        print(f"[DEBUG] No accessible pages found, redirecting to /no_permission")
    except Exception as e:
        print(f"[DEBUG] Exception in get_first_accessible_page: {e}")
        pass
    
    # 如果没有任何页面权限,跳转到无权限页面
    return '/no_permission'

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'error': '用户名和密码不能为空'}), 400
    
    # 对密码进行哈希处理
    password_hash = hashlib.sha256(password.encode()).hexdigest()
    
    # 查询用户
    conn = get_db()
    cursor = conn.cursor(); cursor.execute(convert_query_placeholders("""
        SELECT u.id, u.username, u.role, u.is_active
        FROM users u
        WHERE u.username = ? AND u.password_hash = ? AND u.is_active = TRUE
    """), (username, password_hash))
    
    user = cursor.fetchone()
    conn.close()
    
    if user:
        # 登录成功，设置session（持久 Cookie，见 PERMANENT_SESSION_LIFETIME）
        session.permanent = True
        session['user_id'] = user['id'] if USE_POSTGRES else user[0]
        session['username'] = user['username'] if USE_POSTGRES else user[1]
        session['role'] = user['role'] if USE_POSTGRES else user[2]
        
        # 获取第一个有权限的页面
        user_id = user['id'] if USE_POSTGRES else user[0]
        user_role = user['role'] if USE_POSTGRES else user[2]
        redirect_url = get_first_accessible_page(user_id, user_role)
        
        return jsonify({
            'success': True,
            'redirect': redirect_url,
            'user': {
                'id': user['id'] if USE_POSTGRES else user[0],
                'username': user['username'] if USE_POSTGRES else user[1],
                'role': user['role'] if USE_POSTGRES else user[2]
            }
        })
    else:
        return jsonify({'error': '用户名或密码错误'}), 401

# 用户登出API
@app.route('/api/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'success': True})

# 检查用户登录状态
@app.route('/api/check_login')
def check_login():
    if 'user_id' in session:
        return jsonify({
            'logged_in': True,
            'user': {
                'id': session['user_id'],
                'username': session['username'],
                'role': session['role']
            }
        })
    else:
        return jsonify({'logged_in': False})

# 获取用户权限
@app.route('/api/user_permissions')
def get_user_permissions():
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    user_id = session['user_id']
    
    # 查询用户权限
    conn = get_db()
    cursor = conn.cursor(); cursor.execute("""
        SELECT page_name, can_view, can_edit, can_delete
        FROM user_permissions
        WHERE user_id = ?
    """, (user_id,))
    
    permissions = {}
    for row in cursor.fetchall():
        page_name, can_view, can_edit, can_delete = row
        permissions[page_name] = {
            'can_view': bool(can_view),
            'can_edit': bool(can_edit),
            'can_delete': bool(can_delete)
        }
    
    conn.close()
    return jsonify(permissions)

# 用户权限装饰器
def require_permission(page_name, permission_type='view'):
    def decorator(func):
        @functools.wraps(func)
        def wrapper(*args, **kwargs):
            # 检查是否登录
            if 'user_id' not in session:
                return jsonify({'error': '未登录'}), 401
            
            # 管理员和Boss自动拥有所有权限
            if session.get('role') in ('admin', 'boss'):
                return func(*args, **kwargs)
            conn = get_db()
            query = f"""
                SELECT up.can_{permission_type}
                FROM user_permissions up
                WHERE up.user_id = ? AND up.page_name = ?
            """
            cursor = conn.cursor(); cursor.execute(query, (session['user_id'], page_name))
            result = cursor.fetchone()
            conn.close()
            
            check_val = result[f'can_{permission_type}'] if USE_POSTGRES else result[0]
            if not result or not check_val:
                return jsonify({'error': '权限不足'}), 403
            
            return func(*args, **kwargs)
        return wrapper
    return decorator


# 创建新用户（仅管理员）
@app.route('/api/users', methods=['POST'])
def create_user():
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    # 只有管理员可以创建用户
    if session.get('role') not in ('admin', 'boss'):
        return jsonify({'error': '权限不足'}), 403
    
    data = request.json
    username = data.get('username')
    password = data.get('password')
    role = data.get('role', 'user')
    is_active = data.get('is_active', True)
    
    if not username or not password:
        return jsonify({'error': '用户名和密码不能为空'}), 400
    
    # 对密码进行哈希处理
    password_hash = hashlib.sha256(password.encode()).hexdigest()
    
    try:
        conn = get_db()
        cursor = conn.cursor(); cursor.execute("""
            INSERT INTO users (username, password_hash, role, is_active)
            VALUES (?, ?, ?, ?)
        """, (username, password_hash, role, is_active))
        
        user_id = cursor.lastrowid
        conn.commit()
        
        
        # 为新用户设置默认权限（无权限）
        pages = ['index', 'sorting', 'history', 'statistics', 'logs', 'sorting-schedule']
        for page in pages:
            cursor.execute("""
                INSERT INTO user_permissions (user_id, page_name, can_view, can_edit, can_delete)
                VALUES (?, ?, 0, 0, 0)
            """, (user_id, page))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'user_id': user_id})
    except sqlite3.IntegrityError:
        return jsonify({'error': '用户名已存在'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# 获取当前登录用户的权限
@app.route('/api/user_permissions', methods=['GET'])
def get_current_user_permissions():
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    user_id = session['user_id']
    
    try:
        conn = get_db()
        cursor = conn.cursor(); cursor.execute("""
            SELECT page_name, can_view, can_edit, can_delete
            FROM user_permissions
            WHERE user_id = ?
        """, (user_id,))
        
        permissions = {}
        for row in cursor.fetchall():
            page_name, can_view, can_edit, can_delete = row
            permissions[page_name] = {
                'can_view': bool(can_view),
                'can_edit': bool(can_edit),
                'can_delete': bool(can_delete)
            }
        
        conn.close()
        return jsonify(permissions)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 用户权限管理(仅管理员) - GET获取权限, PUT更新权限
@app.route('/api/users/<int:user_id>/permissions', methods=['GET', 'PUT'])
def manage_user_permissions(user_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    # 只有管理员可以管理用户权限
    if session.get('role') not in ('admin', 'boss'):
        return jsonify({'error': '权限不足'}), 403
    
    # GET - 获取指定用户的权限
    if request.method == 'GET':
        try:
            conn = get_db()
            cursor = conn.cursor(); cursor.execute("""
                SELECT page_name, can_view, can_edit, can_delete
                FROM user_permissions
                WHERE user_id = ?
            """, (user_id,))
            
            permissions = {}
            for row in cursor.fetchall():
                page_name, can_view, can_edit, can_delete = row
                permissions[page_name] = {
                    'can_view': bool(can_view),
                    'can_edit': bool(can_edit),
                    'can_delete': bool(can_delete)
                }
            
            conn.close()
            return jsonify(permissions)
        except Exception as e:
            return jsonify({'error': str(e)}), 500
    
    # PUT - 更新用户权限
    elif request.method == 'PUT':
        data = request.json
        permissions = data.get('permissions', {})
        
        try:
            conn = get_db()
            
            # 删除旧权限
            conn.cursor().execute("DELETE FROM user_permissions WHERE user_id = ?", (user_id,))
            
            # 插入新权限
            for page_name, perms in permissions.items():
                conn.cursor().execute("""
                    INSERT INTO user_permissions 
                    (user_id, page_name, can_view, can_edit, can_delete)
                    VALUES (?, ?, ?, ?, ?)
                """, (
                    user_id, 
                    page_name, 
                    perms.get('can_view', False),
                    perms.get('can_edit', False),
                    perms.get('can_delete', False)
                ))
            
            conn.commit()
            
            # 如果修改的是当前登录用户的权限,立即刷新session中的权限
            immediate_effect = False
            if session.get('user_id') == user_id:
                # 重新加载权限到session
                cursor = conn.cursor(); cursor.execute("""
                    SELECT page_name, can_view, can_edit, can_delete
                    FROM user_permissions
                    WHERE user_id = ?
                """, (user_id,))
                
                # 更新session中的权限
                session_permissions = {}
                for row in cursor.fetchall():
                    page_name, can_view, can_edit, can_delete = row
                    session_permissions[page_name] = {
                        'can_view': bool(can_view),
                        'can_edit': bool(can_edit),
                        'can_delete': bool(can_delete)
                    }
                
                # 将权限存储到session中(如果需要的话)
                # 注意:当前实现中权限是实时从数据库查询的,这里只是示例
                immediate_effect = True
            
            conn.close()
            
            return jsonify({'success': True, 'immediate_effect': immediate_effect})
        except Exception as e:
            return jsonify({'error': str(e)}), 500

# 更新用户信息（仅管理员）
@app.route('/api/users/<int:user_id>', methods=['PUT'])
def update_user(user_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    # 只有管理员可以更新用户信息
    if session.get('role') not in ('admin', 'boss'):
        return jsonify({'error': '权限不足'}), 403
    
    data = request.json
    role = data.get('role')
    is_active = data.get('is_active')
    
    try:
        conn = get_db()
        
        if role is not None:
            conn.cursor().execute("UPDATE users SET role = ? WHERE id = ?", (role, user_id))
        
        if is_active is not None:
            conn.cursor().execute("UPDATE users SET is_active = ? WHERE id = ?", (is_active, user_id))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 验证管理员密码
@app.route('/api/admin/verify_password', methods=['POST'])
def verify_admin_password():
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    # 只有管理员或boss可以进行此操作
    if session.get('role') not in ('admin', 'boss'):
        return jsonify({'error': '权限不足'}), 403
        
    data = request.json
    password = data.get('password')
    
    if not password:
        return jsonify({'success': False, 'error': '密码不能为空'}), 400
        
    # 对输入的密码进行哈希处理
    password_hash = hashlib.sha256(password.encode()).hexdigest()
    
    # 验证当前登录管理员的密码
    user_id = session.get('user_id')
    conn = get_db()
    cursor = conn.cursor(); cursor.execute(convert_query_placeholders("""
        SELECT id FROM users WHERE id = ? AND password_hash = ?
    """), (user_id, password_hash))
    
    user = cursor.fetchone()
    conn.close()
    
    if user:
        return jsonify({'success': True})
    else:
        return jsonify({'success': False, 'error': '管理员密码错误'})

# 获取系统配置 (仅管理员)
@app.route('/api/admin/system_config', methods=['GET'])
def get_system_config():
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if session.get('role') not in ('admin', 'boss'):
        return jsonify({'error': '权限不足'}), 403
    
    key = request.args.get('key')
    if not key:
        return jsonify({'error': '缺少配置键'}), 400
        
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(convert_query_placeholders("SELECT config_value FROM system_config WHERE config_key = ?"), (key,))
        row = cursor.fetchone()
        conn.close()
        
        if row:
            # SQLite Row and Postgres DictCursor both support key access
            return jsonify({'success': True, 'value': row['config_value']})
        else:
            return jsonify({'success': True, 'value': None})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# 更新系统配置 (仅管理员)
@app.route('/api/admin/system_config', methods=['POST'])
def update_system_config():
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if session.get('role') not in ('admin', 'boss'):
        return jsonify({'error': '权限不足'}), 403
    
    data = request.json
    key = data.get('key')
    value = data.get('value')
    description = data.get('description', '')
    
    if not key:
        return jsonify({'error': '缺少配置键'}), 400
        
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # 使用 INSERT OR REPLACE (SQLite) or ON CONFLICT (PostgreSQL)
        if USE_POSTGRES:
            sql = """
                INSERT INTO system_config (config_key, config_value, description)
                VALUES (%s, %s, %s)
                ON CONFLICT (config_key) DO UPDATE SET config_value = EXCLUDED.config_value, updated_at = CURRENT_TIMESTAMP
            """
            cursor.execute(sql, (key, value, description))
        else:
            sql = "INSERT OR REPLACE INTO system_config (config_key, config_value, description) VALUES (?, ?, ?)"
            cursor.execute(sql, (key, value, description))
            
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '配置已更新'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


def _admin_gofo_dms_guard():
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    if session.get('role') not in ('admin', 'boss'):
        return jsonify({'error': '权限不足'}), 403
    if _gofo_dms_auth is None:
        return jsonify({
            'success': False,
            'error': 'Gofo DMS 登录未就绪：请在服务器执行 pip install pycryptodome',
        }), 503
    return None


@app.route('/api/admin/gofo/captcha', methods=['POST'])
def admin_gofo_captcha():
    err = _admin_gofo_dms_guard()
    if err is not None:
        return err
    try:
        s = _gofo_dms_auth.new_dms_session()
        base = _gofo_dms_auth.get_gofo_api_base()
        uuid, img_b64 = _gofo_dms_auth.fetch_captcha(s, base)
        session['gofo_dms_captcha_uuid'] = uuid
        session['gofo_dms_captcha_cookies'] = _gofo_dms_auth.session_cookies_dict(s)
        session.modified = True
        raw = (img_b64 or "").strip()
        if raw.startswith("data:"):
            img_url = raw
        else:
            img_url = "data:image/png;base64," + raw
        return jsonify({'success': True, 'img': img_url})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/gofo/login_token', methods=['POST'])
def admin_gofo_login_token():
    err = _admin_gofo_dms_guard()
    if err is not None:
        return err
    data = request.json or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    captcha_code = (data.get('captcha') or '').strip()
    plain_password = bool(data.get('plain_password'))
    if not username or not password:
        return jsonify({'success': False, 'error': '请填写 DMS 账号和密码'}), 400
    if not captcha_code:
        return jsonify({'success': False, 'error': '请填写验证码'}), 400
    uuid = session.get('gofo_dms_captcha_uuid')
    cookies = session.get('gofo_dms_captcha_cookies')
    if not uuid or not isinstance(cookies, dict):
        return jsonify({'success': False, 'error': '请先点击「获取验证码」'}), 400

    s = _gofo_dms_auth.new_dms_session()
    _gofo_dms_auth.apply_cookies_dict(s, cookies)
    base = _gofo_dms_auth.get_gofo_api_base()
    try:
        body = _gofo_dms_auth.login_dms(
            s,
            base=base,
            uuid=uuid,
            username=username,
            password_plain=password,
            captcha_code=captcha_code,
            plain_password=plain_password,
        )
    except (requests.RequestException, ValueError) as e:
        return jsonify({'success': False, 'error': str(e)}), 502

    if body.get('code') != 200:
        msg = body.get('msg') or body.get('message') or str(body)
        return jsonify({'success': False, 'error': msg, 'detail': body}), 400

    token = _gofo_dms_auth.extract_jwt_from_login_body(body)
    if not token:
        return jsonify({'success': False, 'error': '登录成功但未解析到 JWT', 'detail': body}), 500

    session.pop('gofo_dms_captcha_uuid', None)
    session.pop('gofo_dms_captcha_cookies', None)
    session.modified = True

    try:
        conn = get_db()
        cursor = conn.cursor()
        key = 'gofo_admin_token'
        desc = 'Gofo API Admin Token'
        val = token.strip()
        if USE_POSTGRES:
            sql = """
                INSERT INTO system_config (config_key, config_value, description)
                VALUES (%s, %s, %s)
                ON CONFLICT (config_key) DO UPDATE SET config_value = EXCLUDED.config_value, updated_at = CURRENT_TIMESTAMP
            """
            cursor.execute(sql, (key, val, desc))
        else:
            sql = "INSERT OR REPLACE INTO system_config (config_key, config_value, description) VALUES (?, ?, ?)"
            cursor.execute(sql, (key, val, desc))
        conn.commit()
        conn.close()
    except Exception as e:
        return jsonify({'success': False, 'error': f'已登录但写入配置失败: {e}'}), 500

    return jsonify({'success': True, 'message': 'Token 已保存'})


# 删除用户（仅管理员）
@app.route('/api/users/<int:user_id>', methods=['DELETE'])
def delete_user(user_id):
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    # 只有管理员可以删除用户
    if session.get('role') not in ('admin', 'boss'):
        return jsonify({'error': '权限不足'}), 403
    
    # 不能删除自己
    if session.get('user_id') == user_id:
        return jsonify({'error': '不能删除自己'}), 400
    
    try:
        conn = get_db()
        conn.cursor().execute("DELETE FROM users WHERE id = ?", (user_id,))
        conn.commit()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 新增API：设置揽收预估数据
@app.route('/api/pickup_forecast', methods=['POST'])
def set_pickup_forecast():
    try:
        data = request.json
        forecast_date = data.get('date')
        forecast_amount = data.get('amount')
        
        if not forecast_date or forecast_amount is None:
            return jsonify({'error': '请提供日期和预估数量'}), 400
        
        conn = get_db()
        
        # 检查是否已存在该日期的预估数据
        cursor = conn.cursor(); cursor.execute(convert_query_placeholders("SELECT id FROM pickup_forecast WHERE forecast_date = ?"), (forecast_date,))
        existing_record = cursor.fetchone()
        
        if existing_record:
            # 更新现有记录
            conn.cursor().execute("""UPDATE pickup_forecast 
                SET forecast_amount = ?, updated_at = CURRENT_TIMESTAMP 
                WHERE forecast_date = ?""", (forecast_amount, forecast_date))
        else:
            # 插入新记录
            conn.cursor().execute("""INSERT INTO pickup_forecast 
                (forecast_date, forecast_amount) 
                VALUES (?, ?)""", (forecast_date, forecast_amount))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 新增API：获取揽收预估数据
@app.route('/api/pickup_forecast')
def get_pickup_forecast():
    try:
        # 获取日期参数，默认为今天
        date_str = request.args.get('date')
        
        if not date_str:
            # 获取系统当前日期
            date_str = datetime.now(LA_TZ).strftime('%Y-%m-%d')
        
        conn = get_db()
        cursor = conn.cursor(); cursor.execute(convert_query_placeholders("SELECT forecast_amount FROM pickup_forecast WHERE forecast_date = ?"), (date_str,))
        record = cursor.fetchone()
        conn.close()
        
        if record:
            return jsonify({'amount': record[0]})
        else:
            return jsonify({'amount': 0})  # 如果没有预估数据，返回0
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# 新增API：获取当天排班配置
@app.route('/api/current-schedule')
def get_current_schedule():
    """
    返回当前业务日的排班配置
    业务日定义：05:00到次日05:00
    """
    try:
        # 获取当前洛杉矶时间
        now_la = datetime.now(LA_TZ)
        
        # 计算业务日期
        if now_la.hour < 5:
            business_date = now_la.date() - timedelta(days=1)
        else:
            business_date = now_la.date()
        
        # 获取星期几 (0=周一, 6=周日)
        day_of_week = business_date.weekday()
        
        # 星期名称
        day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        day_name = day_names[day_of_week]
        
        # 从数据库读取配置（优先 V2 当天保存；未保存则返回 0，不回退旧配置）
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute(
            convert_query_placeholders("SELECT plan_json FROM sorting_schedule_daily_plan WHERE plan_date = ? LIMIT 1"),
            (business_date.strftime('%Y-%m-%d'),),
        )
        row = cursor.fetchone()
        conn.close()
        
        # 默认配置
        defaults = {
            "manual": {"capacity": 3000, "hoursPerShift": 9, "schedule": [3, 3, 3, 3, 3, 3, 3]},
            "machine": {"capacity": 4500, "hoursPerShift": 6, "schedule": [4, 4, 4, 4, 4, 4, 4]},
            "night": {"capacity": 4500, "hoursPerShift": 6, "schedule": [4, 4, 4, 4, 4, 4, 4]}
        }
        
        daily_plan = {}
        if row:
            raw_plan = row['plan_json'] if hasattr(row, 'keys') else row[0]
            daily_plan = _parse_config_json_from_db(raw_plan, {})
            config = _v2_plan_to_schedule_config(daily_plan, business_date) or {
                "manual": {"piecesPerPersonPerHour": 310, "defaultShiftHours": 8, "shiftsByDay": [[] for _ in range(7)]},
                "machine": {"capacity": 3100, "hoursPerShift": 8, "schedule": [0] * 7, "startTimes": [17.5] * 7},
                "night": {"capacity": 0, "hoursPerShift": 0, "schedule": [0] * 7},
            }
            # 与产能计划看板「总人数」一致（产线+分拣台投入人数之和）
            manual_people = _v2_plan_total_headcount(daily_plan)
        else:
            config = {
                "manual": {"piecesPerPersonPerHour": 310, "defaultShiftHours": 8, "shiftsByDay": [[] for _ in range(7)]},
                "machine": {"capacity": 3100, "hoursPerShift": 8, "schedule": [0] * 7, "startTimes": [17.5] * 7},
                "night": {"capacity": 0, "hoursPerShift": 0, "schedule": [0] * 7},
            }
            manual_people = 0

        machine_schedule = config.get('machine', {}).get('schedule', defaults['machine']['schedule'])
        night_schedule = config.get('night', {}).get('schedule', defaults['night']['schedule'])
        machine_lanes = machine_schedule[day_of_week] if day_of_week < len(machine_schedule) else 0
        night_lanes = night_schedule[day_of_week] if day_of_week < len(night_schedule) else 0

        _mcfg = config.get('manual', {}) if isinstance(config.get('manual'), dict) else {}
        manual_cap = _manual_ppp_from_config(_mcfg)
        try:
            machine_cap = float(config.get('machine', {}).get('capacity', 4500))
            night_cap = float(config.get('night', {}).get('capacity', 4500))
        except (TypeError, ValueError):
            machine_cap = 4500.0
            night_cap = 4500.0

        machine_hours = config.get('machine', {}).get('hoursPerShift', 6)
        night_hours = config.get('night', {}).get('hoursPerShift', 6)

        def calc_time_str(start_hour, start_minute, duration_hours):
            start_time = datetime.now().replace(hour=start_hour, minute=start_minute, second=0, microsecond=0)
            end_time = start_time + timedelta(hours=duration_hours)
            return f"{start_time.strftime('%H:%M')}-{end_time.strftime('%H:%M')}"

        shifts_today = _manual_shifts_for_weekday(_mcfg, day_of_week)
        if shifts_today:
            segs = []
            for sh in shifts_today:
                sr = float(sh['start'])
                dur = float(sh['hours'])
                hh = int(sr)
                mm = int(round((sr - hh) * 60))
                if mm >= 60:
                    mm = 59
                segs.append(calc_time_str(hh, mm, dur))
            manual_time_str = '；'.join(segs) if len(segs) <= 5 else ('；'.join(segs[:4]) + '…')
        else:
            manual_time_str = '--'
        _msched_full = config.get('machine', {}) or {}
        _mlines = _msched_full.get('machineLines')
        if isinstance(_mlines, list) and len(_mlines) > 0:
            _mseg = []
            for _ln in _mlines:
                if not isinstance(_ln, dict):
                    continue
                try:
                    _sr = float(_ln.get('start', 17.5))
                    _dur = float(_ln.get('hours', machine_hours))
                except (TypeError, ValueError):
                    continue
                _hh = int(_sr)
                _mm = int(round((_sr - _hh) * 60))
                if _mm >= 60:
                    _mm = 59
                _mseg.append(calc_time_str(_hh, _mm, _dur))
            machine_time_str = '；'.join(_mseg) if _mseg else calc_time_str(18, 0, machine_hours)
        else:
            machine_time_str = calc_time_str(18, 0, machine_hours)
        night_time_str = calc_time_str(0, 0, 6) # 夜班固定 00:00-06:00 (6小时) based on request, or use night_hours if dynamic? 
        # User request says "0点到6点", implying 6 hours. But config might say 12 hours (shared with machine).
        # However, the display on statistics page is for "Night Shift Sorting", which is distinct.
        # User said "Each card display time should take value from shift duration".
        # If Night Shift duration in config is 12, then 00:00 + 12 = 12:00. This might be wrong if it's strictly Night Shift.
        # But previous logic was fixed 23:30-05:00.
        # Let's use 00:00 as start. And use 6 hours as the "Night Shift" specific duration for display, 
        # or verify if we should use the configured `night.hoursPerShift`.
        # The user request says "Take value from shift duration".
        # Let's assume night shift duration should be used.
        # But wait, in sorting-schedule.html defaults, night hoursPerShift is 12 (shared). 
        # But for *Night Sorting* specific card, it usually implies the late night shift.
        # Let's stick to the user's specific "0点到6点" for now as the base, unless duration is configurable *specifically* for night.
        # Actually, looking at `sorting-schedule.html`, night shift hours are displayed as "12 (Shared)".
        # Use 00:00 start + 6 hours fixed for now as per specific request "Night shift is 00:00 to 06:00", 
        # UNLESS the user explicitly wants the 12 hour shift displayed. 
        # "每个卡片显示的时间要取值班次时长" -> "Each card's time should take value from shift duration".
        # If I use 12 hours, it becomes 00:00-12:00. That seems too long for "Night Shift" card which usually means the grapple shift.
        # Let's look at the context. "Night Shift Sorting" 23:30-05:00 was 5.5 hours.
        # User "Night shift is 00:00 to 06:00". That is 6 hours.
        # If I use the configured '12', it contradicts. 
        # Update: I will use 6 hours hardcoded for night for now as per "0点到6点" request override, 
        # OR better, if I can find a specific night shift duration config. 
        # In `sorting-schedule.html`: `night: { hoursPerShift: 12 }`.
        # I'll use 6 hours for Night Shift display specifically to match "00:00-06:00" request, 
        # but for Manual and Machine, I will use their configured hours.
        # Wait, "Machine" is 18:00 + 12 = 06:00. That matches 18:00-06:00.
        # "Night" card in stats page... is it "Machine Night" or "Strictly Night"?
        # It says "Night Sorting".
        # Let's use manual_hours and machine_hours. 
        # For Night, I will use a calculated 6 hours (00:00-06:00) to match the request exactly, 
        # ignoring the "12 hours" shared config for this specific card display to avoid confusion.
        night_time_str = "00:00-06:00"
        return jsonify({
            'business_date': business_date.strftime('%Y-%m-%d'),
            'day_of_week': day_of_week,
            'day_name': day_name,
            'manual': {
                'people': manual_people,
                'lanes': manual_people,
                'piecesPerPersonPerHour': manual_cap,
                'capacity': manual_cap,
                'time': manual_time_str
            },
            'machine': {
                'lanes': machine_lanes,
                'capacity': machine_cap,
                'time': machine_time_str
            },
            'night': {
                'lanes': night_lanes,
                'capacity': night_cap,
                'time': night_time_str
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# Freight Rates API
# ============================================================================

@app.route('/api/freight-rates', methods=['GET'])
def get_freight_rates():
    """获取当前运费 (由于配置可能包含未来生效日期，这里取已生效或即将生效的最新配置)"""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        # 获取每个路线的最新运费 (OrderBy effective_date DESC, id DESC 确保取最新的一条)
        # 兼容 SQLite 和 PostgreSQL
        query = """
            SELECT route_code, rate
            FROM freight_rates fr1
            WHERE id = (
                SELECT id
                FROM freight_rates fr2
                WHERE fr2.route_code = fr1.route_code
                ORDER BY effective_date DESC, id DESC
                LIMIT 1
            )
        """
        
        cursor.execute(query)
        rows = cursor.fetchall()
        conn.close()
        
        # 转换为字典格式
        rates = {}
        for row in rows:
            if USE_POSTGRES:
                route_code = row['route_code']
                rate = row['rate']
            else:
                route_code = row[0]
                rate = row[1]
            rates[route_code] = float(rate)
        
        return jsonify(rates)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/freight-rates/history', methods=['GET'])
def get_freight_rate_history():
    """获取运费历史记录"""
    try:
        route_code = request.args.get('route_code')
        if not route_code:
            return jsonify({'error': 'route_code is required'}), 400
        
        conn = get_db()
        cursor = conn.cursor()
        
        # 使用 ? 作为占位符，后面会通过 convert_query_placeholders 转换
        query = """
            SELECT id, route_code, rate, effective_date, created_by, created_at, notes
            FROM freight_rates
            WHERE route_code = ?
            ORDER BY effective_date DESC, id DESC
            LIMIT 100
        """
        
        cursor.execute(convert_query_placeholders(query), (route_code,))
        rows = cursor.fetchall()
        conn.close()
        
        history = []
        for row in rows:
            if USE_POSTGRES:
                history.append({
                    'id': row['id'],
                    'route_code': row['route_code'],
                    'rate': float(row['rate']),
                    'effective_date': str(row['effective_date']) if row['effective_date'] else None,
                    'created_by': row['created_by'],
                    'created_at': row['created_at'].isoformat() if row['created_at'] else None,
                    'notes': row['notes']
                })
            else:
                history.append({
                    'id': row[0],
                    'route_code': row[1],
                    'rate': float(row[2]),
                    'effective_date': row[3],
                    'created_by': row[4],
                    'created_at': row[5],
                    'notes': row[6]
                })
        
        return jsonify(history)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/freight-rates', methods=['POST'])
def update_freight_rate():
    """更新运费(创建新的历史记录)"""
    try:
        # 检查权限
        if 'user_id' not in session:
            return jsonify({'error': 'Unauthorized'}), 401
        
        data = request.json
        route_code = data.get('route_code')
        rate = data.get('rate')
        effective_date = data.get('effective_date')
        notes = data.get('notes', '')
        
        if not route_code or rate is None:
            return jsonify({'error': 'route_code and rate are required'}), 400
        
        # 获取用户名
        username = session.get('username', 'unknown')
        
        # 统一生效日期格式为 YYYY-MM-DD 字符串，避免 SQLite 存储 datetime 对象导致的格式不一致
        if effective_date:
            try:
                # 兼容多种格式，最终转为 YYYY-MM-DD
                if 'T' in effective_date:
                    effective_date_str = effective_date.split('T')[0]
                else:
                    effective_date_str = effective_date[:10]
            except:
                effective_date_str = datetime.now(LA_TZ).strftime('%Y-%m-%d')
        else:
            effective_date_str = datetime.now(LA_TZ).strftime('%Y-%m-%d')
        
        conn = get_db()
        cursor = conn.cursor()
        
        placeholder = get_placeholder()
        query = f"""
            INSERT INTO freight_rates (route_code, rate, effective_date, created_by, notes)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
        """
        
        cursor.execute(query, (route_code, float(rate), effective_date_str, username, notes))
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Freight rate updated successfully'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# Outbound Records API
# ============================================================================

def _normalize_outbound_route_type(route_type):
    """禁止持久化 route_type=regular：与 get_route_type 兜底一致，统一记为 branch。"""
    if str(route_type or '').strip().lower() == 'regular':
        return 'branch'
    return route_type


@app.route('/api/outbound/save', methods=['POST'])
def save_outbound_record():
    """保存出库记录 (独立记录模式)"""
    try:
        if 'user_id' not in session:
            return jsonify({'error': 'Unauthorized'}), 401
        
        data = request.json
        date_str = data.get('date')
        route_code = data.get('route_code')
        route_type = _normalize_outbound_route_type(data.get('route_type'))
        vehicle_count = 1  # 一行一车次，忽略请求体中的 vehicle_count
        cost = float(data.get('cost', 0))
        notes = data.get('notes', '')
        
        if not all([date_str, route_code, route_type]):
            return jsonify({'error': 'Missing required fields'}), 400
            
        username = session.get('username', 'unknown')
        la_now = datetime.now(LA_TZ)
        created_at_utc = la_now.astimezone(pytz.UTC).strftime('%Y-%m-%d %H:%M:%SZ')
        
        conn = get_db()
        cursor = conn.cursor()
        
        placeholder = get_placeholder()
        
        # Insert new record (No check for existing)
        insert_query = f"""
            INSERT INTO outbound_records (record_date, route_code, route_type, vehicle_count, cost, notes, created_by, created_at, updated_at)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
        """
        # updated_at is initially same as created_at
        cursor.execute(insert_query, (date_str, route_code, route_type, vehicle_count, cost, notes, username, created_at_utc, created_at_utc))
        
        # Get ID
        if USE_POSTGRES:
            cursor.execute("SELECT LASTVAL()")
            record_id = cursor.fetchone()[0]
        else:
            record_id = cursor.lastrowid
            
        # Log insert
        log_query = f"""
            INSERT INTO operation_logs (operation_type, table_name, record_id, old_data, new_data, operator, created_at)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
        """
        cursor.execute(log_query, ('INSERT', 'outbound_records', record_id, None, f"Created {route_code}", username, la_now))
    
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': 'Record saved successfully'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def _format_outbound_record_date(val):
    """Serialize outbound record_date as YYYY-MM-DD for JSON (avoids ISO/T breaking frontend)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if 'T' in s:
        return s.split('T')[0][:10]
    return s[:10] if len(s) >= 10 else s


@app.route('/api/outbound/records', methods=['GET'])
def get_outbound_records():
    """获取出库记录 (支持单日和时间范围查询)"""
    try:
        # 检查查询模式
        start_date_str = request.args.get('start_date')
        end_date_str = request.args.get('end_date')
        single_date_str = request.args.get('date')
        
        conn = get_db()
        cursor = conn.cursor()
        
        if start_date_str and end_date_str:
            # 时间范围查询模式 - 按 record_date 过滤
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            
            if start_date > end_date:
                conn.close()
                return jsonify({'error': 'start_date must be before or equal to end_date'}), 400
            
            query = """
                SELECT id, record_date, route_code, route_type, vehicle_count, cost, notes, created_by, created_at, updated_at
                FROM outbound_records
                WHERE record_date >= ? AND record_date <= ?
                ORDER BY record_date DESC, created_at DESC
            """
            cursor.execute(convert_query_placeholders(query), (str(start_date), str(end_date)))
            
        elif single_date_str:
            # 单日查询模式 - 按 record_date 精确匹配
            query = """
                SELECT id, record_date, route_code, route_type, vehicle_count, cost, notes, created_by, created_at, updated_at
                FROM outbound_records
                WHERE record_date = ?
                ORDER BY created_at DESC
            """
            cursor.execute(convert_query_placeholders(query), (single_date_str,))
            
        else:
            conn.close()
            return jsonify({'error': 'date or start_date/end_date parameters required'}), 400
        
        rows = cursor.fetchall()
        conn.close()
        
        records = []
        for row in rows:
            # Use created_at for the "Entry Time" as it is now aligned with record_date
            if USE_POSTGRES:
                time_val = row['created_at']
            else:
                # created_at is column 8 (index 8)
                time_val = row[8]

            # Format time for display
            formatted_time = '-'
            if time_val:
                try:
                    if isinstance(time_val, str):
                        # If it's a "base" time from our migration (00:00:00), 
                        # just show the date part to avoid TZ shifts
                        if ' 00:00:00' in time_val:
                            dt_str = time_val.split(' ')[0]
                            dt = datetime.strptime(dt_str, '%Y-%m-%d')
                            formatted_time = dt.strftime('%m-%d') + ' 00:00'
                        else:
                            # Standard format/conversion
                            clean_time_val = time_val.replace('Z', '').replace('T', ' ')
                            try:
                                if '.' in clean_time_val:
                                    dt = datetime.strptime(clean_time_val, '%Y-%m-%d %H:%M:%S.%f')
                                else:
                                    dt = datetime.strptime(clean_time_val, '%Y-%m-%d %H:%M:%S')
                            except ValueError:
                                dt = datetime.strptime(clean_time_val[:19], '%Y-%m-%d %H:%M:%S')
                            
                            # For non-zero times, we still do the LA_TZ conversion if intended,
                            # but for simplicity and "realness", we can just show the raw hours
                            dt_la = dt.replace(tzinfo=pytz.UTC).astimezone(LA_TZ)
                            formatted_time = dt_la.strftime('%m-%d %H:%M')
                    elif isinstance(time_val, datetime):
                        dt_la = time_val.replace(tzinfo=pytz.UTC).astimezone(LA_TZ)
                        formatted_time = dt_la.strftime('%m-%d %H:%M')
                except Exception as e:
                    print(f"Time formatting error: {e}")


            if USE_POSTGRES:
                records.append({
                    'id': row['id'],
                    'record_date': _format_outbound_record_date(row['record_date']),
                    'route_code': row['route_code'],
                    'route_type': row['route_type'],
                    'vehicle_count': row['vehicle_count'],
                    'cost': float(row['cost']) if row['cost'] else 0,
                    'notes': row['notes'],
                    'created_by': row['created_by'],
                    'created_at': formatted_time
                })
            else:
                records.append({
                    'id': row[0],
                    'record_date': _format_outbound_record_date(row[1]),
                    'route_code': row[2],
                    'route_type': row[3],
                    'vehicle_count': row[4],
                    'cost': float(row[5]) if row[5] else 0,
                    'notes': row[6],
                    'created_by': row[7],
                    'created_at': formatted_time
                })
        
        return jsonify(records)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/outbound/update', methods=['PUT'])
def update_outbound_record():
    """更新出库记录"""
    try:
        if 'user_id' not in session:
            return jsonify({'error': 'Unauthorized'}), 401
        
        data = request.json
        record_id = data.get('id')
        route_code = data.get('route_code')
        route_type = _normalize_outbound_route_type(data.get('route_type'))
        vehicle_count = 1  # 一行一车次
        cost = data.get('cost', 0)
        notes = data.get('notes', '')
        
        if not record_id:
            return jsonify({'error': 'Record ID required'}), 400
        
        username = session.get('username', 'unknown')
        
        conn = get_db()
        cursor = conn.cursor()
        
        # 获取原记录用于日志
        cursor.execute(convert_query_placeholders("SELECT * FROM outbound_records WHERE id = ?"), (record_id,))
        old_record = cursor.fetchone()
        
        if not old_record:
            conn.close()
            return jsonify({'error': 'Record not found'}), 404
        
        # 更新记录
        placeholder = get_placeholder()
        query = f"""
            UPDATE outbound_records 
            SET route_code = {placeholder}, route_type = {placeholder}, vehicle_count = {placeholder}, 
                cost = {placeholder}, notes = {placeholder}
            WHERE id = {placeholder}
        """
        cursor.execute(query, (route_code, route_type, vehicle_count, cost, notes, record_id))
        
        # 记录操作日志
        log_query = f"""
            INSERT INTO operation_logs (operation_type, table_name, record_id, old_data, new_data, operator, created_at)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
        """
        
        la_now = datetime.now(LA_TZ)
        
        # 序列化旧数据
        if USE_POSTGRES:
            old_record_dict = dict(old_record)
            for k, v in old_record_dict.items():
                if isinstance(v, (datetime, date)):
                    old_record_dict[k] = v.isoformat()
            old_data = json.dumps(old_record_dict)
        else:
            old_data = json.dumps(list(old_record) if old_record else [])
            
        # 序列化新数据
        new_data_dict = {
            'route_code': route_code,
            'route_type': route_type,
            'vehicle_count': vehicle_count,
            'cost': cost,
            'notes': notes
        }
        new_data = json.dumps(new_data_dict)
        
        cursor.execute(log_query, ('UPDATE', 'outbound_records', record_id, old_data, new_data, username, la_now))
        
        conn.commit()
        
        return jsonify({'success': True, 'message': 'Record updated successfully'})
    except Exception as e:
        if conn:
            conn.rollback()
        print(f"Error updating record: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        if conn:
            conn.close()


@app.route('/api/outbound/delete', methods=['DELETE'])
def delete_outbound_record():
    """删除出库记录"""
    try:
        if 'user_id' not in session:
            return jsonify({'error': 'Unauthorized'}), 401
        
        data = request.json
        record_id = data.get('id')
        
        if not record_id:
            return jsonify({'error': 'Record ID required'}), 400
        
        username = session.get('username', 'unknown')
        
        conn = get_db()
        cursor = conn.cursor()
        
        # 获取原记录用于日志
        cursor.execute(convert_query_placeholders("SELECT * FROM outbound_records WHERE id = ?"), (record_id,))
        old_record = cursor.fetchone()
        
        if not old_record:
            conn.close()
            return jsonify({'error': 'Record not found'}), 404
        
        # 删除记录
        cursor.execute(convert_query_placeholders("DELETE FROM outbound_records WHERE id = ?"), (record_id,))
        
        # 记录操作日志
        placeholder = get_placeholder()
        log_query = f"""
            INSERT INTO operation_logs (operation_type, table_name, record_id, old_data, operator, created_at)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
        """
        
        la_now = datetime.now(LA_TZ)
        
        # 序列化旧数据
        if USE_POSTGRES:
            old_record_dict = dict(old_record)
            for k, v in old_record_dict.items():
                if isinstance(v, (datetime, date)):
                    old_record_dict[k] = v.isoformat()
            old_data = json.dumps(old_record_dict)
        else:
            old_data = json.dumps(list(old_record) if old_record else [])
            
        cursor.execute(log_query, ('DELETE', 'outbound_records', record_id, old_data, username, la_now))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Pending Outbound Shipments API
@app.route('/api/outbound/pending', methods=['GET'])
def get_pending_outbound():
    """Get aggregated pending outbound shipments for a specific date"""
    try:
        date_str = request.args.get('date')
        if not date_str:
            return jsonify({'error': 'Date required'}), 400
            
        conn = get_db()
        cursor = conn.cursor()
        
        placeholder = get_placeholder()
        query = f"""
            SELECT route_code, SUM(quantity) as total_quantity 
            FROM pending_shipments 
            WHERE record_date = {placeholder} 
            GROUP BY route_code 
            ORDER BY route_code
        """
        
        cursor.execute(query, (date_str,))
        rows = cursor.fetchall()
        
        records = []
        for row in rows:
            records.append({
                'route_code': row[0],
                'quantity': row[1]
            })
            
        conn.close()
        return jsonify({'records': records})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/outbound/pending/save', methods=['POST'])
def save_pending_outbound():
    """Save a pending outbound shipment record"""
    try:
        if 'user_id' not in session:
            return jsonify({'error': 'Unauthorized'}), 401
            
        data = request.json
        record_date = data.get('date')
        route_code = data.get('route_code')
        quantity = data.get('quantity')
        
        if not all([record_date, route_code, quantity]):
             return jsonify({'error': 'Missing required fields'}), 400
             
        username = session.get('username', 'unknown')
        la_now = datetime.now(LA_TZ) # Store as LA time string or UTC? Database usually UTC.
        # SQLite creates as UTC by default if using CURRENT_TIMESTAMP, but here we insert manually?
        # Actually existing tables utilize default CURRENT_TIMESTAMP. 
        # But we want to store it explicitly to be safe? 
        # Let's use UTC for created_at
        created_at_utc = datetime.now(pytz.UTC)
        
        conn = get_db()
        cursor = conn.cursor()
        
        placeholder = get_placeholder()
        query = f"""
            INSERT INTO pending_shipments (record_date, route_code, quantity, created_at)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder})
        """
        cursor.execute(query, (record_date, route_code, quantity, created_at_utc))
        
        # Log operation
        if USE_POSTGRES:
            cursor.execute("SELECT LASTVAL()")
            record_id = cursor.fetchone()[0]
        else:
            record_id = cursor.lastrowid
            
        log_query = f"""
            INSERT INTO operation_logs (operation_type, table_name, record_id, old_data, new_data, operator, created_at)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
        """
        new_data = json.dumps(data)
        cursor.execute(log_query, ('INSERT', 'pending_shipments', record_id, None, new_data, username, la_now))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/outbound/pending/delete', methods=['DELETE'])
def delete_pending_outbound():
    """Delete all pending outbound shipment records for a route on a date"""
    try:
        if 'user_id' not in session:
            return jsonify({'error': 'Unauthorized'}), 401
            
        data = request.json
        route_code = data.get('route_code')
        record_date = data.get('date')
        
        if not route_code or not record_date:
            return jsonify({'error': 'Route and Date required'}), 400
            
        username = session.get('username', 'unknown')
        conn = get_db()
        cursor = conn.cursor()
        
        placeholder = get_placeholder()
        
        # Get count for logs
        check_query = f"SELECT COUNT(*) FROM pending_shipments WHERE route_code = {placeholder} AND record_date = {placeholder}"
        cursor.execute(check_query, (route_code, record_date))
        count = cursor.fetchone()[0]
        
        if count == 0:
            conn.close()
            return jsonify({'error': 'Records not found'}), 404
            
        # Delete
        delete_query = f"DELETE FROM pending_shipments WHERE route_code = {placeholder} AND record_date = {placeholder}"
        cursor.execute(delete_query, (route_code, record_date))
        
        # Log
        log_query = f"""
            INSERT INTO operation_logs (operation_type, table_name, record_id, old_data, operator, created_at)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
        """
        la_now = datetime.now(LA_TZ)
        
        old_data = json.dumps({'route_code': route_code, 'record_date': record_date, 'deleted_count': count})
            
        cursor.execute(log_query, ('DELETE', 'pending_shipments', 0, old_data, username, la_now))
        
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==========================================
# 耗材进销存管理 API (Consumables Inventory)
# ==========================================

@app.route('/api/consumables', methods=['GET'])
def get_consumables():
    """获取所有耗材状态"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        query = f"""
            SELECT c.id, c.name, COALESCE(m.unit, c.unit) as unit, c.safety_stock, c.current_stock, c.lead_time_days, c.updated_at,
                   m.price, m.batch
            FROM consumables c
            LEFT JOIN production_consumable_master m ON c.name = m.name
            WHERE c.safety_stock >= 0 AND NOT (c.safety_stock = 0 AND c.current_stock = 0)
            ORDER BY c.name ASC
        """
        cursor.execute(query)
        rows = cursor.fetchall()
        
        consumables_list = []
        for row in rows:
            # Handle row access based on DB Type
            if hasattr(row, 'keys'):
                # PostgreSQL dict-like row
                c_data = dict(row)
            else:
                # SQLite tuple
                c_data = {
                    'id': row[0],
                    'name': row[1],
                    'unit': row[2],
                    'safety_stock': row[3],
                    'current_stock': row[4],
                    'lead_time_days': row[5],
                    'updated_at': row[6],
                    'price': row[7],
                    'batch': row[8]
                }
            consumables_list.append(c_data)
            
        conn.close()
        return jsonify({'success': True, 'data': consumables_list})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/consumables/master_list', methods=['GET'])
def get_consumables_master_list():
    """获取生产耗材主数据列表(用于下拉选择)"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("SELECT id, name, price, batch, unit FROM production_consumable_master ORDER BY name ASC")
        rows = cursor.fetchall()
        
        items = []
        for row in rows:
            if hasattr(row, 'keys'):
                items.append(dict(row))
            else:
                items.append({
                    'id': row[0],
                    'name': row[1],
                    'price': row[2],
                    'batch': row[3],
                    'unit': row[4]
                })
        conn.close()
        return jsonify({'success': True, 'data': items})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# --- Admin API: Production Consumables Configuration ---

@app.route('/api/admin/production_consumables', methods=['GET'])
def get_production_consumables():
    if 'user_id' not in session:
        return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
    result = cursor.fetchone()
    current_role = result['role'] if USE_POSTGRES else result[0]
    if current_role not in ('admin', 'boss'):
        conn.close()
        return jsonify({'error': '无权访问', 'success': False}), 403

    cursor.execute("SELECT * FROM production_consumable_master ORDER BY id DESC")
    items = []
    for row in cursor.fetchall():
        items.append(dict(row))
    conn.close()
    return jsonify(items)

@app.route('/api/admin/production_consumables/add', methods=['POST'])
def add_production_consumable():
    if 'user_id' not in session:
        return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
    result = cursor.fetchone()
    current_role = result['role'] if USE_POSTGRES else result[0]
    if current_role not in ('admin', 'boss'):
        conn.close()
        return jsonify({'error': '无权访问', 'success': False}), 403

    data = request.json
    name = data.get('name')
    price = data.get('price')
    batch = data.get('batch') # Optional
    unit = data.get('unit') # Optional

    if not name or price is None:
        return jsonify({'success': False, 'error': '品名和价格为必填项'}), 400
        
    try:
        price = float(price)
    except ValueError:
        return jsonify({'success': False, 'error': '价格必须为数字'}), 400

    username = session.get('username', 'system')
    la_now = datetime.now(LA_TZ)
    la_now_str = la_now.strftime('%Y-%m-%d %H:%M:%S') 
    
    placeholder = get_placeholder()
    try:
        cursor.execute(f"""
            INSERT INTO production_consumable_master (name, price, batch, unit, created_at, updated_at) 
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
        """, (name, price, batch, unit, la_now_str, la_now_str))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'success': False, 'error': '该品名已存在'}), 400
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500
        
    conn.close()
    return jsonify({'success': True})

@app.route('/api/admin/production_consumables/update', methods=['POST'])
def update_production_consumable():
    if 'user_id' not in session:
        return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
    result = cursor.fetchone()
    current_role = result['role'] if USE_POSTGRES else result[0]
    if current_role not in ('admin', 'boss'):
        conn.close()
        return jsonify({'error': '无权访问', 'success': False}), 403

    data = request.json
    item_id = data.get('id')
    name = data.get('name')
    price = data.get('price')
    batch = data.get('batch')
    unit = data.get('unit')

    if not item_id or not name or price is None:
        return jsonify({'success': False, 'error': 'ID、品名和价格为必填项'}), 400
        
    try:
        price = float(price)
    except ValueError:
        return jsonify({'success': False, 'error': '价格必须为数字'}), 400

    username = session.get('username', 'system')
    la_now_str = datetime.now(LA_TZ).strftime('%Y-%m-%d %H:%M:%S')
    
    placeholder = get_placeholder()
    try:
        cursor.execute(f"""
            UPDATE production_consumable_master 
            SET name = {placeholder}, price = {placeholder}, batch = {placeholder}, unit = {placeholder}, updated_at = {placeholder}
            WHERE id = {placeholder}
        """, (name, price, batch, unit, la_now_str, item_id))
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'success': False, 'error': '该品名已存在'}), 400
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500
        
    conn.close()
    return jsonify({'success': True})

@app.route('/api/admin/production_consumables/delete', methods=['POST'])
def delete_production_consumable():
    if 'user_id' not in session:
        return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
    result = cursor.fetchone()
    current_role = result['role'] if USE_POSTGRES else result[0]
    if current_role not in ('admin', 'boss'):
        conn.close()
        return jsonify({'error': '无权访问', 'success': False}), 403

    data = request.json
    item_id = data.get('id')

    if not item_id:
        return jsonify({'success': False, 'error': 'ID 不能为空'}), 400

    username = session.get('username', 'system')
    placeholder = get_placeholder()
    try:
        # Check if the name exists in consumables table
        cursor.execute(f"SELECT name FROM production_consumable_master WHERE id = {placeholder}", (item_id,))
        row = cursor.fetchone()
        if row:
            master_name = row['name'] if hasattr(row, 'keys') else row[0]
            cursor.execute(f"SELECT COUNT(*) as count FROM consumables WHERE name = {placeholder}", (master_name,))
            count_res = cursor.fetchone()
            count = count_res['count'] if hasattr(count_res, 'keys') else count_res[0]
            if count > 0:
                conn.close()
                return jsonify({'success': False, 'error': f'请先删除耗材管理页面中依赖此品名({master_name})的项，然后再删除基础配置！'}), 400

        cursor.execute(f"DELETE FROM production_consumable_master WHERE id = {placeholder}", (item_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500
        
    conn.close()
    return jsonify({'success': True})

# --- Admin API: Labor and Equipment Cost Configuration ---

def require_admin(cursor):
    """Helper to check if current user is admin"""
    cursor.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
    result = cursor.fetchone()
    current_role = result['role'] if USE_POSTGRES else result[0]
    return current_role in ('admin', 'boss')

# 1. 人工计时配置 (labor_hourly)
@app.route('/api/admin/config/labor_hourly', methods=['GET'])
def get_labor_hourly():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    cursor.execute("SELECT * FROM config_labor_hourly ORDER BY id DESC")
    items = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(items)

@app.route('/api/admin/config/labor_hourly', methods=['POST'])
def save_labor_hourly():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    
    data = request.json
    item_id = data.get('id')
    company_name = data.get('company_name')
    hourly_rate = data.get('hourly_rate')
    if not company_name or hourly_rate is None: return jsonify({'success': False, 'error': '劳务公司和时薪为必填项'}), 400
    try: hourly_rate = float(hourly_rate)
    except: return jsonify({'success': False, 'error': '时薪必须为数字'}), 400

    la_now_str = datetime.now(LA_TZ).strftime('%Y-%m-%d %H:%M:%S')
    placeholder = get_placeholder()
    try:
        if item_id:
            cursor.execute(f"UPDATE config_labor_hourly SET company_name={placeholder}, hourly_rate={placeholder}, updated_at={placeholder} WHERE id={placeholder}", 
                          (company_name, hourly_rate, la_now_str, item_id))
        else:
            cursor.execute(f"INSERT INTO config_labor_hourly (company_name, hourly_rate, updated_at) VALUES ({placeholder}, {placeholder}, {placeholder})", 
                          (company_name, hourly_rate, la_now_str))
        conn.commit()
    except sqlite3.IntegrityError: return jsonify({'success': False, 'error': '该公司配置已存在'}), 400
    except Exception as e: conn.rollback(); return jsonify({'success': False, 'error': str(e)}), 500
    finally: conn.close()
    return jsonify({'success': True})

@app.route('/api/admin/config/labor_hourly/delete', methods=['POST'])
def delete_labor_hourly():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    item_id = request.json.get('id')
    if not item_id: return jsonify({'success': False, 'error': 'ID为空'}), 400
    try:
        cursor.execute(f"DELETE FROM config_labor_hourly WHERE id = {get_placeholder()}", (item_id,))
        conn.commit()
    except Exception as e: conn.rollback(); return jsonify({'success': False, 'error': str(e)}), 500
    finally: conn.close()
    return jsonify({'success': True})

# 2. 计件费用配置 (labor_piece)
@app.route('/api/admin/config/labor_piece', methods=['GET'])
def get_labor_piece():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    cursor.execute("SELECT * FROM config_labor_piece ORDER BY id DESC")
    items = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(items)

@app.route('/api/admin/config/labor_piece', methods=['POST'])
def save_labor_piece():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    
    data = request.json
    item_id = data.get('id')
    operation_name = data.get('operation_name')
    piece_rate = data.get('piece_rate')
    if not operation_name or piece_rate is None: return jsonify({'success': False, 'error': '作业类型和计件单价为必填项'}), 400
    try: piece_rate = float(piece_rate)
    except: return jsonify({'success': False, 'error': '计件单价必须为数字'}), 400

    la_now_str = datetime.now(LA_TZ).strftime('%Y-%m-%d %H:%M:%S')
    placeholder = get_placeholder()
    try:
        if item_id:
            cursor.execute(f"UPDATE config_labor_piece SET operation_name={placeholder}, piece_rate={placeholder}, updated_at={placeholder} WHERE id={placeholder}", 
                          (operation_name, piece_rate, la_now_str, item_id))
        else:
            cursor.execute(f"INSERT INTO config_labor_piece (operation_name, piece_rate, updated_at) VALUES ({placeholder}, {placeholder}, {placeholder})", 
                          (operation_name, piece_rate, la_now_str))
        conn.commit()
    except sqlite3.IntegrityError: return jsonify({'success': False, 'error': '该作业类型配置已存在'}), 400
    except Exception as e: conn.rollback(); return jsonify({'success': False, 'error': str(e)}), 500
    finally: conn.close()
    return jsonify({'success': True})

@app.route('/api/admin/config/labor_piece/delete', methods=['POST'])
def delete_labor_piece():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    item_id = request.json.get('id')
    if not item_id: return jsonify({'success': False, 'error': 'ID为空'}), 400
    try:
        cursor.execute(f"DELETE FROM config_labor_piece WHERE id = {get_placeholder()}", (item_id,))
        conn.commit()
    except Exception as e: conn.rollback(); return jsonify({'success': False, 'error': str(e)}), 500
    finally: conn.close()
    return jsonify({'success': True})

# 3. 设备维护配置 (equipment_hourly)
@app.route('/api/admin/config/equipment_hourly', methods=['GET'])
def get_equipment_hourly():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    cursor.execute("SELECT * FROM config_equipment_hourly ORDER BY id DESC")
    items = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(items)

@app.route('/api/admin/config/equipment_hourly', methods=['POST'])
def save_equipment_hourly():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    
    data = request.json
    item_id = data.get('id')
    equipment_name = data.get('equipment_name')
    hourly_rate = data.get('hourly_rate')
    if not equipment_name or hourly_rate is None: return jsonify({'success': False, 'error': '设备类型和时薪为必填项'}), 400
    try: hourly_rate = float(hourly_rate)
    except: return jsonify({'success': False, 'error': '时薪必须为数字'}), 400

    la_now_str = datetime.now(LA_TZ).strftime('%Y-%m-%d %H:%M:%S')
    placeholder = get_placeholder()
    try:
        if item_id:
            cursor.execute(f"UPDATE config_equipment_hourly SET equipment_name={placeholder}, hourly_rate={placeholder}, updated_at={placeholder} WHERE id={placeholder}", 
                          (equipment_name, hourly_rate, la_now_str, item_id))
        else:
            cursor.execute(f"INSERT INTO config_equipment_hourly (equipment_name, hourly_rate, updated_at) VALUES ({placeholder}, {placeholder}, {placeholder})", 
                          (equipment_name, hourly_rate, la_now_str))
        conn.commit()
    except sqlite3.IntegrityError: return jsonify({'success': False, 'error': '该设备类型配置已存在'}), 400
    except Exception as e: conn.rollback(); return jsonify({'success': False, 'error': str(e)}), 500
    finally: conn.close()
    return jsonify({'success': True})

@app.route('/api/admin/config/equipment_hourly/delete', methods=['POST'])
def delete_equipment_hourly():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    item_id = request.json.get('id')
    if not item_id: return jsonify({'success': False, 'error': 'ID为空'}), 400
    try:
        cursor.execute(f"DELETE FROM config_equipment_hourly WHERE id = {get_placeholder()}", (item_id,))
        conn.commit()
    except Exception as e: conn.rollback(); return jsonify({'success': False, 'error': str(e)}), 500
    finally: conn.close()
    return jsonify({'success': True})

# ====== 新增: 成本核算系统 API ======

# 1. 流向工序计件配置 (config_labor_price_flow)
@app.route('/api/admin/config/labor_price_flow', methods=['GET'])
def get_labor_price_flow():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    cursor.execute("SELECT * FROM config_labor_price_flow ORDER BY id DESC")
    items = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(items)

@app.route('/api/admin/config/labor_price_flow', methods=['POST'])
def save_labor_price_flow():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    
    data = request.json
    item_id = data.get('id')
    direction = data.get('direction')
    operation_name = data.get('operation_name')
    piece_rate = data.get('piece_rate')
    if not direction or not operation_name or piece_rate is None: return jsonify({'success': False, 'error': '流向、作业类型和计件单价为必填项'}), 400
    try: piece_rate = float(piece_rate)
    except: return jsonify({'success': False, 'error': '计件单价必须为数字'}), 400

    la_now_str = datetime.now(LA_TZ).strftime('%Y-%m-%d %H:%M:%S')
    placeholder = get_placeholder()
    try:
        if item_id:
            cursor.execute(f"UPDATE config_labor_price_flow SET direction={placeholder}, operation_name={placeholder}, piece_rate={placeholder}, updated_at={placeholder} WHERE id={placeholder}", 
                          (direction, operation_name, piece_rate, la_now_str, item_id))
        else:
            cursor.execute(f"INSERT INTO config_labor_price_flow (direction, operation_name, piece_rate, updated_at) VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder})", 
                          (direction, operation_name, piece_rate, la_now_str))
        conn.commit()
    except sqlite3.IntegrityError: return jsonify({'success': False, 'error': '该流向上的该作业类型配置已存在'}), 400
    except Exception as e: conn.rollback(); return jsonify({'success': False, 'error': str(e)}), 500
    finally: conn.close()
    return jsonify({'success': True})

@app.route('/api/admin/config/labor_price_flow/delete', methods=['POST'])
def delete_labor_price_flow():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    item_id = request.json.get('id')
    if not item_id: return jsonify({'success': False, 'error': 'ID为空'}), 400
    try:
        cursor.execute(f"DELETE FROM config_labor_price_flow WHERE id = {get_placeholder()}", (item_id,))
        conn.commit()
    except Exception as e: conn.rollback(); return jsonify({'success': False, 'error': str(e)}), 500
    finally: conn.close()
    return jsonify({'success': True})

# 2. 耗材分摊规则 (config_consumable_split)
@app.route('/api/admin/config/consumable_split', methods=['GET'])
def get_consumable_split():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    cursor.execute("SELECT * FROM config_consumable_split ORDER BY id DESC")
    items = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify(items)

@app.route('/api/admin/config/consumable_split', methods=['POST'])
def save_consumable_split():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    
    data = request.json
    item_id = data.get('id')
    consumable_name = data.get('consumable_name')
    split_method = data.get('split_method', 'weight')
    weight_json = data.get('weight_json', '{}')
    
    if not consumable_name: return jsonify({'success': False, 'error': '耗材名称为必填项'}), 400

    la_now_str = datetime.now(LA_TZ).strftime('%Y-%m-%d %H:%M:%S')
    placeholder = get_placeholder()
    try:
        if item_id:
            cursor.execute(f"UPDATE config_consumable_split SET consumable_name={placeholder}, split_method={placeholder}, weight_json={placeholder}, updated_at={placeholder} WHERE id={placeholder}", 
                          (consumable_name, split_method, weight_json, la_now_str, item_id))
        else:
            cursor.execute(f"INSERT INTO config_consumable_split (consumable_name, split_method, weight_json, updated_at) VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder})", 
                          (consumable_name, split_method, weight_json, la_now_str))
        conn.commit()
    except sqlite3.IntegrityError: return jsonify({'success': False, 'error': '该耗材规则已存在'}), 400
    except Exception as e: conn.rollback(); return jsonify({'success': False, 'error': str(e)}), 500
    finally: conn.close()
    return jsonify({'success': True})

@app.route('/api/admin/config/consumable_split/delete', methods=['POST'])
def delete_consumable_split():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    item_id = request.json.get('id')
    if not item_id: return jsonify({'success': False, 'error': 'ID为空'}), 400
    try:
        cursor.execute(f"DELETE FROM config_consumable_split WHERE id = {get_placeholder()}", (item_id,))
        conn.commit()
    except Exception as e: conn.rollback(); return jsonify({'success': False, 'error': str(e)}), 500
    finally: conn.close()
    return jsonify({'success': True})
    return jsonify({'success': True})

# In-memory store for sync status since the daemon shares this state
_sync_status_cache = {
    'Outsource': {'last_sync': None, 'success': False},
    'Gofo': {'last_sync': None, 'success': False}
}

def update_sync_status(source, success_flag):
    tz = LA_TZ
    now = datetime.now(tz)
    _sync_status_cache[source] = {
        'last_sync': now.strftime('%Y-%m-%d %H:%M:%S'),
        'success': success_flag
    }

@app.route('/api/admin/labor_sync_status', methods=['GET'])
def get_labor_sync_status():
    if 'user_id' not in session:
        return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db()
    cursor = conn.cursor()
    if not require_admin(cursor):
        conn.close()
        return jsonify({'error': '无权访问', 'success': False}), 403
    conn.close()
    return jsonify({
        'success': True,
        'data': _sync_status_cache
    })

# 3. 生产人工数据源自动与手动同步与导出
@app.route('/api/admin/outsource/sync', methods=['POST'])
def sync_outsource_data():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    conn.close()
    try:
        data = request.json
        link = data.get('link', '').strip()
        print(f"[DEBUG API] /api/admin/outsource/sync CALLED with link: {link}")
        
        # Save the link using a local sqlite connection
        if link:
            conn2 = get_db(); cur2 = conn2.cursor()
            cur2.execute("INSERT OR REPLACE INTO system_config (config_key, config_value) VALUES ('last_feishu_link', ?)", (link,))
            conn2.commit()
            conn2.close()
            
        import calc_outsource_finance
        import importlib
        importlib.reload(calc_outsource_finance)
        
        # If no link provided, try to load the last one
        if not link:
            conn2 = get_db(); cur2 = conn2.cursor()
            cur2.execute("SELECT config_value FROM system_config WHERE config_key = 'last_feishu_link'")
            row = cur2.fetchone()
            conn2.close()
            if row:
                link = row[0]
                print(f"[DEBUG API] Using saved link: {link}")
        
        result = calc_outsource_finance.run_sync(link)
        print(f"[DEBUG API] /api/admin/outsource/sync RESULT: {result}")
        
        # Log manual execution
        update_sync_status('Outsource', result.get('success', False))
        
        return jsonify(result)
    except Exception as e:
        print(f"[DEBUG API] /api/admin/outsource/sync ERROR: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/gofo/sync', methods=['POST'])
def sync_gofo_piece_rate():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    conn.close()

    try:
        data = request.json or {}
        target_date_str = data.get('target_date', '')
        
        # Mark sync as in-progress immediately
        _sync_status_cache['Gofo'] = {
            'last_sync': datetime.now(LA_TZ).strftime('%Y-%m-%d %H:%M:%S (Processing...)'),
            'success': False  # Default to false until finished
        }
        
        import threading
        
        def background_sync_task(date_param):
            try:
                import calc_gofo_piece_rate
                import importlib
                importlib.reload(calc_gofo_piece_rate)
                
                result = calc_gofo_piece_rate.fetch_and_summarize_gofo_piece_rate(date_param if date_param else None)
                
                # Log execution result
                update_sync_status('Gofo', result.get('success', False))
            except Exception as e:
                # Update status with failure on thread exception
                _sync_status_cache['Gofo'] = {
                    'last_sync': datetime.now(LA_TZ).strftime('%Y-%m-%d %H:%M:%S (Failed)'),
                    'success': False
                }
                print(f"[DEBUG API] GOFO Background thread ERROR: {e}")

        # Start thread
        thread = threading.Thread(target=background_sync_task, args=(target_date_str,))
        thread.daemon = True
        thread.start()
        
        return jsonify({'success': True, 'message': '后台正在抓取Gofo数据，请稍后查看同步状态...'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/gofo/cno_narrowbelt_sync_day', methods=['POST'])
def admin_gofo_cno_narrowbelt_sync_day():
    """提交后台任务：洛杉矶日历日从 0 点起拉取 CNO 窄带 operatelog（当日截至今、历史日拉满 24h）。

    同步 HTTP 若跑满易超时导致 net::ERR_CONNECTION_RESET，故与计件拉取一样改为守护线程执行。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db()
    cursor = conn.cursor()
    if not require_admin(cursor):
        conn.close()
        return jsonify({'error': '无权访问', 'success': False}), 403
    conn.close()

    data = request.json or {}
    date_str = (data.get('date') or '').strip()[:10]
    if not date_str:
        date_str = datetime.now(LA_TZ).strftime('%Y-%m-%d')

    _start_cno_operlog_hourly_sync_background(date_str, log_prefix='CnoNarrowbeltAdminSync')
    return jsonify({
        'success': True,
        'async': True,
        'date': date_str,
        'message': (
            f'已提交后台拉取 {date_str}（洛杉矶日历日 0 点起；Operalog 切片多，可能需数分钟）。'
            f'请留意服务端日志 [CnoNarrowbeltAdminSync]，完成后刷新统计页。'
        ),
    })


@app.route('/api/admin/gofo/backfill_sorting_range', methods=['POST'])
def admin_gofo_backfill_sorting_range():
    """按日期区间从 Gofo 重抓 hourly 集包数据并写回 sorting_records（更正统计图等）。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db()
    cursor = conn.cursor()
    if not require_admin(cursor):
        conn.close()
        return jsonify({'error': '无权访问', 'success': False}), 403
    conn.close()

    data = request.json or {}
    start_date = (data.get('start_date') or '').strip()
    end_date = (data.get('end_date') or '').strip()
    if not start_date or not end_date:
        return jsonify({'success': False, 'error': '请提供 start_date 与 end_date（YYYY-MM-DD）'}), 400
    try:
        result = perform_gofo_backfill_range(start_date, end_date)
        return jsonify(result)
    except ValueError as e:
        return jsonify({'success': False, 'error': str(e)}), 400
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/outsource/export', methods=['GET'])
def export_outsource_data():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    
    import csv
    import io
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'inbound.db')
    
    if not os.path.exists(db_path):
        return jsonify({'success': False, 'error': '未找到核算数据文件，请先同步飞书数据'}), 404
        
    try:
        conn = get_db()
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM daily_cost_summary ORDER BY Record_Date DESC")
        rows = cursor.fetchall()
        conn.close()
        
        output = io.StringIO()
        if len(rows) > 0:
            writer = csv.writer(output)
            writer.writerow(rows[0].keys())
            for row in rows:
                writer.writerow(row)
        else:
            output.write("没有查到明细数据\n")
            
        from flask import Response
        output.seek(0)
        return Response(
            # Append BOM for Excel compatibility
            '\uFEFF' + output.read(),
            mimetype="text/csv",
            headers={"Content-disposition": "attachment; filename=outsource_finance_details.csv"}
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/statistics/waybill-trends', methods=['GET'])
def get_waybill_trends():
    """获取飞书运单票数趋势 (按流向聚合)"""
    try:
        days = int(request.args.get('days', 30))
        days = max(1, min(days, 366))
        end_raw = request.args.get('end_date') or request.args.get('date')
        if end_raw:
            try:
                end_d = datetime.strptime(str(end_raw)[:10], '%Y-%m-%d').date()
            except ValueError:
                end_d = datetime.now(LA_TZ).date()
        else:
            end_d = datetime.now(LA_TZ).date()
        start_d = end_d - timedelta(days=days - 1)
        start_date = start_d.strftime('%Y-%m-%d')
        end_date = end_d.strftime('%Y-%m-%d')
        conn = get_db()
        cursor = conn.cursor()
        
        # 聚合查询：按日期和流向汇总票数（窗口 [start_date, end_date]）
        query = """
            SELECT record_date, destination as direction, SUM(tickets_count) as total_tickets
            FROM feishu_raw_data
            WHERE record_date >= ? AND record_date <= ?
            GROUP BY record_date, destination
            ORDER BY record_date ASC
        """
        
        cursor.execute(convert_query_placeholders(query), (start_date, end_date))
        rows = cursor.fetchall()
        conn.close()
        
        # 转换数据格式供 Chart.js 使用
        data_by_date = {} # { date: { direction: tickets } }
        all_directions = set()
        all_dates = []
        
        for row in rows:
            if USE_POSTGRES:
                r_date = str(row['record_date']).split(' ')[0][:10]
                direction = row['direction']
                tickets = row['total_tickets']
            else:
                r_date = str(row[0]).split(' ')[0][:10]
                direction = row[1]
                tickets = row[2]
            
            if r_date not in data_by_date:
                data_by_date[r_date] = {}
                all_dates.append(r_date)
            
            # Normalize destinations (Aggregated view)
            if direction:
                direction = direction.upper()
                if direction.endswith('.H'): direction = direction[:-2]
                if "LAS" in direction: direction = "LAS"
                elif "PHX" in direction: direction = "PHX"
                elif "ATL.G" in direction: direction = "ATL"
                
            data_by_date[r_date][direction] = data_by_date[r_date].get(direction, 0) + tickets
            all_directions.add(direction)
            
        # 确保日期排序
        all_dates.sort()
        sorted_directions = sorted(list(all_directions))
        
        datasets = []
        for direction in sorted_directions:
            dataset = {
                'label': direction,
                'data': [data_by_date[d].get(direction, 0) for d in all_dates]
            }
            datasets.append(dataset)
            
        return jsonify({
            'labels': [d[5:] for d in all_dates], # 只显示 MM-DD
            'datasets': datasets
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/statistics/weekly-waybill-trends', methods=['GET'])
def get_weekly_waybill_trends():
    """获取飞书运单票数每周趋势 (按流向堆叠)"""
    try:
        end_raw = request.args.get('end_date') or request.args.get('date')
        end_filter = None
        if end_raw:
            try:
                end_filter = datetime.strptime(str(end_raw)[:10], '%Y-%m-%d').strftime('%Y-%m-%d')
            except ValueError:
                end_filter = None
        conn = get_db()
        cursor = conn.cursor()
        
        # 聚合查询：按日期和流向汇总票数
        query = """
            SELECT record_date, destination as direction, SUM(tickets_count) as total_tickets
            FROM feishu_raw_data
            GROUP BY record_date, destination
            ORDER BY record_date ASC
        """
        
        cursor.execute(convert_query_placeholders(query))
        rows = cursor.fetchall()
        conn.close()
        
        if not rows:
            return jsonify({'labels': [], 'datasets': []})

        ws_raw = request.args.get('week_start')
        we_raw = request.args.get('week_end')
        week_mon_min = None
        week_mon_max = None
        if ws_raw and we_raw:
            try:
                d0 = datetime.strptime(str(ws_raw)[:10], '%Y-%m-%d').date()
                d1 = datetime.strptime(str(we_raw)[:10], '%Y-%m-%d').date()
                week_mon_min = d0 - timedelta(days=d0.weekday())
                week_mon_max = d1 - timedelta(days=d1.weekday())
                if week_mon_min > week_mon_max:
                    week_mon_min, week_mon_max = week_mon_max, week_mon_min
            except ValueError:
                week_mon_min = week_mon_max = None
            
        # 按周聚合数据
        def get_week_label(date_str):
            date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
            weekday = date_obj.weekday()
            week_start = date_obj - timedelta(days=weekday)
            week_end = week_start + timedelta(days=6)
            return week_start, f"W{week_start.isocalendar()[1]} ({week_start.strftime('%m/%d')})"

        weeks_data = {} # { week_start_date: { direction: total_tickets } }
        week_labels = {} # { week_start_date: "WXX (MM/DD)" }
        all_directions = set()
        
        for row in rows:
            if USE_POSTGRES:
                r_date = str(row['record_date']).split(' ')[0][:10]
                direction = row['direction']
                tickets = row['total_tickets']
            else:
                r_date = str(row[0]).split(' ')[0][:10]
                direction = row[1]
                tickets = row[2]
            
            if not direction: continue
            if end_filter and r_date > end_filter:
                continue
                
            week_start, label = get_week_label(r_date)
            if week_start not in weeks_data:
                weeks_data[week_start] = {}
                week_labels[week_start] = label
            
            # Normalize destinations (Aggregated view)
            if direction:
                direction = direction.upper()
                if direction.endswith('.H'): direction = direction[:-2]
                if "LAS" in direction: direction = "LAS"
                elif "PHX" in direction: direction = "PHX"
                elif "ATL.G" in direction: direction = "ATL"
                
            weeks_data[week_start][direction] = weeks_data[week_start].get(direction, 0) + (tickets or 0)
            all_directions.add(direction)
            
        # 排序周
        sorted_week_starts = sorted(weeks_data.keys())
        if week_mon_min is not None and week_mon_max is not None:
            sorted_week_starts = [w for w in sorted_week_starts if week_mon_min <= w <= week_mon_max]
        sorted_directions = sorted(list(all_directions))
        if not sorted_week_starts:
            return jsonify({'labels': [], 'datasets': [], 'wow_rates': []})

        datasets = []
        weekly_totals = []
        for w in sorted_week_starts:
            # Calculate total for this week across all directions
            total = sum(weeks_data[w].get(direction, 0) for direction in sorted_directions)
            weekly_totals.append(total)

        for direction in sorted_directions:
            dataset = {
                'label': direction,
                'data': [weeks_data[w].get(direction, 0) for w in sorted_week_starts]
            }
            datasets.append(dataset)
            
        # Calculate WoW % (current - previous) / previous * 100
        wow_rates = [0] # First week has no WoW reference
        for i in range(1, len(weekly_totals)):
            prev = weekly_totals[i-1]
            curr = weekly_totals[i]
            if prev > 0:
                rate = round(((curr - prev) / prev) * 100, 2)
            else:
                rate = 0
            wow_rates.append(rate)

        return jsonify({
            'labels': [week_labels[w] for w in sorted_week_starts],
            'datasets': datasets,
            'wow_rates': wow_rates
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/gofo/export', methods=['GET'])
def export_gofo_data():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    
    import csv
    import io
    import os
    db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'inbound.db')
    
    if not os.path.exists(db_path):
        return jsonify({'success': False, 'error': '未找到核算数据文件，请先同步数据'}), 404
        
    try:
        conn = get_db()
        cursor = conn.cursor()
        if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
        
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Check if table exists
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='gofo_piece_rate_summary'")
        if not cursor.fetchone():
            conn.close()
            return jsonify({'success': False, 'error': 'Gofo数据表不存在，请先执行拉取'}), 404
            
        cursor.execute("SELECT * FROM gofo_piece_rate_summary ORDER BY Record_Date DESC")
        rows = cursor.fetchall()
        conn.close()
        
        output = io.StringIO()
        if len(rows) > 0:
            writer = csv.writer(output)
            writer.writerow(rows[0].keys())
            for row in rows:
                writer.writerow(row)
        else:
            output.write("没有查到Gofo计件汇总数据\n")
            
        from flask import Response
        output.seek(0)
        return Response(
            '\uFEFF' + output.read(),
            mimetype="text/csv",
            headers={"Content-disposition": "attachment; filename=gofo_piece_rate_summary.csv"}
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
# ==================================

# ==============================================================
# 成本核算业务数据 API
# ==============================================================

@app.route('/api/admin/cost_accounting/preview', methods=['GET'])
def get_cost_accounting_preview():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    directions_str = request.args.get('directions', '')
    
    if not start_date or not end_date:
        conn.close()
        return jsonify({'success': False, 'error': '缺失日期参数'}), 400
        
    directions = [d.strip() for d in directions_str.split(',') if d.strip() and d.strip() not in ('篮筐', '托盘')]
    
    # 1. 从新的 feishu_raw_data 取票数和箱数
    # 取出独立发车数据后进行聚合 (SUM) 作为单票成本核算的源头
    feishu_query = """
        SELECT record_date, destination as direction, SUM(tickets_count) as total_pieces, SUM(boxes_count) as total_boxes
        FROM feishu_raw_data
        WHERE record_date >= ? AND record_date <= ?
    """
    params = [start_date, end_date]
    if directions:
        placeholders = ','.join(['?'] * len(directions))
        feishu_query += f" AND destination IN ({placeholders})"
        params.extend(directions)
    feishu_query += " GROUP BY record_date, destination"
    
    try:
        cursor.execute(convert_query_placeholders(feishu_query), tuple(params))
        feishu_rows = cursor.fetchall()
        
        # 汇总到 direction 层级供总表使用
        feishu_data = {}
        pieces_daily = {}
        pieces_detailed = {} # { date: { direction: pieces } }
        
        for row in feishu_rows:
            d = row['direction']
            rdate = row['record_date']
            
            if rdate:
                pieces_daily[rdate] = pieces_daily.get(rdate, 0) + row['total_pieces']
                if rdate not in pieces_detailed: pieces_detailed[rdate] = {}
                pieces_detailed[rdate][d] = pieces_detailed[rdate].get(d, 0) + row['total_pieces']

            if not d: continue
            if d not in feishu_data:
                feishu_data[d] = {'total_pieces': 0, 'total_boxes': 0}
            feishu_data[d]['total_pieces'] += row['total_pieces']
            feishu_data[d]['total_boxes'] += row['total_boxes']
            
        pieces_data = {k: v['total_pieces'] for k, v in feishu_data.items()}
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': f'查询飞书聚合票数失败: {str(e)}'}), 500

    # 2. 查询运费数据 (transport cost), 剔除篮筐和托盘
    transport_query = """
        SELECT record_date, route_code as direction, SUM(cost) as total_transport_cost, SUM(vehicle_count) as total_vehicles
        FROM outbound_records 
        WHERE record_date >= ? AND record_date <= ?
        AND route_code NOT IN ('篮筐', '托盘')
        AND route_type NOT IN ('篮筐', '托盘')
    """
    transport_params = [start_date, end_date]
    if directions:
        placeholders = ','.join(['?'] * len(directions))
        transport_query += f" AND route_code IN ({placeholders})"
        transport_params.extend(directions)
    transport_query += " GROUP BY record_date, route_code"
    
    try:
        cursor.execute(convert_query_placeholders(transport_query), tuple(transport_params))
        transport_rows = cursor.fetchall()
        
        transport_data = {}
        transport_daily = {}
        transport_detailed = {} # { date: { direction: cost } }
        
        for row in transport_rows:
            d = row['direction']
            rdate = row['record_date']
            cost = row['total_transport_cost']
            vehicles = row['total_vehicles']
            
            if rdate:
                transport_daily[rdate] = transport_daily.get(rdate, 0) + cost
                if rdate not in transport_detailed: transport_detailed[rdate] = {}
                transport_detailed[rdate][d] = transport_detailed[rdate].get(d, 0) + cost
                
            if not d: continue
            if d not in transport_data:
                transport_data[d] = {'cost': 0, 'vehicles': 0}
            transport_data[d]['cost'] += cost
            transport_data[d]['vehicles'] += vehicles
            
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': f'查询运费失败: {str(e)}'}), 500
    
    # 3. 计算人工费 (labor cost) purely from pulled datasets
    try:
        labor_cost_data = {}
        labor_daily = {}  # { date: [ {operation, pieces, rate, subtotal} ] }
            
        # Fetch Hourly and Piece Rates from the main DB (daily_cost_summary)
        # This unified table contains both Feishu-mapped hourly and Gofo-mapped piece rates.
        import datetime
        try:
            l_query = """SELECT Record_Date, Agency_Name, Hourly_Cost_USD, Piece_Cost_USD 
                         FROM daily_cost_summary 
                         WHERE Agency_Name != '【当日总计】'"""
            
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='daily_cost_summary'")
            if cursor.fetchone():
                cursor.execute(l_query)
                l_rows = cursor.fetchall()
            else:
                l_rows = []
                
            # Parse start_date and end_date
            s_date = datetime.datetime.strptime(start_date, '%Y-%m-%d').date() if start_date else None
            e_date = datetime.datetime.strptime(end_date, '%Y-%m-%d').date() if end_date else None
            
            for rdate_str, agency, h_cost, p_cost in l_rows:
                # Date format normalization
                try:
                    dt = datetime.datetime.strptime(rdate_str, '%m/%d/%y').date()
                except:
                    try:
                        dt = datetime.datetime.strptime(rdate_str, '%Y-%m-%d').date()
                    except:
                        continue
                        
                if s_date and dt < s_date: continue
                if e_date and dt > e_date: continue
                    
                formatted_rdate = dt.strftime('%Y-%m-%d')
                
                if formatted_rdate not in labor_daily:
                    labor_daily[formatted_rdate] = []
                
                if h_cost and h_cost > 0:
                    labor_daily[formatted_rdate].append({
                        'operation_name': f"时薪-{agency}",
                        'subtotal': h_cost,
                        'is_hourly': True
                    })
                    
                    if 'Hourly' not in labor_cost_data:
                        labor_cost_data['Hourly'] = []
                    labor_cost_data['Hourly'].append({
                        'operation_name': f"时薪-{agency}",
                        'piece_rate': 0,
                        'pieces': 0,
                        'subtotal': h_cost,
                        'is_hourly': True
                    })
                
                if p_cost and p_cost > 0:
                    labor_daily[formatted_rdate].append({
                        'operation_name': f"计件-{agency}",
                        'subtotal': p_cost,
                        'is_hourly': False
                    })
                    
                    if 'Gofo计件' not in labor_cost_data:
                        labor_cost_data['Gofo计件'] = []
                    labor_cost_data['Gofo计件'].append({
                        'operation_name': f"计件-{agency}",
                        'piece_rate': 0,
                        'pieces': 0,
                        'subtotal': p_cost,
                        'is_hourly': False
                    })

        except Exception as e:
            print("Read outsourced data error:", e)
    except Exception as e:
        conn.close()
        return jsonify({'success': False, 'error': f'计算人工费失败: {str(e)}'}), 500

    # 4. 汇总该周期的耗材消耗 (consumable cost)
    try:
        consumable_query = """
            SELECT c.name, SUM(t.quantity) as total_qty, m.price, DATE(t.created_at) as cdate
            FROM inventory_transactions t
            JOIN consumables c ON t.consumable_id = c.id
            LEFT JOIN production_consumable_master m ON c.name = m.name
            WHERE t.type = 'OUT' AND t.quantity > 0 AND DATE(t.created_at) >= ? AND DATE(t.created_at) <= ?
            GROUP BY cdate, c.name
        """
        cursor.execute(convert_query_placeholders(consumable_query), (start_date, end_date))
        transactions = [dict(row) for row in cursor.fetchall()]
        
        # [PATCH REFINED] 确保每一天都有最少 80 卷缠绕膜的消耗进入核算
        # 1. 计算日期范围内的天数
        from datetime import datetime
        try:
            d1 = datetime.strptime(start_date, '%Y-%m-%d')
            d2 = datetime.strptime(end_date, '%Y-%m-%d')
            num_days = (d2 - d1).days + 1
            if num_days <= 0: num_days = 1
            mandatory_total_qty = 80.0 * num_days
            
            # 2. 检查现有的缠绕膜记录总量
            actual_sf_qty = sum(t['total_qty'] for t in transactions if '缠绕膜' in t['name'])
            
            # 3. 如果不足或者没有，则补齐差额（确保统计表中每一行最终都能分到这 80 卷/天 的基本费用）
            if actual_sf_qty < mandatory_total_qty:
                diff_qty = mandatory_total_qty - actual_sf_qty
                # 获取单价
                cursor.execute("SELECT price FROM production_consumable_master WHERE name = '缠绕膜'")
                sf_p = cursor.fetchone()
                sf_price = sf_p['price'] if sf_p else 6.5
                
                # 找到有生产的日期（优先）或 start_date
                active_dates = sorted(pieces_daily.keys())
                f_date = active_dates[0] if active_dates else start_date
                
                transactions.append({
                    'name': '缠绕膜 (系统补全)',
                    'total_qty': diff_qty,
                    'price': sf_price,
                    'cdate': f_date
                })
        except: pass

        
        # 5. 应用耗材分摊规则
        split_query = "SELECT consumable_name, split_method, weight_json FROM config_consumable_split"
        cursor.execute(split_query)
        split_configs = {row['consumable_name']: dict(row) for row in cursor.fetchall()}
        
        consumable_cost_data = {} 
        consumables_daily = {}
        
        for t in transactions:
            name = t['name']
            cdate = t['cdate']
            qty = t['total_qty'] or 0
            price = t['price'] or 0
            total_amount = qty * price
            if total_amount == 0: continue
            
            if cdate not in consumables_daily:
                consumables_daily[cdate] = []
            
            consumables_daily[cdate].append({
                'name': name,
                'total_qty': qty,
                'price': price,
                'total_amount': total_amount
            })
            
            # 如果是系统补全的耗材，由于 config 里没有它，强制使用 average 分摊
            if '系统补全' in name:
                cfg = {'split_method': 'average'}
            else:
                cfg = split_configs.get(name)

            if cfg and cfg['split_method'] == 'weight':
                try:
                    weights = json.loads(cfg['weight_json'])
                    if directions:
                        weights = {k: v for k, v in weights.items() if k in directions}
                        
                    total_weight = sum(weights.values())
                    if total_weight > 0:
                        for d, w in weights.items():
                            if d not in consumable_cost_data: consumable_cost_data[d] = []
                            split_amount = total_amount * (w / total_weight)
                            
                            # Note: The directional consumable_cost_data aggregates across all days in the period
                            # It is possible a single direction+consumable pair gets added multiple times. We will group it.
                            exist_item = next((i for i in consumable_cost_data[d] if i['name'] == name), None)
                            if exist_item:
                                exist_item['total_qty'] += qty * (w / total_weight)
                                exist_item['total_amount'] += split_amount
                                exist_item['split_amount'] += split_amount
                            else:
                                consumable_cost_data[d].append({
                                    'name': name,
                                    'total_qty': qty * (w / total_weight), # Estimated qty allocated to here
                                    'price': price,
                                    'total_amount': split_amount, # for final
                                    'split_ratio': w / total_weight,
                                    'split_amount': split_amount
                                })
                except: pass
            elif cfg and cfg['split_method'] == 'average':
                target_dirs = directions if directions else list(pieces_data.keys())
                count = len(target_dirs)
                if count > 0:
                    for d in target_dirs:
                        if d not in consumable_cost_data: consumable_cost_data[d] = []
                        split_amount = total_amount / count
                        
                        exist_item = next((i for i in consumable_cost_data[d] if i['name'] == name), None)
                        if exist_item:
                            exist_item['total_qty'] += qty / count
                            exist_item['total_amount'] += split_amount
                            exist_item['split_amount'] += split_amount
                        else:
                            consumable_cost_data[d].append({
                                'name': name,
                                'total_qty': qty / count,
                                'price': price,
                                'total_amount': split_amount,
                                'split_ratio': 1.0 / count,
                                'split_amount': split_amount
                            })
    except Exception as e:
        conn.close()
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'计算耗材费失败: {str(e)}'}), 500
    
    conn.close()
    
    result = {
        'directions': list(set(list(pieces_data.keys()) + list(transport_data.keys()))),
        'pieces': pieces_data,
        'transport': transport_data,
        'labor_daily': labor_daily,
        'consumables_daily': consumables_daily,
        'pieces_daily': pieces_daily,
        'transport_daily': transport_daily,
        'pieces_detailed': pieces_detailed,
        'transport_detailed': transport_detailed,
        'labor': labor_cost_data,           # Retained for backward compat
        'consumables': consumable_cost_data # Retained for backward compat
    }
    
    if directions:
        result['directions'] = [d for d in result['directions'] if d in directions]
        
    return jsonify({'success': True, 'data': result})


def _feishu_sync_core(link=''):
    """Shared sync logic called by HTTP route and scheduled job."""
    import re
    DEFAULT_TOKEN = "SvBYstNvyhvh8ptbq29cMc7Ln8c"
    DEFAULT_SHEET = "24dfdb"
    spreadsheet_token = DEFAULT_TOKEN
    sheet_id = DEFAULT_SHEET
    if link:
        m_token = re.search(r'/sheets/([A-Za-z0-9]+)', link)
        if m_token: spreadsheet_token = m_token.group(1)
        m_sheet = re.search(r'[?&]sheet=([A-Za-z0-9]+)', link)
        if m_sheet: sheet_id = m_sheet.group(1)

    import json
    try:
        token = feishu_auth.feishu_tenant_access_token()

        url_sheet = (f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
                     f"{spreadsheet_token}/values/{sheet_id}!A1:AG30000?valueRenderOption=FormattedValue")
        res_sheet = requests.get(url_sheet, headers={"Authorization": f"Bearer {token}"})
        if res_sheet.status_code != 200:
            return False, f'读取飞书表格失败: {res_sheet.text}'

        values = res_sheet.json().get("data", {}).get("valueRange", {}).get("values", [])
        if not values or len(values) < 2:
            return False, '飞书表格中没有数据'

        header_row = values[0]
        col = {str(h).strip(): i for i, h in enumerate(header_row) if h}
        def find_col(*names):
            for n in names:
                if n in col: return col[n]
            return None

        COL_DATE     = find_col('发车日期', '日期')
        COL_TASK     = find_col('任务编码', '任务号')
        COL_ORIGIN   = find_col('始发地')
        COL_DEST     = find_col('目的地')
        COL_BOXES    = find_col('装车箱数-合并', '装车箱数', '装车筱数-合并', '装车筱数')
        COL_TICKETS  = find_col('装车总票数-合并', '装车总票数')
        COL_VOL_RATE = find_col('装载率（%）', '装载率', '体积装载率')
        COL_BOX_RATE = find_col('车型核载箱数', '车型核载票数', '车型核载筱数', '核载箱数')

        req_cols = [('发车日期', COL_DATE), ('任务编码', COL_TASK),
                    ('始发地', COL_ORIGIN), ('目的地', COL_DEST),
                    ('装车箱数-合并', COL_BOXES), ('装车总票数-合并', COL_TICKETS)]
        missing = [name for name, idx in req_cols if idx is None]
        if missing:
            return False, f'表格缺少必需列: {", ".join(missing)}，请检查表格结构'

        conn = get_db(); cursor = conn.cursor()
        sync_count = 0
        placeholder = get_placeholder()
        for i in range(1, len(values)):
            row = values[i]
            task_code = str(row[COL_TASK] if len(row) > COL_TASK else "").strip()
            if not task_code: continue
            start_location = str(row[COL_ORIGIN] if len(row) > COL_ORIGIN else "").strip()
            if "CNO.H" not in start_location: continue

            record_date_raw = str(row[COL_DATE] if len(row) > COL_DATE else "").strip()
            destination = str(row[COL_DEST] if len(row) > COL_DEST else "").strip()

            record_date = None
            for fmt in ('%Y/%m/%d', '%Y-%m-%d', '%m/%d/%y', '%m/%d/%Y', '%d/%m/%Y', '%Y年%m月%d日'):
                try:
                    from datetime import datetime as _dt
                    record_date = _dt.strptime(record_date_raw, fmt).strftime('%Y-%m-%d')
                    break
                except (ValueError, TypeError):
                    continue
            if not record_date:
                digits = ''.join(c for c in record_date_raw if c.isdigit())
                if len(digits) == 8:
                    record_date = f"{digits[:4]}-{digits[4:6]}-{digits[6:8]}"
                else:
                    continue

            if destination.endswith('.H'):
                destination = destination[:-2]
            # No longer excluding .G routes as per user request for complete data
            # if destination.endswith('.G'):
            #     continue
            if not record_date or not destination:
                continue

            try: boxes = int(float(str(row[COL_BOXES]).strip() if len(row) > COL_BOXES and row[COL_BOXES] else "0"))
            except: boxes = 0
            try: tickets = int(float(str(row[COL_TICKETS]).strip() if len(row) > COL_TICKETS and row[COL_TICKETS] else "0"))
            except: tickets = 0

            vol_rate = str(row[COL_VOL_RATE] if COL_VOL_RATE and len(row) > COL_VOL_RATE else "").strip()
            box_rate = str(row[COL_BOX_RATE] if COL_BOX_RATE and len(row) > COL_BOX_RATE else "").strip()
            la_now_str = datetime.now(LA_TZ).strftime('%Y-%m-%d %H:%M:%S')

            insert_query = f"""
                INSERT INTO feishu_raw_data
                (task_code, record_date, destination, boxes_count, tickets_count, volume_load_rate, box_load_rate, updated_at)
                VALUES ({placeholder},{placeholder},{placeholder},{placeholder},{placeholder},{placeholder},{placeholder},{placeholder})
                ON CONFLICT(task_code) DO UPDATE SET
                    record_date=excluded.record_date, destination=excluded.destination,
                    boxes_count=excluded.boxes_count, tickets_count=excluded.tickets_count,
                    volume_load_rate=excluded.volume_load_rate, box_load_rate=excluded.box_load_rate,
                    updated_at=excluded.updated_at
            """
            cursor.execute(insert_query, (task_code, record_date, destination, boxes, tickets, vol_rate, box_rate, la_now_str))
            if cursor.rowcount > 0: sync_count += 1

        conn.commit(); conn.close()
        return True, f'成功同步 {sync_count} 条记录'
    except Exception as e:
        import traceback
        return False, f'同步失败: {str(e)}\n{traceback.format_exc()}'


@app.route('/api/admin/feishu/sync', methods=['POST'])
def sync_feishu_data():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    conn.close()
    data = request.json or {}
    link = data.get('link', '').strip()
    ok, msg = _feishu_sync_core(link=link)
    if ok:
        return jsonify({'success': True, 'message': msg})
    else:
        return jsonify({'success': False, 'error': msg}), 500


def _get_system_config_value(cursor, config_key: str, default: str = '') -> str:
    cursor.execute(
        convert_query_placeholders(
            "SELECT config_value FROM system_config WHERE config_key = ?"
        ),
        (config_key,),
    )
    row = cursor.fetchone()
    if not row:
        return default
    val = _db_row_get(row, 'config_value', None)
    if val is None:
        val = _db_row_get(row, 0, default)
    return str(val or default).strip()


def _set_system_config_value(cursor, config_key: str, config_value: str) -> None:
    desc = 'auto-saved'
    if USE_POSTGRES:
        cursor.execute(
            """
            INSERT INTO system_config (config_key, config_value, description)
            VALUES (%s, %s, %s)
            ON CONFLICT (config_key) DO UPDATE SET
                config_value = EXCLUDED.config_value,
                updated_at = CURRENT_TIMESTAMP
            """,
            (config_key, config_value, desc),
        )
    else:
        cursor.execute(
            "INSERT OR REPLACE INTO system_config (config_key, config_value, description) VALUES (?, ?, ?)",
            (config_key, config_value, desc),
        )


@app.route('/api/admin/labor/last_link', methods=['GET'])
def get_labor_last_link():
    """返回上次同步使用过的生产人工飞书表格链接。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db()
    cursor = conn.cursor()
    if not require_admin(cursor):
        conn.close()
        return jsonify({'error': '无权访问', 'success': False}), 403
    try:
        link = _get_system_config_value(cursor, 'last_labor_sheet_link', '')
        return jsonify({'success': True, 'link': link})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/admin/labor/sync', methods=['POST'])
def sync_labor_data():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    
    data = request.json or {}
    link = (data.get('link') or '').strip()
    if not link:
        link = _get_system_config_value(cursor, 'last_labor_sheet_link', '')
    if link:
        try:
            _set_system_config_value(cursor, 'last_labor_sheet_link', link)
            conn.commit()
        except Exception:
            pass
    conn.close()
    
    # 调用 calc_outsource_finance 中的 run_sync
    # 注意: run_sync 内部会打印日志并返回 dict
    try:
        res = calc_outsource_finance.run_sync(link=link)
        if res.get('success'):
            return jsonify({'success': True, 'message': res.get('message', '同步成功')})
        else:
            return jsonify({'success': False, 'error': res.get('error', '同步失败')}), 500
    except Exception as e:
        import traceback
        return jsonify({'success': False, 'error': f"执行同步脚本失败: {str(e)}", 'traceback': traceback.format_exc()}), 500


@app.route('/api/admin/labor/data', methods=['GET'])
def get_labor_data():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # 从 daily_cost_summary 表中读取数据
    query = "SELECT * FROM daily_cost_summary WHERE Agency_Name != '【当日总计】'"
    params = []
    if start_date and end_date:
        query += " AND Record_Date >= ? AND Record_Date <= ?"
        params.extend([start_date, end_date])
    query += " ORDER BY Record_Date DESC, Agency_Name ASC"
    
    try:
        cursor.execute(convert_query_placeholders(query), tuple(params))
        items = [dict(row) for row in cursor.fetchall()]
        return jsonify({'success': True, 'data': items})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/admin/labor/weekly-summary', methods=['GET'])
def get_weekly_labor_summary():
    """获取每周人工成本汇总 (按公司堆叠)"""
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    
    try:
        # 查询日常成本摘要中的人工数据，排除占比很小的 OTHERS 和 Unknown
        query = "SELECT Record_Date, Agency_Name, Total_Cost_USD FROM daily_cost_summary WHERE Agency_Name NOT IN ('【当日总计】', 'OTHERS', 'Unknown', 'Others', 'UNKNOWN') ORDER BY Record_Date ASC"
        cursor.execute(convert_query_placeholders(query))
        rows = cursor.fetchall()
        conn.close()
        
        if not rows:
            return jsonify({'labels': [], 'datasets': []})
            
        # 按周聚合数据
        def get_week_label(date_str):
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                weekday = date_obj.weekday()
                week_start = date_obj - timedelta(days=weekday)
                return week_start, f"W{week_start.isocalendar()[1]} ({week_start.strftime('%m/%d')})"
            except:
                return None, None

        weeks_data = {} # { week_start_date: { agency: total_cost } }
        week_labels = {} # { week_start_date: "WXX (MM/DD)" }
        all_agencies = set()
        
        for row in rows:
            if USE_POSTGRES:
                r_date = str(row['record_date'])
                agency = row['agency_name']
                cost = float(row['total_cost_usd'] or 0)
            else:
                r_date = str(row[0])
                agency = row[1]
                cost = float(row[2] or 0)
            
            if not agency or not r_date: continue
            
            # 合并 UNS 和 A-SHARE (同一家公司)
            if agency.upper() == 'UNS':
                agency = 'A-SHARE'
                
            week_start, label = get_week_label(r_date)
            if not week_start: continue
            
            if week_start not in weeks_data:
                weeks_data[week_start] = {}
                week_labels[week_start] = label
            
            weeks_data[week_start][agency] = weeks_data[week_start].get(agency, 0) + cost
            all_agencies.add(agency)
            
        # 排序周
        sorted_week_starts = sorted(weeks_data.keys())
        sorted_agencies = sorted(list(all_agencies))
        
        datasets = []
        for agency in sorted_agencies:
            dataset = {
                'name': agency,
                'type': 'bar',
                'stack': 'total',
                'data': [round(weeks_data[w].get(agency, 0), 2) for w in sorted_week_starts]
            }
            datasets.append(dataset)
            
        return jsonify({
            'success': True,
            'labels': [week_labels[w] for w in sorted_week_starts],
            'datasets': datasets
        })
    except Exception as e:
        if 'conn' in locals(): conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500


def _is_waybill_comparison_excluded_gofo_destin(dest):
    """运单对比集包量：排除本站 CNO / CNO01（含 .H / .G 后缀）。"""
    if dest is None or not str(dest).strip():
        return False
    r = str(dest).upper().strip()
    if r.endswith('.H'):
        r = r[:-2]
    if r.endswith('.G'):
        r = r[:-2]
    return r in ('CNO', 'CNO01')


_GOFO_COMPARISON_DESTIN_MERGE = {
    'TUC01': 'LAS',
    'IFP01': 'LAS',
    'CGZ01': 'PHX',
}


def _normalize_gofo_comparison_direction(dest):
    """CNO 集包 destin_name → 运单对比流向；站点合并见 _GOFO_COMPARISON_DESTIN_MERGE。"""
    if dest is None or not str(dest).strip():
        return 'OTHER'
    r = str(dest).upper().strip()
    if r.endswith('.H'):
        r = r[:-2]
    if r.endswith('.G'):
        r = r[:-2]
    merged = _GOFO_COMPARISON_DESTIN_MERGE.get(r)
    if merged:
        return merged
    return _normalize_comparison_direction(dest)


def _normalize_comparison_direction(dest):
    """运单对比：将飞书目的地规范为可与 CNO 集包流向比对的代码。"""
    if dest is None or not str(dest).strip():
        return 'OTHER'
    r = str(dest).upper().strip()
    if r.endswith('.H'):
        r = r[:-2]
    if r.endswith('.G'):
        r = r[:-2]
    if 'LAS' in r:
        return 'LAS'
    if 'PHX' in r:
        return 'PHX'
    if 'LAX' in r:
        return 'LAX'
    if 'SEA' in r:
        return 'SEA'
    if 'MCO' in r:
        return 'MCO'
    return r or 'OTHER'


def _feishu_raw_tickets_total(cursor, start_date, end_date):
    """与「数据采集与核算 → 飞书运单数据源」相同：feishu_raw_data.tickets_count 区间合计。"""
    cursor.execute(
        convert_query_placeholders("""
            SELECT COALESCE(SUM(tickets_count), 0) AS total
            FROM feishu_raw_data
            WHERE record_date >= ? AND record_date <= ?
        """),
        (start_date, end_date),
    )
    row = cursor.fetchone()
    return int((row['total'] if row else 0) or 0)


@app.route('/api/admin/feishu/data', methods=['GET'])
def get_feishu_data():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # 聚合查询：按发车日期和目的地求和，展示给前端
    query = """
        SELECT record_date, destination, 
               SUM(boxes_count) as boxes_count, 
               SUM(tickets_count) as tickets_count,
               MAX(volume_load_rate) as volume_load_rate, 
               MAX(box_load_rate) as box_load_rate
        FROM feishu_raw_data
    """
    params = []
    if start_date and end_date:
        query += " WHERE record_date >= ? AND record_date <= ?"
        params.extend([start_date, end_date])
    query += " GROUP BY record_date, destination ORDER BY record_date DESC, destination ASC"
    
    try:
        cursor.execute(convert_query_placeholders(query), tuple(params))
        items = [dict(row) for row in cursor.fetchall()]
        return jsonify({'success': True, 'data': items})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


@app.route('/api/admin/cost_accounting/save', methods=['POST'])
def save_cost_accounting():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    
    data = request.json
    start_date = data.get('period_start')
    end_date = data.get('period_end')
    records = data.get('records', [])
    
    if not start_date or not end_date or not records:
        conn.close()
        return jsonify({'success': False, 'error': '数据不完整'}), 400
        
    la_now_str = datetime.now(LA_TZ).strftime('%Y-%m-%d %H:%M:%S')
    placeholder = get_placeholder()
    
    try:
        # [REFINEMENT] Overwrite logic: Delete existing records for the target dates and directions to prevent duplication
        # First, collect all unique dates being saved
        save_dates = set()
        for rec in records:
            r_start = rec.get('period_start', start_date)
            if r_start: save_dates.add(r_start)
        
        if save_dates:
            placeholders = ','.join(['?'] * len(save_dates))
            # Delete daily records for these specific days
            delete_query = f"DELETE FROM cost_main WHERE period_start IN ({placeholders}) AND period_start = period_end"
            cursor.execute(delete_query, tuple(save_dates))
            print(f"   - Cleared {cursor.rowcount} daily records.")
            
            # [ADDITION] Also clear old-style aggregate records that exactly match this range
            if start_date and end_date and start_date != end_date:
                del_agg = f"DELETE FROM cost_main WHERE period_start = {placeholder} AND period_end = {placeholder}"
                cursor.execute(del_agg, (start_date, end_date))
                print(f"   - Cleared {cursor.rowcount} aggregate records for range {start_date} to {end_date}.")

        for rec in records:
            direction = rec.get('direction')
            # Use record-specific dates if provided, fallback to period defaults
            r_start = rec.get('period_start', start_date)
            r_end = rec.get('period_end', end_date)
            
            transport_cost = float(rec.get('transport_cost', 0))
            labor_cost = float(rec.get('labor_cost', 0))
            consumable_cost = float(rec.get('consumable_cost', 0))
            other_cost = float(rec.get('other_cost', 0))
            total_cost = float(rec.get('total_cost', 0))
            total_pieces = int(rec.get('total_pieces', 0))
            unit_cost = float(rec.get('unit_cost', 0))
            time_labor_cost = float(rec.get('time_labor_cost', 0))
            piece_labor_cost = float(rec.get('piece_labor_cost', 0))
            
            insert_query = f"""
                INSERT INTO cost_main 
                (period_start, period_end, direction, total_transport_cost, total_labor_cost, 
                 total_consumable_cost, total_other_cost, total_cost, total_pieces, unit_cost, 
                 total_time_labor_cost, total_piece_labor_cost, created_at, updated_at)
                VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder},
                        {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder},
                        {placeholder}, {placeholder}, {placeholder}, {placeholder})
            """
            cursor.execute(insert_query, (
                r_start, r_end, direction, transport_cost, labor_cost, 
                consumable_cost, other_cost, total_cost, total_pieces, unit_cost,
                time_labor_cost, piece_labor_cost, la_now_str, la_now_str
            ))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()
        
    return jsonify({'success': True})

@app.route('/api/admin/cost_accounting/history', methods=['GET'])
def get_cost_accounting_history():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    placeholder = get_placeholder()
    
    query = "SELECT * FROM cost_main"
    params = []
    conditions = []
    if start_date and end_date:
        # Overlap: any record whose accounting period overlaps the selected range
        conditions.append(f"period_start <= {placeholder}")
        params.append(end_date)
        conditions.append(f"period_end >= {placeholder}")
        params.append(start_date)
    elif start_date:
        conditions.append(f"period_end >= {placeholder}")
        params.append(start_date)
    elif end_date:
        conditions.append(f"period_start <= {placeholder}")
        params.append(end_date)
    if conditions:
        query += " WHERE " + " AND ".join(conditions)
    query += " ORDER BY period_end DESC, direction ASC"
    
    cursor.execute(query, tuple(params))
    items = [dict(row) for row in cursor.fetchall()]
    conn.close()
    return jsonify({'success': True, 'data': items})

@app.route('/api/admin/cost_accounting/summary', methods=['GET'])
def get_cost_accounting_summary():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    placeholder = get_placeholder()
    
    if not start_date or not end_date:
        conn.close()
        return jsonify({'success': False, 'error': '缺少日期参数'}), 400
    
    query = f"""
        SELECT
            SUM(total_pieces) as total_pieces,
            SUM(total_transport_cost) as total_transport_cost,
            SUM(total_time_labor_cost) as total_time_labor_cost,
            SUM(total_piece_labor_cost) as total_piece_labor_cost,
            SUM(total_consumable_cost) as total_consumable_cost,
            SUM(total_labor_cost) as total_labor_cost,
            SUM(total_cost) as total_cost,
            COUNT(DISTINCT period_end) as period_count
        FROM cost_main
        WHERE period_start <= {placeholder} AND period_end >= {placeholder}
    """
    cursor.execute(query, (end_date, start_date))
    row = cursor.fetchone()
    conn.close()
    
    if not row:
        return jsonify({'success': True, 'data': {
            'total_pieces': 0, 'total_transport_cost': 0.0,
            'total_time_labor_cost': 0.0, 'total_piece_labor_cost': 0.0,
            'total_consumable_cost': 0.0, 'total_labor_cost': 0.0,
            'total_cost': 0.0, 'period_count': 0
        }})
    
    data = dict(row)
    # Normalize destinations (Consolidated view)
    if 'direction' in data and data['direction']:
        dir_val = str(data['direction']).upper()
        if "ATL.G" in dir_val: 
            data['direction'] = "ATL"
        elif dir_val.endswith('.H'):
            data['direction'] = dir_val[:-2]

    # Ensure all values are numbers not None
    for key in data:
        if data[key] is None:
            data[key] = 0
    
    return jsonify({'success': True, 'data': data})


@app.route('/api/admin/cost_accounting/delete', methods=['POST'])
def delete_cost_accounting():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    
    data = request.json
    record_id = data.get('id')
    if not record_id:
        conn.close()
        return jsonify({'success': False, 'error': 'Missing ID'}), 400

    try:
        placeholder = get_placeholder()
        cursor.execute(f"DELETE FROM cost_main WHERE id = {placeholder}", (record_id,))
        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()
        
    return jsonify({'success': True})


@app.route('/api/admin/cost_accounting/waybill_comparison', methods=['GET'])
def get_waybill_comparison():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    if not require_admin(cursor): conn.close(); return jsonify({'error': '无权访问', 'success': False}), 403
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    if not start_date or not end_date:
        conn.close()
        return jsonify({'success': False, 'error': '缺失日期参数'}), 400

    try:
        # 装载执行监控 KPI：与飞书运单数据源页同一表、同一日期区间的 tickets_count 合计
        total_feishu_tickets = _feishu_raw_tickets_total(cursor, start_date, end_date)

        # 1. 飞书：与 /api/admin/feishu/data 相同聚合，再规范化流向以便与 Gofo 对比
        f_query = """
            SELECT record_date, destination, SUM(tickets_count) as total_tickets
            FROM feishu_raw_data
            WHERE record_date >= ? AND record_date <= ?
            GROUP BY record_date, destination
        """
        cursor.execute(convert_query_placeholders(f_query), (start_date, end_date))
        f_rows = cursor.fetchall()
        
        feishu_data = {}  # { (date, route): tickets }
        for row in f_rows:
            r_date = str(row['record_date']).split(' ')[0][:10]
            dest = row['destination']
            t_cnt = row['total_tickets'] or 0
            norm_dest = _normalize_comparison_direction(dest)
            key = (r_date, norm_dest)
            feishu_data[key] = feishu_data.get(key, 0) + t_cnt
            
        # 2. CNO 集包量
        g_query = """
            SELECT record_date, destin_name, SUM(waybill_cnt) as total_waybills
            FROM gofo_center_collect_stats
            WHERE record_date >= ? AND record_date <= ?
            GROUP BY record_date, destin_name
        """
        cursor.execute(convert_query_placeholders(g_query), (start_date, end_date))
        g_rows = cursor.fetchall()

        gofo_data = {}  # { (date, route): waybills }，不含 CNO / CNO01
        for row in g_rows:
            dest = row['destin_name']
            if _is_waybill_comparison_excluded_gofo_destin(dest):
                continue
            r_date = str(row['record_date']).split(' ')[0][:10]
            w_cnt = row['total_waybills'] or 0
            norm_dest = _normalize_gofo_comparison_direction(dest)
            key = (r_date, norm_dest)
            gofo_data[key] = gofo_data.get(key, 0) + w_cnt

        total_gofo_tickets = int(sum(gofo_data.values()))
            
        # 3. Align and Compare
        all_keys = sorted(list(set(feishu_data.keys()) | set(gofo_data.keys())), key=lambda x: (x[0], x[1]), reverse=True)
        
        records = []

        for key in all_keys:
            f_val = feishu_data.get(key, 0)
            g_val = gofo_data.get(key, 0)
            diff = f_val - g_val
            abs_diff = abs(diff)

            # calculate diff pct
            if f_val > 0:
                diff_pct = round((diff / f_val) * 100, 2)
            elif g_val > 0:
                diff_pct = -100.0
            else:
                diff_pct = 0.0

            records.append({
                'record_date': key[0],
                'direction': key[1],
                'feishu_tickets': f_val,
                'gofo_tickets': g_val,
                'diff': diff,
                'abs_diff': abs_diff,
                'diff_pct': diff_pct
            })

        net_diff = total_feishu_tickets - total_gofo_tickets
        if total_gofo_tickets > 0:
            diff_pct_of_gofo = round((net_diff / total_gofo_tickets) * 100, 2)
        elif total_feishu_tickets > 0:
            diff_pct_of_gofo = None
        else:
            diff_pct_of_gofo = 0.0

        summary = {
            'total_feishu_tickets': total_feishu_tickets,
            'total_gofo_tickets': total_gofo_tickets,
            'net_diff': net_diff,
            'diff_pct_of_gofo': diff_pct_of_gofo,
            'feishu_data_source': 'feishu_raw_data',
        }
        
        return jsonify({
            'success': True,
            'summary': summary,
            'records': records
        })
        
    except Exception as e:
        import traceback
        print("Waybill comparison API error:", e)
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500
    finally:
        conn.close()


def get_route_type(route_code):
    """
    Classify route_code into route_type (trunk or branch).
    未匹配干线/支线关键词时记为 branch，不再使用 regular。
    """
    trunk_routes = ['MIA', 'MCO', 'DFW', 'ATL', 'ORD', 'IAH', 'EWR', 'CLT', 'JFK']
    branch_routes = ['LAX', 'DEN', 'SFO', 'LAS', 'PHX']

    
    clean_code = str(route_code).strip().upper()
    
    # Priority 1: Exact matches
    if clean_code in trunk_routes:
        return 'trunk'
    if clean_code in branch_routes:
        return 'branch'
        
    # Priority 2: Keyword matches (e.g., "MIA DROP", "DFW - IAH")
    for trunk in trunk_routes:
        if trunk in clean_code:
            return 'trunk'
    for branch in branch_routes:
        if branch in clean_code:
            return 'branch'
            
    return 'branch'


def _normalize_truck_destination_key(raw):
    """
    与写入 outbound_records 时的 route_code 规则一致，用于聚合键。
    若多个不同的 TO 原文规范后相同，应合并为一组，否则会出现 DELETE 删不到、INSERT 叠行（outbound 条数 > 源表行数）。
    """
    raw_s = str(raw or "")
    norm = raw_s.strip().upper()
    if not norm:
        return ""
    norm = norm.replace("\u2013", "-").replace("\u2014", "-").replace("\u2212", "-")
    # 旧表写 ATL.G 表示 ATL 枢纽；仅替换片段，不要把「ATL.G-ATLG」整条误收成「ATL」。
    if "ATL.G" in norm:
        norm = norm.replace("ATL.G", "ATL")
    if norm.endswith(".H"):
        norm = norm[:-2]
    # 拉斯维加斯支线：表内可能写 LAV、LAV（往返）、LAV DROP、带不可见字符/导出差异等；凡以 LAV 开头一律归为 LAV，
    # 避免出现 outbound 里多出「LAV（往返）」而你在表里只看到「LAV」、或条件未命中时整段原文入库。
    # LAX 不以 LAV 开头，不会误合并。
    if norm.startswith("LAV"):
        return "LAV"
    # 与前端 route-distribution 一致：ORD-ORD / ATL-ATL 等自环收成单点。
    m = re.match(r"^([A-Z0-9()]+)-\1$", norm)
    if m:
        norm = m.group(1)
    return norm.strip()


def _purge_auto_synced_outbound_for_dates(cursor, dates):
    """
    删除指定日期下由卡车表同步写入的 outbound（不动手工录入）。
    在重写 outbound 之前调用，杜绝历史叠行、部分失败残留。
    """
    if not dates:
        return
    dates = list(dict.fromkeys(dates))  # 保序去重
    placeholder = get_placeholder()
    in_list = ",".join([placeholder] * len(dates))
    q = (
        f"DELETE FROM outbound_records WHERE created_by = 'System' "
        f"AND notes = 'Auto-synced from Google Sheets' AND record_date IN ({in_list})"
    )
    cursor.execute(convert_query_placeholders(q), dates)


def _verify_truck_outbound_sync(cursor, merged):
    """
    校验：每个 record_date 下，自动同步 outbound 行数 = merged 中该车次的展开条数之和。
    merged: {(pickup_date_str, norm_route): {'count': n, 'cost': ...}, ...}
    """
    from collections import defaultdict

    exp_by_date = defaultdict(int)
    for (_pd, _nk), agg in merged.items():
        exp_by_date[str(_pd)[:10]] += int(agg["count"])

    q = """
        SELECT record_date, COUNT(*) AS c
        FROM outbound_records
        WHERE created_by = 'System' AND notes = 'Auto-synced from Google Sheets'
        GROUP BY record_date
    """
    cursor.execute(convert_query_placeholders(q))
    ob_by_date = {}
    for row in cursor.fetchall():
        rd = row[0]
        c = row[1]
        rd_s = rd.strftime("%Y-%m-%d") if hasattr(rd, "strftime") else str(rd)[:10]
        ob_by_date[rd_s] = int(c)

    for d, exp in sorted(exp_by_date.items()):
        got = ob_by_date.get(d, 0)
        if got != exp:
            return False, f"校验失败 {d}: 自动同步 outbound 行数={got}，期望={exp}（应与 truck_bookings 有效行一致）"

    return True, ""


def _sync_truck_bookings_core(sheet_url, clear_before_sync=True):
    """
    从 Google 表格 CSV 同步 truck_bookings，并聚合写入 outbound_records。
    clear_before_sync=True（默认）：先清空 outbound_records 与 truck_bookings，再全量导入（避免增量残留导致统计偏差）。
    """
    try:
        conn = get_db()
        cursor = conn.cursor()

        if clear_before_sync:
            cursor.execute("DELETE FROM outbound_records")
            cursor.execute("DELETE FROM truck_bookings")
            # 与下方 truck_bookings / outbound 写入同一事务提交，失败则一并回滚
            print("[TruckSync] Cleared outbound_records and truck_bookings (pending commit).")
        
        if 'docs.google.com/spreadsheets' in sheet_url:
            import re
            if '/edit' in sheet_url:
                base_url = re.sub(r'/edit.*', '', sheet_url)
                gid = '0'
                gid_match = re.search(r'gid=(\d+)', sheet_url)
                if gid_match: gid = gid_match.group(1)
                sheet_url = f"{base_url}/export?format=csv&gid={gid}"
                
        import requests, csv, io
        from datetime import datetime
        print(f"[TruckSync] Fetching from: {sheet_url}")
        response = requests.get(sheet_url, timeout=15)
        response.raise_for_status()
        
        csv_content = response.content.decode('utf-8-sig')
        f_csv = io.StringIO(csv_content)
        reader = csv.reader(f_csv)
        rows = list(reader)
        
        if not rows:
            return False, '表格内容为空', 0
            
        start_index = -1
        header_row = None
        for i, row in enumerate(rows):
            if row and 'DATE' in [c.upper() for c in row]:
                start_index = i + 1
                header_row = row
                break
                
        if start_index == -1:
            return False, '未找到 "DATE" 表头列', 0

        # Resolve column indexes by header names instead of fixed positions.
        # This prevents silent data corruption when sheet columns shift.
        def _norm_header(v):
            return str(v or "").strip().lower()

        header_map = {_norm_header(v): idx for idx, v in enumerate(header_row or [])}

        def _find_col(*aliases):
            for a in aliases:
                key = _norm_header(a)
                if key in header_map:
                    return header_map[key]
            return None

        date_col = _find_col("date")
        to_col = _find_col("to", "destination")
        pickup_col = _find_col("pickup #", "pickup#", "pickup no", "pickup no.")
        vendor_col = _find_col("vender", "vendor")
        cost_col = _find_col("$", "cost")
        driver_col = _find_col("driver info")
        pickup_time_col = _find_col("pickup time")

        if date_col is None or to_col is None:
            return False, '表头缺少 DATE 或 TO 列，无法同步', 0

        # K 列 MT# 规则（与 standalone/route-distribution/app.py 口径一致）：
        # 一行只有当对应"MT 列"以 'MT' 开头时才算实际发车。MT 列优先取
        # 「无表头但每行写 MT2026...」的那列（即 2026 年 4 月起的 K 列），
        # 找不到再退回到带 MT# 表头的列。这样旧表（MT 在 M）也能继续同步，新表（MT 在 K）也工作。
        body_rows_for_mt = rows[start_index:]
        exclude_for_mt = {c for c in (date_col, to_col, pickup_col, vendor_col, cost_col, driver_col, pickup_time_col) if c is not None}
        mt_col = _resolve_mt_column(header_row or [], body_rows_for_mt, exclude_for_mt)

        print(
            f"[TruckSync] Header mapping: DATE={date_col}, TO={to_col}, "
            f"PICKUP={pickup_col}, VENDOR={vendor_col}, COST={cost_col}, MT={mt_col}",
            flush=True,
        )
        if mt_col < 0:
            print(
                "[TruckSync][WARN] 未在表中定位到 MT 列，本次同步将放弃 K-MT 过滤、按原 TO-非空逻辑处理。",
                flush=True,
            )

        synced_count = 0
        seen_pickup_no = set()
        missing_pickup_no_count = 0
        temp_pickup_no_count = 0
        duplicate_nonempty_pickup_no_count = 0
        duplicate_pickup_no_renamed_count = 0
        skipped_no_mt_count = 0
        placeholder = get_placeholder()
        
        for row in rows[start_index:]:
            if len(row) < 3: continue
            raw_date = row[date_col].strip() if len(row) > date_col else ""
            if not raw_date or raw_date.upper() == 'DATE': continue

            # K-列 MT# 过滤：只有 MT 列以 "MT" 开头的行才算一次实际发车。
            # mt_col < 0 时（极旧表头解析失败）跳过过滤、退回原行为，已在上面 warn。
            if mt_col >= 0:
                mt_cell = row[mt_col].strip() if len(row) > mt_col else ""
                if not _MT_FULL_RE.match(mt_cell):
                    skipped_no_mt_count += 1
                    continue
            
            try:
                m_part, d_part = 0, 0
                if '-' in raw_date:
                    parts = raw_date.split('-')
                    if len(parts) >= 2: m_part, d_part = int(parts[0]), int(parts[1])
                elif '/' in raw_date:
                    parts = raw_date.split('/')
                    if len(parts) >= 2: m_part, d_part = int(parts[0]), int(parts[1])
                
                if 1 <= m_part <= 5: 
                    full_date = f"2026-{m_part:02d}-{d_part:02d}"
                elif 6 <= m_part <= 12: 
                    full_date = f"2025-{m_part:02d}-{d_part:02d}"
                else: 
                    continue # Skip invalid month parts (like '4350' or '50')
                
                # Double check validity by parsing
                datetime.strptime(full_date, '%Y-%m-%d')
            except Exception as e: 
                print(f"[TruckSync] Skipping invalid row due to date error: {raw_date} -> {e}")
                continue # Skip rows that don't have a valid date in col 0
                
            destination = row[to_col].strip() if len(row) > to_col else ""
            vendor = row[vendor_col].strip() if (vendor_col is not None and len(row) > vendor_col) else ""
            pickup_no = row[pickup_col].strip() if (pickup_col is not None and len(row) > pickup_col) else ""
            
            cost = 0.0
            def parse_float(val):
                if not val: return 0.0
                try: return float(str(val).replace('$', '').replace(',', '').strip())
                except: return 0.0
                
            cost = parse_float(row[cost_col]) if (cost_col is not None and len(row) > cost_col) else 0.0
            
            if not pickup_no and not destination: continue
            if not pickup_no:
                missing_pickup_no_count += 1
                import hashlib
                # Build a stronger stable fallback key from the full row payload.
                # This avoids accidental merges when multiple trucks share similar vendor/time.
                full_row_key = "|".join(str(cell).strip() for cell in row)
                raw_key = f"{full_date}|{full_row_key}"
                pickup_no = "TEMP_" + hashlib.md5(raw_key.encode("utf-8")).hexdigest()[:16]
                temp_pickup_no_count += 1
            elif pickup_no in seen_pickup_no:
                duplicate_nonempty_pickup_no_count += 1

            # Keep every row from sheet: disambiguate any duplicate pickup_no (non-empty or TEMP).
            if pickup_no in seen_pickup_no:
                base_no = pickup_no
                seq = 2
                while pickup_no in seen_pickup_no:
                    pickup_no = f"{base_no}__DUP{seq}"
                    seq += 1
                duplicate_pickup_no_renamed_count += 1

            seen_pickup_no.add(pickup_no)
            
            
            la_now = datetime.now(LA_TZ).strftime('%Y-%m-%d %H:%M:%S')
            # In full rebuild mode we must preserve every sheet row.
            # Do not upsert by pickup_no, otherwise duplicated pickup_no in source will be merged away.
            cursor.execute(
                f"""
                INSERT INTO truck_bookings (pickup_date, destination, vendor, pickup_no, cost, updated_at)
                VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
                """,
                (full_date, destination, vendor, pickup_no, cost, la_now),
            )
            synced_count += 1
            
        # Step 2: 按「规范化 route_code」合并 truck_bookings 再写入 outbound_records。
        # 旧逻辑：SQL GROUP BY 原始 destination，但 INSERT 用 norm_r_code；DELETE 用原始 r_code。
        # 当「LAV」与「LAV（往返）DROP」等规范后同为 LAV 时，会删不干净、叠插入，导致 outbound 行数 > 源表行数。
        from collections import defaultdict
        cursor.execute(
            convert_query_placeholders(
                "SELECT pickup_date, destination, cost FROM truck_bookings "
                "WHERE pickup_date IS NOT NULL AND destination IS NOT NULL AND destination != ''"
            )
        )
        tb_rows = cursor.fetchall()
        merged = defaultdict(lambda: {"count": 0, "cost": 0.0})
        for tb in tb_rows:
            pd, dest, cost = tb[0], tb[1], tb[2]
            if isinstance(pd, (datetime, date)):
                pd = pd.strftime("%Y-%m-%d")
            else:
                pd = str(pd)[:10]
            nk = _normalize_truck_destination_key(dest)
            if not nk:
                continue
            merged[(pd, nk)]["count"] += 1
            merged[(pd, nk)]["cost"] += float(cost or 0)

        # 先按「到货日」清空这些日期的自动同步 outbound，再统一插入（避免与历史行叠加；全量清空后本步多为空操作）
        date_set = set()
        for tb in tb_rows:
            pd0 = tb[0]
            if isinstance(pd0, (datetime, date)):
                date_set.add(pd0.strftime("%Y-%m-%d"))
            else:
                date_set.add(str(pd0)[:10])
        if date_set:
            _purge_auto_synced_outbound_for_dates(cursor, sorted(date_set))

        current_time = datetime.now(LA_TZ).strftime("%Y-%m-%d %H:%M:%S")
        insert_query = """
                INSERT INTO outbound_records (record_date, route_code, route_type, vehicle_count, cost, notes, created_by, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
        for (r_date, norm_r_code), agg in sorted(merged.items(), key=lambda it: (it[0][0], it[0][1])):
            v_int = int(agg["count"])
            if v_int <= 0:
                continue
            t_cost = agg["cost"]
            r_type = get_route_type(norm_r_code)
            record_entry_time = f"{r_date} 00:00:00"
            total_f = float(t_cost or 0)
            unit_cost = round(total_f / v_int, 2)
            accrued = 0.0
            for i in range(v_int):
                if i == v_int - 1:
                    row_cost = round(total_f - accrued, 2)
                else:
                    row_cost = unit_cost
                    accrued += row_cost
                cursor.execute(
                    convert_query_placeholders(insert_query),
                    (r_date, norm_r_code, r_type, 1, row_cost, "Auto-synced from Google Sheets", "System", record_entry_time, current_time),
                )

        ok_verify, verify_msg = _verify_truck_outbound_sync(cursor, merged)
        if not ok_verify:
            conn.rollback()
            print(f"[TruckSync][VERIFY FAIL] {verify_msg}", flush=True)
            return False, verify_msg, 0

        conn.commit()
        print(
            f"[TruckSync] Synced rows: {synced_count}; unique pickup_no: {len(seen_pickup_no)}",
            flush=True,
        )
        print(
            f"[TruckSync] pickup_no diagnostics: missing={missing_pickup_no_count}, temp_generated={temp_pickup_no_count}",
            flush=True,
        )
        print(
            f"[TruckSync] pickup_no diagnostics: duplicate_nonempty={duplicate_nonempty_pickup_no_count}",
            flush=True,
        )
        print(
            f"[TruckSync] pickup_no diagnostics: renamed_duplicates={duplicate_pickup_no_renamed_count}",
            flush=True,
        )
        print(
            f"[TruckSync] K-MT filter: skipped_rows_without_mt={skipped_no_mt_count}",
            flush=True,
        )
        return True, "Success", synced_count
        
    except Exception as e:
        import traceback
        print(f"[TruckSync][ERROR] {e}", flush=True)
        traceback.print_exc()
        if 'conn' in locals() and conn: conn.rollback()
        return False, str(e), 0
    finally:
        if 'conn' in locals() and conn: conn.close()


# ---------------------------------------------------------------------------
# Helpers shared by the K-column MT filter that _sync_truck_bookings_core uses
# below. Kept module-level so they're easy to unit-test / reuse.
# ---------------------------------------------------------------------------
_MT_FULL_RE = re.compile(r"^MT\d{6,}", re.IGNORECASE)


def _resolve_mt_column(header_row, body_rows, exclude_idxs):
    """
    Find the column that holds the per-row MT# value.
    Strategy (matches standalone/route-distribution/app.py):
      1. Among columns whose HEADER cell is blank, pick the one with the most
         "MT######" prefix hits — that's column K in the live 2026 sheet.
      2. If no blank-header column has any hits, fall back to a column whose
         header matches MT# / MT / 发车运单号 — preserves legacy sheets.
      3. Otherwise return -1 (caller errors out).
    """
    if not body_rows:
        return -1

    def _hits(idx):
        n = 0
        for r in body_rows:
            if idx < len(r) and _MT_FULL_RE.match(str(r[idx]).strip()):
                n += 1
        return n

    max_col = max(len(r) for r in body_rows)
    blank_best, blank_hits = -1, 0
    for j in range(max_col):
        if j in exclude_idxs:
            continue
        hcell = str(header_row[j]).strip() if j < len(header_row) else ""
        if hcell:
            continue
        h = _hits(j)
        if h > blank_hits:
            blank_best, blank_hits = j, h

    if blank_best >= 0 and blank_hits > 0:
        return blank_best

    mt_aliases = ("mt#", "mt #", "mt", "发车运单号", "运单号", "运单")
    for j, cell in enumerate(header_row):
        if j in exclude_idxs:
            continue
        s = str(cell or "").strip().lower()
        for a in mt_aliases:
            if s == a or s.startswith(a) or a in s:
                return j
    return -1


@app.route('/api/admin/truck_bookings/sync', methods=['POST'])
def sync_truck_bookings():
    if 'user_id' not in session: return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db(); cursor = conn.cursor()
    cursor.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
    result = cursor.fetchone()
    current_role = result['role'] if USE_POSTGRES else result[0]
    conn.close()
    
    if current_role not in ('admin', 'boss'):
        return jsonify({'error': '无权访问', 'success': False}), 403
    
    data = request.json or {}
    sheet_url = data.get('url')
    if not sheet_url:
        return jsonify({'success': False, 'error': 'URL不能为空'}), 400

    # Force full rebuild to avoid stale incremental leftovers from edited/malformed sheet rows.
    # Keep request flag only for logging visibility.
    requested_clear = bool(data.get('clear_before_sync', False))
    clear_before_sync = True
    print(
        f"[TruckSync] Sync mode: full_rebuild=true (client requested clear_before_sync={requested_clear})",
        flush=True,
    )
    success, msg, count = _sync_truck_bookings_core(sheet_url, clear_before_sync=clear_before_sync)
    if not success:
        return jsonify({'success': False, 'error': msg}), 500
    return jsonify({'success': True, 'count': count, 'clear_before_sync': clear_before_sync})


@app.route('/api/admin/outbound/clear', methods=['POST'])
def admin_clear_outbound_and_truck():
    """仅清空 outbound_records 与 truck_bookings（需 admin/boss）。用于手工或其它方式重新导入前。"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录', 'success': False}), 401
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM users WHERE id = ?", (session['user_id'],))
    result = cursor.fetchone()
    conn.close()
    current_role = result['role'] if USE_POSTGRES else result[0]
    if current_role not in ('admin', 'boss'):
        return jsonify({'error': '无权访问', 'success': False}), 403
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM outbound_records")
        outbound_deleted = cursor.rowcount if hasattr(cursor, 'rowcount') else -1
        cursor.execute("DELETE FROM truck_bookings")
        truck_deleted = cursor.rowcount if hasattr(cursor, 'rowcount') else -1
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '已清空 outbound_records 与 truck_bookings'})
    except Exception as e:
        if 'conn' in locals() and conn:
            conn.rollback()
            conn.close()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/consumables/save', methods=['POST'])
def save_consumable():
    """新增或修改耗材"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    # 鉴传是否有页面操作权限（这里默认同 consumbales）
    if not check_page_permission('consumables'):
        return jsonify({'error': '无权限操作耗材数据'}), 403
    
    data = request.json
    c_id = data.get('id')
    name = data.get('name')
    unit = data.get('unit')
    safety_stock = float(data.get('safety_stock', 0))
    lead_time_days = int(data.get('lead_time_days', 0))
    current_stock = float(data.get('current_stock', 0)) # Create only
    
    if not name or not unit:
        return jsonify({'success': False, 'error': '缺失必填字段：名称或单位'}), 400
        
    try:
        conn = get_db()
        cursor = conn.cursor()
        placeholder = get_placeholder()
        la_now = datetime.now(LA_TZ)
        
        if c_id:
            # 修改
            update_query = f"""
                UPDATE consumables 
                SET name = {placeholder}, unit = {placeholder}, safety_stock = {placeholder}, 
                    lead_time_days = {placeholder}, updated_at = {placeholder}
                WHERE id = {placeholder}
            """
            cursor.execute(update_query, (name, unit, safety_stock, lead_time_days, la_now, c_id))
        else:
            # 新增
            insert_query = f"""
                INSERT INTO consumables (name, unit, safety_stock, current_stock, lead_time_days, created_at, updated_at)
                VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
            """
            # SQLite does not error uniquely out of the box unless unique constraint is hit
            cursor.execute(insert_query, (name, unit, safety_stock, current_stock, lead_time_days, la_now, la_now))
            
        conn.commit()
        conn.close()
        return jsonify({'success': True})
    except Exception as e:
        if 'UNIQUE' in str(e) or 'unique constraint' in str(e).lower():
            return jsonify({'success': False, 'error': '耗材名称已存在'}), 400
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/consumables/transaction', methods=['POST'])
def consumable_transaction():
    """记录耗材流水 (领用 OUT / 采购 IN)"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
        
    if not check_page_permission('consumables'):
        return jsonify({'error': '无操作权限'}), 403
        
    data = request.json
    consumable_id = data.get('consumable_id')
    consumable_name = data.get('consumable_name')
    trans_type = data.get('type') # 'IN' or 'OUT'
    quantity = float(data.get('quantity', 0))
    related_warehouse_volume = data.get('related_warehouse_volume') # 关联发车件数等，入库则可能为空
    entry_date_str = data.get('entry_date')
    
    if not (consumable_id or consumable_name) or not trans_type or quantity <= 0 or not entry_date_str:
        return jsonify({'success': False, 'error': '缺失必要参数或数量无效'}), 400
        
    if trans_type not in ('IN', 'OUT', 'RENT', 'REPAIR', 'ENERGY'):
        return jsonify({'success': False, 'error': '无效的流水类型'}), 400
        

         
    username = session.get('username', 'unknown')
    la_now = datetime.now(LA_TZ)
    la_now_str = la_now.strftime('%Y-%m-%d %H:%M:%S') 
    
    try:
        conn = get_db()
        cursor = conn.cursor()
        placeholder = get_placeholder()
        
        # 解析 consumable_name (如果是从基础配置列表直接选取的)
        if consumable_name and not consumable_id:
            cursor.execute(f"SELECT id FROM consumables WHERE name = {placeholder}", (consumable_name,))
            exist_row = cursor.fetchone()
            if exist_row:
                consumable_id = exist_row['id'] if hasattr(exist_row, 'keys') else exist_row[0]
            elif trans_type in ('IN', 'OUT'):
                # 只有库存类交易才自动创建consumables记录
                cursor.execute(f"SELECT unit FROM production_consumable_master WHERE name = {placeholder}", (consumable_name,))
                master_row = cursor.fetchone()
                if not master_row:
                    conn.close()
                    return jsonify({'success': False, 'error': '品名不存在于基础配置中'}), 404
                unit = master_row['unit'] if hasattr(master_row, 'keys') else master_row[0]
                if not unit: unit = ''
                
                cursor.execute(f"""
                    INSERT INTO consumables (name, unit, safety_stock, current_stock, lead_time_days, created_at, updated_at) 
                    VALUES ({placeholder}, {placeholder}, 0, 0, 0, {placeholder}, {placeholder})
                """, (consumable_name, unit, la_now_str, la_now_str))
                consumable_id = cursor.lastrowid
        
        # For expense-only types without a consumable record, skip stock logic
        if trans_type in ('RENT', 'REPAIR', 'ENERGY') and not consumable_id:
            # Directly insert transaction without stock management
            related_vol = None
            record_qty = quantity
            
            la_now = datetime.now(LA_TZ)
            try:
                entry_date = datetime.strptime(entry_date_str, '%Y-%m-%d').date()
                target_datetime = datetime.combine(entry_date, la_now.time())
                target_datetime_tz = LA_TZ.localize(target_datetime)
                db_created_at = target_datetime_tz.astimezone(LA_TZ) if hasattr(target_datetime_tz, 'astimezone') else target_datetime
            except ValueError:
                conn.close()
                return jsonify({'success': False, 'error': '无效的日期格式，需为 YYYY-MM-DD'}), 400
            
            # Find or create a minimal consumable record for FK reference
            cursor.execute(f"SELECT unit FROM production_consumable_master WHERE name = {placeholder}", (consumable_name,))
            master_row = cursor.fetchone()
            if not master_row:
                conn.close()
                return jsonify({'success': False, 'error': '品名不存在于基础配置中'}), 404
            unit = master_row['unit'] if hasattr(master_row, 'keys') else master_row[0]
            if not unit: unit = ''
            
            # Create a hidden consumable record (will NOT show in inventory since it's expense-only)
            cursor.execute(f"""
                INSERT INTO consumables (name, unit, safety_stock, current_stock, lead_time_days, created_at, updated_at) 
                VALUES ({placeholder}, {placeholder}, -1, 0, 0, {placeholder}, {placeholder})
            """, (consumable_name, unit, la_now.strftime('%Y-%m-%d %H:%M:%S'), la_now.strftime('%Y-%m-%d %H:%M:%S')))
            consumable_id = cursor.lastrowid
            
            cursor.execute(f"""
                INSERT INTO inventory_transactions (consumable_id, type, quantity, related_warehouse_volume, operator, created_at)
                VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
            """, (consumable_id, trans_type, record_qty, related_vol, username, db_created_at))
            
            conn.commit()
            conn.close()
            return jsonify({'success': True, 'new_stock': 0})

        # 1. 查询当前库存和安全线
        cursor.execute(f"SELECT current_stock, safety_stock, name, unit FROM consumables WHERE id = {placeholder}", (consumable_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({'success': False, 'error': '未找到耗材'}), 404
            
        current_stock = row['current_stock'] if hasattr(row, 'keys') else row[0]
        safety_stock = row['safety_stock'] if hasattr(row, 'keys') else row[1]
        name = row['name'] if hasattr(row, 'keys') else row[2]
        unit = row['unit'] if hasattr(row, 'keys') else row[3]
        
        # 2. 计算新库存与校验
        next_stock = current_stock
        if trans_type == 'OUT':
            consumed_qty = current_stock - quantity
            next_stock = quantity
            related_vol = int(related_warehouse_volume) if related_warehouse_volume else None
            record_qty = consumed_qty
        elif trans_type == 'IN':
            next_stock += quantity
            related_vol = None
            record_qty = quantity
        else:  # RENT/REPAIR/ENERGY with existing consumable record
            related_vol = None
            record_qty = quantity
            
        # 3. 扣减 / 增加库存
        la_now = datetime.now(LA_TZ)
        cursor.execute(f"UPDATE consumables SET current_stock = {placeholder}, updated_at = {placeholder} WHERE id = {placeholder}", 
                      (next_stock, la_now, consumable_id))
                      
        # 解析 entry_date_str 到洛杉矶时间的 datetime
        try:
            entry_date = datetime.strptime(entry_date_str, '%Y-%m-%d').date()
            target_datetime = datetime.combine(entry_date, la_now.time())
            target_datetime_tz = LA_TZ.localize(target_datetime)
            db_created_at = target_datetime_tz.astimezone(LA_TZ) if hasattr(target_datetime_tz, 'astimezone') else target_datetime
        except ValueError:
            conn.close()
            return jsonify({'success': False, 'error': '无效的日期格式，需为 YYYY-MM-DD'}), 400

        # 4. 插入流水 (使用选定的日期)
        cursor.execute(f"""
            INSERT INTO inventory_transactions (consumable_id, type, quantity, related_warehouse_volume, operator, created_at)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
        """, (consumable_id, trans_type, record_qty, related_vol, username, db_created_at))
        
        conn.commit()
        conn.close()
        
        response_data = {'success': True, 'new_stock': next_stock}
        # 预警逻辑
        if next_stock <= safety_stock:
            response_data['warning'] = f"⚠️ [库存预警] {name} 库存低于或等于安全线！当前结余: {next_stock} {unit}"
            
        return jsonify(response_data)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/consumables/transactions/list', methods=['GET'])
def get_consumables_transactions():
    """获取耗材流水记录 (支持时间范围查询)"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    if not start_date_str or not end_date_str:
        return jsonify({'success': False, 'error': '缺失起止日期参数'}), 400
        
    try:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # 将 start 和 end 转换为时间戳包含当天的所有记录
        start_dt = datetime.combine(start_date, datetime.min.time())
        end_dt = datetime.combine(end_date, datetime.max.time())
        
        # Convert to strings for database querying
        start_str = start_dt.strftime('%Y-%m-%d %H:%M:%S')
        end_str = end_dt.strftime('%Y-%m-%d %H:%M:%S')
        
    except ValueError:
        return jsonify({'success': False, 'error': '无效的日期格式，需为 YYYY-MM-DD'}), 400

    try:
        conn = get_db()
        cursor = conn.cursor()
        placeholder = get_placeholder()
        
        # 由于 created_at 在 SQLite 默认是带有时区的 UTC 字符串，或者通过代码手动存的。我们查询时只需做字符串层面的比较
        query = f"""
            SELECT t.id, t.type, t.quantity, t.related_warehouse_volume, t.operator, t.created_at,
                   c.name as consumable_name, COALESCE(m.unit, c.unit) as unit, m.price as unit_price
            FROM inventory_transactions t
            JOIN consumables c ON t.consumable_id = c.id
            LEFT JOIN production_consumable_master m ON c.name = m.name
            WHERE t.created_at >= {placeholder} AND t.created_at <= {placeholder}
            ORDER BY t.created_at DESC
        """
        cursor.execute(query, (start_str, end_str))
        rows = cursor.fetchall()
        
        records = []
        for row in rows:
            if hasattr(row, 'keys'):
                # PostgreSQL dict-like row
                c_data = dict(row)
            else:
                # SQLite tuple
                c_data = {
                    'id': row[0],
                    'type': row[1],
                    'quantity': row[2],
                    'related_warehouse_volume': row[3],
                    'operator': row[4],
                    'created_at': row[5],
                    'consumable_name': row[6],
                    'unit': row[7],
                    'unit_price': row[8]
                }
                
            # Formatting time for frontend display
            time_val = c_data['created_at']
            if isinstance(time_val, str):
                clean_time = time_val.replace('Z', '').replace('T', ' ')
                c_data['created_at'] = clean_time[:16] # Return up to minutes YYYY-MM-DD HH:MM
            elif isinstance(time_val, datetime):
                c_data['created_at'] = time_val.strftime('%Y-%m-%d %H:%M')
                
            records.append(c_data)
            
        conn.close()
        return jsonify({'success': True, 'data': records})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/consumables/transactions/update', methods=['POST'])
def update_consumable_transaction():
    """编辑耗材流水记录"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
    
    data = request.json
    trans_id = data.get('id')
    new_qty = data.get('quantity')
    new_related_vol = data.get('related_warehouse_volume')
    new_date_str = data.get('entry_date')  # We allow changing the date, it will update created_at
    operator = session.get('username', 'system')
    
    if not all([trans_id, new_qty, new_date_str]):
        return jsonify({'success': False, 'error': '参数不完整'}), 400
        
    try:
        new_qty = float(new_qty)
        entry_date = datetime.strptime(new_date_str, '%Y-%m-%d').date()
        # 简单取当天的当前时间，或直接午夜
        target_dt = datetime.combine(entry_date, datetime.now(LA_TZ).time())
        db_created_at = LA_TZ.localize(target_dt).astimezone(LA_TZ) if hasattr(LA_TZ.localize(target_dt), 'astimezone') else target_dt
    except ValueError:
        return jsonify({'success': False, 'error': '数据格式错误'}), 400

    new_related_vol_val = None
    if new_related_vol and str(new_related_vol).strip() != '':
        try:
            new_related_vol_val = int(new_related_vol)
        except ValueError:
            pass

    try:
        conn = get_db()
        cursor = conn.cursor()
        placeholder = get_placeholder()
        
        # 1. 查找原记录
        cursor.execute(f"SELECT * FROM inventory_transactions WHERE id = {placeholder}", (trans_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({'success': False, 'error': '记录不存在'}), 404
            
        old_record = dict(row) if hasattr(row, 'keys') else {
            'id': row[0], 'consumable_id': row[1], 'type': row[2], 
            'quantity': row[3], 'related_warehouse_volume': row[4], 
            'operator': row[5], 'created_at': str(row[6])
        }
        
        consumable_id = old_record['consumable_id']
        old_qty = old_record['quantity']
        trans_type = old_record['type']
        
        # 2. 查找耗材当前库存
        cursor.execute(f"SELECT current_stock, name FROM consumables WHERE id = {placeholder}", (consumable_id,))
        c_row = cursor.fetchone()
        if not c_row:
            conn.close()
            return jsonify({'success': False, 'error': '关联的耗材不存在'}), 404
            
        current_stock = float(c_row['current_stock'] if hasattr(c_row, 'keys') else c_row[0])
        name = c_row['name'] if hasattr(c_row, 'keys') else c_row[1]
        
        # 3. 计算库存差异
        # if OUT, old_qty was subtracted. To change it, we add back old_qty, then subtract new_qty
        # if IN, old_qty was added. To change it, we subtract old_qty, then add new_qty
        if trans_type == 'OUT':
            new_stock = current_stock + old_qty - new_qty
        elif trans_type == 'IN':
            new_stock = current_stock - old_qty + new_qty
        else:  # RENT or REPAIR - no stock impact
            new_stock = current_stock
            
        if new_stock < 0:
            conn.close()
            return jsonify({'success': False, 'error': f'库存不足! {name} 修改后库存将会变成负数 ({new_stock})'}), 400
            
        # 4. 更新耗材库存
        la_now = datetime.now(LA_TZ)
        cursor.execute(f"UPDATE consumables SET current_stock = {placeholder}, updated_at = {placeholder} WHERE id = {placeholder}", 
                      (new_stock, la_now, consumable_id))
                      
        # 5. 更新流水记录
        cursor.execute(f"""
            UPDATE inventory_transactions 
            SET quantity = {placeholder}, related_warehouse_volume = {placeholder}, created_at = {placeholder}
            WHERE id = {placeholder}
        """, (new_qty, new_related_vol_val, db_created_at, trans_id))
        
        # 6. 记录操作日志
        import json
        new_record = old_record.copy()
        new_record['quantity'] = new_qty
        new_record['related_warehouse_volume'] = new_related_vol_val
        new_record['created_at'] = str(db_created_at)
        
        cursor.execute(f"""
            INSERT INTO operation_logs (operation_type, table_name, record_id, old_data, new_data, operator, created_at)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
        """, ('UPDATE', 'inventory_transactions', trans_id, json.dumps(old_record, ensure_ascii=False), json.dumps(new_record, ensure_ascii=False), operator, la_now))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '修改成功'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/consumables/transactions/delete', methods=['POST'])
def delete_consumable_transaction():
    """删除耗材流水记录(并回退库存)"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
        
    data = request.json
    trans_id = data.get('id')
    operator = session.get('username', 'system')
    
    if not trans_id:
        return jsonify({'success': False, 'error': '缺失记录ID'}), 400
        
    try:
        conn = get_db()
        cursor = conn.cursor()
        placeholder = get_placeholder()
        
        # 1. 查找原记录
        cursor.execute(f"SELECT * FROM inventory_transactions WHERE id = {placeholder}", (trans_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return jsonify({'success': False, 'error': '记录不存在'}), 404
            
        old_record = dict(row) if hasattr(row, 'keys') else {
            'id': row[0], 'consumable_id': row[1], 'type': row[2], 
            'quantity': row[3], 'related_warehouse_volume': row[4], 
            'operator': row[5], 'created_at': str(row[6])
        }
        
        consumable_id = old_record['consumable_id']
        old_qty = old_record['quantity']
        trans_type = old_record['type']
        
        # 2. 查找耗材当前库存
        cursor.execute(f"SELECT current_stock, name FROM consumables WHERE id = {placeholder}", (consumable_id,))
        c_row = cursor.fetchone()
        if not c_row:
            conn.close()
            return jsonify({'success': False, 'error': '关联的耗材不存在'}), 404
            
        current_stock = float(c_row['current_stock'] if hasattr(c_row, 'keys') else c_row[0])
        name = c_row['name'] if hasattr(c_row, 'keys') else c_row[1]
        
        # 3. 回退库存
        # if OUT, it was subtracted, so add it back
        # if IN, it was added, so subtract it
        # RENT/REPAIR don't affect stock, no rollback needed
        if trans_type == 'OUT':
            new_stock = current_stock + old_qty
        elif trans_type == 'IN':
            new_stock = current_stock - old_qty
        else:  # RENT or REPAIR
            new_stock = current_stock
            
        if new_stock < 0:
            conn.close()
            return jsonify({'success': False, 'error': f'库存不足! {name} 回退后库存将会变成负数 ({new_stock})'}), 400
            
        # 4. 更新耗材库存
        la_now = datetime.now(LA_TZ)
        cursor.execute(f"UPDATE consumables SET current_stock = {placeholder}, updated_at = {placeholder} WHERE id = {placeholder}", 
                      (new_stock, la_now, consumable_id))
                      
        # 5. 删除流水记录
        cursor.execute(f"DELETE FROM inventory_transactions WHERE id = {placeholder}", (trans_id,))
        
        # 6. 记录操作日志
        import json
        cursor.execute(f"""
            INSERT INTO operation_logs (operation_type, table_name, record_id, old_data, new_data, operator, created_at)
            VALUES ({placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder}, {placeholder})
        """, ('DELETE', 'inventory_transactions', trans_id, json.dumps(old_record, ensure_ascii=False), None, operator, la_now))
        
        conn.commit()
        conn.close()
        
        return jsonify({'success': True, 'message': '删除并回退库存成功'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/consumables/analytics', methods=['GET'])
def consumable_analytics():
    """采购决策支持：根据预测货量计算推荐采购量"""
    if 'user_id' not in session:
        return jsonify({'error': '未登录'}), 401
        
    consumable_id = request.args.get('consumable_id')
    target_volume = request.args.get('target_volume', type=int)
    
    if not consumable_id or not target_volume:
         return jsonify({'success': False, 'error': '需提供 consumable_id 及 target_volume 目标货量'}), 400
         
    try:
        conn = get_db()
        cursor = conn.cursor()
        placeholder = get_placeholder()
        
        # 1. 获取耗材元属性
        cursor.execute(f"SELECT current_stock, safety_stock FROM consumables WHERE id = {placeholder}", (consumable_id,))
        item = cursor.fetchone()
        if not item:
            conn.close()
            return jsonify({'success': False, 'error': '未找到耗材'}), 404
            
        current_stock = item['current_stock'] if hasattr(item, 'keys') else item[0]
        safety_stock = item['safety_stock'] if hasattr(item, 'keys') else item[1]
        
        # 2. 获取最近 30 天 (通过原生 SQL 统计 SUM(qty) 和 SUM(vol))
        # 我们用数据库的当前时间往前推 30 天
        # 为了兼容 SQLite 和 Postgres，我们用 Python 计算日期字符串然后传参
        past_30_days = (datetime.now(LA_TZ) - timedelta(days=30)).strftime('%Y-%m-%d %H:%M:%S')
        
        query = f"""
            SELECT SUM(quantity) as total_qty, SUM(related_warehouse_volume) as total_vol
            FROM inventory_transactions
            WHERE consumable_id = {placeholder}
            AND type = 'OUT'
            AND created_at >= {placeholder}
        """
        cursor.execute(query, (consumable_id, past_30_days))
        stats = cursor.fetchone()
        
        conn.close()
        
        total_qty = (stats['total_qty'] if hasattr(stats, 'keys') else stats[0]) or 0
        total_vol = (stats['total_vol'] if hasattr(stats, 'keys') else stats[1]) or 0
        
        if total_vol <= 0:
            return jsonify({
                'success': True, 
                'data': {
                    'note': '过去 30 天暂无消耗记录或对应货量记录，无法精准预测',
                    'coefficient': 0,
                    'predicted_demand': 0,
                    'suggested_purchase': max(0, safety_stock - current_stock) 
                }
            })
            
        # 3. 计算消耗系数
        consumption_coefficient = float(total_qty) / float(total_vol)
        
        # 4. 预测需求与采购缺口
        predicted_demand = target_volume * consumption_coefficient
        shortfall = (predicted_demand + safety_stock) - current_stock
        
        return jsonify({
            'success': True,
            'data': {
                'coefficient': consumption_coefficient,
                'predicted_demand': round(predicted_demand, 2),
                'suggested_purchase': max(0, round(shortfall, 2))
            }
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

def auto_sync_labor_data_job():
    """Background loop that waits until exactly 12:00:00 every day and runs data syncs."""
    print("[AutoSync] Labor data auto-sync job started. Waiting for 12:00:00 PM (LA Time).")
    tz = LA_TZ
    
    while True:
        now = datetime.now(tz)
        # Calculate next 12:00:00 PM
        target = now.replace(hour=12, minute=0, second=0, microsecond=0)
        
        if now >= target:
            # If past 12pm, target 12pm tomorrow
            target += timedelta(days=1)
            
        wait_seconds = (target - now).total_seconds()
        print(f"[AutoSync] Scheduled next daily labor sync for {target.strftime('%Y-%m-%d %H:%M:%S')} (in {wait_seconds:.1f} seconds)")
        
        time.sleep(wait_seconds)
        print("[AutoSync] Waking up. Triggering scheduled 12:00:00 PM daily labor sync...")
        
        try:
            # Run Outsource Sync
            import calc_outsource_finance
            import importlib
            importlib.reload(calc_outsource_finance)
            res_outsource = calc_outsource_finance.run_sync('')
            update_sync_status('Outsource', res_outsource.get('success', False))
            if res_outsource.get('success'):
                print("[AutoSync] Outsource finance synchronized successfully.")
            else:
                print(f"[AutoSync] Outsource sync failed: {res_outsource.get('error')}")
                
        except Exception as e:
            print(f"[AutoSync] Exception during scheduled Outsource sync: {e}")
            update_sync_status('Outsource', False)
            
        try:
            # Run Gofo Sync
            import calc_gofo_piece_rate
            import importlib
            importlib.reload(calc_gofo_piece_rate)
            res_gofo = calc_gofo_piece_rate.fetch_and_summarize_gofo_piece_rate(None)
            update_sync_status('Gofo', res_gofo.get('success', False))
            if res_gofo.get('success'):
                print("[AutoSync] Gofo piece-rate synchronized successfully.")
            else:
                print(f"[AutoSync] Gofo sync failed: {res_gofo.get('error')}")
                
        except Exception as e:
            print(f"[AutoSync] Exception during scheduled Gofo sync: {e}")
            update_sync_status('Gofo', False)


# 每日正午自动同步飞书运单数据
def daily_feishu_sync_job():
    import time as _time
    while True:
        try:
            now_la = datetime.now(LA_TZ)
            next_noon = now_la.replace(hour=12, minute=0, second=0, microsecond=0)
            if now_la >= next_noon:
                next_noon = next_noon + timedelta(days=1)
            wait_sec = (next_noon - now_la).total_seconds()
            print(f"[AutoSync] Scheduled next daily Feishu sync for {next_noon.strftime('%Y-%m-%d %H:%M:%S')} (in {wait_sec:.1f} seconds)")
            _time.sleep(wait_sec)
            # 直接调用核心同步函数，无需 HTTP 认证
            try:
                ok, msg = _feishu_sync_core(link='')
                print(f"[AutoSync] {datetime.now(LA_TZ).strftime('%Y-%m-%d %H:%M')} - {'OK' if ok else 'FAIL'}: {msg}")
            except Exception as e:
                print(f"[AutoSync] ERROR: {e}")
        except Exception:
            _time.sleep(3600)

# 飞书「分拣」表：每个整点有量时追加 4 行（A/B/C/D 各一行，记录该时段件数）
_CNO_NARROWBELT_FEISHU_HEADER_FALLBACK = [
    "No",
    "Date",
    "time_slot",
    "line_code",
    "line_label",
    "pieces",
]

# 飞书「元数据」表模板表头（仅当线上第 1 行损坏时回退；正常以表格第 1 行为准）
_CNO_LABOR_FEISHU_TEMPLATE_HEADER_FALLBACK = [
    "No",
    "Date",
    "company_code",
    "group_no",
    "pay_type",
    "day_total",
] + [f"{((17 + i) % 24):02d}:00" for i in range(24)] + [0, 0]


def _anchor_date_to_excel_serial(anchor) -> int:
    from datetime import date as date_cls

    if isinstance(anchor, datetime):
        d = anchor.date()
    elif isinstance(anchor, str):
        d = datetime.strptime(str(anchor)[:10], "%Y-%m-%d").date()
    else:
        d = anchor
    return (d - date_cls(1899, 12, 30)).days


def _normalize_feishu_header_row(header_row: list, fallback: list) -> list:
    if not header_row:
        return list(fallback)
    row = list(header_row)
    while row and (row[-1] is None or str(row[-1]).strip() == ""):
        row.pop()
    if not row or str(row[0]).strip() != "No":
        return list(fallback)
    fb = fallback
    if len(row) < len(fb):
        row.extend([fb[i] for i in range(len(row), len(fb))])
    return row


def _normalize_feishu_template_header(header_row: list) -> list:
    return _normalize_feishu_header_row(
        header_row, _CNO_LABOR_FEISHU_TEMPLATE_HEADER_FALLBACK
    )


def _normalize_feishu_narrowbelt_header(header_row: list) -> list:
    row = _normalize_feishu_header_row(
        header_row, _CNO_NARROWBELT_FEISHU_HEADER_FALLBACK
    )
    names = {str(h).strip().lower() for h in row if h is not None}
    if "time_slot" in names and "line_code" in names:
        return row
    return list(_CNO_NARROWBELT_FEISHU_HEADER_FALLBACK)


def _feishu_header_col_index(header: list, name: str) -> int:
    for i, h in enumerate(header):
        if str(h).strip() == name:
            return i
    return -1


def _parse_feishu_sheet_int(v):
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(str(v).strip()))
    except (TypeError, ValueError):
        return None


def _build_cno_labor_feishu_row(
    no: int,
    excel_date: int,
    r: dict,
    header: list,
    label_to_idx: dict,
) -> list:
    """单行：列顺序与表头一致。"""
    ncols = len(header)
    row = [""] * ncols
    hourly = r.get("hourly") or []
    for col_idx, h in enumerate(header):
        hstr = str(h).strip() if h is not None else ""
        if hstr == "No":
            row[col_idx] = no
        elif hstr == "Date":
            row[col_idx] = excel_date
        elif hstr == "company_code":
            row[col_idx] = r.get("company") or ""
        elif hstr == "group_no":
            row[col_idx] = r.get("group_no") or ""
        elif hstr == "pay_type":
            row[col_idx] = r.get("pay_type") or ""
        elif hstr == "day_total":
            row[col_idx] = int(r.get("total") or 0)
        elif hstr in label_to_idx:
            j = label_to_idx[hstr]
            row[col_idx] = int(hourly[j]) if j < len(hourly) else 0
        elif hstr in ("0", ""):
            row[col_idx] = 0
        else:
            row[col_idx] = ""
    return row


def _cno_labor_feishu_should_append_new_row(
    r: dict, label_to_idx: dict
) -> bool:
    """与统计页一致：day_total > 0 即写入（避免 17:00 为 0 的小组被漏掉）。"""
    if int(r.get("total") or 0) <= 0:
        return False
    if os.getenv("FEISHU_CNO_LABOR_APPEND_REQUIRE_17", "").strip() in (
        "1",
        "true",
        "yes",
    ):
        j = label_to_idx.get("17:00")
        if j is None:
            return int((r.get("hourly") or [0])[0] or 0) > 0
        hourly = r.get("hourly") or []
        return int(hourly[j]) > 0 if j < len(hourly) else False
    return True


def _merge_cno_labor_feishu_sheet_body(
    template_header: list,
    matrix: dict,
    existing_rows: list,
    *,
    replace_operating_day: bool = False,
) -> tuple:
    """
    合并表内历史行与当前运营日矩阵；replace_operating_day 时整批重写该运营日（与统计页行数一致）。
    追加时 No = 全表最大 No + 1，Date = 当日运营锚点 Excel 序列。
    返回 (combined_rows, max_no_used, num_appended_today, matrix_row_count)
    """
    header = _normalize_feishu_template_header(template_header)
    ncols = len(header)
    labels = matrix.get("labels") or []
    label_to_idx = {str(s).strip(): i for i, s in enumerate(labels)}
    excel_today = _anchor_date_to_excel_serial(matrix.get("date") or "")
    matrix_rows = matrix.get("rows") or []

    i_no = _feishu_header_col_index(header, "No")
    i_date = _feishu_header_col_index(header, "Date")
    i_cc = _feishu_header_col_index(header, "company_code")
    i_gn = _feishu_header_col_index(header, "group_no")
    i_pt = _feishu_header_col_index(header, "pay_type")

    def pad(cells):
        r = list(cells or [])
        if len(r) < ncols:
            r.extend([""] * (ncols - len(r)))
        return r[:ncols]

    max_no = 0
    historical = []
    today_no_by_key = {}
    today_row_by_key = {}

    for cells in existing_rows:
        c = pad(cells)
        if i_no < 0:
            continue
        n = _parse_feishu_sheet_int(c[i_no])
        if n is None:
            continue
        max_no = max(max_no, n)
        d = _parse_feishu_sheet_int(c[i_date]) if i_date >= 0 else None
        if d == excel_today and i_cc >= 0 and i_gn >= 0 and i_pt >= 0:
            if replace_operating_day:
                continue
            key2 = (str(c[i_cc]).strip(), str(c[i_gn]).strip(), str(c[i_pt]).strip())
            today_no_by_key[key2] = n
            today_row_by_key[key2] = c
            continue
        historical.append(c)

    num_append = 0
    today_block = []
    used_keys = set()
    for r in matrix_rows:
        key2 = (
            str(r.get("company") or "").strip(),
            str(r.get("group_no") or "").strip(),
            str(r.get("pay_type") or "").strip(),
        )
        if replace_operating_day:
            max_no += 1
            no = max_no
            num_append += 1
        elif key2 in today_no_by_key:
            no = today_no_by_key[key2]
        elif _cno_labor_feishu_should_append_new_row(r, label_to_idx):
            max_no += 1
            no = max_no
            today_no_by_key[key2] = no
            num_append += 1
        else:
            continue
        today_block.append(
            _build_cno_labor_feishu_row(no, excel_today, r, header, label_to_idx)
        )
        used_keys.add(key2)

    if not replace_operating_day:
        for key2, old_cells in today_row_by_key.items():
            if key2 not in used_keys:
                historical.append(pad(old_cells))

    historical.sort(
        key=lambda c: (
            _parse_feishu_sheet_int(c[i_date]) if i_date >= 0 else 0,
            _parse_feishu_sheet_int(c[i_no]) if i_no >= 0 else 0,
        )
    )
    combined = [pad(c) for c in historical] + today_block
    return combined, max_no, num_append, len(matrix_rows)


def _feishu_trim_trailing_empty_rows(rows):
    out = list(rows or [])
    while out:
        last = out[-1]
        if last and any(
            str(x).strip() != "" for x in last if x is not None
        ):
            break
        out.pop()
    return out


def _feishu_count_existing_data_rows(
    tenant_token: str, spreadsheet_token: str, sheet_id: str, data_start_row: int
) -> int:
    """根据 A 列「No」估算已有数据行数（用于清除旧行）。"""
    col_a = feishu_auth.feishu_sheet_read_values(
        tenant_token,
        spreadsheet_token,
        f"{sheet_id}!A{data_start_row}:A500",
    )
    n = 0
    for cells in col_a:
        if not cells:
            continue
        v = cells[0]
        if v is None or str(v).strip() == "":
            break
        n += 1
    return n


_CNO_NARROWBELT_LINE_LABELS = {
    "A": "产线 A（AA）",
    "B": "产线 B（AB）",
    "C": "产线 C（AC）",
    "D": "产线 D（AD）",
}


def _narrowbelt_hourly_line_stats(hourly_arr):
    vals = [int(x or 0) for x in (hourly_arr or [])]
    total = sum(vals)
    active = sum(1 for v in vals if v > 0)
    mean_active = (total / active) if active > 0 else 0.0
    return total, active, mean_active


def _build_cno_narrowbelt_stats_rows(matrix):
    """与 statistics 页 cnoNarrowbeltHourlyStatsTable 一致（A–D 四条产线）。"""
    lines = (matrix or {}).get("lines") or {}
    out = []
    for code in ("A", "B", "C", "D"):
        total, active, mean_active = _narrowbelt_hourly_line_stats(lines.get(code))
        out.append(
            {
                "line_code": code,
                "line_label": _CNO_NARROWBELT_LINE_LABELS.get(code, code),
                "total": total,
                "active_slots": active,
                "mean_active": round(mean_active, 1),
            }
        )
    return out


def _build_cno_narrowbelt_feishu_slot_row(
    no, excel_date, time_slot, line_code, pieces, header
):
    """单产线、单整点一行（17:00 有量时 A/B/C/D 各追加一行）。"""
    ncols = len(header)
    row = [""] * ncols
    label = _CNO_NARROWBELT_LINE_LABELS.get(line_code, line_code)
    for col_idx, h in enumerate(header):
        hstr = str(h).strip() if h is not None else ""
        if hstr == "No":
            row[col_idx] = no
        elif hstr == "Date":
            row[col_idx] = excel_date
        elif hstr in ("time_slot", "time_slot_la"):
            row[col_idx] = time_slot
        elif hstr in ("line_code", "line"):
            row[col_idx] = line_code
        elif hstr in ("line_label", "line_name", "产线"):
            row[col_idx] = label
        elif hstr in (
            "pieces",
            "total_pieces",
            "total",
            "day_total",
            "合计(件)",
            "合计",
        ):
            row[col_idx] = int(pieces or 0)
        elif hstr in ("0", ""):
            row[col_idx] = 0
    return row


def _feishu_narrowbelt_template_kind(header: list) -> str:
    names = {str(h).strip().lower() for h in header if h is not None}
    if "time_slot" in names and "line_code" in names and "pieces" in names:
        return "slot_lines"
    if "time_slot" in names or "time_slot_la" in names:
        return "hourly_wide"
    return "legacy_summary"


def _narrowbelt_slot_has_pieces(matrix, slot_label: str) -> bool:
    labels = matrix.get("labels") or []
    lines = matrix.get("lines") or {}
    if slot_label not in labels:
        return False
    idx = labels.index(slot_label)
    for code in ("A", "B", "C", "D"):
        arr = lines.get(code) or []
        if idx < len(arr) and int(arr[idx] or 0) > 0:
            return True
    return False


def _narrowbelt_line_pieces_at_slot(matrix, line_code: str, slot_label: str) -> int:
    labels = matrix.get("labels") or []
    lines = matrix.get("lines") or {}
    if slot_label not in labels:
        return 0
    idx = labels.index(slot_label)
    arr = lines.get(line_code) or []
    return int(arr[idx]) if idx < len(arr) else 0


def _narrowbelt_slot_la_start(anchor_date, slot_label: str, window_mode: str):
    """运营日锚点 + 整点标签 → 该时段开始的洛杉矶时刻。"""
    if isinstance(anchor_date, str):
        anchor_date = datetime.strptime(str(anchor_date)[:10], "%Y-%m-%d").date()
    elif isinstance(anchor_date, datetime):
        anchor_date = anchor_date.date()
    h = int(str(slot_label)[:2])
    d = anchor_date
    if window_mode == "seventeen" and h < 17:
        d = anchor_date + timedelta(days=1)
    elif window_mode == "business" and h < 5:
        d = anchor_date + timedelta(days=1)
    return datetime.combine(d, dtime(h, 0), tzinfo=LA_TZ)


def _narrowbelt_slot_has_started(anchor_date, slot_label: str, window_mode: str) -> bool:
    """仅当洛杉矶当前时间 ≥ 该整点开始时，才允许写入飞书。"""
    return datetime.now(LA_TZ) >= _narrowbelt_slot_la_start(
        anchor_date, slot_label, window_mode
    )


def _norm_narrowbelt_slot_label(val) -> str:
    """与库内 time_slot、矩阵 labels 对齐（HH:00）。"""
    s = str(val or "").strip()
    if not s:
        return ""
    if len(s) >= 8 and s[2] == ":" and s[5] == ":":
        s = f"{int(s[:2]):02d}:{s[3:5]}"
    elif ":" in s:
        parts = s.split(":")
        try:
            s = f"{int(parts[0]):02d}:{str(parts[1])[:2]}"
        except (ValueError, IndexError):
            return ""
    elif s.isdigit():
        try:
            h = int(s)
            if 0 <= h <= 23:
                s = f"{h:02d}:00"
        except ValueError:
            return ""
    if len(s) == 5 and s[2] == ":":
        return s
    return ""


def _narrowbelt_row_slot_key(cells, header, excel_today):
    i_date = _feishu_header_col_index(header, "Date")
    i_slot = _feishu_header_col_index(header, "time_slot")
    if i_slot < 0:
        i_slot = _feishu_header_col_index(header, "time_slot_la")
    i_lc = _feishu_header_col_index(header, "line_code")
    if i_lc < 0:
        i_lc = _feishu_header_col_index(header, "line")
    if i_date < 0 or i_slot < 0 or i_lc < 0:
        return None
    d = _parse_feishu_sheet_int(cells[i_date])
    if d != excel_today:
        return None
    slot = _norm_narrowbelt_slot_label(cells[i_slot])
    lc = str(cells[i_lc] or "").strip().upper()
    if not slot or not lc:
        return None
    return (slot, lc)


def _build_cno_narrowbelt_feishu_hourly_rows(
    matrix, header, *, only_started: bool = False
):
    """与 export CSV 一致（每整点一行）；only_started 时仅含洛杉矶已开始的整点。"""
    labels = matrix.get("labels") or []
    lines = matrix.get("lines") or {}
    excel_date = _anchor_date_to_excel_serial(matrix.get("date") or "")
    anchor_date = matrix.get("date") or ""
    wm = matrix.get("stats_window") or "seventeen"
    ncols = len(header)
    i_no = _feishu_header_col_index(header, "No")
    i_date = _feishu_header_col_index(header, "Date")
    i_slot = _feishu_header_col_index(header, "time_slot")
    if i_slot < 0:
        i_slot = _feishu_header_col_index(header, "time_slot_la")
    line_cols = {
        "A": _feishu_header_col_index(header, "line_a_pieces"),
        "B": _feishu_header_col_index(header, "line_b_pieces"),
        "C": _feishu_header_col_index(header, "line_c_pieces"),
        "D": _feishu_header_col_index(header, "line_d_pieces"),
    }
    if line_cols["A"] < 0:
        line_cols = {
            "A": _feishu_header_col_index(header, "line_a"),
            "B": _feishu_header_col_index(header, "line_b"),
            "C": _feishu_header_col_index(header, "line_c"),
            "D": _feishu_header_col_index(header, "line_d"),
        }
    out = []
    row_no = 0
    for idx, lab in enumerate(labels):
        if only_started and not _narrowbelt_slot_has_started(
            anchor_date, lab, wm
        ):
            continue
        row_no += 1
        row = [""] * ncols
        if i_no >= 0:
            row[i_no] = row_no
        if i_date >= 0:
            row[i_date] = excel_date
        if i_slot >= 0:
            row[i_slot] = lab
        for code, ci in line_cols.items():
            if ci >= 0:
                arr = lines.get(code) or []
                row[ci] = int(arr[idx]) if idx < len(arr) else 0
        out.append(row)
    return out


def _build_cno_narrowbelt_feishu_today_slot_lines(
    matrix,
    header,
    excel_today: int,
    *,
    only_started: bool,
    only_with_pieces: bool,
    start_no: int,
) -> tuple:
    """生成运营日 slot_lines 行块；(rows, max_no, slot_labels_written)。"""
    labels = matrix.get("labels") or []
    anchor_date = matrix.get("date") or ""
    wm = matrix.get("stats_window") or "seventeen"
    line_codes = ("A", "B", "C", "D")
    max_no = start_no
    out = []
    slots_written = []
    for slot in labels:
        if only_started and not _narrowbelt_slot_has_started(
            anchor_date, slot, wm
        ):
            continue
        if only_with_pieces and not _narrowbelt_slot_has_pieces(matrix, slot):
            continue
        slots_written.append(slot)
        for lc in line_codes:
            max_no += 1
            pcs = _narrowbelt_line_pieces_at_slot(matrix, lc, slot)
            out.append(
                _build_cno_narrowbelt_feishu_slot_row(
                    max_no, excel_today, slot, lc, pcs, header
                )
            )
    return out, max_no, slots_written


def _merge_cno_narrowbelt_feishu_sheet_body(
    template_header: list,
    matrix: dict,
    existing_rows: list,
    *,
    replace_operating_day: bool = True,
    reset_operating_day: bool | None = None,
) -> tuple:
    """
    合并飞书表体。replace_operating_day=True（默认）时整批重写当前运营日：
    写入洛杉矶已开始的全部整点 × A/B/C/D（含 0 件），与统计页矩阵一致，避免漏 12:00 等时段。
    replace_operating_day=False 时沿用增量：仅「已开始且任一线有量」且表中尚无四行时追加。
    """
    if reset_operating_day is not None:
        replace_operating_day = bool(reset_operating_day)

    header = _normalize_feishu_narrowbelt_header(template_header)
    kind = _feishu_narrowbelt_template_kind(header)
    excel_today = _anchor_date_to_excel_serial(matrix.get("date") or "")
    anchor_date = matrix.get("date") or ""
    wm = matrix.get("stats_window") or "seventeen"
    ncols = len(header)
    labels = matrix.get("labels") or []

    def pad(cells):
        r = list(cells or [])
        if len(r) < ncols:
            r.extend([""] * (ncols - len(r)))
        return r[:ncols]

    i_no = _feishu_header_col_index(header, "No")
    i_date = _feishu_header_col_index(header, "Date")
    line_codes = ("A", "B", "C", "D")

    max_no = 0
    historical = []
    for cells in existing_rows:
        c = pad(cells)
        if i_no >= 0:
            n = _parse_feishu_sheet_int(c[i_no])
            if n is not None:
                max_no = max(max_no, n)
        sk = _narrowbelt_row_slot_key(c, header, excel_today)
        if sk and replace_operating_day:
            continue
        if sk:
            continue
        historical.append(c)

    skipped_future_slots = []
    for slot in labels:
        if _narrowbelt_slot_has_pieces(matrix, slot) and not _narrowbelt_slot_has_started(
            anchor_date, slot, wm
        ):
            skipped_future_slots.append(slot)

    if kind == "hourly_wide":
        today_block = _build_cno_narrowbelt_feishu_hourly_rows(
            matrix, header, only_started=True
        )
        for row in today_block:
            n = _parse_feishu_sheet_int(row[i_no]) if i_no >= 0 else None
            if n is not None:
                max_no = max(max_no, n)
        historical.sort(
            key=lambda c: (
                _parse_feishu_sheet_int(c[i_date]) if i_date >= 0 else 0,
                _parse_feishu_sheet_int(c[i_no]) if i_no >= 0 else 0,
            )
        )
        meta = {
            "appended_slots": [],
            "updated_slots": [],
            "skipped_future_slots": skipped_future_slots,
            "replace_operating_day": replace_operating_day,
        }
        return [pad(c) for c in historical] + today_block, max_no, len(
            today_block
        ), meta

    if replace_operating_day:
        today_out, max_no, rewritten_slots = _build_cno_narrowbelt_feishu_today_slot_lines(
            matrix,
            header,
            excel_today,
            only_started=True,
            only_with_pieces=False,
            start_no=max_no,
        )
        meta = {
            "appended_slots": [],
            "updated_slots": rewritten_slots,
            "rewritten_slots": rewritten_slots,
            "skipped_future_slots": skipped_future_slots,
            "replace_operating_day": True,
        }
    else:
        today_by_key = {}
        for cells in existing_rows:
            c = pad(cells)
            sk = _narrowbelt_row_slot_key(c, header, excel_today)
            if not sk or i_no < 0:
                continue
            n = _parse_feishu_sheet_int(c[i_no])
            if n is None:
                continue
            today_by_key[sk] = (n, c)

        today_out = []
        appended_slots = []
        updated_slots = []
        for slot in labels:
            if not _narrowbelt_slot_has_started(anchor_date, slot, wm):
                continue
            if not _narrowbelt_slot_has_pieces(matrix, slot):
                continue
            keys = [(slot, lc) for lc in line_codes]
            had_all = all(k in today_by_key for k in keys)
            slot_new = False
            for lc in line_codes:
                pcs = _narrowbelt_line_pieces_at_slot(matrix, lc, slot)
                if (slot, lc) in today_by_key:
                    no, _old = today_by_key[(slot, lc)]
                else:
                    max_no += 1
                    no = max_no
                    slot_new = True
                today_out.append(
                    _build_cno_narrowbelt_feishu_slot_row(
                        no, excel_today, slot, lc, pcs, header
                    )
                )
            if had_all:
                updated_slots.append(slot)
            elif slot_new:
                appended_slots.append(slot)
        meta = {
            "appended_slots": appended_slots,
            "updated_slots": updated_slots,
            "skipped_future_slots": skipped_future_slots,
            "replace_operating_day": False,
        }

    def _sort_today_row(cells):
        sk = _narrowbelt_row_slot_key(cells, header, excel_today)
        slot_ord = labels.index(sk[0]) if sk and sk[0] in labels else 999
        line_ord = line_codes.index(sk[1]) if sk and sk[1] in line_codes else 9
        return (slot_ord, line_ord)

    today_out.sort(key=_sort_today_row)
    historical.sort(
        key=lambda c: (
            _parse_feishu_sheet_int(c[i_date]) if i_date >= 0 else 0,
            _parse_feishu_sheet_int(c[i_no]) if i_no >= 0 else 0,
        )
    )
    return [pad(c) for c in historical] + today_out, max_no, len(today_out), meta


def feishu_sync_cno_narrowbelt_sheet_once(
    stats_window: str = "seventeen",
    count_mode: str = "raw",
    reset_operating_day: bool | None = None,
    replace_operating_day: bool | None = None,
):
    """
    将 statistics 窄带分时数据写入飞书 eEZ3Ly。
    默认 replace_operating_day=True：当前运营日按矩阵整批重写（含已开始的无量整点），
    避免增量模式漏写 12:00–14:00 等时段。未到洛杉矶整点的不写入。
    """
    spreadsheet_token = os.getenv(
        "FEISHU_CNO_NARROWBELT_SPREADSHEET_TOKEN",
        os.getenv(
            "FEISHU_CNO_LABOR_GROUP_HOURLY_SPREADSHEET_TOKEN",
            "Kg5Mwy0TViWEvokr3AScfTrCnbg",
        ),
    ).strip()
    sheet_id_cfg = os.getenv("FEISHU_CNO_NARROWBELT_SHEET_ID", "eEZ3Ly").strip()
    sheet_title = os.getenv("FEISHU_CNO_NARROWBELT_SHEET_TITLE", "分拣").strip()
    cfg_last_key = "feishu_cno_narrowbelt_sheet_last_synced_at"

    wm = stats_window
    cm = count_mode
    anchor = _default_stats_request_date(wm)
    matrix = _build_cno_narrowbelt_hourly_series(anchor, wm, cm)
    synced_at = datetime.now(LA_TZ).strftime("%Y-%m-%d %H:%M:%S")
    data_start_row = int(os.getenv("FEISHU_CNO_NARROWBELT_DATA_START_ROW", "2"))

    tenant_token = feishu_auth.feishu_tenant_access_token()
    sheet_id = feishu_auth.feishu_sheet_resolve_sheet_id(
        tenant_token,
        spreadsheet_token,
        sheet_id=sheet_id_cfg,
        sheet_title=sheet_title,
    )

    header_rows = feishu_auth.feishu_sheet_read_values(
        tenant_token, spreadsheet_token, f"{sheet_id}!A1:AF1"
    )
    raw_header = (header_rows[0] if header_rows else None) or []
    has_header = raw_header and any(
        str(x).strip() for x in raw_header if x is not None
    )
    template_header = _normalize_feishu_narrowbelt_header(raw_header)
    ncols = len(template_header)
    end_col = feishu_auth.feishu_sheet_col_letter(ncols - 1)
    legacy_header = has_header and _feishu_narrowbelt_template_kind(
        _normalize_feishu_header_row(raw_header, _CNO_NARROWBELT_FEISHU_HEADER_FALLBACK)
    ) == "legacy_summary"

    if not has_header or legacy_header:
        feishu_auth.feishu_sheet_write_values(
            tenant_token,
            spreadsheet_token,
            sheet_id,
            [template_header],
            start_row=1,
            start_col=1,
        )

    max_body = int(os.getenv("FEISHU_CNO_LABOR_SHEET_MAX_BODY_ROWS", "3000"))
    prev_contiguous = _feishu_count_existing_data_rows(
        tenant_token, spreadsheet_token, sheet_id, data_start_row
    )
    body_raw = feishu_auth.feishu_sheet_read_values(
        tenant_token,
        spreadsheet_token,
        f"{sheet_id}!A{data_start_row}:{end_col}{data_start_row + max_body - 1}",
    )
    body_trim = _feishu_trim_trailing_empty_rows(body_raw)
    prev_filled = max(prev_contiguous, len(body_trim))

    if replace_operating_day is None:
        if reset_operating_day is not None:
            replace_operating_day = bool(reset_operating_day)
        else:
            env_rep = os.getenv(
                "FEISHU_CNO_NARROWBELT_REPLACE_OPERATING_DAY", ""
            ).strip().lower()
            if env_rep in ("0", "false", "no"):
                replace_operating_day = False
            elif env_rep in ("1", "true", "yes"):
                replace_operating_day = True
            else:
                legacy = os.getenv(
                    "FEISHU_CNO_NARROWBELT_RESET_OPERATING_DAY", ""
                ).strip().lower() in ("1", "true", "yes")
                replace_operating_day = legacy if legacy else True

    data_values, max_no, today_rows, slot_meta = _merge_cno_narrowbelt_feishu_sheet_body(
        template_header,
        matrix,
        body_trim,
        replace_operating_day=replace_operating_day,
    )
    new_rows = len(data_values)
    kind = _feishu_narrowbelt_template_kind(template_header)

    result = {}
    if new_rows:
        result = feishu_auth.feishu_sheet_write_values(
            tenant_token,
            spreadsheet_token,
            sheet_id,
            data_values,
            start_row=data_start_row,
            start_col=1,
        )

    cleared = 0
    if prev_filled > new_rows:
        blank = [[""] * ncols for _ in range(prev_filled - new_rows)]
        feishu_auth.feishu_sheet_write_values(
            tenant_token,
            spreadsheet_token,
            sheet_id,
            blank,
            start_row=data_start_row + new_rows,
            start_col=1,
        )
        cleared = prev_filled - new_rows

    conn = get_db()
    cursor = conn.cursor()
    _set_system_config_value(cursor, cfg_last_key, synced_at)
    conn.commit()
    conn.close()
    return {
        "spreadsheet_token": spreadsheet_token,
        "sheet_id": sheet_id,
        "template_kind": kind,
        "operating_day": matrix.get("date"),
        "count_mode": cm,
        "stats_window": wm,
        "data_rows": new_rows,
        "today_rows": today_rows,
        "appended_slots": slot_meta.get("appended_slots") or [],
        "updated_slots": slot_meta.get("updated_slots") or [],
        "skipped_future_slots": slot_meta.get("skipped_future_slots") or [],
        "replace_operating_day": slot_meta.get("replace_operating_day", replace_operating_day),
        "rewritten_slots": slot_meta.get("rewritten_slots") or [],
        "max_no": max_no,
        "cleared_old_rows": cleared,
        "updated_range": result.get("updatedRange"),
        "template_preserved": bool(has_header) and not legacy_header,
    }


def feishu_sync_cno_labor_group_hourly_sheet_once(
    stats_window: str = "seventeen",
    count_mode: str = "raw",
    anchor_date=None,
    replace_operating_day: bool = True,
):
    """
    将 CNO 小组分时矩阵写入飞书电子表格（保留表头）。
    默认 replace_operating_day=True：当前运营日按矩阵整批重写，行数与统计页一致。
    历史运营日行保留；No 在整表内递增。
    """
    spreadsheet_token = os.getenv(
        "FEISHU_CNO_LABOR_GROUP_HOURLY_SPREADSHEET_TOKEN",
        "Kg5Mwy0TViWEvokr3AScfTrCnbg",
    ).strip()
    sheet_id_cfg = os.getenv("FEISHU_CNO_LABOR_GROUP_HOURLY_SHEET_ID", "e7b9e6").strip()
    sheet_title = os.getenv(
        "FEISHU_CNO_LABOR_GROUP_HOURLY_SHEET_TITLE", "元数据"
    ).strip()
    cfg_last_key = "feishu_wiki_cno_labor_group_hourly_meta_last_synced_at"

    wm = stats_window
    cm = count_mode
    if anchor_date is None:
        anchor = _default_stats_request_date(wm)
    elif isinstance(anchor_date, datetime):
        anchor = anchor_date.date()
    elif isinstance(anchor_date, str):
        anchor = datetime.strptime(str(anchor_date)[:10], "%Y-%m-%d").date()
    else:
        anchor = anchor_date
    matrix = _build_cno_labor_group_hourly_matrix(anchor, wm, cm)
    synced_at = datetime.now(LA_TZ).strftime("%Y-%m-%d %H:%M:%S")
    data_start_row = int(
        os.getenv("FEISHU_CNO_LABOR_GROUP_HOURLY_DATA_START_ROW", "2")
    )

    tenant_token = feishu_auth.feishu_tenant_access_token()
    sheet_id = feishu_auth.feishu_sheet_resolve_sheet_id(
        tenant_token,
        spreadsheet_token,
        sheet_id=sheet_id_cfg,
        sheet_title=sheet_title,
    )

    # 读取第 1 行表头，仅更新数据区（不改表头/列布局）
    header_rows = feishu_auth.feishu_sheet_read_values(
        tenant_token, spreadsheet_token, f"{sheet_id}!A1:AF1"
    )
    template_header = _normalize_feishu_template_header(
        (header_rows[0] if header_rows else None) or []
    )
    ncols = len(template_header)
    end_col = feishu_auth.feishu_sheet_col_letter(ncols - 1)
    max_body = int(os.getenv("FEISHU_CNO_LABOR_SHEET_MAX_BODY_ROWS", "3000"))

    prev_contiguous = _feishu_count_existing_data_rows(
        tenant_token, spreadsheet_token, sheet_id, data_start_row
    )
    body_raw = feishu_auth.feishu_sheet_read_values(
        tenant_token,
        spreadsheet_token,
        f"{sheet_id}!A{data_start_row}:{end_col}{data_start_row + max_body - 1}",
    )
    body_trim = _feishu_trim_trailing_empty_rows(body_raw)
    prev_filled = max(prev_contiguous, len(body_trim))

    data_values, max_no_used, num_appended, matrix_row_count = (
        _merge_cno_labor_feishu_sheet_body(
            template_header,
            matrix,
            body_trim,
            replace_operating_day=replace_operating_day,
        )
    )
    new_rows = len(data_values)

    result = {}
    if new_rows:
        result = feishu_auth.feishu_sheet_write_values(
            tenant_token,
            spreadsheet_token,
            sheet_id,
            data_values,
            start_row=data_start_row,
            start_col=1,
        )

    cleared = 0
    if prev_filled > new_rows:
        blank = [[""] * ncols for _ in range(prev_filled - new_rows)]
        feishu_auth.feishu_sheet_write_values(
            tenant_token,
            spreadsheet_token,
            sheet_id,
            blank,
            start_row=data_start_row + new_rows,
            start_col=1,
        )
        cleared = prev_filled - new_rows

    conn = get_db()
    cursor = conn.cursor()
    _set_system_config_value(cursor, cfg_last_key, synced_at)
    conn.commit()
    conn.close()
    return {
        "spreadsheet_token": spreadsheet_token,
        "sheet_id": sheet_id,
        "header_cols": ncols,
        "data_rows": new_rows,
        "matrix_rows": matrix_row_count,
        "operating_day_rows": len(matrix.get("rows") or []),
        "max_no": max_no_used,
        "appended_today": num_appended,
        "replace_operating_day": replace_operating_day,
        "cleared_old_rows": cleared,
        "updated_range": result.get("updatedRange"),
        "template_preserved": True,
    }


# 每小时同步：CNO 小组分时明细 → 飞书电子表格（Wiki 链接即表格 token）
def feishu_wiki_sync_cno_labor_group_hourly_metadata_job():
    import time as _time

    def _run_once():
        nb = feishu_sync_cno_narrowbelt_sheet_once()
        print(
            f"[AutoSync] Feishu narrowbelt OK "
            f"range={nb.get('updated_range')} rows={nb.get('data_rows')} "
            f"day={nb.get('operating_day')}"
        )
        info = feishu_sync_cno_labor_group_hourly_sheet_once()
        print(
            f"[AutoSync] Feishu labor group OK "
            f"range={info.get('updated_range')} data_rows={info.get('data_rows')} "
            f"appended_today={info.get('appended_today')}"
        )

    print("[AutoSync] Feishu 表格（窄带按线 + 小组分时）后台任务启动")
    while True:
        try:
            now_la = datetime.now(LA_TZ)
            next_run = now_la.replace(minute=15, second=0, microsecond=0)
            if next_run <= now_la:
                next_run += timedelta(hours=1)
            wait_sec = (next_run - now_la).total_seconds()
            print(
                f"[AutoSync] Feishu sheet meta next run "
                f"{next_run.strftime('%Y-%m-%d %H:%M:%S')} (wait {wait_sec:.0f}s)"
            )
            _time.sleep(wait_sec)
            _run_once()
        except Exception as e:
            print(f"[AutoSync] Feishu sheet meta sync FAIL: {e}")
            _time.sleep(1800)

# 每小时自动同步 Gofo 集包数据 (方案 1)

def truck_booking_hourly_sync_job():
    """Background loop that runs truck booking sync every hour on the hour."""
    import time
    from datetime import datetime, timedelta
    print("[AutoSync] Truck booking hourly sync job started.")
    # The default URL used for background syncs
    default_url = "https://docs.google.com/spreadsheets/d/1sEjOb1Yy7ap_B6LpHNHxIgF21vzDCvD1uSqSBIY9oTA/edit?gid=0#gid=0"
    
    while True:
        try:
            now = datetime.now()
            # Calculate next hour on the hour (e.g., 14:00:00)
            next_run = now.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
            wait_seconds = (next_run - now).total_seconds()
            
            print(f"[AutoSync] Scheduled next truck booking sync for {next_run.strftime('%Y-%m-%d %H:%M:%S')} (in {wait_seconds:.1f} seconds)")
            time.sleep(wait_seconds)
            
            print(f"[AutoSync] Waking up. Triggering scheduled truck booking sync at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}...")
            success, msg, count = _sync_truck_bookings_core(default_url)
            
            if success:
                print(f"[AutoSync] SUCCESS: Synced {count} truck booking records.")
            else:
                print(f"[AutoSync] ERROR: Truck booking sync failed: {msg}")
                
        except Exception as e:
            print(f"[AutoSync] Loop Error in truck booking sync: {e}")
            time.sleep(60)

def _maybe_backfill_center_collect_initial():
    """首次上线（或表为空）时自动回补最近 30 天集包数据（目的中心+站点×时段），幂等。"""
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name='gofo_center_collect_stats'"
        )
        exists = cur.fetchone() is not None
        has_recent = False
        if exists:
            cutoff = (datetime.now() - timedelta(days=1)).strftime('%Y-%m-%d')
            cur.execute(
                convert_query_placeholders(
                    "SELECT COUNT(*) AS n FROM gofo_center_collect_stats WHERE record_date >= ?"
                ),
                (cutoff,),
            )
            has_recent = (cur.fetchone()['n'] or 0) > 0
        conn.close()
    except Exception as e:
        print(f"[GofoAutoSync] center collect backfill check failed: {e}")
        return

    if has_recent:
        return  # 最近 24 小时内有数据，认为已初始化过
    try:
        import sync_center_collect
        print("[GofoAutoSync] center_collect first-run: backfilling last 30 days ...")
        res = sync_center_collect.fetch_center_collect_backfill(days=30)
        print(
            f"[GofoAutoSync] center_collect backfill done: "
            f"total_stored_rows={res.get('total_stored_rows')}"
        )
    except Exception as e:
        print(f"[GofoAutoSync] center_collect backfill failed: {e}")


def daily_packing_operlog_hourly_sync_job():
    """每小时刷新 daily_packing_operlog_daily（逐条），供统计页「逐条（日志）」图表。"""
    import sync_daily_packing_operlog as _dp_oper

    log_prefix = "DailyPackingOperlog"
    print(f"[{log_prefix}] Background job started at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    def job():
        try:
            res = _dp_oper.run_hourly_operlog_sync()
            if res.get("skipped"):
                print(f"[{log_prefix}] skipped: {res.get('reason')}")
                return
            n = len(res.get("results") or [])
            print(f"[{log_prefix}] finished {n} anchor/window sync(s)")
            try:
                broadcast_update("refresh_stats")
            except Exception:
                pass
        except Exception as e:
            print(f"[{log_prefix}] ERROR: {e}")

    try:
        print(f"[{log_prefix}] Bootstrap run...")
        job()
    except Exception as e:
        print(f"[{log_prefix}] Bootstrap error: {e}")

    while True:
        try:
            now = datetime.now()
            # 整点后 10 分钟执行，避开 Gofo 整点同步高峰
            next_run = now.replace(minute=10, second=0, microsecond=0)
            if next_run <= now:
                next_run += timedelta(hours=1)
            wait_seconds = (next_run - now).total_seconds()
            print(
                f"[{log_prefix}] Next run {next_run.strftime('%Y-%m-%d %H:%M:%S')} "
                f"(wait {wait_seconds:.0f}s)"
            )
            if wait_seconds > 0:
                time.sleep(wait_seconds)
            job()
        except Exception as e:
            print(f"[{log_prefix}] Loop error: {e}")
            time.sleep(60)


def gofo_hourly_sync_job():
    """执行每小时 Gofo 同步的任务函数"""
    print(f"[GofoAutoSync] Background job started at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    try:
        _maybe_backfill_center_collect_initial()
    except Exception as e:
        print(f"[GofoAutoSync] center_collect backfill guard error: {e}")
    
    # 定义任务内容
    def job():
        try:
            now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            print(f"[GofoAutoSync] Triggering scheduled sync at {now_str}...")
            result = perform_gofo_hourly_sync()

            try:
                import sync_cno_narrowbelt_hourly as _cno_nb

                nb = _cno_nb.sync_today_la_hours()
                if nb.get("errors"):
                    print(
                        f"[GofoAutoSync] cno narrowbelt partial errors: {nb.get('errors')}"
                    )
                else:
                    print(
                        f"[GofoAutoSync] cno narrowbelt ok date={nb.get('date')} "
                        f"hours={nb.get('hours_attempted')}"
                    )
            except Exception as e:
                print(f"[GofoAutoSync] cno narrowbelt failed: {e}")
            
            # Sync the center checkin stats at the same time
            try:
                import sync_center_checkin
                print(f"[GofoAutoSync] Triggering center checkin sync at {now_str}...")
                sync_center_checkin.fetch_center_checkin_data()
            except Exception as e:
                print(f"[GofoAutoSync] Center checkin sync failed: {e}")

            # 抓上一个完整整点的「集包运单数（目的中心）」快照
            try:
                import sync_center_collect
                print(f"[GofoAutoSync] Triggering center collect sync at {now_str}...")
                cc_res = sync_center_collect.fetch_latest_completed_hour()
                if cc_res.get('success'):
                    print(
                        f"[GofoAutoSync] center collect ok: "
                        f"{cc_res.get('date')} {cc_res.get('hour')} "
                        f"stored={cc_res.get('stored_rows')} centers={cc_res.get('centers')}"
                    )
                else:
                    print(f"[GofoAutoSync] center collect returned error: {cc_res.get('error')}")
            except Exception as e:
                print(f"[GofoAutoSync] Center collect sync failed: {e}")

            # TMS 短驳运输任务（已完成 + CNO.H）当天数据同步
            try:
                print(f"[GofoAutoSync] Triggering TMS shuttle sync at {now_str}...")
                ts_res = _tms_shuttle_sync_date(_tms_shuttle_la_calendar_date())
                if ts_res.get('success'):
                    print(
                        f"[GofoAutoSync] tms shuttle ok: {ts_res.get('date')} "
                        f"fetched={ts_res.get('fetched')} stored={ts_res.get('stored')}"
                    )
                else:
                    print(f"[GofoAutoSync] tms shuttle returned error: {ts_res.get('error')}")
            except Exception as e:
                print(f"[GofoAutoSync] TMS shuttle sync failed: {e}")

            synced_count = result.get('synced_count', 0)
            pieces = result.get('pieces', 0)
            hour = result.get('synced_hour')
            manual_today = result.get('manual_today', 0)
            device_today = result.get('device_today', 0)
            
            # 更新全域狀態
            update_gofo_sync_status(
                "success", 
                f"Auto sync successful: {synced_count} hours",
                synced_count,
                pieces,
                hour,
                manual_today,
                device_today
            )
            
            # 记录到历史记录
            log_gofo_sync_event("auto", "success", f"Auto sync completed: {synced_count} hours", synced_count, pieces, hour, manual_today, device_today)
            
            print(f"[GofoAutoSync] SUCCESS: Synced {synced_count} hours. Last hour: {hour}, Pieces: {pieces}, Manual: {manual_today}, Device: {device_today}")
            
            # 廣播更新以便前端實時刷新
            try:
                broadcast_update('refresh_stats')
            except:
                pass
                
        except Exception as e:
            error_msg = str(e)
            print(f"[GofoAutoSync] ERROR: {error_msg}")
            update_gofo_sync_status("error", f"Auto sync error: {error_msg}")
            log_gofo_sync_event("auto", "error", f"Auto sync failed: {error_msg}")

    # 启动后先立即执行一次，避免服务重启后要等到下一个整点才有新数据。
    try:
        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        print(f"[GofoAutoSync] Bootstrap run at {now_str}...")
        job()
    except Exception as e:
        print(f"[GofoAutoSync] Bootstrap Error: {e}")

    # 循环执行
    while True:
        try:
            now = datetime.now()
            # 下一整点执行（例如 13:25 -> 14:00）
            next_run = now.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)
            wait_seconds = (next_run - now).total_seconds()
            print(f"[GofoAutoSync] Scheduled next sync for {next_run.strftime('%Y-%m-%d %H:%M:%S')} (waiting {wait_seconds:.1f}s)")
            
            if wait_seconds > 0:
                time.sleep(wait_seconds)
            
            now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            print(f"[GofoAutoSync] Waking up. Triggering scheduled sync at {now_str}...")
            job()
            
        except Exception as e:
            print(f"[GofoAutoSync] Loop Error: {e}")
            time.sleep(60)

# 模块级别初始化 - Gunicorn 导入模块时会执行
# 这确保数据库在应用启动时被初始化
def _cc_norm_date(v):
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if len(s) >= 10 and len(s) > 4 and s[4] == "-":
        return s[:10]
    return s


def _cc_norm_hour_slot(v):
    """统一为 HH:00，避免 9:00 / 09:00 与库内键不一致导致缺数。"""
    if v is None:
        return "00:00"
    s = str(v).strip()
    if not s:
        return "00:00"
    if ":" in s:
        parts = s.split(":")
        try:
            h = int(parts[0]) % 24
            return f"{h:02d}:00"
        except ValueError:
            return s
    if s.isdigit():
        return f"{int(s) % 24:02d}:00"
    return s


def _cc_time_point_key(rd, rh):
    return f"{_cc_norm_date(rd)} {_cc_norm_hour_slot(rh)}"


def _stats_single_day_cc_axis(anchor_yyyy_mm_dd: str, window_mode: str):
    """签入/集包单日图 24 时点：calendar=当日 00–23；business=05–次日 04；seventeen=17–次日 16（服务器本地锚点日）。"""
    d = datetime.strptime(anchor_yyyy_mm_dd.strip(), '%Y-%m-%d').date()
    ds = d.strftime('%Y-%m-%d')
    next_ds = (d + timedelta(days=1)).strftime('%Y-%m-%d')
    time_points = []
    labels = []
    if window_mode == 'business':
        h0 = 5
    elif window_mode == 'seventeen':
        h0 = 17
    else:
        h0 = None
    if h0 is not None:
        for i in range(24):
            h = (h0 + i) % 24
            rh = f"{h:02d}:00"
            labels.append(rh)
            rd_use = ds if h >= h0 else next_ds
            time_points.append(_cc_time_point_key(rd_use, rh))
    else:
        for h in range(24):
            rh = f"{h:02d}:00"
            labels.append(rh)
            time_points.append(_cc_time_point_key(ds, rh))
    return time_points, labels


@app.route('/api/center_checkin_trend', methods=['GET'])
def api_center_checkin_trend():
    """获取所有目的站点的签入数趋势数据（按小时）。

    查询参数：
      默认（无 date、无 days）：仅 **洛杉矶日历日当天** record_date。
      date: YYYY-MM-DD，只显示该日各时点。
      days: all/full/0 — 库内最近 max_points 个时间点；正整数 — record_date >= 今天−days（回溯若干天）。
      max_points: 仅在 days=all 或按天回溯模式下限制横轴点数，默认 10000，上限 20000。
    """
    try:
        conn = get_db()
        cursor = conn.cursor()
        window_mode = _parse_stats_window_param(request.args.get('stats_window'))

        date_param = (request.args.get("date") or "").strip()
        days_raw = (request.args.get("days") or "").strip().lower()
        try:
            max_points = int(request.args.get("max_points", "10000"))
        except ValueError:
            max_points = 10000
        max_points = min(max(max_points, 10), 20000)

        filter_date = None  # 单日筛选时非 None，仅查该日
        dates_result = []

        if days_raw in ("all", "full", "0"):
            cursor.execute(
                convert_query_placeholders(
                    """
                SELECT DISTINCT record_date, record_hour 
                FROM gofo_center_checkin_stats 
                WHERE record_date IS NOT NULL AND record_hour IS NOT NULL
                ORDER BY record_date DESC, record_hour DESC 
                LIMIT ?
                """
                ),
                (max_points,),
            )
            dates_result = list(reversed(cursor.fetchall()))
        elif date_param:
            try:
                datetime.strptime(date_param, "%Y-%m-%d")
            except ValueError:
                conn.close()
                return jsonify({"success": False, "error": "日期格式无效，请使用 YYYY-MM-DD"}), 400
            filter_date = date_param
            anchor_d = datetime.strptime(filter_date, "%Y-%m-%d").date()
            rh_clause, rh_binds = _record_date_hour_window_sql_binds(window_mode, anchor_d)
            cursor.execute(
                convert_query_placeholders(
                    f"""
                SELECT DISTINCT record_date, record_hour 
                FROM gofo_center_checkin_stats 
                WHERE record_date IS NOT NULL AND record_hour IS NOT NULL
                  AND ({rh_clause})
                ORDER BY record_date ASC, record_hour ASC
                """
                ),
                rh_binds,
            )
            dates_result = cursor.fetchall()
        elif days_raw.isdigit():
            days = int(days_raw)
            days = min(max(days, 1), 3650)
            cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
            cursor.execute(
                convert_query_placeholders(
                    """
                SELECT DISTINCT record_date, record_hour 
                FROM gofo_center_checkin_stats 
                WHERE record_date IS NOT NULL AND record_hour IS NOT NULL
                  AND record_date >= ?
                ORDER BY record_date ASC, record_hour ASC
                """
                ),
                (cutoff,),
            )
            dates_result = cursor.fetchall()
            if len(dates_result) > max_points:
                dates_result = dates_result[-max_points:]
        else:
            anchor_d = _default_stats_request_date(window_mode)
            filter_date = anchor_d.strftime('%Y-%m-%d')
            rh_clause, rh_binds = _record_date_hour_window_sql_binds(window_mode, anchor_d)
            cursor.execute(
                convert_query_placeholders(
                    f"""
                SELECT DISTINCT record_date, record_hour 
                FROM gofo_center_checkin_stats 
                WHERE record_date IS NOT NULL AND record_hour IS NOT NULL
                  AND ({rh_clause})
                ORDER BY record_hour ASC
                """
                ),
                rh_binds,
            )
            dates_result = cursor.fetchall()

        time_points = []
        labels = []
        if filter_date:
            time_points, labels = _stats_single_day_cc_axis(filter_date, window_mode)
        else:
            for row in dates_result:
                rd = _cc_norm_date(row["record_date"])
                rh = _cc_norm_hour_slot(row["record_hour"])
                dt_str = _cc_time_point_key(row["record_date"], row["record_hour"])
                time_points.append(dt_str)
                dp = rd.split("-")
                short_date = f"{dp[1]}-{dp[2]}" if len(dp) == 3 else rd
                labels.append(f"{short_date} {rh}")
            
        if not time_points:
            return jsonify({"success": True, "dates": [], "series": []})
        
        # 提取目的站点（单日模式只取当天有数据的站点）
        if filter_date:
            anchor_d = datetime.strptime(filter_date, '%Y-%m-%d').date()
            rh_clause, rh_binds = _record_date_hour_window_sql_binds(window_mode, anchor_d)
            cursor.execute(
                convert_query_placeholders(
                    f"""
                SELECT DISTINCT target_site_name FROM gofo_center_checkin_stats 
                WHERE target_site_name IS NOT NULL AND ({rh_clause})
                """
                ),
                rh_binds,
            )
        else:
            cursor.execute(
                "SELECT DISTINCT target_site_name FROM gofo_center_checkin_stats WHERE target_site_name IS NOT NULL"
            )
        sites_result = cursor.fetchall()
        sites = [row["target_site_name"] for row in sites_result]
        
        # 构建 series
        series = []
        for site in sites:
            if filter_date:
                anchor_d = datetime.strptime(filter_date, '%Y-%m-%d').date()
                rh_clause, rh_binds = _record_date_hour_window_sql_binds(window_mode, anchor_d)
                cursor.execute(
                    convert_query_placeholders(
                        f"""
                SELECT record_date, record_hour, check_in_waybill_cnt 
                FROM gofo_center_checkin_stats 
                WHERE target_site_name = ? AND ({rh_clause})
                ORDER BY record_date ASC, record_hour ASC
                """
                    ),
                    (site,) + rh_binds,
                )
            else:
                cursor.execute(
                    convert_query_placeholders(
                        """
                SELECT record_date, record_hour, check_in_waybill_cnt 
                FROM gofo_center_checkin_stats 
                WHERE target_site_name = ? 
                ORDER BY record_date ASC, record_hour ASC
                """
                    ),
                    (site,),
                )
            site_data = cursor.fetchall()

            data_map = {}
            for row in site_data:
                k = _cc_time_point_key(row["record_date"], row["record_hour"])
                v = int(row["check_in_waybill_cnt"] or 0)
                data_map[k] = data_map.get(k, 0) + v
            
            # 补齐每个时间点的数据
            data_points = []
            for tp in time_points:
                data_points.append(data_map.get(tp, 0))
                
            series.append({
                "label": site,
                "data": data_points
            })
            
        conn.close()
        
        return jsonify({
            "success": True,
            "dates": labels, # Use friendly labels for the X-axis
            "series": series
        })
    except Exception as e:
        print(f"Error fetching center checkin trend: {e}")
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/admin/center_checkin/sync', methods=['POST'])
def api_admin_center_checkin_sync():
    """从 GoFO DMS 拉取 CNO.H 签入看板数据并写入 gofo_center_checkin_stats（管理员）。"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': '未登录'}), 401
    conn = get_db()
    cursor = conn.cursor()
    if not require_admin(cursor):
        conn.close()
        return jsonify({'success': False, 'error': '无权访问'}), 403
    conn.close()
    payload = request.get_json(silent=True) or {}
    date_str = (payload.get('date') or '').strip()
    if not date_str:
        date_str = datetime.now().strftime('%Y-%m-%d')
    try:
        import sync_center_checkin
        result = sync_center_checkin.fetch_center_checkin_data(date_str)
        if result.get('success'):
            return jsonify({
                'success': True,
                'message': result.get('message', ''),
                'count': result.get('count', 0),
            })
        return jsonify({'success': False, 'error': result.get('error', '同步失败')}), 500
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/center_collect_trend', methods=['GET'])
def api_center_collect_trend():
    """获取"集包运单数（目的中心）"趋势：按小时 + 按目的中心堆叠。

    查询参数（与 /api/center_checkin_trend 对齐）：
      date:       YYYY-MM-DD（默认：LA 时区今天）
      days:       正整数 ⇒ 最近 N 天；all/full/0 ⇒ 库里全部（限 max_points）
      max_points: 横轴点数上限，默认 10000
      destin_type: 1=中心(默认) / 2=站点 / all=不过滤
    """
    try:
        conn = get_db()
        cursor = conn.cursor()
        window_mode = _parse_stats_window_param(request.args.get('stats_window'))

        date_param = (request.args.get("date") or "").strip()
        days_raw = (request.args.get("days") or "").strip().lower()
        destin_type_raw = (request.args.get("destin_type") or "1").strip().lower()
        try:
            max_points = int(request.args.get("max_points", "10000"))
        except ValueError:
            max_points = 10000
        max_points = min(max(max_points, 10), 20000)

        dt_filter_sql = ""
        dt_params: list = []
        if destin_type_raw in ("1", "center", "centers"):
            dt_filter_sql = " AND destin_type = ?"
            dt_params = [1]
        elif destin_type_raw in ("2", "site", "sites"):
            dt_filter_sql = " AND destin_type = ?"
            dt_params = [2]
        # "all" 不加过滤

        filter_date = None
        if days_raw in ("all", "full", "0"):
            sql = (
                "SELECT DISTINCT record_date, record_hour "
                f"FROM {'gofo_center_collect_stats'} "
                "WHERE record_date IS NOT NULL AND record_hour IS NOT NULL"
                + dt_filter_sql
                + " ORDER BY record_date DESC, record_hour DESC LIMIT ?"
            )
            cursor.execute(convert_query_placeholders(sql), tuple(dt_params + [max_points]))
            dates_result = list(reversed(cursor.fetchall()))
        elif date_param:
            try:
                datetime.strptime(date_param, "%Y-%m-%d")
            except ValueError:
                conn.close()
                return jsonify({"success": False, "error": "日期格式无效，请使用 YYYY-MM-DD"}), 400
            filter_date = date_param
            anchor_d = datetime.strptime(filter_date, '%Y-%m-%d').date()
            rh_clause, rh_binds = _record_date_hour_window_sql_binds(window_mode, anchor_d)
            sql = (
                f"SELECT DISTINCT record_date, record_hour FROM gofo_center_collect_stats "
                f"WHERE record_date IS NOT NULL AND record_hour IS NOT NULL AND ({rh_clause})"
                + dt_filter_sql + " ORDER BY record_date ASC, record_hour ASC"
            )
            cursor.execute(convert_query_placeholders(sql), tuple(list(rh_binds) + dt_params))
            dates_result = cursor.fetchall()
        elif days_raw.isdigit():
            days = min(max(int(days_raw), 1), 3650)
            cutoff = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
            sql = (
                "SELECT DISTINCT record_date, record_hour FROM gofo_center_collect_stats "
                "WHERE record_date >= ?" + dt_filter_sql +
                " ORDER BY record_date ASC, record_hour ASC"
            )
            cursor.execute(convert_query_placeholders(sql), tuple([cutoff] + dt_params))
            dates_result = cursor.fetchall()
            if len(dates_result) > max_points:
                dates_result = dates_result[-max_points:]
        else:
            anchor_d = _default_stats_request_date(window_mode)
            filter_date = anchor_d.strftime('%Y-%m-%d')
            rh_clause, rh_binds = _record_date_hour_window_sql_binds(window_mode, anchor_d)
            sql = (
                f"SELECT DISTINCT record_date, record_hour FROM gofo_center_collect_stats "
                f"WHERE record_date IS NOT NULL AND record_hour IS NOT NULL AND ({rh_clause})"
                + dt_filter_sql + " ORDER BY record_hour ASC"
            )
            cursor.execute(convert_query_placeholders(sql), tuple(list(rh_binds) + dt_params))
            dates_result = cursor.fetchall()

        time_points = []
        labels = []
        if filter_date:
            time_points, labels = _stats_single_day_cc_axis(filter_date, window_mode)
        else:
            for row in dates_result:
                rd = _cc_norm_date(row["record_date"])
                rh = _cc_norm_hour_slot(row["record_hour"])
                dt_str = _cc_time_point_key(row["record_date"], row["record_hour"])
                time_points.append(dt_str)
                dp = rd.split("-")
                short_date = f"{dp[1]}-{dp[2]}" if len(dp) == 3 else rd
                labels.append(f"{short_date} {rh}")

        if not time_points:
            conn.close()
            return jsonify({"success": True, "dates": [], "series": []})

        # 抽目的名单
        if filter_date:
            anchor_d = datetime.strptime(filter_date, '%Y-%m-%d').date()
            rh_clause, rh_binds = _record_date_hour_window_sql_binds(window_mode, anchor_d)
            sql_names = (
                f"SELECT DISTINCT destin_name FROM gofo_center_collect_stats "
                f"WHERE destin_name IS NOT NULL AND ({rh_clause})" + dt_filter_sql
            )
            cursor.execute(convert_query_placeholders(sql_names), tuple(list(rh_binds) + dt_params))
        else:
            sql_names = (
                "SELECT DISTINCT destin_name FROM gofo_center_collect_stats "
                "WHERE destin_name IS NOT NULL" + dt_filter_sql
            )
            cursor.execute(convert_query_placeholders(sql_names), tuple(dt_params))
        names = [r["destin_name"] for r in cursor.fetchall() if r["destin_name"]]
        names.sort()

        series = []
        for nm in names:
            if filter_date:
                anchor_d = datetime.strptime(filter_date, '%Y-%m-%d').date()
                rh_clause, rh_binds = _record_date_hour_window_sql_binds(window_mode, anchor_d)
                sql_data = (
                    f"SELECT record_date, record_hour, waybill_cnt, package_cnt "
                    f"FROM gofo_center_collect_stats "
                    f"WHERE destin_name = ? AND ({rh_clause})" + dt_filter_sql +
                    " ORDER BY record_date ASC, record_hour ASC"
                )
                cursor.execute(
                    convert_query_placeholders(sql_data),
                    tuple([nm] + list(rh_binds) + dt_params),
                )
            else:
                sql_data = (
                    "SELECT record_date, record_hour, waybill_cnt, package_cnt "
                    "FROM gofo_center_collect_stats "
                    "WHERE destin_name = ?" + dt_filter_sql +
                    " ORDER BY record_date ASC, record_hour ASC"
                )
                cursor.execute(
                    convert_query_placeholders(sql_data),
                    tuple([nm] + dt_params),
                )
            rows = cursor.fetchall()
            dmap: dict = {}
            for r in rows:
                k = _cc_time_point_key(r["record_date"], r["record_hour"])
                dmap[k] = dmap.get(k, 0) + int(r["waybill_cnt"] or 0)
            data_points = [dmap.get(tp, 0) for tp in time_points]
            if sum(data_points) == 0:
                continue  # 不画全 0 的系列
            series.append({"label": nm, "data": data_points})

        # 最大系列（按总量降序），便于图例顺序稳定
        series.sort(key=lambda s: -sum(s["data"]))

        conn.close()
        return jsonify({
            "success": True,
            "dates": labels,
            "series": series,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"success": False, "error": str(e)}), 500


@app.route('/api/admin/center_collect/sync', methods=['POST'])
def api_admin_center_collect_sync():
    """手动触发「集包运单数（目的中心）」同步。

    Body (JSON, 全部可选):
      date: YYYY-MM-DD  指定日期（默认 LA 时区今天）
      hour: 0..23       只抓该小时；省略则抓该日所有已完成的整点
    """
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': '未登录'}), 401
    conn = get_db()
    cursor = conn.cursor()
    if not require_admin(cursor):
        conn.close()
        return jsonify({'success': False, 'error': '无权访问'}), 403
    conn.close()

    body = request.get_json(silent=True) or {}
    la_tz = pytz.timezone("America/Los_Angeles")
    date_str = (body.get('date') or '').strip() or datetime.now(la_tz).strftime('%Y-%m-%d')
    hour_raw = body.get('hour')
    try:
        import sync_center_collect
        if hour_raw is None or hour_raw == '':
            result = sync_center_collect.fetch_center_collect_day(date_str)
        else:
            try:
                h = int(hour_raw)
            except (TypeError, ValueError):
                return jsonify({'success': False, 'error': 'hour 必须是 0~23 的整数'}), 400
            result = sync_center_collect.fetch_center_collect_hour(date_str, h)
        if result.get('success'):
            return jsonify({
                'success': True,
                'count': int(result.get('stored_rows') or 0),
                'message': (
                    f"{date_str}"
                    + (f" {h:02d}:00" if hour_raw not in (None, '') else "")
                    + f" 已入库 {result.get('stored_rows')} 行"
                ),
                'detail': result,
            })
        return jsonify({'success': False, 'error': result.get('error') or '同步失败', 'detail': result}), 500
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/center_collect/backfill', methods=['POST'])
def api_admin_center_collect_backfill():
    """回补最近 N 天（默认 7，上限 93）。每次按小时抓取，中心与站点写入同表 gofo_center_collect_stats。"""
    if 'user_id' not in session:
        return jsonify({'success': False, 'error': '未登录'}), 401
    conn = get_db()
    cursor = conn.cursor()
    if not require_admin(cursor):
        conn.close()
        return jsonify({'success': False, 'error': '无权访问'}), 403
    conn.close()

    body = request.get_json(silent=True) or {}
    try:
        days = int(body.get('days') or 7)
    except (TypeError, ValueError):
        days = 7
    days = max(1, min(days, 93))
    try:
        import sync_center_collect
        result = sync_center_collect.fetch_center_collect_backfill(days=days)
        return jsonify({
            'success': bool(result.get('success')),
            'days': days,
            'count': int(result.get('total_stored_rows') or 0),
            'message': f"回补 {days} 天完成，入库 {result.get('total_stored_rows')} 行",
            'detail': result,
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


print("[模块加载] 开始初始化应用...")
try:
    initialize_app()
except Exception as e:
    print(f"[模块加载] 警告: 初始化失败 - {e}")
    import traceback
    traceback.print_exc()

if __name__ == '__main__':
    # init_db() # initialize_app already calls init_db
    
    # 从环境变量获取主机和端口配置
    host = os.environ.get('HOST', HOST)
    port = int(os.environ.get('PORT', PORT))
    
    print(f"Starting server on {host}:{port}")
    app.run(debug=True, host=host, port=port)