import sqlite3
from flask import Flask, request, jsonify, send_file
import os
import sys
from datetime import datetime, timedelta
import pytz
import threading
import time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import json

app = Flask(__name__)

# 获取正确的数据库路径
def get_db_path():
    # 如果是打包后的exe环境，数据库在同级目录下
    if getattr(sys, 'frozen', False):
        # 打包后的exe环境
        return os.path.join(os.path.dirname(sys.executable), 'inbound.db')
    else:
        # 开发环境
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'inbound.db')

DB_PATH = get_db_path()

# 获取正确的静态文件目录
def get_static_dir():
    if getattr(sys, 'frozen', False):
        # 打包后的exe环境 - 静态文件被打包到exe中，需要使用特殊方法访问
        # PyInstaller会将数据文件放在_sys_meipass目录中
        if hasattr(sys, '_MEIPASS'):
            return os.path.join(sys._MEIPASS, 'static')
        else:
            # 备用方案
            return os.path.join(os.path.dirname(sys.executable), 'static')
    else:
        # 开发环境 - 使用脚本所在目录的static子目录
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static')

# 定义洛杉矶时区
LA_TZ = pytz.timezone('America/Los_Angeles')

def init_db():
    need = not os.path.exists(DB_PATH)
    conn = sqlite3.connect(DB_PATH)
    if need:
        conn.execute("""CREATE TABLE inbound_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            dock_no INTEGER,
            vehicle_type TEXT,
            vehicle_no TEXT,
            unit TEXT,
            load_amount INTEGER,
            pieces INTEGER,
            time_slot TEXT,
            shift_type TEXT,
            remark TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
        
        # 创建分拣记录表
        conn.execute("""CREATE TABLE sorting_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            sorting_time DATETIME,
            pieces INTEGER,
            remark TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
        
        # 创建操作日志表
        conn.execute("""CREATE TABLE operation_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            operation_type TEXT,  -- 'edit' 或 'delete'
            table_name TEXT,      -- 'inbound_records' 或 'sorting_records'
            record_id INTEGER,    -- 被操作记录的ID
            old_data TEXT,        -- 修改前的数据（JSON格式）
            new_data TEXT,        -- 修改后的数据（JSON格式）
            operator TEXT,        -- 操作人（可选）
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );""")
    conn.commit()
    conn.close()

def convert_utc_to_la(utc_time_str):
    """直接返回时间字符串，因为数据库中存储的已经是洛杉矶时间"""
    return utc_time_str

def daily_reset_check():
    """每日重置检查函数"""
    while True:
        try:
            # 获取洛杉矶当前时间
            la_now = datetime.now(LA_TZ)
            
            # 如果是午夜（0点）附近，执行重置
            if la_now.hour == 0 and la_now.minute == 0:
                print(f"执行每日重置: {la_now}")
                perform_daily_reset()
                
                # 等待一分钟，避免重复执行
                time.sleep(60)
            else:
                # 每分钟检查一次
                time.sleep(60)
        except Exception as e:
            print(f"每日重置检查出错: {e}")
            time.sleep(60)

def perform_daily_reset():
    """执行每日重置 - 仅记录日志，不删除历史数据"""
    try:
        print("每日重置检查完成 - 历史数据已永久保存")
    except Exception as e:
        print(f"每日重置执行出错: {e}")

@app.route('/')
def index():
    # 直接读取并返回HTML文件内容
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'index.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404

@app.route('/sorting')
def sorting():
    # 返回分拣录入页面
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'sorting.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404

@app.route('/history')
def history():
    # 返回历史记录查询页面
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'history.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404

@app.route('/statistics')
def statistics():
    # 返回统计页面
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'statistics.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404

@app.route('/logs')
def logs_page():
    # 返回操作日志查询页面
    static_dir = get_static_dir()
    file_path = os.path.join(static_dir, 'logs.html')
    if os.path.exists(file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        return content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    else:
        return f"File not found: {file_path}", 404

@app.route('/api/record', methods=['POST'])
def record():
    data = request.json
    vt = data.get("vehicle_type","")
    if vt=="26英尺":
        data["unit"]="托盘"
        data["load_amount"]=12
        data["pieces"]=12*344
    elif vt=="Car":
        data["unit"]="篮筐"
        data["load_amount"]=1
        data["pieces"]=1*172
    elif vt=="Van":
        data["unit"]="篮筐"
        data["load_amount"]=9
        data["pieces"]=9*172
    elif vt=="53英尺":
        data.setdefault("unit","托盘")
        # 当车辆类型为53英尺时，如果输入了装载量，则自动计算件数
        load_amount = data.get("load_amount", 0)
        if load_amount and load_amount > 0:
            data["pieces"] = load_amount * 344
        # 如果已有件数但没有装载量，也可以反向计算装载量
        elif data.get("pieces") and data["pieces"] > 0:
            data["load_amount"] = data["pieces"] // 344

    conn=sqlite3.connect(DB_PATH)
    # 获取当前洛杉矶时间
    la_tz = pytz.timezone('America/Los_Angeles')
    current_la_time = datetime.now(la_tz)
    current_la_time_str = current_la_time.strftime('%Y-%m-%d %H:%M:%S')
    
    # 自动判断班次类型：17点之前是早班，17点之后是晚班
    if current_la_time.hour < 17:
        shift_type = "早班"
    else:
        shift_type = "晚班"
    
    conn.execute("""INSERT INTO inbound_records
        (dock_no, vehicle_type, vehicle_no, unit, load_amount, pieces, time_slot, shift_type, remark, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (data.get("dock_no"), data.get("vehicle_type"), data.get("vehicle_no"),
         data.get("unit"), data.get("load_amount"), data.get("pieces"),
         data.get("time_slot"), shift_type, data.get("remark"), current_la_time_str))
    conn.commit()
    conn.close()
    return jsonify({"success":True})

@app.route('/api/record/<int:record_id>', methods=['PUT'])
def update_record(record_id):
    data = request.json
    
    # 获取当前洛杉矶时间并自动判断班次类型
    la_tz = pytz.timezone('America/Los_Angeles')
    current_la_time = datetime.now(la_tz)
    
    # 自动判断班次类型：17点之前是早班，17点之后是晚班
    if current_la_time.hour < 17:
        shift_type = "早班"
    else:
        shift_type = "晚班"
    
    # 对于53英尺车辆，如果输入了装载量，则自动计算件数
    vt = data.get("vehicle_type", "")
    if vt == "53英尺":
        data.setdefault("unit", "托盘")
        load_amount = data.get("load_amount", 0)
        if load_amount and load_amount > 0:
            data["pieces"] = load_amount * 344
        # 如果已有件数但没有装载量，也可以反向计算装载量
        elif data.get("pieces") and data["pieces"] > 0:
            data["load_amount"] = data["pieces"] // 344
    
    conn = None
    try:
        conn = sqlite3.connect(DB_PATH)
        
        # 获取修改前的数据
        old_record_cur = conn.execute("SELECT * FROM inbound_records WHERE id=?", (record_id,))
        old_record = old_record_cur.fetchone()
        
        cursor = conn.execute("""UPDATE inbound_records SET
            dock_no=?, vehicle_type=?, vehicle_no=?, unit=?, load_amount=?, pieces=?, time_slot=?, shift_type=?, remark=?
            WHERE id=?""",
            (data.get("dock_no"), data.get("vehicle_type"), data.get("vehicle_no"),
             data.get("unit"), data.get("load_amount"), data.get("pieces"),
             data.get("time_slot"), shift_type, data.get("remark"), record_id))
        
        # 如果记录被成功更新，记录日志
        if cursor.rowcount > 0:
            # 获取修改后的数据
            new_record_cur = conn.execute("SELECT * FROM inbound_records WHERE id=?", (record_id,))
            new_record = new_record_cur.fetchone()
            
            # 记录操作日志
            column_names = [description[0] for description in old_record_cur.description]
            old_data = dict(zip(column_names, old_record)) if old_record else {}
            new_data = dict(zip(column_names, new_record)) if new_record else {}
            
            # 删除游标对象，避免序列化错误
            if 'cursor' in old_data:
                del old_data['cursor']
            if 'cursor' in new_data:
                del new_data['cursor']
            
            conn.execute("""INSERT INTO operation_logs 
                (operation_type, table_name, record_id, old_data, new_data)
                VALUES (?, ?, ?, ?, ?)""",
                ('edit', 'inbound_records', record_id, 
                 json.dumps(old_data, default=str), 
                 json.dumps(new_data, default=str)))
            
            conn.commit()
            if conn:
                conn.close()
            return jsonify({"success": True})
        else:
            conn.commit()
            if conn:
                conn.close()
            return jsonify({"success": False, "error": "记录未找到"}), 404
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass

@app.route('/api/record/<int:record_id>', methods=['DELETE'])
def delete_record(record_id):
    conn = None
    try:
        conn = sqlite3.connect(DB_PATH)
        
        # 获取删除前的数据
        old_record_cur = conn.execute("SELECT * FROM inbound_records WHERE id=?", (record_id,))
        old_record = old_record_cur.fetchone()
        
        cursor = conn.execute("DELETE FROM inbound_records WHERE id=?", (record_id,))
        
        # 如果记录被成功删除，记录日志
        if cursor.rowcount > 0:
            # 记录操作日志
            column_names = [description[0] for description in old_record_cur.description]
            old_data = dict(zip(column_names, old_record)) if old_record else {}
            
            # 删除游标对象，避免序列化错误
            if 'cursor' in old_data:
                del old_data['cursor']
            
            conn.execute("""INSERT INTO operation_logs 
                (operation_type, table_name, record_id, old_data, new_data)
                VALUES (?, ?, ?, ?, ?)""",
                ('delete', 'inbound_records', record_id, 
                 json.dumps(old_data, default=str), 
                 json.dumps({})))  # 删除操作没有新数据
            
            conn.commit()
            if conn:
                conn.close()
            return jsonify({"success": True})
        else:
            conn.commit()
            if conn:
                conn.close()
            return jsonify({"success": False, "error": "记录未找到"}), 404
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass

@app.route('/api/list')
def list_data():
    conn=sqlite3.connect(DB_PATH)
    
    # 获取洛杉矶当前日期
    la_tz = pytz.timezone('America/Los_Angeles')
    la_today = datetime.now(la_tz).date()
    
    # 计算次日日期
    next_date = la_today + timedelta(days=1)
    
    # 构建日期范围查询条件（使用洛杉矶时区时间进行计算）
    # 当天05:00:00的时间（洛杉矶时间）
    today_5am_la = la_tz.localize(datetime.combine(la_today, datetime.min.time().replace(hour=5)))
    
    # 次日05:00:00的时间（洛杉矶时间，用于上限）
    next_5am_la = la_tz.localize(datetime.combine(next_date, datetime.min.time().replace(hour=5)))
    
    # 转换为UTC时间用于数据库查询
    today_5am_utc = today_5am_la.astimezone(pytz.utc)
    next_5am_utc = next_5am_la.astimezone(pytz.utc)
    
    # 查询属于当前业务日的记录（查询当天05:00之后到次日05:00之前的所有记录）
    cur=conn.execute("""
        SELECT id, dock_no, vehicle_type, vehicle_no, unit, load_amount,
               pieces, time_slot, shift_type, remark, created_at
        FROM inbound_records 
        WHERE 
            created_at >= ? AND created_at < ?
        ORDER BY created_at DESC""", (
            today_5am_utc.strftime('%Y-%m-%d %H:%M:%S'), 
            next_5am_utc.strftime('%Y-%m-%d %H:%M:%S')
        ))
    rows=[{
        "id":r[0], "dock_no":r[1], "vehicle_type":r[2], "vehicle_no":r[3],
        "unit":r[4], "load_amount":r[5], "pieces":r[6],
        "time_slot":r[7], "shift_type":r[8], "remark":r[9],
        "created_at":convert_utc_to_la(r[10])  # 转换为洛杉矶时间
    } for r in cur.fetchall()]
    conn.close()
    return jsonify(rows)

# 新增API：获取按时间段分组的入库数据

@app.route('/api/inbound_hourly')
def inbound_hourly_data():
    # 获取日期参数，默认为今天
    date_str = request.args.get('date')
    
    conn=sqlite3.connect(DB_PATH)
    
    # 获取洛杉矶当前日期
    la_tz = pytz.timezone('America/Los_Angeles')
    
    if date_str:
        # 如果提供了日期参数，使用指定日期
        try:
            request_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            conn.close()
            return jsonify({"error": "日期格式无效，请使用YYYY-MM-DD格式"}), 400
    else:
        # 如果没有提供日期参数，使用今天
        request_date = datetime.now(la_tz).date()
    
    # 计算次日日期
    next_date = request_date + timedelta(days=1)
    
    # 构建日期范围查询条件（使用洛杉矶时区时间进行计算）
    # 当天05:00:00的时间（洛杉矶时间）
    request_date_5am_la = la_tz.localize(datetime.combine(request_date, datetime.min.time().replace(hour=5)))
    
    # 次日05:00:00的时间（洛杉矶时间，用于上限）
    next_date_5am_la = la_tz.localize(datetime.combine(next_date, datetime.min.time().replace(hour=5)))
    
    # 转换为UTC时间用于数据库查询
    request_date_5am_utc = request_date_5am_la.astimezone(pytz.utc)
    next_date_5am_utc = next_date_5am_la.astimezone(pytz.utc)
    
    # 查询入库记录，按时间段分组（查询当天05:00之后到次日05:00之前的所有记录）
    cur=conn.execute("""
        SELECT time_slot, SUM(pieces) as total_pieces
        FROM inbound_records 
        WHERE 
            created_at >= ? AND created_at < ? AND time_slot IS NOT NULL
        GROUP BY time_slot
        ORDER BY time_slot""", (
            request_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S'), 
            next_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S')
        ))
    rows=[{
        "time_slot": r[0],
        "total_pieces": r[1] if r[1] else 0
    } for r in cur.fetchall()]
    return jsonify(rows)

# 新增API：获取按时间段分组的分拣数据
@app.route('/api/sorting_hourly')
def sorting_hourly_data():
    # 获取日期参数，默认为今天
    date_str = request.args.get('date')
    
    conn=sqlite3.connect(DB_PATH)
    
    # 获取洛杉矶当前日期
    la_tz = pytz.timezone('America/Los_Angeles')
    
    if date_str:
        # 如果提供了日期参数，使用指定日期
        try:
            request_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            conn.close()
            return jsonify({"error": "日期格式无效，请使用YYYY-MM-DD格式"}), 400
    else:
        # 如果没有提供日期参数，使用今天
        request_date = datetime.now(la_tz).date()
    
    # 计算次日日期
    next_date = request_date + timedelta(days=1)
    
    # 构建日期范围查询条件（使用洛杉矶时区时间进行计算）
    # 当天05:00:00的时间（洛杉矶时间）
    request_date_5am_la = la_tz.localize(datetime.combine(request_date, datetime.min.time().replace(hour=5)))
    
    # 次日05:00:00的时间（洛杉矶时间，用于上限）
    next_date_5am_la = la_tz.localize(datetime.combine(next_date, datetime.min.time().replace(hour=5)))
    
    # 转换为UTC时间用于数据库查询
    request_date_5am_utc = request_date_5am_la.astimezone(pytz.utc)
    next_date_5am_utc = next_date_5am_la.astimezone(pytz.utc)
    
    # 查询分拣记录，按时间段分组（按照业务日逻辑查询，查询当天05:00之后到次日05:00之前的所有记录）
    cur=conn.execute("""SELECT time_slot, SUM(pieces) as total_pieces
                        FROM sorting_records 
                        WHERE 
                            created_at >= ? AND created_at < ? AND time_slot IS NOT NULL
                        GROUP BY time_slot
                        ORDER BY time_slot""", (
                            request_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S'), 
                            next_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S')
                        ))
    rows=[{
        "time_slot": r[0],
        "total_pieces": r[1] if r[1] else 0
    } for r in cur.fetchall()]
    return jsonify(rows)

@app.route('/api/history')
def get_history():
    date_str = request.args.get('date')
    
    if not date_str:
        return jsonify({"error": "请提供日期参数"}), 400
    
    conn = sqlite3.connect(DB_PATH)
    
    try:
        # 解析请求的日期
        request_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # 获取洛杉矶当前日期
        la_tz = pytz.timezone('America/Los_Angeles')
        la_today = datetime.now(la_tz).date()
        
        # 检查是否选择了未来日期（明天或以后）
        if request_date > la_today:
            conn.close()
            return jsonify({
                "error": "不能查询未来日期的数据",
                "records": [],
                "sorting_records": [],
                "stats": {
                    "total_vehicles": 0,
                    "total_pieces": 0,
                    "record_count": 0,
                    "vehicle_stats": []
                }
            }), 400
        
        # 计算次日日期
        next_date = request_date + timedelta(days=1)
        
        # 构建日期范围查询条件（使用洛杉矶时区时间进行计算）
        # 当天05:00:00的时间（洛杉矶时间）
        request_date_5am_la = la_tz.localize(datetime.combine(request_date, datetime.min.time().replace(hour=5)))
        
        # 次日05:00:00的时间（洛杉矶时间，用于上限）
        next_date_5am_la = la_tz.localize(datetime.combine(next_date, datetime.min.time().replace(hour=5)))
        
        # 转换为UTC时间用于数据库查询
        request_date_5am_utc = request_date_5am_la.astimezone(pytz.utc)
        next_date_5am_utc = next_date_5am_la.astimezone(pytz.utc)
        
        # 查询指定日期的入库记录（查询当天05:00之后到次日05:00之前的所有记录）
        inbound_query = """
            SELECT id, dock_no, vehicle_type, vehicle_no, unit, load_amount,
                   pieces, time_slot, shift_type, remark, created_at
            FROM inbound_records 
            WHERE 
                created_at >= ? AND created_at < ?
            ORDER BY created_at DESC
        """
        inbound_cur = conn.execute(inbound_query, (
            request_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S'), 
            next_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S')
        ))
        inbound_rows = [{
            "id": r[0], "dock_no": r[1], "vehicle_type": r[2], "vehicle_no": r[3],
            "unit": r[4], "load_amount": r[5], "pieces": r[6],
            "time_slot": r[7], "shift_type": r[8], "remark": r[9],
            "created_at": convert_utc_to_la(r[10])  # 转换为洛杉矶时间
        } for r in inbound_cur.fetchall()]
        
        # 查询指定日期的分拣记录（按照业务日逻辑查询，查询当天05:00之后到次日05:00之前的所有记录）
        sorting_query = """
            SELECT id, sorting_time, pieces, remark, created_at, time_slot
            FROM sorting_records 
            WHERE 
                created_at >= ? AND created_at < ?
            ORDER BY created_at DESC
        """
        sorting_cur = conn.execute(sorting_query, (
            request_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S'), 
            next_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S')
        ))
        sorting_rows = [{
            "id": r[0], "sorting_time": r[1], "pieces": r[2], "remark": r[3],
            "created_at": convert_utc_to_la(r[4]), "time_slot": r[5]
        } for r in sorting_cur.fetchall()]
        
        # 计算统计信息
        total_vehicles = len(inbound_rows)
        total_pieces = sum(record.get("pieces", 0) for record in inbound_rows)
        record_count = len(inbound_rows) + len(sorting_rows)
        
        # 各车型统计
        vehicle_stats = {}
        for record in inbound_rows:
            vehicle_type = record.get("vehicle_type", "未知")
            if vehicle_type not in vehicle_stats:
                vehicle_stats[vehicle_type] = {"count": 0, "pieces": 0}
            vehicle_stats[vehicle_type]["count"] += 1
            vehicle_stats[vehicle_type]["pieces"] += record.get("pieces", 0)
        
        # 转换为列表格式
        vehicle_stats_list = [
            {
                "vehicle_type": vt,
                "count": stats["count"],
                "total_pieces": stats["pieces"]
            }
            for vt, stats in vehicle_stats.items()
        ]
        
        conn.close()
        
        return jsonify({
            "records": inbound_rows,
            "sorting_records": sorting_rows,
            "stats": {
                "total_vehicles": total_vehicles,
                "total_pieces": total_pieces,
                "record_count": record_count,
                "vehicle_stats": vehicle_stats_list
            }
        })
    except Exception as e:
        conn.close()
        return jsonify({"error": f"处理历史记录查询时出错: {str(e)}"}), 500

@app.route('/api/logs')
def get_operation_logs():
    """获取操作日志列表"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 查询所有操作日志，按时间倒序排列
        cursor.execute("""
            SELECT id, operation_type, table_name, record_id, old_data, new_data, created_at 
            FROM operation_logs 
            ORDER BY created_at DESC
        """)
        
        logs = cursor.fetchall()
        result = []
        
        for log in logs:
            log_id, operation_type, table_name, record_id, old_data, new_data, created_at = log
            
            # 解析JSON数据
            try:
                old_data_parsed = json.loads(old_data) if old_data else {}
            except:
                old_data_parsed = {"raw_data": old_data}
                
            try:
                new_data_parsed = json.loads(new_data) if new_data else {}
            except:
                new_data_parsed = {"raw_data": new_data}
            
            result.append({
                "id": log_id,
                "operation_type": operation_type,
                "table_name": table_name,
                "record_id": record_id,
                "old_data": old_data_parsed,
                "new_data": new_data_parsed,
                "created_at": created_at
            })
        
        conn.close()
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/logs/<int:log_id>')
def get_operation_log_detail(log_id):
    """获取单个操作日志详情"""
    try:
        conn = sqlite3.connect(DB_PATH)
        cursor = conn.cursor()
        
        # 查询指定ID的操作日志
        cursor.execute("""
            SELECT id, operation_type, table_name, record_id, old_data, new_data, created_at 
            FROM operation_logs 
            WHERE id = ?
        """, (log_id,))
        
        log = cursor.fetchone()
        if not log:
            conn.close()
            return jsonify({"error": "日志未找到"}), 404
        
        log_id, operation_type, table_name, record_id, old_data, new_data, created_at = log
        
        # 解析JSON数据
        try:
            old_data_parsed = json.loads(old_data) if old_data else {}
        except:
            old_data_parsed = {"raw_data": old_data}
            
        try:
            new_data_parsed = json.loads(new_data) if new_data else {}
        except:
            new_data_parsed = {"raw_data": new_data}
        
        result = {
            "id": log_id,
            "operation_type": operation_type,
            "table_name": table_name,
            "record_id": record_id,
            "old_data": old_data_parsed,
            "new_data": new_data_parsed,
            "created_at": created_at
        }
        
        conn.close()
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/api/sorting', methods=['POST'])
def add_sorting_record():
    data = request.json
    
    conn=sqlite3.connect(DB_PATH)
    
    # 获取当前洛杉矶时间
    la_tz = pytz.timezone('America/Los_Angeles')
    current_la_time = datetime.now(la_tz)
    current_la_time_str = current_la_time.strftime('%Y-%m-%d %H:%M:%S')
    
    conn.execute("""INSERT INTO sorting_records
        (sorting_time, pieces, remark, time_slot, created_at)
        VALUES (?, ?, ?, ?, ?)""",
        (data.get("sorting_time"), data.get("pieces"), data.get("remark"), data.get("time_slot"), current_la_time_str))
    conn.commit()
    conn.close()
    return jsonify({"success":True})

@app.route('/api/sorting', methods=['GET'])
def get_sorting_records():
    conn=sqlite3.connect(DB_PATH)
    cur=conn.execute("""SELECT id, sorting_time, pieces, remark, created_at, time_slot
                        FROM sorting_records ORDER BY created_at DESC""")
    rows=[{
        "id":r[0], "sorting_time":r[1], "pieces":r[2], "remark":r[3],
        "created_at":convert_utc_to_la(r[4]), "time_slot":r[5]
    } for r in cur.fetchall()]
    return jsonify(rows)

@app.route('/api/sorting/<int:record_id>', methods=['DELETE'])
def delete_sorting_record(record_id):
    conn = None
    try:
        conn = sqlite3.connect(DB_PATH)
        
        # 获取删除前的数据
        old_record_cur = conn.execute("SELECT * FROM sorting_records WHERE id=?", (record_id,))
        old_record = old_record_cur.fetchone()
        
        cursor = conn.execute("DELETE FROM sorting_records WHERE id=?", (record_id,))
        
        # 如果记录被成功删除，记录日志
        if cursor.rowcount > 0:
            # 记录操作日志
            column_names = [description[0] for description in old_record_cur.description]
            old_data = dict(zip(column_names, old_record)) if old_record else {}
            
            # 删除游标对象，避免序列化错误
            old_data.pop('cursor', None)
            
            conn.execute("""INSERT INTO operation_logs 
                (operation_type, table_name, record_id, old_data, new_data)
                VALUES (?, ?, ?, ?, ?)""",
                ('delete', 'sorting_records', record_id, 
                 json.dumps(old_data, default=str), 
                 json.dumps({})))  # 删除操作没有新数据
        
            conn.commit()
            conn.close()
            return jsonify({"success": True})
        else:
            conn.commit()
            conn.close()
            return jsonify({"success": False, "error": "记录未找到"}), 404
    except Exception as e:
        if conn:
            conn.rollback()
            conn.close()
        return jsonify({"success": False, "error": str(e)}), 500
    finally:
        if conn:
            try:
                conn.close()
            except:
                pass

@app.route('/api/stats')
def get_statistics():
    # 获取日期参数，默认为今天
    date_str = request.args.get('date')
    
    conn=sqlite3.connect(DB_PATH)
    
    # 获取洛杉矶当前日期
    la_tz = pytz.timezone('America/Los_Angeles')
    
    if date_str:
        # 如果提供了日期参数，使用指定日期
        try:
            request_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            conn.close()
            return jsonify({"error": "日期格式无效，请使用YYYY-MM-DD格式"}), 400
    else:
        # 如果没有提供日期参数，使用今天
        request_date = datetime.now(la_tz).date()
    
    # 计算次日日期
    next_date = request_date + timedelta(days=1)
    
    # 构建日期范围查询条件（使用洛杉矶时区时间进行计算）
    # 当天05:00:00的时间（洛杉矶时间）
    request_date_5am_la = la_tz.localize(datetime.combine(request_date, datetime.min.time().replace(hour=5)))
    
    # 次日05:00:00的时间（洛杉矶时间，用于上限）
    next_date_5am_la = la_tz.localize(datetime.combine(next_date, datetime.min.time().replace(hour=5)))
    
    # 转换为UTC时间用于数据库查询
    request_date_5am_utc = request_date_5am_la.astimezone(pytz.utc)
    next_date_5am_utc = next_date_5am_la.astimezone(pytz.utc)
    
    # 查询属于指定业务日的记录（查询当天05:00之后到次日05:00之前的所有记录）
    records_query = """
        SELECT id, created_at, vehicle_type, time_slot FROM inbound_records 
        WHERE 
            created_at >= ? AND created_at < ?
    """
    records_cur = conn.execute(records_query, (
        request_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S'), 
        next_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S')
    ))
    records = records_cur.fetchall()
    
    # 总车次和总货物量（查询当天05:00之后到次日05:00之前的所有记录）
    total_query = """
        SELECT COUNT(*) as total_vehicles, SUM(pieces) as total_pieces 
        FROM inbound_records 
        WHERE 
            created_at >= ? AND created_at < ?
    """
    total_cur = conn.execute(total_query, (
        request_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S'), 
        next_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S')
    ))
    total_result = total_cur.fetchone()
    total_vehicles = total_result[0] if total_result[0] else 0
    total_pieces = int(total_result[1]) if total_result[1] else 0
    
    # 各车型统计（查询当天05:00之后到次日05:00之前的所有记录）
    vehicle_stats_query = """
        SELECT vehicle_type, COUNT(*) as count, SUM(pieces) as total_pieces 
        FROM inbound_records 
        WHERE 
            created_at >= ? AND created_at < ?
        GROUP BY vehicle_type
    """
    vehicle_stats_cur = conn.execute(vehicle_stats_query, (
        request_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S'), 
        next_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S')
    ))
    vehicle_stats = [{
        "vehicle_type": r[0],
        "count": r[1],
        "total_pieces": int(r[2]) if r[2] else 0
    } for r in vehicle_stats_cur.fetchall()]
    
    # 初始化统计变量
    vehicles_19_to_20 = 0
    # 统计19:00-20:00时间段各车型到车数量
    vehicles_19_to_20_by_type = {}
    
    # 统计20:00-21:00时间段记录
    vehicles_20_to_21 = 0
    # 统计20:00-21:00时间段各车型到车数量
    vehicles_20_to_21_by_type = {}
    
    # 统计超过24:00的记录（即次日00:00之后的记录）
    vehicles_after_24 = 0
    
    # 处理每条记录，用于统计特定时间段的数据（按录入的时间段time_slot统计）
    for record in records:
        record_id, created_at_str, vehicle_type, time_slot = record
        # 基于录入的时间段进行统计
        if time_slot:
            try:
                time_slot_int = int(time_slot)
                # 检查是否在19:00-20:00之间（录入的时间段为19）
                if time_slot_int == 19:
                    vehicles_19_to_20 += 1
                    
                    # 统计19:00-20:00时间段各车型到车数量
                    if vehicle_type not in vehicles_19_to_20_by_type:
                        vehicles_19_to_20_by_type[vehicle_type] = 0
                    vehicles_19_to_20_by_type[vehicle_type] += 1
                
                # 检查是否在20:00-21:00之间（录入的时间段为20）
                if time_slot_int == 20:
                    vehicles_20_to_21 += 1
                    
                    # 统计20:00-21:00时间段各车型到车数量
                    if vehicle_type not in vehicles_20_to_21_by_type:
                        vehicles_20_to_21_by_type[vehicle_type] = 0
                    vehicles_20_to_21_by_type[vehicle_type] += 1
                
                # 检查是否是超过24:00的记录（录入的时间段为24或更大）
                if time_slot_int >= 24:
                    vehicles_after_24 += 1
            except ValueError:
                # 如果time_slot不是有效的整数，跳过这条记录
                pass
        else:
            # 如果没有录入时间段，仍然使用原来基于创建时间的逻辑作为后备
            if created_at_str:
                try:
                    # 将UTC时间字符串转换为datetime对象
                    utc_time = datetime.strptime(created_at_str, '%Y-%m-%d %H:%M:%S')
                    utc_time = pytz.utc.localize(utc_time)
                    # 转换为洛杉矶时间
                    la_time = utc_time.astimezone(LA_TZ)
                    
                    # 检查是否在19:00-20:00之间（洛杉矶时间）
                    if la_time.hour == 19:
                        vehicles_19_to_20 += 1
                        
                        # 统计19:00-20:00时间段各车型到车数量
                        if vehicle_type not in vehicles_19_to_20_by_type:
                            vehicles_19_to_20_by_type[vehicle_type] = 0
                        vehicles_19_to_20_by_type[vehicle_type] += 1
                    
                    # 检查是否在20:00-21:00之间（洛杉矶时间）
                    if la_time.hour == 20:
                        vehicles_20_to_21 += 1
                        
                        # 统计20:00-21:00时间段各车型到车数量
                        if vehicle_type not in vehicles_20_to_21_by_type:
                            vehicles_20_to_21_by_type[vehicle_type] = 0
                        vehicles_20_to_21_by_type[vehicle_type] += 1
                    
                    # 检查是否是明天的记录（即次日00:00之后的记录）
                    if la_time.date() > request_date:
                        vehicles_after_24 += 1
                except Exception as e:
                    print(f"处理记录 {record_id} 的时间时出错: {e}")
    
    conn.close()
    
    return jsonify({
        "total_vehicles": total_vehicles,
        "total_pieces": total_pieces,
        "vehicle_stats": vehicle_stats,
        "vehicles_19_to_20": vehicles_19_to_20,
        "vehicles_19_to_20_by_type": vehicles_19_to_20_by_type,
        "vehicles_20_to_21": vehicles_20_to_21,
        "vehicles_20_to_21_by_type": vehicles_20_to_21_by_type,
        "vehicles_after_24": vehicles_after_24
    })

@app.route('/api/export_excel')
def export_excel():
    try:
        # 获取查询日期参数，如果没有则使用当天
        date_str = request.args.get('date')
        if not date_str:
            # 获取洛杉矶当前日期
            la_tz = pytz.timezone('America/Los_Angeles')
            date_str = datetime.now(la_tz).strftime('%Y-%m-%d')
        
        conn = sqlite3.connect(DB_PATH)
        
        # 解析请求的日期
        request_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        # 计算次日日期
        next_date = request_date + timedelta(days=1)
        
        # 构建日期范围查询条件（使用洛杉矶时区时间进行计算）
        # 当天05:00:00的时间（洛杉矶时间）
        request_date_5am_la = la_tz.localize(datetime.combine(request_date, datetime.min.time().replace(hour=5)))
        
        # 次日05:00:00的时间（洛杉矶时间，用于上限）
        next_date_5am_la = la_tz.localize(datetime.combine(next_date, datetime.min.time().replace(hour=5)))
        
        # 转换为UTC时间用于数据库查询
        request_date_5am_utc = request_date_5am_la.astimezone(pytz.utc)
        next_date_5am_utc = next_date_5am_la.astimezone(pytz.utc)
        
        # 查询指定日期的入库记录（查询当天05:00之后到次日05:00之前的所有记录）
        inbound_query = """
            SELECT id, dock_no, vehicle_type, vehicle_no, unit, load_amount,
                   pieces, time_slot, shift_type, remark, created_at
            FROM inbound_records 
            WHERE 
                created_at >= ? AND created_at < ?
            ORDER BY created_at DESC
        """
        inbound_cur = conn.execute(inbound_query, (
            request_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S'), 
            next_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S')
        ))
        inbound_rows = [{
            "id": r[0], "dock_no": r[1], "vehicle_type": r[2], "vehicle_no": r[3],
            "unit": r[4], "load_amount": r[5], "pieces": r[6],
            "time_slot": r[7], "shift_type": r[8], "remark": r[9],
            "created_at": convert_utc_to_la(r[10])  # 转换为洛杉矶时间
        } for r in inbound_cur.fetchall()]
        
        # 查询指定日期的分拣记录（按照业务日逻辑查询，查询当天05:00之后到次日05:00之前的所有记录）
        sorting_query = """
            SELECT id, sorting_time, pieces, remark, created_at, time_slot
            FROM sorting_records 
            WHERE 
                created_at >= ? AND created_at < ?
            ORDER BY created_at DESC
        """
        sorting_cur = conn.execute(sorting_query, (
            request_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S'), 
            next_date_5am_utc.strftime('%Y-%m-%d %H:%M:%S')
        ))
        sorting_rows = [{
            "id": r[0], "sorting_time": r[1], "pieces": r[2], "remark": r[3],
            "created_at": convert_utc_to_la(r[4]), "time_slot": r[5]
        } for r in sorting_cur.fetchall()]
        
        conn.close()
        
        # 创建Excel工作簿
        wb = Workbook()
        
        # 创建入库记录工作表
        ws1 = wb.active
        ws1.title = "入库记录"
        
        # 添加表头
        inbound_headers = ['ID', '码头号', '车辆类型', '车牌号', '单位', '装载量', '件数', '时间段', '班次类型', '备注', '创建时间']
        ws1.append(inbound_headers)
        
        # 设置表头样式
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        for col in range(1, len(inbound_headers) + 1):
            cell = ws1.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # 添加入库数据
        for row in inbound_rows:
            # 转换创建时间为洛杉矶时间
            if row[10]:  # created_at字段
                la_time_str = convert_utc_to_la(row[10])
                modified_row = list(row)
                modified_row[10] = la_time_str
                ws1.append(modified_row)
            else:
                ws1.append(row)
        
        # 调整列宽
        for column in ws1.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws1.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # 创建分拣记录工作表
        ws2 = wb.create_sheet("分拣记录")
        
        # 添加表头
        sorting_headers = ['ID', '分拣日期', '件数', '时间段', '备注', '创建时间']
        ws2.append(sorting_headers)
        
        # 设置表头样式
        for col in range(1, len(sorting_headers) + 1):
            cell = ws2.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # 添加分拣数据
        for row in sorting_rows:
            # 转换创建时间为洛杉矶时间
            if row[4]:  # created_at字段
                la_time_str = convert_utc_to_la(row[4])
                modified_row = list(row)
                modified_row[4] = la_time_str
                ws2.append(modified_row)
            else:
                ws2.append(row)
        
        # 调整列宽
        for column in ws2.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws2.column_dimensions[column_letter].width = min(adjusted_width, 50)
        
        # 保存Excel文件
        filename = f"入库统计数据_{date_str}.xlsx"
        filepath = os.path.join(os.path.dirname(__file__), filename)
        wb.save(filepath)
        
        # 返回Excel文件
        return send_file(filepath, as_attachment=True, download_name=filename)
        
    except Exception as e:
        print(f"导出Excel文件出错: {e}")
        return jsonify({"error": f"导出失败: {str(e)}"}), 500

if __name__ == "__main__":
    init_db()
    
    # 启动每日重置检查线程
    reset_thread = threading.Thread(target=daily_reset_check, daemon=True)
    reset_thread.start()
    
    app.run(host="0.0.0.0", port=8080)