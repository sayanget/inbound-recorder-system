"""
每日数据汇总邮件发送脚本
功能：每天自动汇总入库和分拣数据，并发送邮件到指定邮箱
"""

import sqlite3
import os
import sys
from datetime import datetime, timedelta
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import schedule
import time
import zipfile

# 获取正确的数据库路径
def get_db_path():
    if getattr(sys, 'frozen', False):
        return os.path.join(os.path.dirname(sys.executable), 'inbound.db')
    else:
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), 'inbound.db')

# 数据库路径
DB_PATH = os.environ.get('DATABASE_PATH') or get_db_path()

# 洛杉矶时区
LA_TZ = pytz.timezone('America/Los_Angeles')

# 与 single_app 一致：入库录入件数 → 实到件数（统计口径；装载量不乘系数）
INBOUND_PIECES_ACTUAL_FACTOR = float(os.environ.get("INBOUND_PIECES_ACTUAL_FACTOR", "0.76"))


def _py_inbound_actual_pieces(pieces, excluded_pieces=0):
    try:
        p = int(pieces or 0)
    except (TypeError, ValueError):
        p = 0
    try:
        e = int(excluded_pieces or 0)
    except (TypeError, ValueError):
        e = 0
    net = max(0, p - e)
    return float(net * INBOUND_PIECES_ACTUAL_FACTOR)


def _py_inbound_arrival_pieces(vehicle_type, vehicle_no, pieces, excluded_pieces=0):
    if str(vehicle_type or "").strip() == "53英尺" and str(vehicle_no or "").strip() == "G":
        return 0.0
    return _py_inbound_actual_pieces(pieces, excluded_pieces)


# 导入邮件配置
try:
    from email_config import (
        SMTP_SERVER, SMTP_PORT, SENDER_EMAIL, SENDER_PASSWORD, 
        RECIPIENT_EMAIL, REPORT_TIME, DELETE_TEMP_FILE, EMAIL_SUBJECT_PREFIX
    )
    EMAIL_CONFIG = {
        'smtp_server': SMTP_SERVER,
        'smtp_port': SMTP_PORT,
        'sender_email': SENDER_EMAIL,
        'sender_password': SENDER_PASSWORD,
        'recipient_email': RECIPIENT_EMAIL,
    }
except ImportError:
    print("警告: 未找到email_config.py配置文件，使用默认配置")
    print("请复制email_config.py并填写您的邮箱配置")
    EMAIL_CONFIG = {
        'smtp_server': 'smtp.gmail.com',
        'smtp_port': 587,
        'sender_email': 'your_email@gmail.com',
        'sender_password': 'your_app_password',
        'recipient_email': 'sayanget@yahoo.com',
    }
    REPORT_TIME = "08:00"
    DELETE_TEMP_FILE = False
    EMAIL_SUBJECT_PREFIX = "入库系统每日数据汇总"

def get_daily_summary(target_date=None):
    """
    获取指定日期的数据汇总
    如果不指定日期，则获取昨天的数据
    """
    if target_date is None:
        # 获取昨天的日期
        target_date = (datetime.now(LA_TZ) - timedelta(days=1)).date()
    
    conn = sqlite3.connect(DB_PATH)
    
    # 计算查询时间范围
    start_time = datetime.combine(target_date, datetime.min.time())
    end_time = datetime.combine(target_date, datetime.max.time())
    
    # 查询入库记录
    inbound_cursor = conn.execute("""
        SELECT ir.id, ir.dock_no, ir.vehicle_type, ir.vehicle_no, ir.unit, 
               ir.load_amount, ir.pieces, ir.time_slot, ir.shift_type, 
               ir.remark, ir.created_at, u.username as created_by_username,
               COALESCE(ir.excluded_pieces, 0)
        FROM inbound_records ir
        LEFT JOIN users u ON ir.created_by = u.id
        WHERE ir.created_at >= ? AND ir.created_at <= ?
        ORDER BY ir.created_at ASC
    """, (start_time.strftime('%Y-%m-%d %H:%M:%S'), 
          end_time.strftime('%Y-%m-%d %H:%M:%S')))
    
    inbound_records = inbound_cursor.fetchall()
    
    # 查询分拣记录
    sorting_cursor = conn.execute("""
        SELECT id, sorting_time, pieces, remark, time_slot, created_at
        FROM sorting_records
        WHERE sorting_time >= ? AND sorting_time <= ?
        ORDER BY sorting_time ASC
    """, (start_time.strftime('%Y-%m-%d'), 
          end_time.strftime('%Y-%m-%d')))
    
    sorting_records = sorting_cursor.fetchall()
    
    conn.close()
    
    return {
        'date': target_date,
        'inbound_records': inbound_records,
        'sorting_records': sorting_records
    }

def calculate_statistics(summary_data):
    """
    计算统计数据
    """
    inbound_records = summary_data['inbound_records']
    sorting_records = summary_data['sorting_records']
    
    # 入库统计（实到件数口径）
    total_inbound_pieces = 0.0
    for r in inbound_records:
        total_inbound_pieces += _py_inbound_arrival_pieces(r[2], r[3], r[6], r[12])
    total_inbound_pieces = int(round(total_inbound_pieces))
    total_inbound_vehicles = len(inbound_records)
    
    # 按车辆类型统计
    vehicle_stats = {}
    for record in inbound_records:
        vehicle_type = record[2]  # vehicle_type
        ap = _py_inbound_arrival_pieces(record[2], record[3], record[6], record[12])
        
        if vehicle_type not in vehicle_stats:
            vehicle_stats[vehicle_type] = {'count': 0, 'pieces': 0.0}
        
        vehicle_stats[vehicle_type]['count'] += 1
        vehicle_stats[vehicle_type]['pieces'] += ap
    
    # 分拣统计
    total_sorting_pieces = sum(r[2] for r in sorting_records if r[2])  # pieces
    total_sorting_records = len(sorting_records)
    
    for vt in vehicle_stats:
        vehicle_stats[vt]['pieces'] = int(round(vehicle_stats[vt]['pieces']))

    return {
        'total_inbound_pieces': total_inbound_pieces,
        'total_inbound_vehicles': total_inbound_vehicles,
        'vehicle_stats': vehicle_stats,
        'total_sorting_pieces': total_sorting_pieces,
        'total_sorting_records': total_sorting_records
    }

def create_excel_report(summary_data, stats):
    """
    创建Excel报表
    """
    wb = Workbook()
    
    # 设置样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 创建入库记录工作表
    ws_inbound = wb.active
    ws_inbound.title = "入库记录"
    
    # 入库记录表头（件数：录入 / 实到，与系统统计口径一致）
    inbound_headers = ['序号', '道口号', '车辆类型', '车牌号', '单位', '装载量', 
                       '录入件数', '实到件数', '时间段', '班次', '备注', '录入时间', '录入人']
    ws_inbound.append(inbound_headers)
    
    # 设置表头样式
    for col in range(1, len(inbound_headers) + 1):
        cell = ws_inbound.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 填充入库数据
    for idx, record in enumerate(summary_data['inbound_records'], 1):
        entered = int(record[6] or 0)
        actual = int(round(_py_inbound_arrival_pieces(record[2], record[3], record[6], record[12])))
        row_data = [
            idx,  # 序号
            record[1],  # dock_no
            record[2],  # vehicle_type
            record[3],  # vehicle_no
            record[4],  # unit
            record[5],  # load_amount
            entered,
            actual,
            record[7],  # time_slot
            record[8],  # shift_type
            record[9],  # remark
            record[10],  # created_at
            record[11] or '未知'  # created_by_username
        ]
        ws_inbound.append(row_data)
        
        # 设置边框和对齐
        for col in range(1, len(row_data) + 1):
            cell = ws_inbound.cell(idx + 1, col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整列宽
    for col in range(1, len(inbound_headers) + 1):
        ws_inbound.column_dimensions[get_column_letter(col)].width = 15
    
    # 创建分拣记录工作表
    ws_sorting = wb.create_sheet("分拣记录")
    
    # 分拣记录表头
    sorting_headers = ['序号', '分拣日期', '件数', '时间段', '备注', '录入时间']
    ws_sorting.append(sorting_headers)
    
    # 设置表头样式
    for col in range(1, len(sorting_headers) + 1):
        cell = ws_sorting.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # 填充分拣数据
    for idx, record in enumerate(summary_data['sorting_records'], 1):
        row_data = [
            idx,  # 序号
            record[1],  # sorting_time
            record[2],  # pieces
            record[4],  # time_slot
            record[3],  # remark
            record[5]   # created_at
        ]
        ws_sorting.append(row_data)
        
        # 设置边框和对齐
        for col in range(1, len(row_data) + 1):
            cell = ws_sorting.cell(idx + 1, col)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 调整列宽
    for col in range(1, len(sorting_headers) + 1):
        ws_sorting.column_dimensions[get_column_letter(col)].width = 15
    
    # 创建统计汇总工作表
    ws_stats = wb.create_sheet("统计汇总")
    
    # 添加统计数据
    ws_stats.append(['统计项目', '数值'])
    ws_stats.cell(1, 1).fill = header_fill
    ws_stats.cell(1, 1).font = header_font
    ws_stats.cell(1, 2).fill = header_fill
    ws_stats.cell(1, 2).font = header_font
    
    ws_stats.append(['入库总件数(实到)', stats['total_inbound_pieces']])
    ws_stats.append(['入库车辆数', stats['total_inbound_vehicles']])
    ws_stats.append(['分拣总件数', stats['total_sorting_pieces']])
    ws_stats.append(['分拣记录数', stats['total_sorting_records']])
    ws_stats.append([])
    ws_stats.append(['车辆类型统计', ''])
    ws_stats.cell(ws_stats.max_row, 1).font = Font(bold=True)
    
    for vehicle_type, data in stats['vehicle_stats'].items():
        ws_stats.append([f'{vehicle_type}', f"数量: {data['count']}, 实到件数: {data['pieces']}"])
    
    # 调整列宽
    ws_stats.column_dimensions['A'].width = 20
    ws_stats.column_dimensions['B'].width = 30
    
    # 保存文件
    filename = f"daily_report_{summary_data['date'].strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    wb.save(filepath)
    
    return filepath

def compress_database():
    """
    压缩数据库文件
    返回压缩文件路径，如果失败返回None
    """
    try:
        # 检查数据库文件是否存在
        if not os.path.exists(DB_PATH):
            print(f"警告: 数据库文件不存在: {DB_PATH}")
            return None
        
        # 生成压缩文件名（带日期）
        today = datetime.now(LA_TZ).strftime('%Y%m%d')
        zip_filename = f"inbound_backup_{today}.zip"
        zip_filepath = os.path.join(os.path.dirname(DB_PATH), zip_filename)
        
        # 如果压缩文件已存在，先删除
        if os.path.exists(zip_filepath):
            os.remove(zip_filepath)
        
        # 创建ZIP压缩文件
        print(f"正在压缩数据库文件: {DB_PATH}")
        with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(DB_PATH, os.path.basename(DB_PATH))
        
        # 显示压缩信息
        original_size = os.path.getsize(DB_PATH) / (1024 * 1024)
        compressed_size = os.path.getsize(zip_filepath) / (1024 * 1024)
        compression_ratio = (1 - compressed_size / original_size) * 100 if original_size > 0 else 0
        
        print(f"数据库压缩完成: {compressed_size:.2f} MB (压缩率: {compression_ratio:.1f}%)")
        
        return zip_filepath
        
    except Exception as e:
        print(f"压缩数据库时出错: {str(e)}")
        return None

def send_email(subject, body, attachment_paths=None):
    """
    发送邮件（支持多个附件）
    attachment_paths: 可以是单个文件路径（字符串）或文件路径列表
    """
    try:
        print(f"\n[调试] 开始准备邮件...")
        print(f"[调试] 主题: {subject}")
        
        # 创建邮件对象
        msg = MIMEMultipart()
        msg['From'] = EMAIL_CONFIG['sender_email']
        msg['To'] = EMAIL_CONFIG['recipient_email']
        msg['Subject'] = subject
        print(f"[调试] 发件人: {EMAIL_CONFIG['sender_email']}")
        print(f"[调试] 收件人: {EMAIL_CONFIG['recipient_email']}")
        
        # 添加邮件正文
        msg.attach(MIMEText(body, 'html', 'utf-8'))
        print(f"[调试] 已添加邮件正文")
        
        # 处理附件（支持单个或多个）
        if attachment_paths:
            # 如果是字符串，转换为列表
            if isinstance(attachment_paths, str):
                attachment_paths = [attachment_paths]
            
            print(f"[调试] 准备添加 {len(attachment_paths)} 个附件")
            # 添加所有附件
            for attachment_path in attachment_paths:
                if os.path.exists(attachment_path):
                    file_size = os.path.getsize(attachment_path) / 1024
                    print(f"[调试] 正在添加附件: {os.path.basename(attachment_path)} ({file_size:.2f} KB)")
                    with open(attachment_path, 'rb') as f:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(f.read())
                        encoders.encode_base64(part)
                        part.add_header(
                            'Content-Disposition',
                            f'attachment; filename={os.path.basename(attachment_path)}'
                        )
                        msg.attach(part)
                    print(f"[调试] ✓ 已添加附件: {os.path.basename(attachment_path)}")
                else:
                    print(f"[调试] ✗ 附件不存在: {attachment_path}")
        
        # 连接SMTP服务器并发送邮件
        print(f"\n[调试] 开始连接SMTP服务器...")
        print(f"[调试] 服务器: {EMAIL_CONFIG['smtp_server']}")
        print(f"[调试] 端口: {EMAIL_CONFIG['smtp_port']}")
        
        with smtplib.SMTP(EMAIL_CONFIG['smtp_server'], EMAIL_CONFIG['smtp_port'], timeout=30) as server:
            print(f"[调试] ✓ 已连接到SMTP服务器")
            
            print(f"[调试] 正在启动TLS加密...")
            server.starttls()  # 启用TLS加密
            print(f"[调试] ✓ TLS加密已启动")
            
            print(f"[调试] 正在登录邮箱...")
            server.login(EMAIL_CONFIG['sender_email'], EMAIL_CONFIG['sender_password'])
            print(f"[调试] ✓ 登录成功")
            
            print(f"[调试] 正在发送邮件...")
            server.send_message(msg)
            print(f"[调试] ✓ 邮件已发送")
        
        print(f"\n✅ 邮件发送成功: {subject}\n")
        return True
    
    except Exception as e:
        print(f"\n❌ 邮件发送失败: {str(e)}")
        import traceback
        print("\n[调试] 详细错误信息:")
        traceback.print_exc()
        return False

def generate_email_body(summary_data, stats):
    """
    生成邮件正文
    """
    date_str = summary_data['date'].strftime('%Y年%m月%d日')
    
    # 车辆统计表格
    vehicle_stats_html = ""
    for vehicle_type, data in stats['vehicle_stats'].items():
        vehicle_stats_html += f"""
        <tr>
            <td style="padding: 8px; border: 1px solid #ddd;">{vehicle_type}</td>
            <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{data['count']}</td>
            <td style="padding: 8px; border: 1px solid #ddd; text-align: center;">{data['pieces']}</td>
        </tr>
        """
    
    html_body = f"""
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
            .container {{ max-width: 800px; margin: 0 auto; padding: 20px; }}
            h1 {{ color: #4472C4; border-bottom: 3px solid #4472C4; padding-bottom: 10px; }}
            h2 {{ color: #5B9BD5; margin-top: 30px; }}
            table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            th {{ background-color: #4472C4; color: white; padding: 12px; text-align: left; }}
            td {{ padding: 8px; border: 1px solid #ddd; }}
            .summary-box {{ background-color: #f0f4f8; padding: 15px; border-radius: 5px; margin: 20px 0; }}
            .highlight {{ color: #4472C4; font-weight: bold; font-size: 1.2em; }}
        </style>
    </head>
    <body>
        <div class="container">
            <h1>📊 每日数据汇总报告</h1>
            <p><strong>日期：</strong>{date_str}</p>
            
            <div class="summary-box">
                <h2>📈 总体统计</h2>
                <ul>
                    <li>入库总件数(实到): <span class="highlight">{stats['total_inbound_pieces']:,}</span> 件</li>
                    <li>入库车辆数: <span class="highlight">{stats['total_inbound_vehicles']}</span> 辆</li>
                    <li>分拣总件数: <span class="highlight">{stats['total_sorting_pieces']:,}</span> 件</li>
                    <li>分拣记录数: <span class="highlight">{stats['total_sorting_records']}</span> 条</li>
                </ul>
            </div>
            
            <h2>🚚 车辆类型统计</h2>
            <table>
                <thead>
                    <tr>
                        <th>车辆类型</th>
                        <th style="text-align: center;">车辆数量</th>
                        <th style="text-align: center;">实到件数</th>
                    </tr>
                </thead>
                <tbody>
                    {vehicle_stats_html}
                </tbody>
            </table>
            
            <p style="margin-top: 30px; color: #666; font-size: 0.9em;">
                详细数据请查看附件Excel文件。<br>
                此邮件由系统自动发送，请勿回复。
            </p>
        </div>
    </body>
    </html>
    """
    
    return html_body

def send_daily_report():
    """
    发送每日报告
    """
    print(f"开始生成每日报告... {datetime.now(LA_TZ)}")
    
    try:
        # 获取数据汇总
        summary_data = get_daily_summary()
        
        # 计算统计数据
        stats = calculate_statistics(summary_data)
        
        # 创建Excel报表
        excel_path = create_excel_report(summary_data, stats)
        
        # 压缩数据库文件
        print("\n正在准备数据库备份...")
        db_backup_path = compress_database()
        
        # 准备附件列表
        attachments = [excel_path]
        if db_backup_path:
            attachments.append(db_backup_path)
            print("数据库备份将作为附件一同发送")
        else:
            print("数据库备份失败，仅发送Excel报表")
        
        # 生成邮件正文
        date_str = summary_data['date'].strftime('%Y年%m月%d日')
        subject = f"{EMAIL_SUBJECT_PREFIX} - {date_str}"
        body = generate_email_body(summary_data, stats)
        
        # 发送邮件（包含Excel报表和数据库备份）
        success = send_email(subject, body, attachments)
        
        if success:
            print(f"每日报告发送成功: {date_str}")
            
            # 根据配置决定是否删除临时文件
            if DELETE_TEMP_FILE:
                try:
                    os.remove(excel_path)
                    print(f"已删除临时Excel文件: {excel_path}")
                    if db_backup_path:
                        os.remove(db_backup_path)
                        print(f"已删除临时备份文件: {db_backup_path}")
                except Exception as e:
                    print(f"删除临时文件失败: {str(e)}")
            else:
                print(f"Excel报表已保留: {excel_path}")
                if db_backup_path:
                    print(f"数据库备份已保留: {db_backup_path}")
        else:
            print(f"每日报告发送失败: {date_str}")
        
    except Exception as e:
        print(f"生成或发送每日报告时出错: {str(e)}")
        import traceback
        traceback.print_exc()

def schedule_daily_report():
    """
    设置定时任务
    每天在配置的时间发送前一天的数据汇总
    """
    # 设置每天在指定时间执行
    schedule.every().day.at(REPORT_TIME).do(send_daily_report)
    
    print("每日报告定时任务已启动")
    print(f"将在每天 {REPORT_TIME} 发送前一天的数据汇总邮件")
    print(f"收件人: {EMAIL_CONFIG['recipient_email']}")
    print("\n按 Ctrl+C 停止程序")
    
    # 持续运行
    while True:
        schedule.run_pending()
        time.sleep(60)  # 每分钟检查一次

if __name__ == "__main__":
    # 如果需要立即测试发送，取消下面这行的注释
    # send_daily_report()
    
    # 启动定时任务
    schedule_daily_report()
